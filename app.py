"""
╔══════════════════════════════════════════════════════════════╗
║  ATENDE.AI v2.0 — Sistema de Atendente IA para WhatsApp    ║
║  SaaS completo com painel admin + mídia + Mercado Pago      ║
╚══════════════════════════════════════════════════════════════╝

Requisitos:
  pip install flask mercadopago requests openai

Configuração (variáveis de ambiente):
  SECRET_KEY=sua_chave_secreta
  MERCADOPAGO_ACCESS_TOKEN=seu_token_mp
  WHATSAPP_VERIFY_TOKEN=seu_token_verificacao
  ANTHROPIC_API_KEY=sua_chave_anthropic
  OPENAI_API_KEY=sua_chave_openai (fallback para transcrição de áudio)
  GROQ_API_KEY=sua_chave_groq (transcrição de áudio — mais barato)
  BASE_URL=https://seudominio.com
  ADMIN_EMAIL=admin@atende.ai
  ADMIN_PASSWORD=admin123

Rodar:
  python app.py
"""

import os, json, sqlite3, hashlib, secrets, time, re, base64, tempfile, io, random
from datetime import datetime, timedelta, timezone
from functools import wraps
from flask import (
    Flask, request, jsonify, redirect, url_for,
    session, g, make_response, abort, send_file
)
from markupsafe import escape as html_escape


def esc(text):
    """Escape HTML para prevenir XSS. Retorna string vazia se None."""
    if text is None:
        return ""
    return str(html_escape(str(text)))


# Fuso horário de Brasília (UTC-3)
BRAZIL_TZ = timezone(timedelta(hours=-3))

def to_br_time(utc_str, fmt="%H:%M"):
    """Converte string de UTC (formato SQLite 'YYYY-MM-DD HH:MM:SS') para horário de Brasília"""
    if not utc_str:
        return ""
    try:
        # SQLite datetime('now') retorna 'YYYY-MM-DD HH:MM:SS' em UTC
        if "T" in utc_str:
            # ISO format
            dt = datetime.fromisoformat(utc_str.replace("Z", "+00:00"))
        else:
            dt = datetime.strptime(utc_str[:19], "%Y-%m-%d %H:%M:%S")
        # Marca como UTC e converte para Brasília
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        br_dt = dt.astimezone(BRAZIL_TZ)
        return br_dt.strftime(fmt)
    except Exception:
        return utc_str[:16] if len(utc_str) >= 16 else utc_str


def to_br_date(utc_str):
    """Converte UTC para data DD/MM/YYYY em horário de Brasília"""
    return to_br_time(utc_str, "%d/%m/%Y")


def to_br_datetime(utc_str):
    """Converte UTC para data + hora no formato brasileiro"""
    return to_br_time(utc_str, "%d/%m/%Y %H:%M")


def csv_safe(value):
    """Previne CSV Injection - escapa valores que começam com =, +, -, @, tab, CR"""
    if value is None:
        return ""
    s = str(value)
    # Escape dos caracteres perigosos
    if s and s[0] in ('=', '+', '-', '@', '\t', '\r'):
        s = "'" + s
    # Escape de aspas e quebras
    if '"' in s or ',' in s or '\n' in s or '\r' in s:
        s = '"' + s.replace('"', '""') + '"'
    return s

# ─── CONFIG ────────────────────────────────────────────────────
app = Flask(__name__)

# LGPD/Segurança: SECRET_KEY deve ser definida em produção.
# Sem ela, sessões não persistem entre restarts E podem ser forjadas se a chave aleatória
# vazar em logs. Em produção (Railway), exigir env var explícita.
_secret_key_env = os.getenv("SECRET_KEY", "").strip()
_flask_env = os.getenv("FLASK_ENV", "production").lower()
if _secret_key_env:
    if len(_secret_key_env) < 32:
        print(f"⚠️  AVISO: SECRET_KEY tem apenas {len(_secret_key_env)} caracteres. Recomendado: 64+ caracteres.")
    app.secret_key = _secret_key_env
elif _flask_env == "development":
    # Dev local: fallback aleatório com aviso explícito
    app.secret_key = secrets.token_hex(32)
    print("⚠️  DEV: SECRET_KEY não definida, usando fallback aleatório. Sessões serão perdidas a cada restart.")
else:
    # PRODUÇÃO sem SECRET_KEY: falha alto antes de iniciar
    print("=" * 60)
    print("❌ ERRO CRÍTICO: SECRET_KEY não está definida em produção!")
    print("   Defina SECRET_KEY no Railway → Variables com pelo menos 32 caracteres.")
    print("   Gere uma com: python -c \"import secrets; print(secrets.token_hex(32))\"")
    print("=" * 60)
    raise RuntimeError(
        "SECRET_KEY environment variable is required in production. "
        "Set FLASK_ENV=development to bypass this check (local dev only)."
    )

app.config['SESSION_COOKIE_SECURE'] = _flask_env != "development"
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=12)

MERCADOPAGO_ACCESS_TOKEN = os.getenv("MERCADOPAGO_ACCESS_TOKEN", "TEST-xxxx")
WHATSAPP_VERIFY_TOKEN = os.getenv("WHATSAPP_VERIFY_TOKEN", secrets.token_hex(16))
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
BASE_URL = os.getenv("BASE_URL", "https://atendente.online")
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "admin@atende.ai")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")  # Obrigatório configurar
DATABASE = os.getenv("DATABASE_PATH", "/app/data/atendeia.db" if os.path.isdir("/app/data") else "atendeia.db")
MEDIA_FOLDER = os.getenv("MEDIA_PATH", "/app/data/media_files" if os.path.isdir("/app/data") else "media_files")

os.makedirs(MEDIA_FOLDER, exist_ok=True)
# Garante que o diretório do banco existe (para volumes persistentes)
db_dir = os.path.dirname(DATABASE)
if db_dir:
    os.makedirs(db_dir, exist_ok=True)
print(f"[STORAGE] Banco de dados: {DATABASE}")
print(f"[STORAGE] Mídia: {MEDIA_FOLDER}")

# ─── LOGGING ──────────────────────────────────────────────────
DEBUG_MODE = os.getenv("FLASK_ENV", "").lower() == "development" or os.getenv("DEBUG", "").lower() in ("1", "true", "yes")

def log_debug(msg):
    """Log só em modo debug/dev"""
    if DEBUG_MODE:
        print(msg)

# ─── FORCE HTTPS ───────────────────────────────────────────────
@app.before_request
def force_https():
    if request.headers.get('X-Forwarded-Proto', 'http') == 'http' and not request.host.startswith('localhost') and not request.host.startswith('127.'):
        url = request.url.replace('http://', 'https://', 1)
        return redirect(url, code=301)

# ─── SECURITY ──────────────────────────────────────────────────
login_attempts = {}  # {ip: {"count": n, "last": timestamp}}

# ─── RATE LIMITING ─────────────────────────────────────────────
# Usa Redis se disponível (suporta múltiplos workers), senão fallback para memória
_redis_client = None
login_attempts = {}  # fallback em memória

def _get_redis():
    """Retorna cliente Redis se REDIS_URL configurado, senão None"""
    global _redis_client
    if _redis_client is not None:
        return _redis_client if _redis_client != "disabled" else None
    redis_url = os.getenv("REDIS_URL", "")
    if not redis_url:
        _redis_client = "disabled"
        return None
    try:
        import redis
        _redis_client = redis.Redis.from_url(redis_url, decode_responses=True, socket_timeout=2, socket_connect_timeout=2)
        _redis_client.ping()
        safe_log("[REDIS] Conectado com sucesso — rate limit distribuído ativo")
        return _redis_client
    except Exception as e:
        safe_log(f"[REDIS] Falha ao conectar ({e}) — usando rate limit em memória", level="ERROR")
        _redis_client = "disabled"
        return None


@app.before_request
def generate_nonce():
    """Gera nonce CSP único por request para scripts inline"""
    g.csp_nonce = base64.b64encode(secrets.token_bytes(16)).decode('ascii')


@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    # Content Security Policy
    # Handlers inline foram migrados para addEventListener, então script-src
    # pode depender apenas de nonce sem 'unsafe-inline'.
    nonce = getattr(g, 'csp_nonce', '')
    csp = (
        "default-src 'self'; "
        f"script-src 'self' 'nonce-{nonce}'; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com data:; "
        "img-src 'self' data: https:; "
        "connect-src 'self'; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'; "
        "object-src 'none'"
    )
    response.headers['Content-Security-Policy'] = csp
    if request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

def check_rate_limit(ip, max_attempts=5, window=300):
    """Bloqueia login após 5 tentativas em 5 minutos. Usa Redis se disponível."""
    r = _get_redis()
    if r:
        try:
            key = f"ratelimit:login:{ip}"
            count = r.get(key)
            if count and int(count) >= max_attempts:
                return False
            return True
        except Exception as e:
            safe_log(f"[REDIS] Erro no check: {e}", level="ERROR")
            # Fallback para memória
    # Fallback: memória local
    now = time.time()
    if ip in login_attempts:
        data = login_attempts[ip]
        if now - data["last"] > window:
            login_attempts[ip] = {"count": 0, "last": now}
            return True
        if data["count"] >= max_attempts:
            return False
    return True


def record_login_attempt(ip, window=300):
    r = _get_redis()
    if r:
        try:
            key = f"ratelimit:login:{ip}"
            pipe = r.pipeline()
            pipe.incr(key)
            pipe.expire(key, window)
            pipe.execute()
            return
        except Exception as e:
            safe_log(f"[REDIS] Erro no record: {e}", level="ERROR")
    # Fallback memória
    now = time.time()
    if ip not in login_attempts:
        login_attempts[ip] = {"count": 0, "last": now}
    login_attempts[ip]["count"] += 1
    login_attempts[ip]["last"] = now


def reset_login_attempts(ip):
    r = _get_redis()
    if r:
        try:
            r.delete(f"ratelimit:login:{ip}")
            return
        except Exception:
            pass
    if ip in login_attempts:
        del login_attempts[ip]

# ─── LOGO (base64 inline) ─────────────────────────────────────
LOGO_NAV_B64 = "iVBORw0KGgoAAAANSUhEUgAAALsAAAA4CAIAAAAq1h5LAAABCGlDQ1BJQ0MgUHJvZmlsZQAAeJxjYGA8wQAELAYMDLl5JUVB7k4KEZFRCuwPGBiBEAwSk4sLGHADoKpv1yBqL+viUYcLcKakFicD6Q9ArFIEtBxopAiQLZIOYWuA2EkQtg2IXV5SUAJkB4DYRSFBzkB2CpCtkY7ETkJiJxcUgdT3ANk2uTmlyQh3M/Ck5oUGA2kOIJZhKGYIYnBncAL5H6IkfxEDg8VXBgbmCQixpJkMDNtbGRgkbiHEVBYwMPC3MDBsO48QQ4RJQWJRIliIBYiZ0tIYGD4tZ2DgjWRgEL7AwMAVDQsIHG5TALvNnSEfCNMZchhSgSKeDHkMyQx6QJYRgwGDIYMZAKbWPz9HbOBQAAAsbElEQVR42u29d7xtVXU2/Iwx51q773165fZKuZeOoCBFiiIoWAE/SSwosUbfJIoauxI1mBgN1kQlxBJRIRZUREQEhUu73N7L6W33ssqcc3x/7HMLWF4NmrySO/845+z9W3utveZ81hjjGeMZ8yg8iUHMJNJz9Mnnve6v4gjl0V1EhCPjyPgNeCEi4kTy6X//3Rtn5J+3jCe6h9pvH5mbp/DQT+bDAvGIx2eCOzfW0gbQ6TaUADkys0fGb/BKgL/8aQve/g3/pEsIwBED81Qf6sl9nAFhoRR0+d6vEjHkiHV5ig9+cpEMAMoOHlPo6wWrIwbmSBzzKxBp/yAScQDgLIDMwlVeJgtnQUTEAAkcBEeimSOIIYEQhCGcTJPOWGh4OWsTsU0mOhcjLppWVSwEIFIARNwR3DzVgtffy4EpP+n8Duv1cbLLT2cB8rOdLnbOReJ5kbEIqiqYoPpoXJsRQIjRtkZHxv8axBAAgqhE0ug+eF0Llqw8+ZS1xxy7bPiooWQmo7USIWtdUC1PTkxu2Tm2Ydv42Nioa82hsq1VHBOAiOUIbv53IIYAUUxesjNA/4qTnv6855239vSTegYGE14KxAIoJl+zUsQEzRCR6tzE+nUbb73trvse2e5TyYz8IgwDIiVij0z3U59da+3p3EIvv+r8yy9bevKZz7vi+XBKREGMT6LI+ko0OXLWWRtFcWwN+bnjTlo5sHyFJ65a51bHctWaiIP6ESb1lEcMaU0qszjbvexNb3vD2Ze96P6HdqRTavFQpwIlNCvFmlkrVoqYySfnaXisNLtqtfXT+/dc/Yrnn3/mMTs27Zy1R6nW/jhogvhIIPxURQwRQyX6Mn2r3vqet/WuOLpZr/f19oyOTx27elkul+jIJbNZL5/1chmdy6hcRqVSXjLtJZMqldRbdk/nO7ILC2kvl7vw/NO2rd+4cxpePOXi+EgN4akYxxBBRGsf+ePe8K7rlq5ZW602Ojs6envySW3CSnlmem6uWKpXgkq1aRy00jqhM5lkZzbd01MYGujo6enqHuoSg7lSw7FnW9VXXvOxPRvvlel7DTHckSj4KZGPofn8HEOs1rCqcM6zzl107HFzs9VFw722WbvrGz/65br1u/cXK4GFE5CCsxABaTDADGiVSKeSauFg7vijF1549glPP+vUekwm1/X611/8N3+5XRUW2/JeYk/EHmFPTwGvRGBFYr3sYHr5BWENz33xZSrf1VvIbnnol5+4/p/uvuOBydlmSKQ9xQqKrWKn2CkyimMmYe0IURSF01OzGx/df9vt96174JHjj13c0dvtkLj91p/Yo07Xiawr7hCmI67pTx8xxEosMksKV98Qzs5yZfppl1zW29e58ed33Xj9jY1YdFcHp1KSyDhdID9pSDlmIRZA2LPaFy9FflYlPPY80rGCGdlTvOe+R57xtOMWLxn6xX0P7Z0OF/zZG1q7RqW00x2hTn+agw/aFwVHfjb73Hfmzzu5tWN938JF3f395emJr/zbraRJp1LWK5jcIiw7LXHKM23/ak51w0uKkIAdGKThZ2y2K+49xgyeZLuWu0ReJwSdffet2+Al1NHHLMXY/r4e0/3S65A9BiJE+kmWQo+M/znEgEUgnSd2X3CO1xiXUtgxOJDvTD328KbmXIWSvvMynBvInfC0f//oVQ988kXPuPxMyfZxMg/tCTOxkPaRymaWrv3Yu17yjx9+SXbN8cgP2USqNLpvrNyYniiuXL0cJPUt2xY+fYVddYWnFESO8KY/WcSIQJGXXxoVvMb2UZBkOjqN5ZmpMkwAUuSnrN950umLrzy6e23Oe9W5SyTbxawAhrUAlFJCHZdevPqvzhp680m9Vz57jUt2+ql0s1Jbv2Hf1Gyxt6+b0zK9da/OOh48RmUXCizRERvzp2pjBESe9sMQtX0TgFWe12iG1hhyFiAhsAu3bJ57tNqKQT98dJbqZYRNxAFcDOPEhIDZMxbU4QCZrgkIEAFjamy6Vm0UCrlMJlkan5mdruhcRlL97aseWYM/VXYNgalN8XjJTY4CsXK6Xq33DQ2JEBOkVYc/Nf3Q+kve0ezt7l7/6GZdH3WNOqImBALjwtALJu+//d7nxI3h4e7v/mATt0omDNghLs+lMxkn7JGrVau1yaqT0JoIAMH9D7slovmk1J8K2ycQzT+MIgCBiUTkv038qA/NnHNRMOFt3ou5CsSauFEsVwaWLF645tj9Dz+EbgUHBM2xH+4ZiyziahxXgZg8RRByZBuNeHYaqYmf79qGdIp8kfokGjVMT5773DM7enrHRybJObQqjR37aKakwrL7PSdKAEAIJH+Y6IcAIZH22Q6e/79qquXQk/dHxYsc+MZtuAs5cQCISNoI+iOHhodsjAV0fbvs2WRrVaBVn51mh0a9cdmrX3X/9/u3P7y+FZRImp1dHUtW9y9dtnLpisVDQ4P5zkJCkRNbb0RzU1OPPrx18+ad+8b2lGcC7Vod+ez5L33ZpVe9eHquNjU5UyvNQdhtH1X1Utwcp9+9zEQsrEgE4gR/CMAQsUDY48Iw54fQqpvpDSCi3//sRARWEAdxkD/ychG8hOoZTHX162rFjm5pcJJPPLlr9+5GceK/SSBwEDHCUDaoUPVh7ScAWxob85QNgqAZZC555av8ayK2cT6VGBzo6u3pTqXTqSwpRjMEHNJJuBjVeuvyqy6tFOuTUzNzpQYzZbNZTmf3js5YeFvXb4qKM6qr32uVosm7jG3f4e82w+JgHQAGC8g9+XUhTyRSS89WV3/F5Xr1vZ+Lb3ktmMX+3mcWEdiYDlotMPBHcXBKkzVy9RsW/vX7e3Mpet+7xr/wd83P33LMC56b27wTl5/38PRISEzi/rh2Th9+5wK0tt2R7FmjEunpkV0ubLKXDGr18ZHxwcHejo4eL+03Yi+YaXRk4x999luV2aljjl+Ty+Z27tl367fuPfGME/767a+NJNk9uMAvxKVyo1ivlybHmi3TaET3/OhOwCpwPPNwNHo/QAJ7WGFrvneFQA5MsCTiQEQMcWr4BF77Yupcjkf/I9p8C1jB2fbh0u6oEycQIkD4oGGWeSYm7gCNbwcAIAVSAtKDa6N0LxnYiU3zHyARKAILLIkIgUAQwbwoTEAMgESkHVCI8zoX6BOucH2rafOt4YbvHJCP8cFenN/y3BPxIR87r4ymQ1MiAmJiiBORdlmFTnwG+6l6hMTGB+vJnHrOc7NAbfnybL7Lnx4NmNmJiICZASdtrzt/7xCAmAgkAvmvpjYOQwwcgLhVlcmHU+mORrW17Zd3r33uy7dveJQTqlQsmygKculsOpVMeARZdNyJ17/j+s9/5j/h55ErnHfBaa9945WjI/u2bdzdik3vQF+ho2N8qlSvNiPn9u2bCGOAYmUmG3v3GiuHzVE7EvAglsUKRMGByLKGM0zkRNSaPzMXvZkM6NGvSnvBiJmUuJgFDBjyCRawNC9KhyiPbCwAz4sAZd67EZSzsC0BXO9KiLCN7NS2A1GwUs46WAaEPXKGiASOxTqQsGJnHUDEBAgzrOUTXmKe9xHEwNYfMsSRAhwRwdl5MBzA7RP8i2Kytj0BDEBp52x7mTXgAFFKW2vEEoD2lVWCFy5Nhs5FgZseD5vl6K1v2P2Cy/pv+86enY9VRdgaB7BSZK1tm+Q2ugGwInHs5j2naEXGPjnEHLwVE7fq1ZZS3j233rhg2VELlz99bO9ua4xx1jrnrESx3wrCrv7BG7748a2bdyhWy1cuWrRw8JZv3Pn1W38WLVzGNmrc+6lLLj7vlIsviEqVfXsnW41m0mefa3Ft3MbBr4leXcRgZIcokSYXSWUCLiZiB82s3eDRrhHrxixNbWNOgFhsS8RxogOpTrJG10YsQKxJp9jLwIUuKEmiwNkBqY1LVGvDBWK1AIXFksjy3E50LRYmalRQ3AUi2BgAZYfYz3J92kRlgAREfp4TadcqkolUfhg6geJegQMpxQkMnhS3jNeYpPENIM0MZxwE5GWQ7VMulMq4/RV+JgJrBZ7qHlBMVJwSG8YMOOJkjvxUwoS2WYnyvf7wUGJ8MqhMxQAGFiQHjtKWotlpjO8TP+XfctPst786FxRJaZXMk+d5tVJsI9e/MJHJ8e5tAQwIAJO1AqB7WCVSVJyyQc38l/nCb+AlQgQh5ouufsviUy8pFusmqOZzuY7OXD6XTiYSnqZkKlHIFTyPReyOvVPv/j8f4NIWP8E2qMf1SS+38qLXvKk4Or7ll3dXpna6qKq5YUyMx8nFiYjByjvjdbTmxVI4yvkZZercKtlffs784tP67LfRqX8ep7rhpcSEXnVU1yfDf72UupboC96BBWeYdDeJ0SPrwm9cS0zeK7/ncoPe5tvM1E6c/mqbKuig7L75GrPrboLwwqfThe+R/jWSyKnprbaw0Oa79ex2+cTpLqqqNS9Sp7/a9a112tdBkR7+cnjHh7xjL6fLPy7aVz/6qO1d5o57EZi8iQeim6/mJWfTJR+OE3nx8ogjVd7j2Vb05ctArM97hyw/F5kBigOubjPffacZeaStkyeCCCuP/7+3DF/80vzgQhZIdSb+6udrX/mH0UTe+8IPVh+3Sn37ttbotuY1b+zJdsWNmr7+nZO3fWHyzOf0fOp7gx7MfT/GNRdued37l1z7pm5F9JZX7y5WcONNS4XNDddPrz0+dcHzk8kUP/oL9+aX7qzPhsbh3Mu7X/nG3uVrfe3FjaK++XPFL3xkjKGcCH6fePnXdp/IvL8ghsjtX7phyYN3rz33igUrjq/UG6NjM76iVNJPJhOJhOcnZvxEIuGpWqXVk9az+/cEtXbop3qXLE14Xr24q1V8jE0J4qyRJ9pnIohVmQF10YdjneLZEWWrJjdg08P+pZ/knfeqoRPivtWuNkexsEq64ePdhgkaPk2/8humY1BaQo0Zk+x0J17izV4nj/y7GVxrokBWvYROyzsDG1Sle4U6/wO06yxecAZdc3vs59gIBdV48BREdXHMM7viqOpf+D57wd8akKtXgSjMLfKf+0G9817qXhz2LpZyxbvg/ZzJmyhCUJdjn8cnXEGpbvQfK6UZmIhJY/A0mdrAheX84s/EQydz1KTmpPE6ZNH5iUsT9nMXi42IiAh+Wn3waysuuUQJTLPl64Tp7HYf+njv3HT48E8qa08j4eb5L1TD+d4YQTkwhb74De/tve2L0wOLmMjGcBvX1wAcfRInChWH1P794ann5Ib744lG468+XOhMejU0ak2cfY53xRu7P/O3Y69654K3fbDPR6sUxWKof1H4tx/u2bU1uOvbs0opa/EkEXMw8HLECmL3bHxwdmJq9akXDy1cmO0d9JKF0Jh6MRImVqw9rUHJXPKCq16y5f5eW57xVdPLDfm5jrv+9bri9AQRZD5+fGKw1RY9iIvprr+jrd+3M1vJGf+Cj0WnvcJpRUuebu54t54Zc6dfC+bEg/8Sb7rVRk39ws+E3qDev52/9So7tVm/+seW1qCw0PWtkdhQKwRi+epfMFq4+B9Mq+7rpNU5vugDIbKqWlb3fdw9dos69Rr7tNeyI7f/QbX4bDnnHaYR+jt+hNuu5b6j3ZVfi7nL7zvW5I5Cy7IDlfe7r/6Nv+Iic8orXLWl+ta6e29g59HZbwa09/BNduMtJgzVhe813cfp2d347lvtrjv9F/9LtOQcl+rz8gNhcR8pdtZddd2Cp19CY83m/d9X//LeXSedk33z33W30pVzXlAYHw1CMo2GS3rqPX8zOjEq1/1zbyiBYc2+Glrpt8TEjjY9HALUMaRmTVyrRCN7ohe/NjHlmqGWvdtbN7xp5JJXdF74Mm/KxdluLDspe+0H8iVbeuQe+uBrRrv70//43f5CtnHcaam7vo35PNSTQQyBhHg+D+rsUStOOulZl62755F1P/gOQH422z801Nl/VLKjM5PPk5fRvkdaW+vY2s6+5Q1OlMa2T255oDYzefpF54mt3P/j7xEE3ObSjwsASQREUp+OfnEjFp/NKy5FOmN7ViBuiUtxoxTN7uBwSoQoNmbrj+I9d/rHv8zkVnK9RPd/wY4/7C290FIS0Dy+BdmeOBbSPt3zj9Ejn1GpPnXWu1EYjie3q4Hjbd8pHMTe5lvNXR+0AO+9D6e8hloBxrfRyVdHEfu1Wbn7BiLNw2fYEKwC1KZ49cW2GSkb4dt/Ge+/0xOmE17hlENUjed267AoTrMJ7dbvxHt/6i16Zjz4NAQ1fuxb8Y4f6OETbaJbDFN10jZmmclZ6RhOn/nS9MRsoz6Vfv81u8JysHdH85JrMwuPZaOjvmWJqjGckK9/oX7zx6aTXYnpuY6uYTcyXpfY9Cz15upREPL2Tc10p9JdqIRuZE/YKtu+JSjGBpH+0OunNt9T6xjyznpJV+RHkxPheVcWqrbWrOvPfGRubkpOOEeXo1ArvXNL+Jsjk98n8hVSDOMEzGr1aRe96p0fT2YKK9dufvjuldvWPzq+f2Rk+46R7Y8BHpABEuAEUln4KTgHB5DxdLx0yZILX/fqS654aTLf9aVPfvzfPv5uE7WYtXs8qRMiAvmnvc6d/peSW0gQaRQth9KMFDfd7HYiQvcaCUMdzrnqfmIlA6e4KHJBiU76C3Xam226x1nn7XvArPusOvedErUorMnGbxIr7lkhyqPYYnob9awSC7JNt/m7jhQgqu84EwPBJJrT0rmaGqU4tuoF/0I66xI5Ms574PPRyC/pvPdJFGLyETN2D7GSjkUuDigSO7WdiXTfGhNF3JqWyhhY0eJzyDgXVOLlF+kV57nMAhHj13e7n7zHRg32PDLxwmMTznPV2Hvg5/WwHCvF2W4dkio1g7Ep0zFElchFLbnrtpJSWLDCj3wqNbFjiyVFqQFM1119Rk/sihYek27p0DRl46ORTiAzoIqtaP8Ws+OhBit0LtDl0JrQjU7Ys052UyUdNe3r/qknmUS2ix3i738tvOtbZSJ2zh1g3r9Tylo/zrgwwTkS43Shf9kJlenqqlOenevunB3d2zs0eNmrrg3qpbBcnJ4YKc9O18rNRivwtAh88bxkLtfX3dGdzxx11MDQUYP9Q4PJXHZ8csYLas+56tof/eCB1vhDxendTHSokkQs4pLHXRk/870urHkbvuY23OzGH1LP/Xy08AxdGnHlXYqUSw1SFKC4X0rbyVmClqiho6Yae9CZwAUlmnjY7LiNQJJfgSDwSntMZVSc5cKi2Hkc1ml2ixSWIoo5btnauLQJ68DJEodeZdoFMyzaRLFX2S+Tj4lYqe3n/fcGY7/0Bp4m8DiKaGI9bASAOxa6yLKpYHqTiEimj8KmK4248h5y1nlZiVpkGmpkHUxAtk5TD9udd7hWRUA2sgA4ScVAtKGpqRiw1mLBmjTyZnrO27sr6BnwS814dkyN77DWUmFAVWxEddq8Icp0KpvkUsvt3RrGDdu7TBcDcMhbHmvlenTD46DO2zZFpgkBdSxSc00b1fTUXtNyfrHuivuxd1NL+9HcmGz/md3wkxJAwDwLIRzM/dF8JEsHc5KPQ9JhOl+lYI1OZt2SS84699LLnnfCe//iuv17xz22Uave0dltJSp0deYXHrXmtFPB7Hl+IYl0grXHuh2pEEUC5+AiG0dx1Aw1EAT1Zg0t1XvlW29oTD38pX/4gFLKtameCAFu2SWmWtYukHs/auY26QXnSn4BNRuuMoeo7tKDmjNo1ZwRTg/Y6j5XnaLQsIPs/qnb/GUA1LFSLXu2m94gOo0olvKIi6sESG6hxBHbSIq7kT3KBQ1IqPueZqcfo+Ne7rqORrPiamNS3IX6uHj9ZCHrbnTFzQLwgjO9wRMl2+WEKGxgZmc7n0OpPgRNhCVX2c+JvPUKCGoqrKtE3poA1RGOAgjT1IPmwc8RIOkOf9lZrU0/PvH5hdOuKjx22+z+7XGxbl0Ydy3LJDq0l/BPf3nnyFwkLr3hzsYFr+ucLmF0f1SdigHJL/DnqkIejWyJ8wNexca67CZnWCc4PZSYKQqY9m9t5vr1TDOmBvZuiwRO+crv1FMzcWPOH9sczRRtuirlhrrtE3Nz2y0gC473Tnh29/oflrSPnuFkGLtWPYpjskYQiziCEMlhWVW4drgjchAxzGJN8thzzSnvWDs89NJn+bFTS1cvHds/3moErTCILGUSOuVrj0kDXQUVN8tj+xsTxSiIhBNeT4e/vDc91Jerxs5Z+D4zcTLp1ZtRFAeV6ameTv9Nb3l/aa54203/rJjtAVNDJmYTSxzJ2dd7cZP71kRi0CyisJhPeJXb8HUAEoTOL+gX3IoHPopNN3srXxKxz6e/lVc+n2wsw6dg6lH3iw8L+4hDN7OTnAUgmSGJYwrLtjqmZzdpQ3HUoLVXqmXn295jXKMkqQ4q7zGmSVvv1KevDf1O/fwv89wOKE+WPBN3fYCk5QyxrUlprwAq2SHJfmdDVR9FY078tHbORTWT6fde/HV+9GZsvVWteWkkSTnxtWrZeQiLqm8lrE1M33vGG3oKXebMN/Z94WUj2+53Q8f66Ile+KnFWhvuQTnKPvLF4tTWgDu8yWI8PUpR0wJIDvjFKkQwubOVyOp6rOpTrvtYOunlfZFnS3VIzMV9ZvC4ZLEhEsrETgcgWSCTwFxdVSZtdTx+5Pu1/HHdko5f9ume0pjohFq4lvb81D36g6IAi1ZkrnxdNt8jxQmzf084OW6mJ6LZyag45RoladYkCgB7sODqNIGgWKxNP/3P1VnvsmOly0+luVLdiF619phH1t0yPV3MpPKxcfCtiOf5qjuH7333F9/8xege0xn4BagEiJHm3jyeNaRef8Fwd1dXM7BgxwRxUi5XxciqFcPFYviad3zwwXvuGNuznZmdQETshn9L5BaJzprCcooa7qF/TvSsiofPAeeIO5ypqs236GNeCpBTOU50xdW9ctdbvRNfL5mjpPNoccLVott3p0r2IWyIidXclhAgTrCXUmGFKvvINN3+e3jHV/SCC4hSLpHjdZ9Uw2davZLmdgjgNnxBJZL+0udarwM9J8CFGH8MxY16+GRpzLKt2dooAEpmNMHFJV3aZiSmsEKbvqrXXCWApLo4mYqqI/jxO5KnXBuneiQ3QPkhF9fV2AOuZXY/ZoZOThUfaVb3N3/ykfEz3jycXeTrbiNhsjUmO+8sP3LTTMfSVCPwoyLP7gtFhBVJRs8VVVxVjTlTHo333ieZhb4NFJKxr/1mMxmVbX1W0l2JcslHIJVRAyBV8CLjByWa29MCZOOtdV3wV16YS2RdatiZUO/dRNt+WSUCOVr341K5Xn7O1flnXZx63gu7kkiakCpVFOfU1GRjZF+4d19tfCLYtiEa32sbM0TEmp1RZ7wmdfHfVh7b05dtfeDqZePTNXHWxuaG69754tdcs2zJsJftymcok84M96Y+/olv/PChAMuO4Y4OTiZ0KhkTJOUhk3L15pJw5vMvX9bf2d2wErRatXrt9p9u+sZnb/r29z4dxy6Zy3/x05/83PverBRb6wjMcJTup47lIOXK211jkvwO3XOckHVzOxAWRRx3rqL0MOKalHdKXHViSWW5YxknO8S0XH1CGiPkZ5EskPgSzFjT9EAu1UdKi7PSnIJYAqnu4yTZ40pbXXOCswNQPoKKRFUQwTnKDlJ2IStPoppU9rioSuluUgkQpDHjrLBmTnVbUhw3XKskRCLwOhZSrk/iEHO7nGlZcdrz0bWIdVLEolV21Qlrxe9I9q7OzW2qBTUDWGbuXJFKdCiJqbo/aMwGTMwJle7RIhQ3TVQCOEp2+8ojZ9AqGjFQSepckhSS1oQlHyolElNjMkx2ap1mEm7OmTiwXkKlezSUMy00Zo0SGGezPYnCIt/3KWi46ljYmrPahwV0ykFUIukGV9CzXpS89MrMUf06Jek+WpzDKiAA6kBlwpS+c/vsB/56lAhQi89OvOmbZt2DYUsX0u76VywZnSkmtF9Ie5/4m7f3rj75Gec/I9/ZnUzpwYGO9T9b9w83PeQffXykUvAS6OhEHCGdVJ05SulcZ7o0WrkgW/v4n60tNTiKavUw/Oxnb9Wt4mc//85de2uxk72j06+/9PRWda6t6XhCjZ5Yw5lDQg/SAETMIR0KM6Dh4kMqEYDJc2IJB+q2bZ2UuHYyUYgAZmlrOtoFuUNXASmABcJiDm53QwAxi3MHCAS3k42H0wQH4Xm9ynxWAu16kzg6kKYkgMBCB49iBsAEJ+5QooEUk5O20mL+Sykih0P1Qlbteuv8Xk8EEjBgCUSHZGn0eI2O/KoeqB27elmkO1hrFPqkf6HuG1Jb10ltOvazvOp0vuya1AmrU0P+UIaGZquN3fvL69dNb93VnJqx936rrpk9dfF1Yp0pVzjbWSnHzZZdcFTf3q27vvChG0fHa9I5O1tusa5FJlXIt37+4BbO5OJa/W2vO26s6L5+19hHrz3hw7eNnLRAXv2c4Wv/c9pL82MTtZGZWiKViiPbbASNcrU403z3R265+NnP7Mgl+/v7lh930ob77mBmZ9s9cgqQduVYnGvXCyEMuPlFItVeewcHEUIEEIgPQMNZGIBBql2nhDgBmNjOa50EsG5+RdunNY7UfHVLBIgFJIdExyQQOAewEAs5CEEcQQk7SLsiLQRxQu0vP/+OtK9CIEXzWUs4chBipjYvma9OKtHUXnURgQgTQOQAFmEHsfNYoXad0tk2vNxhWfk2+kUE8AAFTorW8Hx4KUqnoNOUzFMmw9m8yhWQykkmr9J5SqWRzXIqpax1XpqVTxBxzt79LZtn2vqz+MtVafwlTGXP3q1bdm1o7d6IVkiXXt6//bvV2qxoHljLa09zxWmBDwCe9x93jN347nNv/dLXRjZO6VXHlupmdq6STPkJY5uNFCvlmiXWXnWqMlsF10oDOsy5+o6tjR91+o3putRaXr2B2NYRu8hUSq1qrbVxd2nj+P13ff+ev//IG/NHLVm68vgN991B87B3Bx+JQ0RO5mvpB54Z6x5XxZifqwPHzmvC5HEqurbI4fGfE5GDdgj28QKo9vFPuMq8TurAOxaHqU8OXM0erv88IAK188hxBEfzsHSHlA0yb4QOGi13SM2XsJSA9uD5SGSQSMFPwfdUOsmpjE6k2U9RIsXpDAo5L59XqZxLZl02q/ykSqSZ2bJnE77zGJGjEDY0EtYQxag3bK0kzRDT++K5SVMvu0bR1YpSL4k0FFtpNR0pbszJD74S77i/xZar0zqVwnUfWvX9/5javb7JTFr6VphcXtsKKbLQ7MsDe/FPX1v/lrdetWPHvt2bZuuDfbVSpd6dFSvT5eZppx637oFvOsinP3U7MgVOZ6/8q++jswDP2z1S9wa6zM4dp6zWKumXKgHZeK5Sn6vGNLw8s3Dprh1brn//jTf860eXrFxxYM7/WPqjP6zyDYf8zoHUo8yvPJ6gb8Hjcdv+wwM86DQyWcrmWCdUJq9yed3V5acyKl/Q+U7q7E5m0l4h73f1+OmMl0iyl5BMgskTKBAJk3WII2OCEMY062FQqiGKXKtqaiU3OhNW5kytJNWqbVYlbCIMpVF3YQtRIEEDiMXGsIYEDEfsCGKdtO00A9AeSLjRkMXHZucm4nhONZqu0KE+8Mnj7/zPfbf/+yyRds5qSmWF2Q33+j1+0FRQShX0TT9pAeWbb37flz73rc/d9JOtW5O9Qz1J7RWn5xYMD7z2Vc+++Yu3hl5egsjZmpfIuFrktC9KxTsePXaBvuryF81Vgzi2zpqp2XK9XEc2H9Sqif7+B9ZtvP7d/3TyGWtYec4Z0P98Oy0dXuI6WCA9ZEfooFDwcEmnHNK5EJLipSjb42cKkimg0O1191Fnpy5kdfeA192b7MjrRJKzqUwm46czLNZ4vtZKsaNmZCMTmziqN9BsulbLTs4266OtVjMO6qbciJo106y7WsOEQdxquqCBMJSghTiADWAMrIWNITGLFTiQEJwQEYGsJcVEJM4SCTkn4iBw4pyxcA5iLRyJiJAkM9wUO3SiOvkCfdN7G8Upl+/jt/zdont+OnLTJ6ZZKWcNQOStOg/XfdselU/tGwl+sFWy/bAReUnblFMXubdftTAZTtz0lR/vK3NHd7azUEhnUwsXDU7t2HTj+z6BTA9nMg4e4Om039+bP/+M5Rc+71lGZaJWy4FFzN0/Xf/Nm79jeocp2ys25sndMrN/8YrFk5tvbVQm8D+9Ud5vF1KTIvaUSimd9nQSibxOF3QmT5kuynboXIcq9CCb4nSaUzl0drLPzmOjPYoiF0UujCRuShhSvRGXSiYK0GrG1WocthBFLghd0HBhEyaEjRGHsCGcgTFgAjlIDAG4Hb4JkeN2Vk1cO55vWzQSiIiIhWuXYNoIEGmHfPOuW55QAjjsJYFAyidKyfAq77LrUk74839ez/TIVe/of+jHxXu/2WLDzs27Tu32P6w33WeGnx2sHko2muHPp22hF+RUltdNqRf+/ciVJ6f+/DUvT1D9/sd2bdw6PjnRKleaCeeOPX716c95Vnd3AU757PyU33fUUKrQVao2xVUFcIIoinaNTEatpqoWbamMevGUU1dc/bE3/uwn99266Xt/CPvyayfid9P0E4iZlGrHqaRZJRX7HntKJZVKQKWUTnoqBZ2A8p32wB4xiXI2hAQlNzUduS0GRkzoTChR6KKWMzGMsSYSZ8VZB0doN5waVgyCwBFcm3qxECkhEVgRdiwOTgQO1om4tlaTpC0PEcA5cXAkcIdLXd2v3C8dUsHOPxIH60Z0KE48KGtv/4eJtCw42T/zirRJ6W13RkiaC67pv/Mbc5t+FDD48AiPmJVa8zy8/d9Mb0IXKPXI3vDOkUjnkUkziai0axlEteccl3rRMzpXLkC5Un1s28SWvZPJRIG0pz2vK5fSHnnaExuLdaxYMdpcpVgOvnXLHXvWb0Szduppy59/xWW9CxY0LIEzn3rry3Y/djczu/Z2MoTDiO18qPCE2/tdgDM/T+D5n6xAIGIwERGYiJgJwiQiokAeOwMB2n9DQRRRktgj8qA0FJFKiEoQE0hBSGDBjsRZMc7FggBirFjnAgshWCB2YkWMiIWIiGv7gfZLiDsQsku7TCK/+QH4NW/SE+bmdzKgv82SEoE1D5yYOOY5vmVHrMYfMMmUNzda339frJjs46XmWojtjh+rm97lvfLvDaN+0oL0QEZ9b1swVbbJBHtNP5WL012373K3b5kd7sBFx6ROXb3sqtWrFNVL1fpcNSzWgkolagRhbBREhKFYw4E0JmcqY/vGELveBQsuu+pFfctWFWfngigmLV4iCQgObvgrT7QN9MQSWFswS21OQgSQImIiEsVgBphZt802QRyxa4ehbt5yz2dDPA3fh09IapVNqo50Mqf8zmSiM5HISybrp7LaT4mmkMKmhK04imGtaYVxLQqqrlmMo2psGpFtWdOytuFcKDa2LhaEIobaZJkOUvTf7gAfr02hAzzqiQK3X33xOxnn36mFqp3ViKuy+Q6T6pRkhuIGj61rVSctM1vnniCHIBArreFl1dl/Jle8xw7mJed8Y/QvdpoHRmyx6TyibEFlOyWTj8lH0yEKc9n46F5/1RCvHPAW9OhcRvk6Ymui2NQbYTOMmnFsjNu2e3r9vY/6hZ7hY45fuyyjFKyhauBKldatH3tTaWIHeykbB4C0TbSIzIMB5NS8eRDy6CApFQExiInICbW7lw5wLg+s4SloBe1ToUNnMiqdTORyiXxadeb9rg6kEl7e0wWffKWTlnUzkWiiVY5bVdNoxJWqNJpRtRGU6qbUjMrNuBra0MaNCE2LwIkRWBEn5ETAcI6cgZAISA54C6EDYD9g8uQJIvj/B8kggUQxvAx7Gd0sGhtIO4F5AOjyOFgTFLRHGrzyWerZ/8ecdJbr1KzhzVR524jduN+MjIsYSqU43aHy3ZLqCFUKzoOxiGKQTXq2M02DuXio08+nVTalfc/kUqyZSSlW5HuiCFFkA+PKTbtvbG5095Zqaa46NxNWZ11Yd3HdxqHEkbgQxrazpjKfrm3n1ghM0D4pD57PqYz2EzqV9HL5RKZDdXcnerrTnV3Jzjylk0h56UzGsopdRC5sVopRpWRa9WZxOpwdM5UpUyubWsXWyzZsoB4iChFbxA7OwAkcQQzEKYEQkbWwThwEBPff2LD639ub+zjr9pv9GB2Wp2bxhPML1epny6mXuaWnSKaTIVyvY2yS9ux1U/tseQ5RTBrsZznTzZksUlnrZy2nLSsYIGZIDHEwQGxAFnKA/zkDF0IiREXEZSaDuApTFxOBHJMiUlCKFMFLkZ9RfgJ+ghIJ8lOcTLGvla/YT/pJTyeZPNZJ0iTKiRc24zgIoygKanG1HtXqcVAPKrOmUZZWFUETQQ1xE3EAG0NCuLidiCMSESZxBDffxyyQdoOGyOOz7v9LNnx8InJ+Wy8Bt/8JhVKkfc72yoKTadk50n8qcsOSLcCSqpapOInZcSqOm9qMhHWYEM6RAL7PXgpelv00JVPQvuiEsE/KBykohrPijLiYnSCui2kIHCkQESkFJqU90lqIAXfAGZGYQEwIsRIFErfENCRqShy6sCVhVUyEOEDcgothY9gYCOFiwIIBGIZFm5mIE9fuqhRxFuLgDqWMn2By5fC5w68e8pSGC/7vbIN+jU9TSrQPYtIJzg5Q1wruXuU6V0p6iJKd4udZJREFaNXRqKBRRKNogyoFFYQtZx25CHACBpi0BjFYgTWYwEQ2EhfCxSIWZOAsxMCF4hxcDBvCRWgvqjWQEGJACrCAhbh5/TG39ycREgsYOAdnRGJYA2dETJuekDhpp+GPbGv0h0+AP5GU0byPIF+0g0qSnyU/T4mCZAY52etS3Uj3c7IHfl68DGsP1sIKjHEmho1hQsQhbORsCyaACSSqwNThBKaGuEYwB1xVBImISJyDbQEWIgQnNhJnIA5E84c5CxDEQiyknb+0EIvH14cOJ57/dyN7ZDx5xBxW3wfIEXwQgwmkQSyKiAjQYJ+UgkqIyrBOOS8DP0s6A+WTSkFpms9ExrAGJnJxFXGNwBLXETfgDJyFNbCRSExtpu3iA/uitNHQZn9C4h6foz88tOCDfcXtTmU6jKL+bmLnI+PJ9kT+6jF84De3e1PazW9gApiEhQyROlDsJbCaP0AOtrG1k9ttIyFtIwEiwMGZg4JSkXY/ynxfyoF1t0eW/v+d8f8D4kuzZ7CXFOkAAAAASUVORK5CYII="

# Planos
PLANS = {
    "starter":  {
        "name": "Starter",
        "price": 97.00,
        "msgs": 500,
        "desc": "Para quem está começando",
        "tagline": "Automatize seu WhatsApp com IA",
        "features": [
            "Atendimento automático com IA (Claude)",
            "500 mensagens/mês",
            "Entende texto, áudio, imagens e PDFs",
            "Responde por áudio (text-to-speech)",
            "Base de conhecimento treinável",
            "Dashboard com conversas em tempo real",
            "Galeria de produtos (sem venda)",
            "Suporte por email"
        ],
        "highlight": False,
        "cta": "Começar grátis"
    },
    "pro": {
        "name": "Profissional",
        "price": 197.00,
        "msgs": 2000,
        "desc": "Para negócios em crescimento",
        "tagline": "Atendimento + Agência Digital",
        "features": [
            "Tudo do Starter",
            "2.000 mensagens/mês",
            "📸 Agência Digital com IA",
            "Posts automáticos para redes sociais",
            "Aprovação via Telegram Bot",
            "Biblioteca de mídia ilimitada",
            "Agenda semanal de publicações",
            "Respostas rápidas",
            "Exportação de conversas",
            "Suporte prioritário"
        ],
        "highlight": True,
        "cta": "Começar teste grátis"
    },
    "business": {
        "name": "Business",
        "price": 397.00,
        "msgs": 10000,
        "desc": "Para empresas que vendem",
        "tagline": "Vendas + Campanhas + Agência",
        "features": [
            "Tudo do Profissional",
            "10.000 mensagens/mês",
            "🛒 Comércio direto no WhatsApp",
            "PIX automático via Mercado Pago",
            "Link de cartão/boleto no chat",
            "Detecção de intenção de compra",
            "📢 Campanhas em massa (broadcast)",
            "CRM com funil de vendas",
            "Gestão de contatos e tags",
            "Analytics avançado",
            "Multicanal (Instagram + Messenger)*",
            "Suporte WhatsApp"
        ],
        "highlight": False,
        "cta": "Começar teste grátis"
    },
    "agency": {
        "name": "Agência",
        "price": 997.00,
        "msgs": 50000,
        "desc": "Para revender como serviço",
        "tagline": "Multi-contas + tudo ilimitado",
        "features": [
            "Tudo do Business",
            "50.000 mensagens/mês",
            "Até 10 contas de cliente",
            "Sub-usuários atendentes",
            "Campanhas ilimitadas",
            "Posts automáticos ilimitados",
            "Modelos de templates para clientes",
            "Dashboard multi-empresa",
            "API dedicada",
            "Relatórios white-label",
            "Suporte telefônico",
            "Onboarding personalizado"
        ],
        "highlight": False,
        "cta": "Falar com vendas"
    }
}
# *Multicanal disponível após aprovação Meta (em andamento)

# ─── DATABASE ──────────────────────────────────────────────────
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db: db.close()

def init_db():
    db = sqlite3.connect(DATABASE)
    db.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        name TEXT NOT NULL,
        company TEXT DEFAULT '',
        phone TEXT DEFAULT '',
        plan TEXT DEFAULT 'starter',
        plan_status TEXT DEFAULT 'trial',
        mp_subscription_id TEXT DEFAULT '',
        msgs_used INTEGER DEFAULT 0,
        msgs_limit INTEGER DEFAULT 500,
        trial_ends_at TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        whatsapp_phone_id TEXT DEFAULT '',
        whatsapp_token TEXT DEFAULT '',
        ai_system_prompt TEXT DEFAULT 'Você é um atendente virtual simpático e prestativo. Responda de forma clara e objetiva.',
        ai_tone TEXT DEFAULT 'profissional',
        ai_greeting TEXT DEFAULT 'Olá! 👋 Como posso ajudar você hoje?',
        business_hours TEXT DEFAULT '08:00-18:00',
        auto_reply_off_hours TEXT DEFAULT 'Nosso horário de atendimento é de 08h às 18h. Deixe sua mensagem!',
        is_active INTEGER DEFAULT 1,
        last_login TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS knowledge_base (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        title TEXT NOT NULL,
        content TEXT NOT NULL,
        category TEXT DEFAULT 'geral',
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS conversations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        customer_phone TEXT NOT NULL,
        customer_name TEXT DEFAULT '',
        status TEXT DEFAULT 'active',
        is_human_takeover INTEGER DEFAULT 0,
        satisfaction_rating INTEGER DEFAULT 0,
        tags TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        last_message_at TEXT DEFAULT (datetime('now')),
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        conversation_id INTEGER NOT NULL,
        sender TEXT NOT NULL,
        content TEXT NOT NULL,
        msg_type TEXT DEFAULT 'text',
        media_url TEXT DEFAULT '',
        external_message_id TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (conversation_id) REFERENCES conversations(id)
    );
    CREATE TABLE IF NOT EXISTS payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        mp_payment_id TEXT DEFAULT '',
        amount REAL NOT NULL,
        status TEXT DEFAULT 'pending',
        plan TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS quick_replies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        shortcut TEXT NOT NULL,
        content TEXT NOT NULL,
        category TEXT DEFAULT 'geral',
        times_used INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS blocked_contacts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        phone TEXT NOT NULL,
        reason TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS api_usage_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        api_name TEXT NOT NULL,
        tokens_in INTEGER DEFAULT 0,
        tokens_out INTEGER DEFAULT 0,
        cost_estimate REAL DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS admin_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        action TEXT NOT NULL,
        details TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS system_settings (
        key TEXT PRIMARY KEY,
        value TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS verification_codes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT NOT NULL,
        code TEXT NOT NULL,
        expires_at TEXT NOT NULL,
        used INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS product_gallery (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        keywords TEXT DEFAULT '',
        description TEXT DEFAULT '',
        file_path TEXT NOT NULL,
        file_type TEXT DEFAULT 'image/jpeg',
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        conversation_id INTEGER,
        customer_phone TEXT,
        customer_name TEXT,
        items TEXT DEFAULT '',
        total REAL DEFAULT 0,
        payment_method TEXT DEFAULT 'pix',
        payment_status TEXT DEFAULT 'pending',
        mp_payment_id TEXT DEFAULT '',
        mp_qr_code TEXT DEFAULT '',
        mp_copy_paste TEXT DEFAULT '',
        mp_checkout_url TEXT DEFAULT '',
        paid_at TEXT DEFAULT '',
        expires_at TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS campaigns (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        message TEXT DEFAULT '',
        template_id INTEGER,
        total_contacts INTEGER DEFAULT 0,
        sent_count INTEGER DEFAULT 0,
        delivered_count INTEGER DEFAULT 0,
        failed_count INTEGER DEFAULT 0,
        status TEXT DEFAULT 'draft',
        scheduled_for TEXT DEFAULT '',
        started_at TEXT DEFAULT '',
        completed_at TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS campaign_contacts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        campaign_id INTEGER NOT NULL,
        phone TEXT NOT NULL,
        name TEXT DEFAULT '',
        status TEXT DEFAULT 'pending',
        error TEXT DEFAULT '',
        sent_at TEXT DEFAULT '',
        variables TEXT DEFAULT '',
        FOREIGN KEY (campaign_id) REFERENCES campaigns(id)
    );
    CREATE TABLE IF NOT EXISTS whatsapp_templates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        content TEXT NOT NULL,
        variables_count INTEGER DEFAULT 0,
        category TEXT DEFAULT 'marketing',
        status TEXT DEFAULT 'pending',
        meta_template_id TEXT DEFAULT '',
        times_used INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS contacts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        phone TEXT NOT NULL,
        name TEXT DEFAULT '',
        email TEXT DEFAULT '',
        tags TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        lifecycle_stage TEXT DEFAULT 'lead',
        last_contact_at TEXT DEFAULT '',
        total_orders INTEGER DEFAULT 0,
        total_spent REAL DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS pipeline_stages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        color TEXT DEFAULT '#6366f1',
        position INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS pipeline_cards (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        stage_id INTEGER NOT NULL,
        conversation_id INTEGER,
        contact_id INTEGER,
        title TEXT NOT NULL,
        value REAL DEFAULT 0,
        notes TEXT DEFAULT '',
        position INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now')),
        moved_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS sub_users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_id INTEGER NOT NULL,
        email TEXT NOT NULL,
        password_hash TEXT NOT NULL,
        name TEXT NOT NULL,
        role TEXT DEFAULT 'agent',
        is_active INTEGER DEFAULT 1,
        last_login TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (owner_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS webhook_errors (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        source TEXT NOT NULL,
        error_type TEXT DEFAULT '',
        error_message TEXT DEFAULT '',
        payload_preview TEXT DEFAULT '',
        resolved INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS processed_webhook_events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source TEXT NOT NULL,
        event_key TEXT NOT NULL,
        user_id INTEGER,
        payload_preview TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        UNIQUE(source, event_key)
    );
    CREATE TABLE IF NOT EXISTS admin_audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        action TEXT NOT NULL,
        target_type TEXT DEFAULT '',
        target_id TEXT DEFAULT '',
        ip_address TEXT DEFAULT '',
        user_agent TEXT DEFAULT '',
        details TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS social_media_library (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        file_path TEXT NOT NULL,
        file_type TEXT DEFAULT 'image/jpeg',
        media_type TEXT DEFAULT 'photo',
        theme TEXT DEFAULT 'geral',
        description TEXT DEFAULT '',
        times_used INTEGER DEFAULT 0,
        last_used_at TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS scheduled_posts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        media_id INTEGER,
        caption TEXT DEFAULT '',
        hashtags TEXT DEFAULT '',
        status TEXT DEFAULT 'pending',
        scheduled_for TEXT DEFAULT '',
        approved_at TEXT DEFAULT '',
        posted_at TEXT DEFAULT '',
        platforms TEXT DEFAULT 'manual',
        telegram_message_id TEXT DEFAULT '',
        rejection_reason TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id),
        FOREIGN KEY (media_id) REFERENCES social_media_library(id)
    );
    CREATE TABLE IF NOT EXISTS consent_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        email TEXT DEFAULT '',
        consent_type TEXT NOT NULL,
        consent_version TEXT NOT NULL,
        accepted INTEGER DEFAULT 1,
        ip_address TEXT DEFAULT '',
        user_agent TEXT DEFAULT '',
        details TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    """)
    # ── MIGRAÇÃO AUTOMÁTICA ──
    migrations = [
        ("users", "is_active", "INTEGER DEFAULT 1"),
        ("users", "last_login", "TEXT DEFAULT ''"),
        ("users", "email_verified", "INTEGER DEFAULT 0"),
        ("users", "instagram_page_id", "TEXT DEFAULT ''"),
        ("users", "instagram_token", "TEXT DEFAULT ''"),
        ("users", "messenger_page_id", "TEXT DEFAULT ''"),
        ("users", "messenger_token", "TEXT DEFAULT ''"),
        ("users", "telegram_bot_token", "TEXT DEFAULT ''"),
        ("users", "telegram_chat_id", "TEXT DEFAULT ''"),
        ("users", "social_post_time", "TEXT DEFAULT '09:00'"),
        ("users", "social_auto_enabled", "INTEGER DEFAULT 0"),
        ("users", "social_post_tone", "TEXT DEFAULT 'profissional'"),
        ("users", "social_business_context", "TEXT DEFAULT ''"),
        ("users", "social_post_days", "TEXT DEFAULT '1,2,3,4,5'"),
        ("users", "social_post_times", "TEXT DEFAULT '09:00'"),
        ("users", "social_last_run", "TEXT DEFAULT ''"),
        ("users", "mp_access_token", "TEXT DEFAULT ''"),
        ("users", "mp_public_key", "TEXT DEFAULT ''"),
        ("users", "commerce_enabled", "INTEGER DEFAULT 0"),
        ("users", "auto_payment_enabled", "INTEGER DEFAULT 1"),
        ("product_gallery", "price", "REAL DEFAULT 0"),
        ("product_gallery", "stock", "INTEGER DEFAULT -1"),
        ("product_gallery", "sku", "TEXT DEFAULT ''"),
        ("product_gallery", "category", "TEXT DEFAULT ''"),
        ("product_gallery", "active", "INTEGER DEFAULT 1"),
        ("conversations", "satisfaction_rating", "INTEGER DEFAULT 0"),
        ("conversations", "tags", "TEXT DEFAULT ''"),
        ("conversations", "notes", "TEXT DEFAULT ''"),
        ("conversations", "channel", "TEXT DEFAULT 'whatsapp'"),
        ("messages", "media_url", "TEXT DEFAULT ''"),
        ("messages", "external_message_id", "TEXT DEFAULT ''"),
        ("verification_codes", "code_type", "TEXT DEFAULT 'signup'"),
        ("users", "deleted_at", "TEXT DEFAULT ''"),
        ("users", "totp_secret", "TEXT DEFAULT ''"),
        ("users", "totp_enabled", "INTEGER DEFAULT 0"),
        ("users", "totp_backup_codes", "TEXT DEFAULT ''"),
        ("users", "mp_webhook_secret", "TEXT DEFAULT ''"),
    ]
    for table, column, col_type in migrations:
        try:
            db.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}")
        except sqlite3.OperationalError:
            pass
    try:
        db.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_processed_webhook_events_unique ON processed_webhook_events(source, event_key)")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("CREATE INDEX IF NOT EXISTS idx_users_whatsapp_phone_id ON users(whatsapp_phone_id)")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("CREATE INDEX IF NOT EXISTS idx_messages_external_message_id ON messages(external_message_id)")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("CREATE INDEX IF NOT EXISTS idx_consent_log_user ON consent_log(user_id, consent_type, created_at)")
    except sqlite3.OperationalError:
        pass
    db.commit()
    db.close()


def _build_fernet_from_key(key_string):
    """Constrói instância Fernet a partir de uma string-chave qualquer.
    Usa SHA-256 da string como base, depois base64 urlsafe (formato Fernet).
    Retorna None se a chave estiver vazia ou cryptography não estiver instalada.

    IMPORTANTE: esta função é determinística — mesma string-chave produz mesma chave Fernet.
    Útil pra migração: podemos criar Fernet "antigo" com SECRET_KEY e Fernet "novo" com
    DATA_ENCRYPTION_KEY no mesmo runtime."""
    if not key_string:
        return None
    try:
        from cryptography.fernet import Fernet
        import hashlib as _h
        key_bytes = _h.sha256(str(key_string).encode()).digest()
        fernet_key = base64.urlsafe_b64encode(key_bytes)
        return Fernet(fernet_key)
    except ImportError:
        return None
    except Exception:
        return None


def _get_fernet():
    """Retorna instância Fernet ATIVA para criptografia de dados em repouso.

    Prioridade da chave:
    1. DATA_ENCRYPTION_KEY (env var) — RECOMENDADO em produção (separada da SECRET_KEY)
    2. SECRET_KEY (env var) — fallback para compatibilidade com dados criptografados antes
       da introdução de DATA_ENCRYPTION_KEY. NÃO troque SECRET_KEY sem migrar antes.

    ⚠️ ATENÇÃO: ativar DATA_ENCRYPTION_KEY enquanto há dados criptografados com SECRET_KEY
    causa falhas silenciosas de descriptografia. SEMPRE rode migrate_recrypt_to_new_data_key()
    antes (chamada automaticamente no startup se necessário).

    Cacheia a instância. Para forçar reload, reinicie o app."""
    if not hasattr(_get_fernet, "_instance"):
        # Tenta a chave dedicada primeiro
        data_key = os.getenv("DATA_ENCRYPTION_KEY", "").strip()
        secret = os.getenv("SECRET_KEY", "").strip() or (
            app.secret_key if isinstance(app.secret_key, str) else (app.secret_key or b"").decode('utf-8', errors='ignore')
        )

        if data_key:
            base = data_key
            _get_fernet._source = "DATA_ENCRYPTION_KEY"
        elif secret:
            base = secret
            _get_fernet._source = "SECRET_KEY (legacy fallback)"
        else:
            base = "fallback-key-please-set-DATA_ENCRYPTION_KEY"
            _get_fernet._source = "FALLBACK (INSEGURO!)"

        _get_fernet._instance = _build_fernet_from_key(base)
        if _get_fernet._instance is None:
            _get_fernet._source = "cryptography não instalada"
    return _get_fernet._instance


def _assert_crypto_available():
    """Em produção, criptografia é OBRIGATÓRIA. Se cryptography não está instalada, falha."""
    is_dev = os.getenv("FLASK_ENV", "").lower() == "development"
    if _get_fernet() is None and not is_dev:
        raise RuntimeError(
            "[CRITICAL] biblioteca 'cryptography' não instalada em produção. "
            "Segredos não podem ser salvos em texto puro. "
            "Instale com: pip install cryptography>=42.0 "
            "ou defina FLASK_ENV=development para testes locais."
        )


def _encrypt_value(plaintext):
    """Criptografa valor usando Fernet (AES-128-CBC + HMAC-SHA256). Falha em produção se cryptography não instalada.
    Idempotente: se já estiver criptografado (fer:v1:), retorna o mesmo valor."""
    if not plaintext:
        return ""
    # IDEMPOTÊNCIA: não re-criptografar valor já criptografado
    if isinstance(plaintext, str) and plaintext.startswith("fer:v1:"):
        return plaintext
    # Também evita re-encrypt de formato antigo (será migrado automaticamente)
    if isinstance(plaintext, str) and plaintext.startswith("enc:v1:"):
        return plaintext
    fernet = _get_fernet()
    if not fernet:
        _assert_crypto_available()
        # Se chegou aqui, está em dev explícito — permite fallback com aviso
        safe_log("[CRYPTO] ⚠️ AVISO DEV: salvando em texto puro (cryptography não instalada)", level="WARN")
        return plaintext
    try:
        token = fernet.encrypt(plaintext.encode('utf-8'))
        return f"fer:v1:{token.decode('ascii')}"
    except Exception as e:
        safe_log(f"[CRYPTO] Erro ao criptografar: {e}", level="ERROR")
        _assert_crypto_available()
        return plaintext


def _decrypt_value(encrypted):
    """Descriptografa valor Fernet. Aceita formato antigo e novo."""
    if not encrypted or not isinstance(encrypted, str):
        return encrypted
    # Formato novo (Fernet)
    if encrypted.startswith("fer:v1:"):
        fernet = _get_fernet()
        if not fernet:
            return ""
        try:
            return fernet.decrypt(encrypted[7:].encode('ascii')).decode('utf-8')
        except Exception as e:
            safe_log(f"[CRYPTO] Erro ao descriptografar Fernet: {e}", level="ERROR")
            return ""
    # Formato antigo (XOR custom — compatibilidade com dados antigos)
    if encrypted.startswith("enc:v1:"):
        return _decrypt_legacy(encrypted)
    # Texto puro (ainda não criptografado)
    return encrypted


# Campos de usuário que contêm tokens/segredos e devem ser criptografados
USER_ENCRYPTED_FIELDS = {
    "whatsapp_token",
    "mp_access_token",
    "instagram_token",
    "messenger_token",
    "telegram_bot_token",
    "totp_secret",
    "mp_webhook_secret",
}


def generate_totp_secret():
    """Gera novo segredo TOTP (base32)"""
    try:
        import pyotp
        return pyotp.random_base32()
    except ImportError:
        safe_log("[2FA] pyotp não instalado")
        return None


def generate_totp_uri(secret, issuer="atendente.online", account="admin"):
    """Gera URI otpauth para QR Code do Google Authenticator/Authy"""
    try:
        import pyotp
        totp = pyotp.TOTP(secret)
        return totp.provisioning_uri(name=account, issuer_name=issuer)
    except ImportError:
        return None


def generate_totp_qr_base64(uri):
    """Gera QR Code em base64 (data URI) para embed em HTML"""
    try:
        import qrcode
        import io as io_mod
        import base64 as b64_mod
        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=8, border=2)
        qr.add_data(uri)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io_mod.BytesIO()
        img.save(buf, format='PNG')
        return "data:image/png;base64," + b64_mod.b64encode(buf.getvalue()).decode('ascii')
    except Exception as e:
        safe_log(f"[2FA] Erro gerando QR: {e}", level="ERROR")
        return None


def verify_totp_code(secret, code):
    """Valida código TOTP (6 dígitos). Aceita ±30s de clock skew."""
    if not secret or not code:
        return False
    try:
        import pyotp
        code = str(code).replace(" ", "").strip()
        if not code.isdigit() or len(code) != 6:
            return False
        totp = pyotp.TOTP(secret)
        return totp.verify(code, valid_window=1)
    except Exception as e:
        safe_log(f"[2FA] Erro validando TOTP: {e}", level="ERROR")
        return False


def generate_backup_codes(n=8):
    """Gera códigos de backup para recuperação se perder celular."""
    import secrets as secrets_mod
    codes = []
    for _ in range(n):
        # Formato: XXXX-XXXX (8 chars alfanuméricos)
        code = secrets_mod.token_hex(4).upper()
        codes.append(f"{code[:4]}-{code[4:]}")
    return codes


def user_2fa_enabled(user):
    """True se o usuário tem TOTP habilitado (opt-in).
    Aceita dict ou sqlite3.Row. Tolerante a colunas ausentes (pré-migração)."""
    if not user:
        return False
    try:
        u = dict(user)
        return bool(u.get("totp_enabled", 0)) and bool(u.get("totp_secret", ""))
    except Exception:
        return False


def is_admin_2fa_enabled():
    """Verifica se admin configurou 2FA.
    MECANISMO DE EMERGÊNCIA: se env var ADMIN_2FA_DISABLE_EMERGENCY=1, desabilita 2FA temporariamente.
    Use APENAS se perder celular e backup codes. Depois, configure de novo."""
    if os.getenv("ADMIN_2FA_DISABLE_EMERGENCY", "").strip() == "1":
        safe_log("[2FA] ⚠️ EMERGÊNCIA: 2FA bypassed via ADMIN_2FA_DISABLE_EMERGENCY", level="WARN")
        return False
    return bool(get_setting("ADMIN_TOTP_SECRET", ""))


def log_admin_action(action, target_type="", target_id="", details=""):
    """Registra ação admin para auditoria"""
    try:
        ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()[:45]
        ua = (request.headers.get('User-Agent') or '')[:200]
        db_conn = sqlite3.connect(DATABASE)
        db_conn.execute(
            """INSERT INTO admin_audit_log (action, target_type, target_id, ip_address, user_agent, details)
               VALUES (?,?,?,?,?,?)""",
            (action[:100], str(target_type)[:50], str(target_id)[:50], ip, ua, str(details)[:500])
        )
        db_conn.commit()
        db_conn.close()
    except Exception as e:
        safe_log(f"[AUDIT] Falha ao registrar: {e}", level="ERROR")


def _safe_payload_preview(payload, limit=500):
    """Gera preview reduzido e menos sensível para armazenar em logs/tabelas."""
    try:
        if not payload:
            return ""
        if isinstance(payload, dict):
            preview = {}
            for key in ("id", "type", "from", "timestamp", "status", "object"):
                if key in payload:
                    preview[key] = payload.get(key)
            if "metadata" in payload and isinstance(payload["metadata"], dict):
                preview["metadata"] = {
                    "phone_number_id": payload["metadata"].get("phone_number_id", "")
                }
            return json.dumps(preview, ensure_ascii=False)[:limit]
        return str(payload)[:limit]
    except Exception:
        return ""


# ============================================================
# LGPD — Versionamento de documentos legais
# ============================================================
# IMPORTANTE: Quando atualizar Política de Privacidade, Termos ou DPA,
# INCREMENTE a versão correspondente. Usuários ativos serão solicitados
# a re-aceitar os documentos atualizados no próximo login.
PRIVACY_POLICY_VERSION = "2026.05.18"
TERMS_OF_SERVICE_VERSION = "2026.05.18"
DPA_VERSION = "2026.05.18"

# Tipos de consentimento registrados na tabela consent_log
CONSENT_TYPES = {
    "privacy_policy": "Política de Privacidade",
    "terms_of_service": "Termos de Serviço",
    "dpa": "Contrato de Operador (DPA)",
    "marketing_email": "Recebimento de comunicações de marketing",
    "data_processing": "Tratamento de dados pessoais",
}


def register_consent(user_id, email, consent_type, version, accepted=True, details=""):
    """Registra consentimento do usuário para fins de comprovação LGPD (Art. 8º).

    Args:
        user_id: ID do usuário (pode ser None se for em momento de signup pré-criação)
        email: email do usuário (sempre registrado, mesmo se user_id for None)
        consent_type: tipo de consentimento (ver CONSENT_TYPES)
        version: versão do documento aceito (ex: PRIVACY_POLICY_VERSION)
        accepted: True se aceitou, False se revogou
        details: informações adicionais (texto livre)

    Returns:
        True se registrou com sucesso, False em caso de erro.
    """
    if consent_type not in CONSENT_TYPES:
        safe_log(f"[CONSENT] Tipo inválido: {consent_type}")
        return False
    try:
        ip = request.remote_addr if request else ""
        ua = request.headers.get("User-Agent", "")[:300] if request else ""
        db = get_db()
        db.execute(
            """INSERT INTO consent_log
               (user_id, email, consent_type, consent_version, accepted, ip_address, user_agent, details)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (user_id, (email or "")[:200], consent_type, version, 1 if accepted else 0, ip, ua, details[:500])
        )
        db.commit()
        return True
    except Exception as e:
        safe_log(f"[CONSENT] Erro ao registrar consentimento: {e}", level="ERROR")
        return False


def get_user_consents(user_id):
    """Retorna todos os consentimentos registrados de um usuário, mais recente primeiro."""
    try:
        db = get_db()
        rows = db.execute(
            """SELECT consent_type, consent_version, accepted, created_at, details
               FROM consent_log
               WHERE user_id = ?
               ORDER BY created_at DESC""",
            (user_id,)
        ).fetchall()
        return [dict(r) for r in rows]
    except Exception as e:
        safe_log(f"[CONSENT] Erro ao buscar consentimentos: {e}", level="ERROR")
        return []


def user_has_accepted_latest(user_id, consent_type, current_version):
    """Verifica se o usuário já aceitou a versão atual de um documento.
    Útil para forçar re-aceitação quando termos são atualizados."""
    try:
        db = get_db()
        row = db.execute(
            """SELECT 1 FROM consent_log
               WHERE user_id = ? AND consent_type = ? AND consent_version = ? AND accepted = 1
               LIMIT 1""",
            (user_id, consent_type, current_version)
        ).fetchone()
        return row is not None
    except Exception:
        return True  # Em caso de erro, não bloqueia o usuário


def user_needs_to_reaccept_terms(user_id):
    """Retorna True se o usuário precisa re-aceitar termos (novo signup ou versão antiga).
    Verifica os 3 documentos legais principais."""
    if not user_id:
        return False
    return not (
        user_has_accepted_latest(user_id, "privacy_policy", PRIVACY_POLICY_VERSION) and
        user_has_accepted_latest(user_id, "terms_of_service", TERMS_OF_SERVICE_VERSION) and
        user_has_accepted_latest(user_id, "dpa", DPA_VERSION)
    )


def mask_email(email):
    """Mascarar email para exibição: cleriston@gmail.com → cle***@gmail.com"""
    if not email or "@" not in email:
        return email or ""
    local, domain = email.split("@", 1)
    if len(local) <= 3:
        return f"{local[0]}***@{domain}"
    return f"{local[:3]}***@{domain}"


def mask_phone(phone):
    """Mascarar telefone: +5588999998888 → +5588****8888"""
    if not phone:
        return ""
    s = str(phone)
    if len(s) <= 6:
        return s
    return s[:6] + "*" * (len(s) - 10) + s[-4:] if len(s) >= 10 else s[:3] + "*" * (len(s) - 3)


# ════════════════════════════════════════════════════════════════
#  LOGGING SEGURO (mascaramento PII)
# ════════════════════════════════════════════════════════════════
#  Substitui safe_log() em pontos sensíveis. Os logs do Railway/produção
#  são logs operacionais; PII (email, telefone, token, conteúdo de
#  mensagem) NÃO deve aparecer em texto puro neles.
#  Princípio LGPD Art. 6º, VII (segurança) e Art. 46 (medidas técnicas).

import re as _re_log

_EMAIL_RE = _re_log.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b')
_PHONE_RE = _re_log.compile(r'\+?\b\d{10,15}\b')
_TOKEN_LIKE_RE = _re_log.compile(r'\b(EA[A-Za-z0-9_-]{20,}|sk-[A-Za-z0-9-]{20,}|gAAAAA[A-Za-z0-9_-]{30,}|TEST-[A-Za-z0-9-]{20,}|APP_USR-[A-Za-z0-9-]{20,})\b')


def _scrub_pii(text):
    """Remove/mascara PII de uma string antes de logar."""
    if not text:
        return text
    s = str(text)
    # Mascara emails
    s = _EMAIL_RE.sub(lambda m: mask_email(m.group(0)), s)
    # Mascara telefones longos (10+ dígitos)
    s = _PHONE_RE.sub(lambda m: mask_phone(m.group(0)), s)
    # Esconde tokens longos (Meta, Anthropic, Fernet, MercadoPago)
    s = _TOKEN_LIKE_RE.sub(lambda m: m.group(0)[:6] + "***[REDACTED]", s)
    return s


def _short_resp_text(resp, n=80):
    """Snippet curto e single-line da resposta de um provedor externo.
    Reduz vazamento de detalhe técnico em logs (ainda passa por _scrub_pii via safe_log)."""
    try:
        txt = (resp.text or "")[:n]
    except Exception:
        return "[no body]"
    # Remove quebras de linha e espaços extras
    txt = " ".join(txt.split())
    if len(txt) >= n:
        txt += "...[truncated]"
    return txt or "[empty]"


def safe_log(*args, level="INFO"):
    """safe_log() seguro que mascara PII automaticamente.
    Use no lugar de safe_log() para logs operacionais.
    Args podem ser quaisquer valores; todos passam por _scrub_pii."""
    try:
        parts = [_scrub_pii(a) for a in args]
        msg = " ".join(str(p) for p in parts)
        # Prefixa com nível
        print(f"[{level}] {msg}")
    except Exception:
        # Fallback ultra-conservador: se algo falhar, não logue NADA do conteúdo
        print(f"[{level}] [log scrub failed]")


def export_user_data_as_json(user_id):
    """Exporta TODOS os dados pessoais do usuário em formato estruturado JSON.
    Cumpre LGPD Art. 18, V (portabilidade) e Art. 18, II (acesso).

    Retorna dict pronto pra serialização ou None em caso de erro.
    """
    try:
        db = get_db()
        # Dados básicos do usuário (sem hash de senha!)
        user_row = db.execute(
            """SELECT id, email, name, company, phone, plan, plan_status,
                      msgs_limit, msgs_used, trial_ends_at, last_login,
                      created_at, is_active, email_verified
               FROM users WHERE id=?""", (user_id,)
        ).fetchone()
        if not user_row:
            return None

        export = {
            "_meta": {
                "export_date": datetime.now().isoformat(),
                "lgpd_basis": "Art. 18, II e V da Lei 13.709/2018",
                "data_subject_id": user_id,
                "controller": "atendente.online — Clériston Almeida Capistrano",
                "format_version": "1.0",
            },
            "cadastro": dict(user_row),
            "consentimentos": [],
            "configuracoes_whatsapp": [],
            "contatos": [],
            "conversas": [],
            "mensagens": [],
            "pedidos": [],
            "pagamentos": [],
            "pipeline_cards": [],
            "knowledge_base": [],
            "quick_replies": [],
            "templates_whatsapp": [],
            "campanhas": [],
            "contatos_bloqueados": [],
            "produtos": [],
            "biblioteca_social": [],
            "posts_agendados": [],
            "uso_api": [],
            "log_admin": [],
        }

        # Helper interno para tabelas simples (1 SELECT)
        def _fetch_table(table, where="user_id=?", params=(user_id,), order=""):
            try:
                sql = f"SELECT * FROM {table} WHERE {where}"
                if order:
                    sql += f" ORDER BY {order}"
                rows = db.execute(sql, params).fetchall()
                return [dict(r) for r in rows]
            except Exception as e:
                safe_log(f"[EXPORT] Erro lendo {table}: {e}", level="ERROR")
                return []

        export["consentimentos"]       = _fetch_table("consent_log", order="created_at DESC")
        export["contatos"]             = _fetch_table("contacts")
        export["conversas"]            = _fetch_table("conversations")
        export["pedidos"]              = _fetch_table("orders")
        export["pagamentos"]           = _fetch_table("payments")
        export["pipeline_cards"]       = _fetch_table("pipeline_cards")
        export["knowledge_base"]       = _fetch_table("knowledge_base")
        export["quick_replies"]        = _fetch_table("quick_replies")
        export["templates_whatsapp"]   = _fetch_table("whatsapp_templates")
        export["campanhas"]            = _fetch_table("campaigns")
        export["contatos_bloqueados"]  = _fetch_table("blocked_contacts")
        export["produtos"]             = _fetch_table("product_gallery")
        export["biblioteca_social"]    = _fetch_table("social_media_library")
        export["posts_agendados"]      = _fetch_table("scheduled_posts")
        export["uso_api"]              = _fetch_table("api_usage_log", order="created_at DESC")

        # Mensagens: pelo conv.user_id (não tem user_id direto)
        try:
            rows = db.execute(
                """SELECT m.* FROM messages m
                   JOIN conversations c ON c.id = m.conversation_id
                   WHERE c.user_id = ?
                   ORDER BY m.created_at""", (user_id,)
            ).fetchall()
            export["mensagens"] = [dict(r) for r in rows]
        except Exception as e:
            safe_log(f"[EXPORT] Erro lendo messages: {e}", level="ERROR")

        # Resumo numérico
        export["_resumo"] = {
            "total_contatos": len(export["contatos"]),
            "total_conversas": len(export["conversas"]),
            "total_mensagens": len(export["mensagens"]),
            "total_pedidos": len(export["pedidos"]),
            "total_pagamentos": len(export["pagamentos"]),
            "total_consentimentos": len(export["consentimentos"]),
        }

        return export
    except Exception as e:
        safe_log(f"[EXPORT] Erro fatal: {e}", level="ERROR")
        return None


def anonymize_user_account(user_id, reason="user_request"):
    """Anonimiza a conta do usuário (LGPD Art. 18, VI - exclusão).

    Estratégia HÍBRIDA:
    - Anonimiza PII em `users` (mas mantém registro pra integridade referencial e obrigação fiscal)
    - DELETA: conversas, mensagens, contatos, knowledge_base, quick_replies,
              campaigns, pipeline_cards, scheduled_posts, social_media_library,
              product_gallery, blocked_contacts
    - MANTÉM (anonimizado): users (linha vazia com flag), payments, orders, audit logs,
              api_usage_log (necessário pra obrigação fiscal/contábil 5 anos)
    - PRESERVA: consent_log (prova legal perpétua de aceites/revogações)

    Importante: usa PRAGMA table_info() para detectar quais colunas existem e só
    atualiza as que realmente fazem parte do schema (evita falha silenciosa).

    Retorna True se executou sem erros críticos.
    """
    if not user_id:
        return False
    try:
        db = get_db()
        anon_email = f"deleted_user_{user_id}@deleted.local"
        deleted_at = datetime.now().isoformat()

        # 1. Registrar a revogação ANTES de anonimizar (precisamos do email original)
        original = db.execute("SELECT email FROM users WHERE id=?", (user_id,)).fetchone()
        original_email = original["email"] if original else ""

        # ANTES de deletar mensagens, captura caminhos de mídia anexada pra apagar depois.
        # (Após DELETE, perdemos os media_url da tabela.)
        media_urls_to_clean = []
        try:
            rows = db.execute(
                """SELECT DISTINCT m.media_url FROM messages m
                   JOIN conversations c ON c.id = m.conversation_id
                   WHERE c.user_id = ? AND m.media_url != ''""", (user_id,)
            ).fetchall()
            for r in rows:
                url = (r["media_url"] or "").strip()
                if url and url.startswith("/") and "media_files" in url:
                    media_urls_to_clean.append(url)
        except Exception as e:
            safe_log(f"[ANONYMIZE] Aviso ao listar media_urls: {e}", level="WARN")

        # Captura também file_path da biblioteca social
        try:
            rows = db.execute(
                "SELECT file_path FROM social_media_library WHERE user_id=? AND file_path != ''",
                (user_id,)
            ).fetchall()
            for r in rows:
                fp = (r["file_path"] or "").strip()
                if fp and "media_files" in fp:
                    media_urls_to_clean.append(fp)
        except Exception:
            pass

        # E captura paths da galeria de produtos
        try:
            rows = db.execute(
                "SELECT file_path FROM product_gallery WHERE user_id=? AND file_path != ''",
                (user_id,)
            ).fetchall()
            for r in rows:
                fp = (r["file_path"] or "").strip()
                if fp and "media_files" in fp:
                    media_urls_to_clean.append(fp)
        except Exception:
            pass

        # Tabelas a DELETAR (conteúdo operacional, sem obrigação legal de manter)
        tables_to_delete = [
            ("messages", "conversation_id IN (SELECT id FROM conversations WHERE user_id=?)"),
            ("conversations", "user_id=?"),
            ("contacts", "user_id=?"),
            ("knowledge_base", "user_id=?"),
            ("quick_replies", "user_id=?"),
            ("blocked_contacts", "user_id=?"),
            ("campaign_contacts", "campaign_id IN (SELECT id FROM campaigns WHERE user_id=?)"),
            ("campaigns", "user_id=?"),
            ("pipeline_cards", "user_id=?"),
            ("pipeline_stages", "user_id=?"),
            ("whatsapp_templates", "user_id=?"),
            ("product_gallery", "user_id=?"),
            ("social_media_library", "user_id=?"),
            ("scheduled_posts", "user_id=?"),
        ]
        delete_summary = {}
        for table, where in tables_to_delete:
            try:
                cur = db.execute(f"DELETE FROM {table} WHERE {where}", (user_id,))
                delete_summary[table] = cur.rowcount or 0
            except Exception as e:
                safe_log(f"[ANONYMIZE] Erro deletando {table}: {e}", level="WARN")

        # ── Anonimizar ORDERS (obrigação fiscal mantém o registro)
        # Detecta dinamicamente quais colunas de PII existem antes de tentar atualizar
        try:
            orders_cols = {row[1] for row in db.execute("PRAGMA table_info(orders)").fetchall()}
            updates = []
            params = []
            if "customer_name" in orders_cols:
                updates.append("customer_name=?"); params.append("Cliente Excluído")
            if "customer_phone" in orders_cols:
                updates.append("customer_phone=?"); params.append("")
            if "customer_email" in orders_cols:  # caso seja adicionada no futuro
                updates.append("customer_email=?"); params.append("")
            if "notes" in orders_cols:
                updates.append("notes=?"); params.append("")
            if updates:
                params.append(user_id)
                db.execute(f"UPDATE orders SET {', '.join(updates)} WHERE user_id=?", params)
        except Exception as e:
            safe_log(f"[ANONYMIZE] Erro anonimizando orders: {e}", level="WARN")

        # ── Anonimizar USERS: detecta colunas REAIS do schema antes de atualizar
        # Mapeia colunas a serem zeradas: name → "Usuário Excluído", senha → "", tokens/PII → ""
        # Mantém apenas o registro mínimo para integridade referencial (payments etc.)
        try:
            user_cols = {row[1] for row in db.execute("PRAGMA table_info(users)").fetchall()}

            # Colunas que recebem valor especial
            special_updates = {
                "email": anon_email,
                "name": "Usuário Excluído",
                "is_active": 0,
                "plan_status": "deleted",
                "email_verified": 0,
            }
            # Colunas a apagar completamente (vão ficar string vazia)
            blank_columns = [
                "password_hash", "company", "phone",
                "whatsapp_phone_id", "whatsapp_token",
                "instagram_page_id", "instagram_token",
                "messenger_page_id", "messenger_token",
                "telegram_bot_token", "telegram_chat_id",
                "mp_access_token", "mp_public_key",
                "mp_subscription_id",
                "ai_system_prompt", "ai_greeting",
                "social_business_context", "social_last_run",
                "auto_reply_off_hours",
            ]

            updates = []
            params = []
            for col, val in special_updates.items():
                if col in user_cols:
                    updates.append(f"{col}=?")
                    params.append(val)
            for col in blank_columns:
                if col in user_cols:
                    updates.append(f"{col}=?")
                    params.append("")
            # deleted_at se existir
            if "deleted_at" in user_cols:
                updates.append("deleted_at=?")
                params.append(deleted_at)

            if not updates:
                safe_log("[ANONYMIZE] ERRO: nenhuma coluna esperada encontrada em users", level="ERROR")
                return False

            params.append(user_id)
            sql = f"UPDATE users SET {', '.join(updates)} WHERE id=?"
            db.execute(sql, params)
            safe_log(f"[ANONYMIZE] Atualizadas {len(updates)} colunas em users para id={user_id}")
        except Exception as e:
            safe_log(f"[ANONYMIZE] ERRO crítico anonimizando users: {e}", level="ERROR")
            return False

        # Limpar mídia física associada (LGPD: dados pessoais não podem permanecer em disco)
        # 1. Arquivos referenciados em messages/social/gallery (capturados ANTES dos DELETEs)
        # 2. Arquivos com prefixo user{id}_ em subpastas social/gallery (varredura por glob)
        # 3. Subpasta legada media_files/user_{id}/ (se existir)
        try:
            import glob, shutil
            media_root = MEDIA_FOLDER if 'MEDIA_FOLDER' in globals() else os.path.join(os.path.dirname(DATABASE), "media_files")
            files_removed = 0

            # (1) Apaga arquivos referenciados explicitamente no DB (capturados antes)
            for path in media_urls_to_clean:
                try:
                    # Resolve para caminho absoluto seguro dentro de media_root
                    abs_path = os.path.abspath(path)
                    abs_root = os.path.abspath(media_root)
                    if abs_path.startswith(abs_root) and os.path.isfile(abs_path):
                        os.remove(abs_path)
                        files_removed += 1
                except Exception as e:
                    safe_log(f"[ANONYMIZE] Falha removendo {os.path.basename(path)}: {e}", level="WARN")

            # (2) Varredura por padrão user{id}_* (catch-all caso o DB não registre tudo)
            patterns = [
                os.path.join(media_root, "social", f"user{user_id}_*"),
                os.path.join(media_root, "gallery", f"user{user_id}_*"),
            ]
            for pat in patterns:
                for path in glob.glob(pat):
                    try:
                        if os.path.isfile(path):
                            os.remove(path)
                            files_removed += 1
                    except Exception as e:
                        safe_log(f"[ANONYMIZE] Falha removendo {os.path.basename(path)}: {e}", level="WARN")

            # (3) Subpasta user_X legada (se algum dia existir)
            legacy_dir = os.path.join(media_root, f"user_{user_id}")
            if os.path.isdir(legacy_dir):
                shutil.rmtree(legacy_dir, ignore_errors=True)
                files_removed += 1

            if files_removed:
                safe_log(f"[ANONYMIZE] {files_removed} arquivo(s) de mídia removido(s) para user {user_id}")
        except Exception as e:
            safe_log(f"[ANONYMIZE] Aviso ao limpar mídia: {e}", level="WARN")

        # Registra a anonimização no consent_log (prova legal) - usa email mascarado, não o original em texto puro
        try:
            ua_short = ""
            if request:
                ua_short = (request.headers.get("User-Agent", "") or "")[:300]
            db.execute(
                """INSERT INTO consent_log
                   (user_id, email, consent_type, consent_version, accepted, ip_address, user_agent, details)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (user_id, anon_email, "data_processing", PRIVACY_POLICY_VERSION, 0,
                 request.remote_addr if request else "",
                 ua_short,
                 f"Conta anonimizada/excluída a pedido do titular em {deleted_at}. Motivo: {reason}. "
                 f"Deletadas: {delete_summary}")
            )
        except Exception as e:
            safe_log(f"[ANONYMIZE] Erro ao registrar log de exclusão: {e}", level="WARN")

        db.commit()
        safe_log(f"[ANONYMIZE] Usuário {user_id} ({mask_email(original_email)}) anonimizado com sucesso")
        return True
    except Exception as e:
        safe_log(f"[ANONYMIZE] Erro fatal: {e}", level="ERROR")
        return False


def generate_deletion_code(user_id, email):
    """Gera código de 6 dígitos pra confirmar exclusão de conta via email.
    Armazena apenas o HASH (HMAC-SHA256 + pepper).
    Cooldown de 60s entre reenvios (anti-flood).
    Código válido por 15 minutos."""
    if not email:
        return None
    try:
        db = get_db()
        # Cooldown: verifica se já foi enviado um há menos de 60s
        recent = db.execute(
            """SELECT created_at FROM verification_codes
               WHERE email=? AND code_type='deletion'
               ORDER BY created_at DESC LIMIT 1""", (email,)
        ).fetchone()
        if recent and recent["created_at"]:
            try:
                last_time = datetime.fromisoformat(recent["created_at"])
                if (datetime.now() - last_time).total_seconds() < 60:
                    safe_log(f"[DELETION_CODE] Cooldown ativo para {mask_email(email)}", level="WARN")
                    return None
            except Exception:
                pass

        code = f"{secrets.randbelow(900000) + 100000}"  # 6 dígitos
        code_hash = hash_verification_code(code)
        # Limpar códigos antigos do mesmo email
        db.execute("DELETE FROM verification_codes WHERE email=? AND code_type='deletion'", (email,))
        db.execute(
            """INSERT INTO verification_codes (email, code, code_type, expires_at)
               VALUES (?, ?, 'deletion', datetime('now', '+15 minutes'))""",
            (email, code_hash)
        )
        db.commit()
        return code  # retorna o código em texto para enviar por email; só o HASH fica no banco
    except Exception as e:
        safe_log(f"[DELETION_CODE] Erro: {e}", level="ERROR")
        return None


def verify_deletion_code(email, code):
    """Verifica se o código de exclusão é válido (compara hash)."""
    if not email or not code:
        return False
    try:
        db = get_db()
        code_hash = hash_verification_code(code)
        row = db.execute(
            """SELECT id FROM verification_codes
               WHERE email=? AND code=? AND code_type='deletion'
                 AND expires_at > datetime('now')
               LIMIT 1""", (email, code_hash)
        ).fetchone()
        if row:
            db.execute("DELETE FROM verification_codes WHERE id=?", (row["id"],))
            db.commit()
            return True
        return False
    except Exception as e:
        safe_log(f"[DELETION_CODE] Erro verificação: {e}", level="ERROR")
        return False


# ════════════════════════════════════════════════════════════════
#  BACKUP AUTOMATIZADO (LGPD Art. 46 — medidas técnicas)
# ════════════════════════════════════════════════════════════════
#  Backup diário do SQLite no próprio volume Railway com:
#  - Cópia consistente via sqlite3.backup() (sem corrupção)
#  - Criptografia opcional Fernet (se DATA_ENCRYPTION_KEY/SECRET_KEY ok)
#  - Rotação automática: mantém últimos N backups, apaga antigos
#  - Hashing SHA-256 pra verificação de integridade

BACKUP_DIR = os.path.join(os.path.dirname(DATABASE), "backups") if "/" in DATABASE else "backups"
BACKUP_RETENTION_DAYS = int(os.getenv("BACKUP_RETENTION_DAYS", "30"))


def perform_database_backup():
    """Faz backup consistente do banco SQLite.
    - Usa sqlite3.backup() (não cp) para evitar corrupção com escritas concorrentes.
    - Salva em /app/data/backups/atendeia-YYYYMMDD-HHMMSS.db
    - Calcula SHA-256 para integridade.
    - Criptografa com Fernet se a chave estiver disponível.
    - Roda rotação: apaga backups com mais de BACKUP_RETENTION_DAYS dias.

    Retorna dict com status ou None em caso de falha grave.
    """
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_filename = f"atendeia-{timestamp}.db"
        backup_path = os.path.join(BACKUP_DIR, backup_filename)

        # 1. Backup consistente via API SQLite (não copia arquivo bruto)
        src = sqlite3.connect(DATABASE)
        dst = sqlite3.connect(backup_path)
        with dst:
            src.backup(dst)
        src.close()
        dst.close()

        # 2. Hash SHA-256 para integridade
        sha = hashlib.sha256()
        with open(backup_path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                sha.update(chunk)
        digest = sha.hexdigest()

        # 3. Criptografar com Fernet — em produção, é OBRIGATÓRIO (fail-closed):
        #    se a chave não existir OU a criptografia falhar, apaga o .db em claro e aborta.
        #    Em dev (FLASK_ENV=development), mantém o backup em claro com aviso.
        encrypted_path = None
        is_dev = os.getenv("FLASK_ENV", "").lower() == "development"
        fernet = _get_fernet()

        if fernet:
            try:
                with open(backup_path, "rb") as f:
                    raw = f.read()
                encrypted = fernet.encrypt(raw)
                encrypted_path = backup_path + ".enc"
                with open(encrypted_path, "wb") as f:
                    f.write(encrypted)
                # Remove a versão não-criptografada
                os.remove(backup_path)
                final_path = encrypted_path
            except Exception as e:
                err_id = secrets.token_hex(6)
                safe_log(f"[BACKUP] Falha ao criptografar id={err_id}: {e}", level="ERROR")
                # Fail-closed em produção: apaga o .db em claro e aborta.
                if not is_dev:
                    try:
                        if os.path.exists(backup_path):
                            os.remove(backup_path)
                    except OSError:
                        pass
                    safe_log(f"[BACKUP] PRODUÇÃO: backup em claro REMOVIDO e operação ABORTADA (err_id={err_id})", level="ERROR")
                    return None
                safe_log("[BACKUP] DEV: mantendo backup em claro (FLASK_ENV=development)", level="WARN")
                final_path = backup_path
        else:
            # Sem chave Fernet: fail-closed em produção.
            if not is_dev:
                try:
                    if os.path.exists(backup_path):
                        os.remove(backup_path)
                except OSError:
                    pass
                safe_log("[BACKUP] PRODUÇÃO: chave de criptografia ausente — backup em claro REMOVIDO e operação ABORTADA. Configure DATA_ENCRYPTION_KEY ou SECRET_KEY.", level="ERROR")
                return None
            safe_log("[BACKUP] DEV: criptografia indisponível, mantendo backup em claro", level="WARN")
            final_path = backup_path

        # 4. Salvar metadata em arquivo .json (sha, tamanho, criptografado?)
        meta_path = final_path + ".meta.json"
        size_bytes = os.path.getsize(final_path)
        metadata = {
            "filename": os.path.basename(final_path),
            "created_at": datetime.now().isoformat(),
            "size_bytes": size_bytes,
            "sha256_unencrypted": digest,
            "encrypted": bool(encrypted_path),
            "source_db": DATABASE,
        }
        with open(meta_path, "w") as f:
            json.dump(metadata, f, indent=2)

        # 5. Rotação: apagar backups antigos
        removed = 0
        cutoff = time.time() - (BACKUP_RETENTION_DAYS * 86400)
        for entry in os.listdir(BACKUP_DIR):
            full = os.path.join(BACKUP_DIR, entry)
            try:
                if os.path.getmtime(full) < cutoff and (entry.startswith("atendeia-") and (entry.endswith(".db") or entry.endswith(".enc") or entry.endswith(".meta.json"))):
                    os.remove(full)
                    removed += 1
            except Exception:
                pass

        size_mb = size_bytes / (1024 * 1024)
        safe_log(f"[BACKUP] OK: {os.path.basename(final_path)} ({size_mb:.2f}MB) — rotação: {removed} arquivos antigos removidos")
        return {
            "ok": True,
            "path": final_path,
            "size_mb": round(size_mb, 2),
            "encrypted": bool(encrypted_path),
            "removed_old": removed,
        }
    except Exception as e:
        safe_log(f"[BACKUP] ERRO: {e}", level="ERROR")
        return None


# ════════════════════════════════════════════════════════════════
#  RETENÇÃO AUTOMÁTICA DE DADOS (LGPD Art. 16)
# ════════════════════════════════════════════════════════════════
#  Apaga permanentemente mensagens antigas conforme plano do usuário:
#    Starter:    90 dias
#    Pro:       180 dias
#    Business:  365 dias
#    Agência:   730 dias

RETENTION_DAYS_BY_PLAN = {
    "starter": 90,
    "pro": 180,
    "business": 365,
    "agency": 730,
}


def perform_retention_cleanup():
    """Apaga mensagens (e conversas vazias) cujo prazo de retenção do plano expirou.
    Roda diariamente. Resultado: respeita LGPD e reduz tamanho do banco.

    Retorna dict com estatísticas ou None em caso de falha.
    """
    try:
        db = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
        total_msgs_deleted = 0
        total_convs_deleted = 0
        by_plan = {}

        for plan, days in RETENTION_DAYS_BY_PLAN.items():
            # Apaga mensagens com mais de N dias para conversas de usuários nesse plano
            cur = db.execute(
                """DELETE FROM messages
                   WHERE conversation_id IN (
                       SELECT c.id FROM conversations c
                       JOIN users u ON u.id = c.user_id
                       WHERE u.plan = ?
                   )
                   AND created_at < datetime('now', ?)""",
                (plan, f"-{days} days")
            )
            msgs_deleted = cur.rowcount or 0
            total_msgs_deleted += msgs_deleted

            # Apaga conversas que ficaram totalmente vazias (sem mensagens)
            cur = db.execute(
                """DELETE FROM conversations
                   WHERE user_id IN (SELECT id FROM users WHERE plan = ?)
                   AND id NOT IN (SELECT DISTINCT conversation_id FROM messages)
                   AND last_message_at < datetime('now', ?)""",
                (plan, f"-{days} days")
            )
            convs_deleted = cur.rowcount or 0
            total_convs_deleted += convs_deleted

            by_plan[plan] = {"msgs": msgs_deleted, "convs": convs_deleted}

        # Limpa também logs antigos
        # Audit log admin: 180 dias
        cur = db.execute("DELETE FROM admin_audit_log WHERE created_at < datetime('now', '-180 days')")
        audit_deleted = cur.rowcount or 0
        # Webhook errors: 60 dias
        cur = db.execute("DELETE FROM webhook_errors WHERE created_at < datetime('now', '-60 days')")
        webhook_errors_deleted = cur.rowcount or 0
        # Processed webhook events: 30 dias (dedupe não precisa de histórico longo)
        try:
            cur = db.execute("DELETE FROM processed_webhook_events WHERE created_at < datetime('now', '-30 days')")
            webhook_events_deleted = cur.rowcount or 0
        except Exception:
            webhook_events_deleted = 0
        # Verification codes expirados
        cur = db.execute("DELETE FROM verification_codes WHERE expires_at < datetime('now')")
        codes_deleted = cur.rowcount or 0

        db.commit()
        db.close()

        result = {
            "ok": True,
            "messages_deleted": total_msgs_deleted,
            "conversations_deleted": total_convs_deleted,
            "by_plan": by_plan,
            "audit_log_deleted": audit_deleted,
            "webhook_errors_deleted": webhook_errors_deleted,
            "webhook_events_deleted": webhook_events_deleted,
            "verification_codes_deleted": codes_deleted,
        }
        safe_log(f"[RETENTION] OK: {total_msgs_deleted} msgs, {total_convs_deleted} convs, {audit_deleted} audit, {codes_deleted} códigos expirados")
        return result
    except Exception as e:
        safe_log(f"[RETENTION] ERRO: {e}", level="ERROR")
        return None


def start_daily_maintenance_scheduler():
    """Inicia thread daemon que roda backup + retenção diariamente às 3h da manhã (BRT/UTC-3).
    Em produção/Railway, o servidor está em UTC, então usa-se UTC+0 03h = 00h BRT.
    Mas como queremos 3h BRT (horário de menor uso), agendamos para 06h UTC."""
    def loop():
        last_run_date = None
        target_hour_utc = int(os.getenv("MAINTENANCE_HOUR_UTC", "6"))  # 06h UTC = 03h BRT

        while True:
            try:
                now = datetime.utcnow()
                today_key = now.strftime("%Y-%m-%d")

                # Roda uma vez por dia, na hora alvo
                if now.hour == target_hour_utc and last_run_date != today_key:
                    safe_log(f"[MAINTENANCE] Iniciando rotina diária às {now.isoformat()} UTC")
                    perform_database_backup()
                    perform_retention_cleanup()
                    last_run_date = today_key
                    safe_log("[MAINTENANCE] Rotina diária concluída")

            except Exception as e:
                safe_log(f"[MAINTENANCE LOOP] Erro: {e}", level="ERROR")

            # Verifica a cada 30 minutos
            time.sleep(1800)

    thread = threading.Thread(target=loop, daemon=True)
    thread.start()
    safe_log("[MAINTENANCE] Scheduler diário (backup + retenção) iniciado — roda às 06h UTC (~03h BRT)")


def register_processed_webhook_event(source, event_key, user_id=None, payload=None):
    """Registra evento processado. Retorna False se já tiver sido processado."""
    if not source or not event_key:
        return False
    try:
        db_conn = sqlite3.connect(DATABASE)
        db_conn.execute(
            """INSERT INTO processed_webhook_events (source, event_key, user_id, payload_preview)
               VALUES (?,?,?,?)""",
            (source[:50], str(event_key)[:255], user_id, _safe_payload_preview(payload))
        )
        db_conn.commit()
        db_conn.close()
        return True
    except sqlite3.IntegrityError:
        return False
    except Exception as e:
        safe_log(f"[WEBHOOK EVENT] Falha ao registrar {source}:{event_key} -> {e}", level="ERROR")
        return False


def resolve_user_by_whatsapp_phone_id(db_conn, phone_number_id):
    if not phone_number_id:
        return None
    db_conn.row_factory = sqlite3.Row
    return db_conn.execute(
        "SELECT * FROM users WHERE whatsapp_phone_id=? AND is_active=1 LIMIT 1",
        (phone_number_id,)
    ).fetchone()


def parse_db_datetime(value):
    if not value:
        return None
    try:
        if "T" in value:
            dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        else:
            dt = datetime.strptime(value[:19], "%Y-%m-%d %H:%M:%S")
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return None


def validate_user_messaging_access(user):
    """Valida se o tenant pode consumir IA e enviar mensagens automáticas."""
    if not user:
        return False, "user_not_found", "Conta não encontrada."
    if not user.get("is_active", 1):
        return False, "user_inactive", "Conta desativada."

    plan_status = (user.get("plan_status") or "").strip().lower()
    if plan_status in {"inactive", "cancelled", "blocked", "suspended"}:
        return False, "plan_inactive", "Seu plano está inativo. Entre em contato para reativar."

    if plan_status == "trial":
        trial_ends_at = parse_db_datetime(user.get("trial_ends_at", ""))
        if trial_ends_at and datetime.now(timezone.utc) > trial_ends_at:
            return False, "trial_expired", "Seu período de teste expirou."

    msgs_limit = int(user.get("msgs_limit") or 0)
    msgs_used = int(user.get("msgs_used") or 0)
    if msgs_limit > 0 and msgs_used >= msgs_limit:
        return False, "message_limit_reached", "Seu limite mensal de mensagens foi atingido."

    return True, "ok", ""


def get_last_customer_message_at(db_conn, user_id, phone, channel="whatsapp"):
    row = db_conn.execute(
        """
        SELECT MAX(m.created_at) as last_customer_message_at
        FROM messages m
        JOIN conversations c ON c.id = m.conversation_id
        WHERE c.user_id=? AND c.customer_phone=? AND c.channel=? AND m.sender='customer'
        """,
        (user_id, phone, channel)
    ).fetchone()
    return parse_db_datetime(row["last_customer_message_at"]) if row and row["last_customer_message_at"] else None


def is_within_customer_care_window(db_conn, user_id, phone, channel="whatsapp", now_utc=None):
    now_utc = now_utc or datetime.now(timezone.utc)
    last_customer_message_at = get_last_customer_message_at(db_conn, user_id, phone, channel)
    if not last_customer_message_at:
        return False
    return (now_utc - last_customer_message_at).total_seconds() <= 24 * 3600


def get_approved_template_for_campaign(db_conn, campaign):
    template_id = campaign["template_id"] if campaign else None
    if template_id:
        tpl = db_conn.execute(
            "SELECT * FROM whatsapp_templates WHERE id=? AND user_id=? LIMIT 1",
            (template_id, campaign["user_id"])
        ).fetchone()
        if tpl and (tpl["status"] or "").lower() == "approved":
            return tpl
    return None


def validate_campaign_contact_window(db_conn, campaign, contact):
    template = get_approved_template_for_campaign(db_conn, campaign)
    within_window = is_within_customer_care_window(db_conn, campaign["user_id"], contact["phone"], "whatsapp")
    if within_window:
        return True, ""
    if template:
        return False, "Contato fora da janela de 24h. Este projeto ainda nao envia template aprovado automaticamente."
    return False, "Contato fora da janela de 24h e sem template aprovado."


def parse_mp_subscription_reference(external_reference):
    parts = (external_reference or "").split("_")
    if len(parts) < 4 or parts[0] != "user" or parts[2] != "plan":
        return None
    try:
        return {"user_id": int(parts[1]), "plan_key": parts[3]}
    except (TypeError, ValueError):
        return None


def validate_mp_signature(request_obj, webhook_secret=None):
    """Valida a assinatura do webhook Mercado Pago.
    Docs: https://www.mercadopago.com.br/developers/en/docs/your-integrations/notifications/webhooks
    
    COMPORTAMENTO:
    - Produção SEM MP_WEBHOOK_SECRET → REJEITA (retorna False)
    - Dev (FLASK_ENV=development) SEM secret → aceita com warning
    - Secret configurado → valida HMAC-SHA256, retorna True/False
    """
    import hmac as hmac_mod, hashlib as hashlib_mod
    if not webhook_secret:
        webhook_secret = get_setting("MP_WEBHOOK_SECRET", os.getenv("MP_WEBHOOK_SECRET", ""))

    is_dev = os.getenv("FLASK_ENV", "").lower() == "development"

    # Sem secret = FALHA em produção, aceita em dev
    if not webhook_secret:
        if is_dev:
            safe_log("[MP SIG] ⚠️ DEV: MP_WEBHOOK_SECRET não configurado — aceito sem validação", level="WARN")
            return True
        # Produção: REJEITA
        safe_log("[MP SIG] ❌ PRODUÇÃO: MP_WEBHOOK_SECRET não configurado. Webhook rejeitado.", level="ERROR")
        safe_log("[MP SIG] Configure em /admin/api-settings → MP Webhook Secret")
        return False

    signature_header = request_obj.headers.get("x-signature", "")
    request_id = request_obj.headers.get("x-request-id", "")

    if not signature_header:
        safe_log("[MP SIG] Rejeitado: x-signature ausente", level="WARN")
        return False

    parts = dict(p.split("=", 1) for p in signature_header.split(",") if "=" in p)
    ts = parts.get("ts", "")
    received_hash = parts.get("v1", "")

    if not ts or not received_hash:
        safe_log("[MP SIG] Rejeitado: header malformado", level="ERROR")
        return False

    # Anti-replay: rejeita ts fora da janela de 5 minutos.
    # MP envia ts como milissegundos desde epoch.
    try:
        ts_int = int(ts)
        # Detecta automaticamente se está em ms ou em segundos
        if ts_int > 10_000_000_000:  # > ano 2286 em segundos -> só pode ser ms
            ts_seconds = ts_int / 1000.0
        else:
            ts_seconds = float(ts_int)
        now_seconds = time.time()
        delta = abs(now_seconds - ts_seconds)
        MP_REPLAY_WINDOW = 300  # 5 minutos (igual ao padrão Slack/Stripe)
        if delta > MP_REPLAY_WINDOW:
            safe_log(f"[MP SIG] Rejeitado: ts fora da janela ({int(delta)}s de diferença, max {MP_REPLAY_WINDOW}s)", level="ERROR")
            return False
    except (ValueError, TypeError) as e:
        safe_log(f"[MP SIG] Rejeitado: ts inválido ({e})", level="ERROR")
        return False

    try:
        data = request_obj.json or {}
        data_id = str(data.get("data", {}).get("id", ""))
    except:
        data_id = ""

    manifest = f"id:{data_id};request-id:{request_id};ts:{ts};"
    expected_hash = hmac_mod.new(
        webhook_secret.encode("utf-8"),
        manifest.encode("utf-8"),
        hashlib_mod.sha256
    ).hexdigest()

    if not hmac_mod.compare_digest(expected_hash, received_hash):
        safe_log(f"[MP SIG] Rejeitado: hash inválido (data_id={data_id})", level="ERROR")
        return False

    return True


def mask_secret(value, show_last=4):
    """Mascara um segredo para exibição segura na UI.
    Exemplo: 'ABC123XYZ789' -> '••••••••••9789' (mostra só últimos 4 chars)"""
    if not value:
        return ""
    # Se veio criptografado, descriptografa primeiro
    if isinstance(value, str) and value.startswith("fer:v1:"):
        try:
            value = _decrypt_value(value)
        except:
            return "••••••••••"
    if not value or len(value) < 6:
        return "••••••••••"
    if len(value) <= show_last + 2:
        return "•" * (len(value) - 2) + value[-2:]
    return "•" * 10 + value[-show_last:]


def decrypt_user_row(user):
    """Retorna uma cópia do user com tokens descriptografados. Aceita dict ou Row."""
    if user is None:
        return None
    # Converte Row para dict para permitir modificação
    try:
        user_dict = dict(user)
    except (TypeError, ValueError):
        return user
    for field in USER_ENCRYPTED_FIELDS:
        if field in user_dict and user_dict[field]:
            user_dict[field] = _decrypt_value(user_dict[field])
    return user_dict


def migrate_encrypt_user_tokens():
    """Migra tokens por usuário em texto puro para formato Fernet criptografado.
    Roda uma vez no startup — identifica linhas não-criptografadas e criptografa."""
    try:
        db = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
        rows = db.execute(f"SELECT id, {', '.join(USER_ENCRYPTED_FIELDS)} FROM users").fetchall()
        fernet = _get_fernet()
        if not fernet:
            db.close()
            return
        migrated = 0
        for row in rows:
            updates = {}
            for field in USER_ENCRYPTED_FIELDS:
                val = row[field]
                if not val:
                    continue
                if isinstance(val, str) and val.startswith("fer:v1:"):
                    continue  # Já criptografado
                # Criptografa
                encrypted = _encrypt_value(val)
                if encrypted != val:
                    updates[field] = encrypted
            if updates:
                set_clause = ", ".join([f"{k}=?" for k in updates.keys()])
                params = list(updates.values()) + [row["id"]]
                db.execute(f"UPDATE users SET {set_clause} WHERE id=?", params)
                migrated += 1
        if migrated > 0:
            db.commit()
            safe_log(f"[CRYPTO] Migração: {migrated} usuário(s) com tokens criptografados")
        db.close()
    except Exception as e:
        safe_log(f"[CRYPTO] Erro na migração de tokens: {e}", level="ERROR")


def _decrypt_legacy(encrypted):
    """Descriptografa formato antigo (XOR + HMAC) para migração"""
    try:
        import hmac
        secret = os.getenv("SECRET_KEY", "") or (app.secret_key if isinstance(app.secret_key, str) else "")
        if isinstance(secret, str):
            secret = secret.encode()
        key = hashlib.sha256(secret or b"default-key").digest()
        blob = base64.b64decode(encrypted[7:])
        iv = blob[:16]
        mac_received = blob[-32:]
        ciphertext = blob[16:-32]
        mac_expected = hmac.new(key, iv + ciphertext, hashlib.sha256).digest()
        if not hmac.compare_digest(mac_received, mac_expected):
            return ""
        keystream = b""
        counter = 0
        while len(keystream) < len(ciphertext):
            keystream += hashlib.sha256(key + iv + counter.to_bytes(4, 'big')).digest()
            counter += 1
        plaintext = bytes(a ^ b for a, b in zip(ciphertext, keystream[:len(ciphertext)]))
        return plaintext.decode('utf-8')
    except Exception:
        return ""


# Chaves/tokens que precisam ser criptografados no banco
SENSITIVE_SETTINGS = {
    "ANTHROPIC_API_KEY", "OPENAI_API_KEY", "GROQ_API_KEY",
    "MERCADOPAGO_ACCESS_TOKEN", "MP_WEBHOOK_SECRET", "RESEND_API_KEY",
    "WHATSAPP_APP_SECRET", "SMTP_PASSWORD",
    "ADMIN_TOTP_SECRET", "ADMIN_BACKUP_CODES"
}


def get_setting(key, default=""):
    """Busca config do banco, se não existir usa variável de ambiente"""
    try:
        db_conn = sqlite3.connect(DATABASE)
        row = db_conn.execute("SELECT value FROM system_settings WHERE key=?", (key,)).fetchone()
        db_conn.close()
        if row and row[0]:
            value = row[0]
            # Descriptografa se for campo sensível
            if key in SENSITIVE_SETTINGS:
                value = _decrypt_value(value)
            return value
    except:
        pass
    return os.getenv(key, default)


def set_setting(key, value):
    """Salva config no banco. Criptografa se for campo sensível."""
    # Criptografa campos sensíveis antes de salvar
    stored_value = _encrypt_value(value) if key in SENSITIVE_SETTINGS else value
    db_conn = sqlite3.connect(DATABASE)
    db_conn.execute("INSERT OR REPLACE INTO system_settings (key, value) VALUES (?, ?)", (key, stored_value))
    db_conn.commit()
    db_conn.close()


# ─── EMAIL ─────────────────────────────────────────────────────
def send_email(to_email, subject, html_body):
    """Envia email via Resend API. Logs mascarados (LGPD)."""
    resend_key = get_setting("RESEND_API_KEY", "")
    from_email = get_setting("RESEND_FROM_EMAIL", "atendente.online <onboarding@resend.dev>")
    masked = mask_email(to_email) if to_email else "(empty)"

    if not resend_key:
        safe_log(f"[EMAIL] RESEND_API_KEY não configurada. Email para {masked} não enviado.", level="WARN")
        return False

    try:
        import requests as req
        resp = req.post("https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
            json={"from": from_email, "to": [to_email], "subject": subject, "html": html_body},
            timeout=15)
        if resp.status_code == 200:
            # Loga só o tipo do email (primeira palavra do subject), nunca código nem email completo
            subject_kind = (subject or "").split(":")[0][:50]
            safe_log(f"[EMAIL] Enviado para {masked} ({subject_kind})")
            return True
        else:
            safe_log(f"[EMAIL] Erro {resp.status_code} ao enviar para {masked}", level="ERROR")
            return False
    except Exception as e:
        safe_log(f"[EMAIL] Erro ao enviar para {masked}: {e}", level="ERROR")
        return False


# ════════════════════════════════════════════════════════════════
#  Helpers de hash para códigos de verificação (LGPD/segurança)
# ════════════════════════════════════════════════════════════════
def hash_verification_code(code):
    """HMAC-SHA256 do código + pepper derivado da SECRET_KEY.
    Se o banco vazar, atacante não consegue usar os códigos diretamente.
    Códigos têm vida curta (15-30 min), mas hash em repouso é boa prática."""
    pepper = (app.secret_key or "").encode() if isinstance(app.secret_key, str) else (app.secret_key or b"")
    return hashlib.sha256(pepper + (str(code) or "").encode()).hexdigest()


def send_verification_code(email):
    """Gera e envia código de verificação. Armazena apenas o HASH."""
    code = str(random.randint(100000, 999999))
    expires = (datetime.now() + timedelta(minutes=30)).isoformat()
    code_hash = hash_verification_code(code)

    db_conn = sqlite3.connect(DATABASE)
    db_conn.execute("DELETE FROM verification_codes WHERE email=?", (email,))
    db_conn.execute("INSERT INTO verification_codes (email, code, expires_at) VALUES (?,?,?)", (email, code_hash, expires))
    db_conn.commit()
    db_conn.close()

    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;padding:32px;background:#111827;border-radius:12px;color:#f0f4f8">
        <div style="text-align:center;margin-bottom:24px">
            <h2 style="color:#34d399;margin:0">atendente.online</h2>
        </div>
        <p>Olá! Seu código de verificação é:</p>
        <div style="text-align:center;margin:24px 0">
            <span style="font-size:36px;font-weight:700;letter-spacing:8px;color:#00c896;background:#1a2235;padding:16px 32px;border-radius:8px;display:inline-block">{code}</span>
        </div>
        <p style="color:#94a3b8;font-size:14px">Este código expira em <strong>30 minutos</strong>.</p>
        <p style="color:#94a3b8;font-size:14px">Se você não solicitou este código, ignore este email.</p>
        <hr style="border:none;border-top:1px solid #243049;margin:24px 0">
        <p style="color:#64748b;font-size:12px;text-align:center">© 2026 atendente.online</p>
    </div>"""

    return send_email(email, f"Seu código de verificação: {code}", html)


def verify_code(email, code):
    """Verifica se o código é válido. Compara hash, não texto puro."""
    if not email or not code:
        return False
    db_conn = sqlite3.connect(DATABASE)
    db_conn.row_factory = sqlite3.Row
    code_hash = hash_verification_code(code)
    row = db_conn.execute(
        "SELECT * FROM verification_codes WHERE email=? AND code=? AND used=0 ORDER BY created_at DESC LIMIT 1",
        (email, code_hash)
    ).fetchone()
    if not row:
        db_conn.close()
        return False
    if datetime.fromisoformat(row["expires_at"]) < datetime.now():
        db_conn.close()
        return False
    db_conn.execute("UPDATE verification_codes SET used=1 WHERE id=?", (row["id"],))
    db_conn.execute("UPDATE users SET email_verified=1 WHERE email=?", (email,))
    db_conn.commit()
    db_conn.close()
    return True

# ─── AUTH ──────────────────────────────────────────────────────
def generate_csrf_token():
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_hex(32)
    return session['_csrf_token']

app.jinja_env.globals['csrf_token'] = generate_csrf_token

def csrf_field():
    return f'<input type="hidden" name="_csrf_token" value="{generate_csrf_token()}">'

@app.before_request
def csrf_protect():
    if request.method == "POST":
        # Endpoints externos (webhooks de Meta/Mercado Pago) validam autenticidade por outros meios
        public_webhooks = [
            '/webhook/',
            '/api/mercadopago/webhook',
        ]
        if any(request.path.startswith(p) for p in public_webhooks):
            return

        # Verifica CSRF token
        token = request.form.get('_csrf_token') or request.headers.get('X-CSRF-Token')
        if not token or token != session.get('_csrf_token'):
            abort(403)

def log_webhook_error(source, user_id, error_type, error_message, payload=None):
    """Registra erro de webhook para revisão posterior no admin.
    Não falha se o DB estiver down — só imprime no log."""
    try:
        preview = _safe_payload_preview(payload)
        db_conn = sqlite3.connect(DATABASE)
        db_conn.execute(
            """INSERT INTO webhook_errors
               (user_id, source, error_type, error_message, payload_preview)
               VALUES (?,?,?,?,?)""",
            (user_id, source, error_type, str(error_message)[:500], preview)
        )
        db_conn.commit()
        db_conn.close()
    except Exception as e:
        safe_log(f"[WEBHOOK ERROR LOG] Falhou: {e}", level="ERROR")


# ════════════════════════════════════════════════════════════════
#  Política de senhas (LGPD/segurança)
# ════════════════════════════════════════════════════════════════
# Senhas curtas (6 chars) eram permitidas — agora exige mínimo razoável
# com requisitos de complexidade que reduzem dicionários comuns.
def validate_password_strength(password):
    """Valida que a senha atende à política mínima. Retorna (ok: bool, msg: str)."""
    if not password:
        return False, "Senha é obrigatória."
    if len(password) < 10:
        return False, "Senha deve ter pelo menos 10 caracteres."
    if len(password) > 128:
        return False, "Senha não pode ter mais de 128 caracteres."
    has_lower = any(c.islower() for c in password)
    has_upper = any(c.isupper() for c in password)
    has_digit = any(c.isdigit() for c in password)
    # Combinação razoável: pelo menos 3 dos 4 (lower, upper, digit, symbol)
    has_symbol = any(not c.isalnum() for c in password)
    score = sum([has_lower, has_upper, has_digit, has_symbol])
    if score < 3:
        return False, "Senha precisa misturar pelo menos 3 de: letras maiúsculas, minúsculas, números e símbolos."
    # Senhas óbvias proibidas
    common = {"senha", "password", "123456", "qwerty", "admin", "abc123", "atendente", "whatsapp"}
    if password.lower() in common:
        return False, "Senha muito comum, escolha uma diferente."
    return True, ""


# Iterações PBKDF2-SHA256 — OWASP 2023 recomenda ≥600k.
# Formato novo do hash: "pbkdf2$<iters>$<salt_hex>$<hash_hex>".
# Formato legado mantido para compatibilidade: "<salt_hex>:<hash_hex>" (100k iters fixas).
PBKDF2_ITERATIONS = 600000


def hash_password(pw, iterations=PBKDF2_ITERATIONS):
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), iterations)
    return f"pbkdf2${iterations}${salt}${h.hex()}"


def _parse_stored_hash(stored):
    """Retorna (iters, salt, hash_hex) ou None se formato inválido."""
    if not stored or not isinstance(stored, str):
        return None
    if stored.startswith("pbkdf2$"):
        try:
            _, iters_s, salt, h = stored.split("$", 3)
            return int(iters_s), salt, h
        except (ValueError, AttributeError):
            return None
    if ":" in stored:
        try:
            salt, h = stored.split(":", 1)
            return 100000, salt, h
        except (ValueError, AttributeError):
            return None
    return None


def needs_password_rehash(stored):
    """True se o hash usa parâmetros abaixo do padrão atual (precisa re-hash no próximo login)."""
    parsed = _parse_stored_hash(stored)
    if not parsed:
        return False
    iters, _, _ = parsed
    return iters < PBKDF2_ITERATIONS


def validate_and_normalize_image(image_bytes, max_width=2048, max_height=2048):
    """Valida que os bytes são uma imagem real usando Pillow + normaliza tamanho.
    Retorna (bytes_normalizados, content_type) ou (None, None) se inválido.
    Protege contra upload de arquivos maliciosos disfarçados de imagem."""
    try:
        from PIL import Image
        import io as io_mod

        # Abre e valida que é imagem real (não só pelo mime type)
        img = Image.open(io_mod.BytesIO(image_bytes))
        img.verify()  # Primeira passada — detecta corrupção

        # Reabre (verify consome o stream)
        img = Image.open(io_mod.BytesIO(image_bytes))

        # Remove metadados EXIF potencialmente sensíveis
        img_format = img.format
        if img_format not in ("JPEG", "PNG", "WEBP"):
            return None, None

        # Redimensiona se muito grande
        if img.width > max_width or img.height > max_height:
            img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)

        # Converte para RGB se for RGBA (para JPEG)
        if img_format == "JPEG" and img.mode in ("RGBA", "LA", "P"):
            img = img.convert("RGB")

        # Re-salva limpo (sem EXIF, sem exploits)
        output = io_mod.BytesIO()
        save_format = "JPEG" if img_format == "JPEG" else "PNG"
        save_kwargs = {"quality": 85, "optimize": True} if save_format == "JPEG" else {"optimize": True}
        img.save(output, format=save_format, **save_kwargs)
        output.seek(0)

        content_type = f"image/{save_format.lower()}"
        return output.read(), content_type
    except Exception as e:
        safe_log(f"[IMAGE VALIDATION] Rejeitado: {e}", level="ERROR")
        return None, None


def check_password(pw, stored):
    parsed = _parse_stored_hash(stored)
    if not parsed:
        return False
    iters, salt, h = parsed
    try:
        computed = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), iters).hex()
        import hmac as hmac_mod
        return hmac_mod.compare_digest(computed, h)
    except (ValueError, AttributeError, TypeError):
        return False


def maybe_upgrade_password_hash(user_id, password, stored):
    """Re-hash transparente: se o hash armazenado usa params antigos, recalcula com PBKDF2_ITERATIONS atual.
    Chame APÓS check_password retornar True. Falha silenciosamente — segurança best-effort, não bloqueia o login."""
    if not needs_password_rehash(stored):
        return
    try:
        new_hash = hash_password(password)
        db = get_db()
        db.execute("UPDATE users SET password_hash=? WHERE id=?", (new_hash, user_id))
        db.commit()
        safe_log(f"[AUTH] Hash de senha atualizado (user_id={user_id})", level="INFO")
    except Exception as e:
        safe_log(f"[AUTH] Falha ao atualizar hash de senha (user_id={user_id}): {e}", level="WARN")

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect("/login")
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE id=?", (session["user_id"],)).fetchone()
        if not user:
            session.clear()
            return redirect("/login")
        if not dict(user).get("is_active", 1):
            session.clear()
            return redirect("/login?error=Conta+desativada.+Entre+em+contato+com+o+suporte.")
        # Descriptografa tokens sensíveis transparentemente
        g.user = decrypt_user_row(user)

        # LGPD: forçar re-aceite de termos atualizados (Sprint 2)
        # Whitelist: rotas que devem funcionar mesmo sem aceite atualizado
        # (a própria página de aceite, logout, exclusão de conta, exportação, documentos legais)
        path = request.path or ""
        lgpd_whitelist = (
            "/conta/aceitar-termos",
            "/conta/excluir",
            "/conta/exportar",
            "/logout",
            "/privacy",
            "/terms",
            "/dpa",
            "/dpo",
        )
        if not any(path.startswith(p) for p in lgpd_whitelist):
            if user_needs_to_reaccept_terms(session["user_id"]):
                return redirect("/conta/aceitar-termos")

        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect("/admin/login")
        return f(*args, **kwargs)
    return decorated

# ─── MEDIA HANDLING ────────────────────────────────────────────

def download_whatsapp_media(media_id, token):
    """Baixa mídia do WhatsApp e retorna o caminho local"""
    try:
        import requests as req
        # Passo 1: pegar a URL do media
        url = f"https://graph.facebook.com/v18.0/{media_id}"
        headers = {"Authorization": f"Bearer {token}"}
        resp = req.get(url, headers=headers, timeout=10)
        if resp.status_code != 200:
            return None
        media_url = resp.json().get("url")
        
        # Passo 2: baixar o arquivo
        resp2 = req.get(media_url, headers=headers, timeout=30)
        if resp2.status_code != 200:
            return None
        
        # Salvar localmente
        content_type = resp2.headers.get("Content-Type", "")
        ext = ".bin"
        if "image" in content_type: ext = ".jpg"
        elif "audio" in content_type or "ogg" in content_type: ext = ".ogg"
        elif "pdf" in content_type: ext = ".pdf"
        elif "document" in content_type: ext = ".doc"
        elif "video" in content_type: ext = ".mp4"
        
        filename = f"{media_id}{ext}"
        filepath = os.path.join(MEDIA_FOLDER, filename)
        with open(filepath, "wb") as f:
            f.write(resp2.content)
        return filepath
    except Exception as e:
        safe_log(f"Erro ao baixar mídia: {e}", level="ERROR")
        return None


def transcribe_audio(filepath):
    """Transcreve áudio usando Groq (primário) ou OpenAI Whisper (fallback)"""
    
    groq_key = get_setting("GROQ_API_KEY")
    openai_key = get_setting("OPENAI_API_KEY")
    
    # Tenta Groq primeiro (mais barato)
    if groq_key:
        try:
            import requests as req
            url = "https://api.groq.com/openai/v1/audio/transcriptions"
            headers = {"Authorization": f"Bearer {groq_key}"}
            with open(filepath, "rb") as audio_file:
                files = {"file": (os.path.basename(filepath), audio_file)}
                data = {"model": "whisper-large-v3", "language": "pt"}
                resp = req.post(url, headers=headers, files=files, data=data, timeout=60)
            if resp.status_code == 200:
                text = resp.json().get("text", "")
                safe_log(f"[GROQ] Áudio transcrito: {text[:80]}...")
                return text if text else "[Não foi possível transcrever]"
            else:
                safe_log(f"Groq API error: {resp.status_code} {_short_resp_text(resp)}", level="ERROR")
        except Exception as e:
            safe_log(f"Groq transcription error: {e}", level="ERROR")
    
    # Fallback para OpenAI
    if openai_key:
        try:
            import requests as req
            url = "https://api.openai.com/v1/audio/transcriptions"
            headers = {"Authorization": f"Bearer {openai_key}"}
            with open(filepath, "rb") as audio_file:
                files = {"file": (os.path.basename(filepath), audio_file)}
                data = {"model": "whisper-1", "language": "pt"}
                resp = req.post(url, headers=headers, files=files, data=data, timeout=60)
            if resp.status_code == 200:
                text = resp.json().get("text", "")
                safe_log(f"[OPENAI] Áudio transcrito: {text[:80]}...")
                return text if text else "[Não foi possível transcrever]"
            else:
                safe_log(f"Whisper API error: {resp.status_code} {_short_resp_text(resp)}", level="ERROR")
        except Exception as e:
            safe_log(f"OpenAI transcription error: {e}", level="ERROR")
    
    safe_log("[AUDIO] Nenhuma API de transcrição configurada")
    return "[Transcrição indisponível — configure GROQ_API_KEY no painel admin]"


def analyze_image_with_claude(filepath, user_question=""):
    """Analisa imagem usando Claude Vision"""
    api_key = get_setting("ANTHROPIC_API_KEY")
    if not api_key:
        return "[Análise de imagem indisponível — configure ANTHROPIC_API_KEY no painel admin]"
    try:
        import requests as req
        with open(filepath, "rb") as f:
            image_data = base64.b64encode(f.read()).decode("utf-8")
        
        ext = os.path.splitext(filepath)[1].lower()
        media_type = {"jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
                      ".gif": "image/gif", ".webp": "image/webp"}.get(ext, "image/jpeg")
        
        prompt = user_question if user_question else "Descreva esta imagem em detalhes. Se for um produto, diga o que é. Se tiver texto, transcreva."
        
        resp = req.post("https://api.anthropic.com/v1/messages",
            headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={
                "model": "claude-sonnet-4-6",
                "max_tokens": 500,
                "messages": [{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_data}},
                    {"type": "text", "text": prompt}
                ]}]
            }, timeout=30)
        
        if resp.status_code == 200:
            return resp.json()["content"][0]["text"]
        return "[Não foi possível analisar a imagem]"
    except Exception as e:
        safe_log(f"Image analysis error: {e}", level="ERROR")
        return "[Erro ao analisar imagem]"


def extract_pdf_text(filepath, max_pages=100):
    """Extrai texto de PDF (até 100 páginas para treinamento)"""
    try:
        try:
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                text = ""
                for page in pdf.pages[:max_pages]:
                    text += (page.extract_text() or "") + "\n"
                return text.strip() if text.strip() else "[PDF sem texto extraível]"
        except ImportError:
            pass

        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(filepath)
            text = ""
            for page in reader.pages[:max_pages]:
                text += (page.extract_text() or "") + "\n"
            return text.strip() if text.strip() else "[PDF sem texto extraível]"
        except ImportError:
            pass

        return "[Instale pdfplumber ou PyPDF2 para ler PDFs: pip install pdfplumber]"
    except Exception as e:
        safe_log(f"PDF extraction error: {e}", level="ERROR")
        return "[Erro ao extrair texto do PDF]"


def extract_spreadsheet_text(file_obj):
    """Extrai texto de planilha XLSX/XLS"""
    try:
        # Tenta openpyxl primeiro (XLSX)
        try:
            from openpyxl import load_workbook
            import io
            content = file_obj.read()
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            text_parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_parts.append(f"\n=== Aba: {sheet_name} ===")
                for row in ws.iter_rows(values_only=True, max_row=1000):
                    row_text = " | ".join(str(c) if c is not None else "" for c in row)
                    if row_text.strip(" |"):  # Pula linhas vazias
                        text_parts.append(row_text)
            return "\n".join(text_parts)
        except ImportError:
            pass

        # Fallback: xlrd (XLS legado)
        try:
            import xlrd
            import io
            content = file_obj.read() if hasattr(file_obj, "read") else file_obj
            wb = xlrd.open_workbook(file_contents=content)
            text_parts = []
            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                text_parts.append(f"\n=== Aba: {sheet_name} ===")
                for row_idx in range(min(sheet.nrows, 1000)):
                    row = [str(sheet.cell_value(row_idx, c)) for c in range(sheet.ncols)]
                    text_parts.append(" | ".join(row))
            return "\n".join(text_parts)
        except ImportError:
            pass

        return "[Instale openpyxl para ler planilhas: pip install openpyxl]"
    except Exception as e:
        safe_log(f"Spreadsheet extraction error: {e}", level="ERROR")
        return f"[Erro ao extrair planilha: {str(e)}]"


def process_whatsapp_media(msg, token):
    """Processa qualquer tipo de mídia recebida no WhatsApp"""
    msg_type = msg.get("type", "text")
    result = {"type": msg_type, "content": "", "description": "", "media_path": ""}
    
    if msg_type == "text":
        result["content"] = msg.get("text", {}).get("body", "")
        result["description"] = result["content"]
        
    elif msg_type == "audio":
        media_id = msg.get("audio", {}).get("id", "")
        if media_id:
            filepath = download_whatsapp_media(media_id, token)
            if filepath:
                result["media_path"] = filepath
                transcription = transcribe_audio(filepath)
                result["content"] = f"🎤 [Áudio transcrito]: {transcription}"
                result["description"] = transcription
            else:
                result["content"] = "🎤 [Áudio recebido — não foi possível baixar]"
                result["description"] = result["content"]
                
    elif msg_type == "image":
        media_id = msg.get("image", {}).get("id", "")
        caption = msg.get("image", {}).get("caption", "")
        if media_id:
            filepath = download_whatsapp_media(media_id, token)
            if filepath:
                result["media_path"] = filepath
                analysis = analyze_image_with_claude(filepath, caption)
                caption_text = f' (legenda: "{caption}")' if caption else ""
                result["content"] = f"📷 [Imagem recebida{caption_text}]: {analysis}"
                result["description"] = analysis
            else:
                result["content"] = "📷 [Imagem recebida — não foi possível baixar]"
                result["description"] = result["content"]
                
    elif msg_type == "document":
        media_id = msg.get("document", {}).get("id", "")
        filename = msg.get("document", {}).get("filename", "documento")
        mime = msg.get("document", {}).get("mime_type", "")
        if media_id:
            filepath = download_whatsapp_media(media_id, token)
            if filepath:
                result["media_path"] = filepath
                if "pdf" in mime:
                    text = extract_pdf_text(filepath)
                    result["content"] = f"📄 [PDF: {filename}]: {text[:2000]}"
                    result["description"] = text[:2000]
                else:
                    result["content"] = f"📄 [Documento: {filename}] — recebido e salvo"
                    result["description"] = f"Documento {filename} recebido"
            else:
                result["content"] = f"📄 [Documento: {filename}] — não foi possível baixar"
                result["description"] = result["content"]
                
    elif msg_type == "video":
        media_id = msg.get("video", {}).get("id", "")
        caption = msg.get("video", {}).get("caption", "")
        result["content"] = f"🎥 [Vídeo recebido]{': ' + caption if caption else ''}"
        result["description"] = result["content"]
        if media_id:
            filepath = download_whatsapp_media(media_id, token)
            if filepath:
                result["media_path"] = filepath
                
    elif msg_type == "location":
        lat = msg.get("location", {}).get("latitude", "")
        lon = msg.get("location", {}).get("longitude", "")
        loc_name = msg.get("location", {}).get("name", "")
        address = msg.get("location", {}).get("address", "")
        loc_text = f"📍 Localização: {loc_name} {address}".strip() if loc_name or address else f"📍 Localização: {lat}, {lon}"
        result["content"] = loc_text
        result["description"] = loc_text
        
    elif msg_type == "contacts":
        contacts = msg.get("contacts", [])
        names = [c.get("name", {}).get("formatted_name", "?") for c in contacts]
        result["content"] = f"👤 [Contato(s) compartilhado(s)]: {', '.join(names)}"
        result["description"] = result["content"]
        
    elif msg_type == "sticker":
        result["content"] = "😀 [Sticker recebido]"
        result["description"] = "Cliente enviou um sticker"
        
    elif msg_type == "reaction":
        emoji = msg.get("reaction", {}).get("emoji", "")
        result["content"] = f"[Reação: {emoji}]"
        result["description"] = result["content"]
        
    else:
        result["content"] = f"[{msg_type}] Tipo de mensagem não suportado"
        result["description"] = result["content"]
    
    return result


# ─── CSS GLOBAL ────────────────────────────────────────────────

GLOBAL_CSS = """
:root {
    --bg:#0a0e14; --bg2:#111827; --bg3:#1a2235; --bg4:#243049;
    --text:#f0f4f8; --text2:#94a3b8; --text3:#64748b;
    --accent:#00c896; --accent2:#34d399; --accent-glow:rgba(0,200,150,0.12);
    --green:#00b894; --green2:#00e6b0; --red:#ef4444; --orange:#f59e0b; --blue:#0ea5e9;
    --radius:12px; --radius-sm:8px;
    --font:'DM Sans',-apple-system,sans-serif; --mono:'JetBrains Mono',monospace;
}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:var(--font);background:var(--bg);color:var(--text);min-height:100vh;-webkit-font-smoothing:antialiased}
a{color:var(--accent2);text-decoration:none}
a:hover{color:#5eead4}

.nav-main{background:rgba(10,14,20,0.9);border-bottom:1px solid rgba(255,255,255,0.06);position:sticky;top:0;z-index:100;backdrop-filter:blur(24px);-webkit-backdrop-filter:blur(24px)}
.nav-inner{max-width:1400px;margin:0 auto;padding:0 20px;height:68px;display:flex;align-items:center;justify-content:space-between;gap:12px}
.logo{font-size:22px;font-weight:700;color:var(--text);letter-spacing:-0.5px}
.logo span{color:var(--accent)}
.nav-logo-img{height:44px;width:auto;display:block;transition:transform 0.2s;flex-shrink:0}
.nav-logo-img:hover{transform:scale(1.03)}
.nav-links{display:flex;gap:2px;flex:1;justify-content:center;overflow-x:auto;overflow-y:hidden;scrollbar-width:none;-ms-overflow-style:none}
.nav-links::-webkit-scrollbar{display:none}
.nav-link{padding:7px 12px;border-radius:var(--radius-sm);color:var(--text2);font-size:13px;font-weight:500;transition:all 0.2s;white-space:nowrap;flex-shrink:0}
.nav-link:hover{color:var(--text);background:var(--bg3)}
.nav-link-accent{color:var(--accent2)!important}
.nav-user{display:flex;align-items:center;gap:8px;font-size:12px;flex-shrink:0}
.user-plan{background:var(--accent-glow);color:var(--accent2);padding:3px 8px;border-radius:20px;font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px}
.user-name{color:var(--text2);font-size:12px;max-width:120px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.btn-logout{color:var(--text3);font-size:11px;padding:3px 6px}

/* Mobile toggle (menu hamburguer) */
.nav-toggle{display:none;background:none;border:none;color:var(--text);font-size:24px;cursor:pointer;padding:4px 8px}

@media(max-width:1200px){
    .nav-link{padding:7px 10px;font-size:12px}
    .user-name{display:none}
}
@media(max-width:900px){
    .nav-inner{padding:0 16px;height:60px}
    .nav-logo-img{height:36px}
    .nav-links{display:none;position:absolute;top:60px;left:0;right:0;background:var(--bg);border-bottom:1px solid rgba(255,255,255,0.08);padding:12px;flex-direction:column;box-shadow:0 8px 24px rgba(0,0,0,0.4);overflow-x:hidden}
    .nav-links.open{display:flex}
    .nav-link{padding:12px 16px;font-size:14px;width:100%;text-align:left}
    .nav-toggle{display:block}
    .user-plan{display:none}
}

.container{max-width:1200px;margin:0 auto;padding:32px 24px}
.page-header{margin-bottom:32px}
.page-header h1{font-size:28px;font-weight:700;letter-spacing:-0.5px}
.page-header p{color:var(--text2);margin-top:4px;font-size:15px}

.card{background:var(--bg2);border:1px solid rgba(255,255,255,0.06);border-radius:var(--radius);padding:24px}
.card-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:16px}
.card-title{font-size:16px;font-weight:600}

.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-bottom:32px}
.grid-3{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;margin-bottom:32px}
.grid-4{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:16px;margin-bottom:32px}
.grid-5{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px;margin-bottom:32px}

.stat-card{background:var(--bg2);border:1px solid rgba(255,255,255,0.06);border-radius:var(--radius);padding:20px 24px}
.stat-card .stat-icon{width:40px;height:40px;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:18px;margin-bottom:12px}
.stat-card .stat-value{font-size:28px;font-weight:700;letter-spacing:-1px}
.stat-card .stat-label{color:var(--text2);font-size:13px;margin-top:2px}
.stat-icon-green{background:rgba(0,184,148,0.15);color:var(--green2)}
.stat-icon-blue{background:rgba(9,132,227,0.15);color:var(--blue)}
.stat-icon-purple{background:var(--accent-glow);color:var(--accent2)}
.stat-icon-orange{background:rgba(243,156,18,0.15);color:var(--orange)}
.stat-icon-red{background:rgba(231,76,60,0.15);color:var(--red)}

.btn{display:inline-flex;align-items:center;gap:8px;padding:10px 20px;border-radius:var(--radius-sm);font-size:14px;font-weight:500;border:none;cursor:pointer;transition:all 0.2s;font-family:var(--font)}
.btn-primary{background:var(--accent);color:white}
.btn-primary:hover{background:#00a87d;transform:translateY(-1px);box-shadow:0 4px 20px rgba(0,200,150,0.3)}
.btn-secondary{background:var(--bg3);color:var(--text);border:1px solid rgba(255,255,255,0.08)}
.btn-secondary:hover{background:var(--bg4)}
.btn-success{background:var(--green);color:white}
.btn-danger{background:var(--red);color:white}
.btn-sm{padding:6px 14px;font-size:13px}
.btn-lg{padding:14px 28px;font-size:16px}
.btn-block{width:100%;justify-content:center}

.form-group{margin-bottom:20px}
.form-label{display:block;font-size:13px;font-weight:500;color:var(--text2);margin-bottom:6px;text-transform:uppercase;letter-spacing:0.5px}
.form-input{width:100%;padding:12px 16px;background:var(--bg3);border:1px solid rgba(255,255,255,0.08);border-radius:var(--radius-sm);color:var(--text);font-size:14px;font-family:var(--font);transition:border-color 0.2s}
.form-input:focus{outline:none;border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-glow)}
textarea.form-input{min-height:120px;resize:vertical;line-height:1.6}

.table-wrap{overflow-x:auto}
table{width:100%;border-collapse:collapse}
th{text-align:left;padding:12px 16px;font-size:12px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:0.5px;border-bottom:1px solid rgba(255,255,255,0.06)}
td{padding:14px 16px;border-bottom:1px solid rgba(255,255,255,0.04);font-size:14px}
tr:hover td{background:rgba(255,255,255,0.02)}

.badge{display:inline-flex;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.3px}
.badge-green{background:rgba(0,184,148,0.15);color:var(--green2)}
.badge-orange{background:rgba(243,156,18,0.15);color:var(--orange)}
.badge-red{background:rgba(231,76,60,0.15);color:var(--red)}
.badge-purple{background:var(--accent-glow);color:var(--accent2)}
.badge-blue{background:rgba(9,132,227,0.15);color:var(--blue)}

.plan-card{background:var(--bg2);border:1px solid rgba(255,255,255,0.06);border-radius:var(--radius);padding:32px;text-align:center;transition:all 0.3s;position:relative}
.plan-card:hover{border-color:var(--accent);transform:translateY(-4px);box-shadow:0 8px 32px rgba(0,200,150,0.12)}
.plan-card.popular{border-color:var(--accent)}
.plan-card.popular::before{content:'MAIS POPULAR';position:absolute;top:-12px;left:50%;transform:translateX(-50%);background:var(--accent);color:white;padding:4px 16px;border-radius:20px;font-size:10px;font-weight:700;letter-spacing:1px}
.plan-name{font-size:20px;font-weight:700;margin-bottom:8px}
.plan-price{font-size:40px;font-weight:700;color:var(--accent2);margin:16px 0}
.plan-price small{font-size:16px;color:var(--text2);font-weight:400}
.plan-desc{color:var(--text2);font-size:14px;margin-bottom:20px}
.plan-features{list-style:none;text-align:left;margin-bottom:24px}
.plan-features li{padding:8px 0;font-size:14px;color:var(--text2);border-bottom:1px solid rgba(255,255,255,0.04)}
.plan-features li::before{content:'✓';color:var(--green2);margin-right:8px;font-weight:700}

.chat-container{display:flex;height:calc(100vh - 160px);gap:0;background:var(--bg2);border-radius:var(--radius);overflow:hidden;border:1px solid rgba(255,255,255,0.06)}
.chat-sidebar{width:320px;border-right:1px solid rgba(255,255,255,0.06);overflow-y:auto}
.chat-sidebar-header{padding:20px;border-bottom:1px solid rgba(255,255,255,0.06)}
.chat-item{padding:16px 20px;border-bottom:1px solid rgba(255,255,255,0.04);cursor:pointer;transition:background 0.2s}
.chat-item:hover{background:var(--bg3)}
.chat-item.active{background:var(--accent-glow);border-left:3px solid var(--accent)}
.chat-item-name{font-weight:600;font-size:14px}
.chat-item-preview{color:var(--text3);font-size:13px;margin-top:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.chat-item-time{color:var(--text3);font-size:11px;float:right}
.chat-main{flex:1;display:flex;flex-direction:column}
.chat-header{padding:16px 24px;border-bottom:1px solid rgba(255,255,255,0.06);display:flex;justify-content:space-between;align-items:center}
.chat-messages{flex:1;overflow-y:auto;padding:24px;display:flex;flex-direction:column;gap:12px}
.msg{max-width:70%;padding:12px 16px;border-radius:16px;font-size:14px;line-height:1.5}
.msg-customer{background:var(--bg4);align-self:flex-start;border-bottom-left-radius:4px}
.msg-bot{background:var(--accent);color:white;align-self:flex-end;border-bottom-right-radius:4px}
.msg-time{font-size:10px;opacity:0.6;margin-top:4px}
.msg-media{font-size:12px;opacity:0.8;font-style:italic}

.auth-container{min-height:100vh;display:flex;align-items:center;justify-content:center;padding:24px}
.auth-card{background:var(--bg2);border:1px solid rgba(255,255,255,0.06);border-radius:16px;padding:40px;width:100%;max-width:420px;box-shadow:0 4px 24px rgba(0,0,0,0.3)}
.auth-card .logo{font-size:28px;text-align:center;display:block;margin-bottom:32px}
.auth-card h2{font-size:22px;margin-bottom:24px;text-align:center}
.auth-divider{text-align:center;color:var(--text3);font-size:13px;margin:20px 0}

.hero{text-align:center;padding:80px 24px 40px;max-width:800px;margin:0 auto}
.hero h1{font-size:48px;font-weight:700;letter-spacing:-1.5px;line-height:1.1;margin-bottom:20px}
.hero h1 .gradient{background:linear-gradient(135deg,#00c896,#0ea5e9);-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.hero p{font-size:18px;color:var(--text2);max-width:560px;margin:0 auto 32px;line-height:1.6}
.hero-badges{display:flex;gap:12px;justify-content:center;margin-bottom:40px;flex-wrap:wrap}
.hero-badge{display:flex;align-items:center;gap:6px;padding:6px 14px;background:rgba(0,200,150,0.08);border:1px solid rgba(0,200,150,0.15);border-radius:20px;font-size:13px;color:var(--accent2);font-weight:500}
.features-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;max-width:1000px;margin:0 auto 80px;padding:0 24px}
.feature-card{background:var(--bg2);border:1px solid rgba(255,255,255,0.06);border-radius:var(--radius);padding:28px;transition:all 0.3s;position:relative;overflow:hidden}
.feature-card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,transparent,var(--accent),transparent);opacity:0;transition:opacity 0.3s}
.feature-card:hover{border-color:rgba(0,200,150,0.2);transform:translateY(-3px);box-shadow:0 8px 32px rgba(0,0,0,0.2)}
.feature-card:hover::before{opacity:1}
.feature-icon{font-size:28px;margin-bottom:14px;width:48px;height:48px;display:flex;align-items:center;justify-content:center;background:rgba(0,200,150,0.08);border-radius:12px}
.feature-card h3{font-size:17px;font-weight:600;margin-bottom:8px}
.feature-card p{font-size:14px;color:var(--text2);line-height:1.6}

.alert{padding:14px 20px;border-radius:var(--radius-sm);margin-bottom:20px;font-size:14px}
.alert-success{background:rgba(0,184,148,0.1);border:1px solid rgba(0,184,148,0.2);color:var(--green2)}
.alert-error{background:rgba(231,76,60,0.1);border:1px solid rgba(231,76,60,0.2);color:var(--red)}
.alert-info{background:var(--accent-glow);border:1px solid rgba(108,92,231,0.2);color:var(--accent2)}

.empty-state{text-align:center;padding:60px 24px;color:var(--text3)}
.empty-state .icon{font-size:48px;margin-bottom:16px}
.empty-state h3{color:var(--text2);margin-bottom:8px}

.usage-bar-bg{background:var(--bg4);border-radius:20px;height:8px;overflow:hidden}
.usage-bar-fill{height:100%;border-radius:20px;transition:width 0.5s ease}

/* ADMIN SPECIFIC */
.admin-nav{background:linear-gradient(135deg,#0a1628,#0a0e14);border-bottom:2px solid var(--accent)}
.admin-badge{background:var(--red);color:white;padding:3px 8px;border-radius:4px;font-size:10px;font-weight:700;letter-spacing:1px;margin-left:8px}
.metric-card{background:var(--bg2);border:1px solid rgba(255,255,255,0.06);border-radius:var(--radius);padding:24px;text-align:center}
.metric-value{font-size:32px;font-weight:700;margin:8px 0 4px}
.metric-label{font-size:13px;color:var(--text2)}
.metric-trend{font-size:12px;margin-top:4px}
.trend-up{color:var(--green2)}
.trend-down{color:var(--red)}

@keyframes fadeIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
.fade-in{animation:fadeIn 0.4s ease-out forwards}
.fade-in-1{animation-delay:0.1s;opacity:0}
.fade-in-2{animation-delay:0.2s;opacity:0}
.fade-in-3{animation-delay:0.3s;opacity:0}
.fade-in-4{animation-delay:0.4s;opacity:0}

@media(max-width:768px){
    .grid-2,.grid-3,.grid-4,.grid-5{grid-template-columns:1fr}
    .features-grid{grid-template-columns:1fr}
    .hero h1{font-size:32px}
    .chat-sidebar{width:100%}
}
"""

# ─── HTML BUILDERS ─────────────────────────────────────────────

def base_html(title, content, user=None):
    nav = ""
    if user:
        plan_name = PLANS.get(user['plan'],{}).get('name','')
        nav = f"""<nav class="nav-main"><div class="nav-inner">
            <a href="/dashboard"><img src="data:image/png;base64,{LOGO_NAV_B64}" alt="atendente.online" class="nav-logo-img"></a>
            <button class="nav-toggle" id="nav-toggle-btn" aria-label="Menu">☰</button>
            <div class="nav-links">
                <a href="/dashboard" class="nav-link">Dashboard</a>
                <a href="/dashboard/conversations" class="nav-link">Conversas</a>
                <a href="/dashboard/training" class="nav-link">Treino</a>
                <a href="/dashboard/gallery" class="nav-link">Galeria</a>
                <a href="/dashboard/commerce" class="nav-link">🛒 Vendas</a>
                <a href="/dashboard/campaigns" class="nav-link">📢 Campanhas</a>
                <a href="/dashboard/contacts" class="nav-link">Contatos</a>
                <a href="/dashboard/pipeline" class="nav-link">Funil</a>
                <a href="/dashboard/social" class="nav-link">📸 Agência</a>
                <a href="/dashboard/settings" class="nav-link">Config</a>
                <a href="/dashboard/billing" class="nav-link nav-link-accent">Plano</a>
            </div>
            <div class="nav-user">
                <span class="user-plan">{plan_name}</span>
                <span class="user-name">{user['name']}</span>
                <a href="/conta/meus-dados" class="btn-logout" title="Meus dados (LGPD)" style="margin-right:6px">🛡️</a>
                <a href="/logout" class="btn-logout">Sair</a>
            </div></div></nav>"""
    return f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"><link rel="icon" type="image/png" href="/favicon.ico"><link rel="apple-touch-icon" href="/apple-touch-icon.png"><meta name="theme-color" content="#6366f1">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{title} — atendente.online</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>{GLOBAL_CSS}</style></head><body>{nav}{content}
<script nonce="{g.csp_nonce}">
// Hamburger menu toggle (CSP-compliant)
(function(){{
    var btn = document.getElementById('nav-toggle-btn');
    if(btn){{
        btn.addEventListener('click', function(){{
            var links = document.querySelector('.nav-links');
            if(links) links.classList.toggle('open');
        }});
    }}
}})();
</script>
</body></html>"""


def admin_html(title, content):
    nav = f"""<nav class="nav-main admin-nav"><div class="nav-inner">
        <a href="/admin" style="display:flex;align-items:center;gap:10px"><img src="data:image/png;base64,{LOGO_NAV_B64}" alt="atendente.online" class="nav-logo-img"><span class="admin-badge">ADMIN</span></a>
        <button class="nav-toggle" id="admin-nav-toggle-btn" aria-label="Menu">☰</button>
        <div class="nav-links">
            <a href="/admin" class="nav-link">Dashboard</a>
            <a href="/admin/users" class="nav-link">Clientes</a>
            <a href="/admin/payments" class="nav-link">Pagamentos</a>
            <a href="/admin/usage" class="nav-link">Uso API</a>
            <a href="/admin/logs" class="nav-link">Logs</a>
            <a href="/admin/audit-log" class="nav-link">📋 Auditoria</a>
            <a href="/admin/backups" class="nav-link">💾 Backups</a>
            <a href="/admin/recrypt-status" class="nav-link">🔐 Recriptografia</a>
            <a href="/admin/webhook-errors" class="nav-link">🚨 Erros</a>
            <a href="/admin/mp-debug" class="nav-link">🧪 MP Debug</a>
            <a href="/admin/2fa" class="nav-link">🔐 2FA</a>
            <a href="/admin/api-settings" class="nav-link nav-link-accent">APIs</a>
        </div>
        <div class="nav-user">
            <span class="user-plan" style="background:rgba(239,68,68,0.15);color:var(--red)">ADMIN</span>
            <a href="/admin/logout" class="btn-logout">Sair</a>
        </div></div></nav>"""
    return f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"><link rel="icon" type="image/png" href="/favicon.ico"><link rel="apple-touch-icon" href="/apple-touch-icon.png"><meta name="theme-color" content="#6366f1">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{title} — Admin Atende.AI</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>{GLOBAL_CSS}</style></head><body>{nav}{content}
<script nonce="{g.csp_nonce}">
// Admin hamburger toggle (CSP-compliant)
(function(){{
    var btn = document.getElementById('admin-nav-toggle-btn');
    if(btn){{
        btn.addEventListener('click', function(){{
            var links = document.querySelector('.admin-nav .nav-links');
            if(links) links.classList.toggle('open');
        }});
    }}
}})();
</script>
</body></html>"""


# ─── HELPER FUNCTIONS ──────────────────────────────────────────

def get_user_stats(user_id):
    db = get_db()
    convos = db.execute("SELECT COUNT(*) as c FROM conversations WHERE user_id=?", (user_id,)).fetchone()["c"]
    msgs = db.execute("SELECT COUNT(*) as c FROM messages m JOIN conversations c ON m.conversation_id=c.id WHERE c.user_id=?", (user_id,)).fetchone()["c"]
    today_msgs = db.execute("SELECT COUNT(*) as c FROM messages m JOIN conversations c ON m.conversation_id=c.id WHERE c.user_id=? AND m.created_at >= date('now')", (user_id,)).fetchone()["c"]
    kb = db.execute("SELECT COUNT(*) as c FROM knowledge_base WHERE user_id=?", (user_id,)).fetchone()["c"]
    return {"conversations": convos, "messages": msgs, "today_messages": today_msgs, "knowledge_items": kb}


def get_admin_stats():
    db = get_db()
    total_users = db.execute("SELECT COUNT(*) as c FROM users").fetchone()["c"]
    active_users = db.execute("SELECT COUNT(*) as c FROM users WHERE plan_status='active'").fetchone()["c"]
    trial_users = db.execute("SELECT COUNT(*) as c FROM users WHERE plan_status='trial'").fetchone()["c"]
    inactive_users = db.execute("SELECT COUNT(*) as c FROM users WHERE plan_status='inactive' OR plan_status='cancelled'").fetchone()["c"]
    
    # MRR
    mrr_rows = db.execute("SELECT plan FROM users WHERE plan_status='active'").fetchall()
    mrr = sum(PLANS.get(r["plan"], {}).get("price", 0) for r in mrr_rows)
    
    # Receita total
    total_revenue = db.execute("SELECT COALESCE(SUM(amount),0) as s FROM payments WHERE status='approved'").fetchone()["s"]
    
    # Conversas e mensagens totais
    total_conversations = db.execute("SELECT COUNT(*) as c FROM conversations").fetchone()["c"]
    total_messages = db.execute("SELECT COUNT(*) as c FROM messages").fetchone()["c"]
    
    # Hoje
    new_users_today = db.execute("SELECT COUNT(*) as c FROM users WHERE created_at >= date('now')").fetchone()["c"]
    msgs_today = db.execute("SELECT COUNT(*) as c FROM messages WHERE created_at >= date('now')").fetchone()["c"]
    payments_today = db.execute("SELECT COALESCE(SUM(amount),0) as s FROM payments WHERE status='approved' AND created_at >= date('now')").fetchone()["s"]
    
    # Por plano
    by_plan = {}
    for key in PLANS:
        count = db.execute("SELECT COUNT(*) as c FROM users WHERE plan=? AND plan_status='active'", (key,)).fetchone()["c"]
        by_plan[key] = count
    
    # Custo estimado de API
    total_api_cost = db.execute("SELECT COALESCE(SUM(cost_estimate),0) as s FROM api_usage_log").fetchone()["s"]
    
    return {
        "total_users": total_users, "active_users": active_users, "trial_users": trial_users,
        "inactive_users": inactive_users, "mrr": mrr, "total_revenue": total_revenue,
        "total_conversations": total_conversations, "total_messages": total_messages,
        "new_users_today": new_users_today, "msgs_today": msgs_today,
        "payments_today": payments_today, "by_plan": by_plan, "total_api_cost": total_api_cost
    }


# ═══════════════════════════════════════════════════════════════
#  ROTAS DO CLIENTE (mesmo de antes, com melhorias)
# ═══════════════════════════════════════════════════════════════

@app.route("/privacy")
def privacy_policy():
    content = f"""
    <div class="container" style="max-width:860px">
        <div class="card" style="margin-top:40px;padding:40px">
            <h1 style="font-size:28px;font-weight:700;margin-bottom:8px">Política de Privacidade</h1>
            <p style="color:var(--text2);margin-bottom:24px;font-size:14px">
                <strong>Versão:</strong> {PRIVACY_POLICY_VERSION} &nbsp;|&nbsp;
                <strong>Última atualização:</strong> 18 de maio de 2026
            </p>

            <div style="background:var(--bg2);padding:16px;border-radius:8px;margin-bottom:24px;font-size:14px;line-height:1.7">
                <strong>📌 Resumo executivo:</strong> O atendente.online é uma plataforma SaaS de atendimento WhatsApp com IA.
                Coletamos dados estritamente necessários para operar o serviço, criptografamos informações sensíveis,
                não vendemos dados a terceiros, e respeitamos integralmente a LGPD (Lei nº 13.709/2018).
            </div>

            <div style="color:var(--text2);font-size:15px;line-height:1.8">

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">1. Identificação do Controlador</h2>
                <p>O serviço atendente.online é operado por:</p>
                <p style="background:var(--bg2);padding:12px;border-radius:6px;margin:12px 0">
                    <strong>Clériston Almeida Capistrano</strong><br>
                    Pessoa física (CNPJ em processo de abertura)<br>
                    Endereço: Quixadá-CE, Brasil<br>
                    E-mail de contato: <a href="mailto:contato@atendente.online">contato@atendente.online</a>
                </p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">2. Encarregado pelo Tratamento de Dados (DPO)</h2>
                <p>Conforme exigido pelo Art. 41 da LGPD, indicamos como Encarregado pelo Tratamento de Dados Pessoais:</p>
                <p style="background:var(--bg2);padding:12px;border-radius:6px;margin:12px 0">
                    <strong>Nome:</strong> Clériston Almeida Capistrano<br>
                    <strong>E-mail:</strong> <a href="mailto:contato@atendente.online">contato@atendente.online</a><br>
                    <strong>Atribuições:</strong> Receber comunicações de titulares e da ANPD,
                    orientar sobre práticas de proteção de dados e executar demais atividades determinadas pelo controlador.
                </p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">3. Dados Pessoais Tratados</h2>

                <h3 style="color:var(--text);font-size:16px;margin:16px 0 8px">3.1. Dados dos Usuários (clientes do SaaS)</h3>
                <p>• <strong style="color:var(--text)">Cadastro:</strong> nome, email, empresa, telefone, senha (armazenada com hash PBKDF2-SHA256).</p>
                <p>• <strong style="color:var(--text)">Pagamento:</strong> processado integralmente pelo Mercado Pago; não armazenamos dados de cartão.</p>
                <p>• <strong style="color:var(--text)">Credenciais de integração:</strong> tokens da Meta, Mercado Pago, Telegram — armazenados criptografados (AES-128 + HMAC-SHA256, Fernet).</p>
                <p>• <strong style="color:var(--text)">Uso:</strong> data/hora de login, quantidade de mensagens processadas, IP, navegador.</p>

                <h3 style="color:var(--text);font-size:16px;margin:16px 0 8px">3.2. Dados de Clientes Finais (interlocutores do WhatsApp)</h3>
                <p>Quando seus clientes (terceiros) interagem com sua conta WhatsApp Business, processamos:</p>
                <p>• Número de telefone, nome de perfil (fornecido pelo WhatsApp).</p>
                <p>• Conteúdo das mensagens trocadas (texto, áudio, imagem, documentos).</p>
                <p>• Mensagens são armazenadas em banco de dados com acesso restrito, isolamento do volume e medidas de segurança da infraestrutura. Backups são criptografados.</p>
                <p style="background:#fff8e1;padding:12px;border-radius:6px;color:#5d4037;margin:8px 0">
                    ⚠️ <strong>Importante:</strong> Para esses dados, atuamos como <strong>Operador</strong> (LGPD Art. 5º, VII),
                    e você (usuário do atendente.online) é o <strong>Controlador</strong>. Você é responsável por obter
                    consentimento dos seus próprios clientes e cumprir as obrigações de controlador previstas na LGPD.
                </p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">4. Bases Legais para o Tratamento (LGPD Art. 7º e 11)</h2>
                <p>Tratamos seus dados com base nas seguintes hipóteses legais:</p>
                <p>• <strong style="color:var(--text)">Execução de contrato</strong> (Art. 7º, V): para prestar o serviço contratado.</p>
                <p>• <strong style="color:var(--text)">Cumprimento de obrigação legal</strong> (Art. 7º, II): para emissão de notas, registros contábeis.</p>
                <p>• <strong style="color:var(--text)">Consentimento</strong> (Art. 7º, I): para envio de comunicações de marketing (sempre opt-in).</p>
                <p>• <strong style="color:var(--text)">Legítimo interesse</strong> (Art. 7º, IX): para segurança da plataforma (logs, auditoria, prevenção de fraude).</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">5. Compartilhamento com Terceiros (Operadores)</h2>
                <p>Compartilhamos dados estritamente necessários com os seguintes operadores:</p>
                <table style="width:100%;border-collapse:collapse;margin:12px 0;font-size:14px">
                    <tr style="background:var(--bg2)">
                        <th style="padding:10px;border:1px solid var(--border);text-align:left">Operador</th>
                        <th style="padding:10px;border:1px solid var(--border);text-align:left">Finalidade</th>
                        <th style="padding:10px;border:1px solid var(--border);text-align:left">País</th>
                    </tr>
                    <tr>
                        <td style="padding:10px;border:1px solid var(--border)"><strong>Meta Platforms</strong></td>
                        <td style="padding:10px;border:1px solid var(--border)">Envio/recebimento de mensagens WhatsApp, Instagram, Messenger</td>
                        <td style="padding:10px;border:1px solid var(--border)">EUA/Irlanda</td>
                    </tr>
                    <tr>
                        <td style="padding:10px;border:1px solid var(--border)"><strong>Anthropic (Claude)</strong></td>
                        <td style="padding:10px;border:1px solid var(--border)">Geração de respostas com IA</td>
                        <td style="padding:10px;border:1px solid var(--border)">EUA</td>
                    </tr>
                    <tr>
                        <td style="padding:10px;border:1px solid var(--border)"><strong>Groq</strong></td>
                        <td style="padding:10px;border:1px solid var(--border)">Transcrição de áudio (Speech-to-Text)</td>
                        <td style="padding:10px;border:1px solid var(--border)">EUA</td>
                    </tr>
                    <tr>
                        <td style="padding:10px;border:1px solid var(--border)"><strong>Mercado Pago</strong></td>
                        <td style="padding:10px;border:1px solid var(--border)">Processamento de pagamentos</td>
                        <td style="padding:10px;border:1px solid var(--border)">Brasil/Argentina</td>
                    </tr>
                    <tr>
                        <td style="padding:10px;border:1px solid var(--border)"><strong>Railway</strong></td>
                        <td style="padding:10px;border:1px solid var(--border)">Hospedagem da aplicação e banco de dados</td>
                        <td style="padding:10px;border:1px solid var(--border)">EUA</td>
                    </tr>
                </table>
                <p>Transferência internacional ocorre conforme Art. 33 da LGPD, com adoção de cláusulas contratuais
                e medidas de segurança equivalentes às brasileiras.</p>
                <p><strong style="color:var(--text)">Não vendemos, alugamos ou cedemos dados pessoais a terceiros</strong> para fins de marketing,
                profiling de terceiros ou qualquer outra finalidade não declarada nesta política.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">6. Segurança da Informação</h2>
                <p>Adotamos medidas técnicas e administrativas alinhadas com padrões internacionais (ISO 27001 como referência):</p>
                <p>• <strong style="color:var(--text)">Em trânsito:</strong> HTTPS/TLS 1.2+ obrigatório em todas as comunicações.</p>
                <p>• <strong style="color:var(--text)">Em repouso — segredos:</strong> tokens de integração (Meta, Mercado Pago, Telegram), chaves de API e demais segredos do sistema são criptografados em nível de aplicação com Fernet (AES-128-CBC + HMAC-SHA256).</p>
                <p>• <strong style="color:var(--text)">Em repouso — mensagens e contatos:</strong> armazenados em banco de dados protegido por controles de acesso, isolamento do volume e medidas de segurança da infraestrutura de hospedagem. O acesso é restrito à própria conta do usuário e à equipe técnica autorizada.</p>
                <p>• <strong style="color:var(--text)">Senhas:</strong> hash PBKDF2-SHA256 com 600.000 iterações + salt aleatório (alinhado às recomendações OWASP 2023).</p>
                <p>• <strong style="color:var(--text)">Acesso administrativo:</strong> 2FA (TOTP) obrigatório + códigos de backup.</p>
                <p>• <strong style="color:var(--text)">Auditoria:</strong> registro de ações sensíveis (audit log).</p>
                <p>• <strong style="color:var(--text)">Proteção:</strong> rate limiting, CSRF tokens, CSP, HSTS, X-Frame-Options.</p>
                <p>• <strong style="color:var(--text)">Backups:</strong> diários, criptografados.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">7. Retenção de Dados</h2>
                <p>Aplicamos o princípio da necessidade (LGPD Art. 6º, III): dados são mantidos pelo tempo estritamente necessário.</p>
                <p>• <strong style="color:var(--text)">Mensagens de WhatsApp/Instagram/Messenger:</strong> conforme o plano contratado:</p>
                <p style="margin-left:20px">– Plano Starter: <strong>3 meses</strong></p>
                <p style="margin-left:20px">– Plano Profissional: <strong>6 meses</strong></p>
                <p style="margin-left:20px">– Plano Business: <strong>12 meses</strong></p>
                <p style="margin-left:20px">– Plano Agência: <strong>24 meses</strong></p>
                <p>• <strong style="color:var(--text)">Dados de cadastro:</strong> enquanto a conta estiver ativa + 30 dias após cancelamento.</p>
                <p>• <strong style="color:var(--text)">Pedidos e transações:</strong> 5 anos (obrigação legal — Código Tributário).</p>
                <p>• <strong style="color:var(--text)">Logs de auditoria:</strong> 6 meses para análise de segurança.</p>
                <p>• <strong style="color:var(--text)">Registros de consentimento:</strong> 5 anos para comprovação perante ANPD.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">8. Direitos do Titular (LGPD Art. 18)</h2>
                <p>Você tem direito a, mediante requisição:</p>
                <p>• <strong style="color:var(--text)">I.</strong> Confirmação da existência de tratamento;</p>
                <p>• <strong style="color:var(--text)">II.</strong> Acesso aos dados;</p>
                <p>• <strong style="color:var(--text)">III.</strong> Correção de dados incompletos, inexatos ou desatualizados;</p>
                <p>• <strong style="color:var(--text)">IV.</strong> Anonimização, bloqueio ou eliminação de dados desnecessários, excessivos ou tratados em desconformidade;</p>
                <p>• <strong style="color:var(--text)">V.</strong> Portabilidade dos dados (exportação em formato estruturado);</p>
                <p>• <strong style="color:var(--text)">VI.</strong> Eliminação dos dados pessoais tratados com consentimento;</p>
                <p>• <strong style="color:var(--text)">VII.</strong> Informação sobre entidades públicas e privadas com as quais compartilhamos dados;</p>
                <p>• <strong style="color:var(--text)">VIII.</strong> Informação sobre a possibilidade de não fornecer consentimento e suas consequências;</p>
                <p>• <strong style="color:var(--text)">IX.</strong> Revogação do consentimento.</p>
                <p style="background:var(--bg2);padding:12px;border-radius:6px;margin:12px 0">
                    💡 <strong>Como exercer:</strong> usuários logados podem acessar diretamente a página
                    <a href="/conta/meus-dados"><strong>Meus Dados</strong></a> para exportar, corrigir ou excluir.
                    Alternativamente, envie solicitação para <a href="mailto:contato@atendente.online">contato@atendente.online</a>
                    com prazo de resposta de até 15 dias úteis.
                </p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">9. Uso de Cookies e Tecnologias Similares</h2>
                <p>Utilizamos cookies estritamente necessários (sessão de login, CSRF token). Não utilizamos cookies de rastreamento
                comportamental de terceiros nem fornecemos dados para redes de publicidade.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">10. Incidentes de Segurança</h2>
                <p>Em caso de incidente que possa acarretar risco ou dano relevante aos titulares, comunicaremos
                à ANPD e aos titulares afetados em prazo razoável, conforme Art. 48 da LGPD.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">11. Alterações desta Política</h2>
                <p>Esta política pode ser atualizada para refletir mudanças legais, técnicas ou de negócio. A versão atual
                é <strong>{PRIVACY_POLICY_VERSION}</strong>. Versões futuras serão notificadas por email e
                pelo painel do sistema, com solicitação de novo aceite quando aplicável.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">12. Foro</h2>
                <p>Fica eleito o foro da Comarca de Quixadá, Estado do Ceará, para dirimir quaisquer controvérsias
                oriundas desta Política, com renúncia expressa a qualquer outro, por mais privilegiado que seja.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">13. Autoridade Nacional</h2>
                <p>Você também pode apresentar reclamação à Autoridade Nacional de Proteção de Dados (ANPD)
                em <a href="https://www.gov.br/anpd" target="_blank" rel="noopener">gov.br/anpd</a>.</p>

            </div>

            <div style="margin-top:32px;text-align:center;display:flex;gap:12px;justify-content:center;flex-wrap:wrap">
                <a href="/" class="btn btn-secondary">← Voltar ao início</a>
                <a href="/terms" class="btn btn-secondary">Ver Termos de Serviço</a>
                <a href="/dpa" class="btn btn-secondary">Ver DPA (Contrato de Operador)</a>
            </div>
        </div>
    </div>"""
    return base_html("Política de Privacidade", content)


@app.route("/terms")
def terms_of_service():
    content = f"""
    <div class="container" style="max-width:860px">
        <div class="card" style="margin-top:40px;padding:40px">
            <h1 style="font-size:28px;font-weight:700;margin-bottom:8px">Termos de Serviço</h1>
            <p style="color:var(--text2);margin-bottom:24px;font-size:14px">
                <strong>Versão:</strong> {TERMS_OF_SERVICE_VERSION} &nbsp;|&nbsp;
                <strong>Última atualização:</strong> 18 de maio de 2026
            </p>

            <div style="background:var(--bg2);padding:16px;border-radius:8px;margin-bottom:24px;font-size:14px;line-height:1.7">
                <strong>📌 Resumo:</strong> Estes termos regem o uso do serviço atendente.online (SaaS de atendimento WhatsApp com IA).
                Ao se cadastrar, você concorda integralmente com estas condições. Leia atentamente — em especial as seções
                de responsabilidades do usuário, conformidade com a Meta e limitações de responsabilidade.
            </div>

            <div style="color:var(--text2);font-size:15px;line-height:1.8">

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">1. Definições</h2>
                <p>• <strong style="color:var(--text)">"Serviço":</strong> a plataforma atendente.online, acessível em https://atendente.online.</p>
                <p>• <strong style="color:var(--text)">"Contratante" ou "Usuário":</strong> pessoa física ou jurídica que se cadastra e contrata o Serviço.</p>
                <p>• <strong style="color:var(--text)">"Contratada":</strong> Clériston Almeida Capistrano, operador do Serviço.</p>
                <p>• <strong style="color:var(--text)">"Cliente Final":</strong> pessoa que interage com o Contratante via WhatsApp/Instagram/Messenger.</p>
                <p>• <strong style="color:var(--text)">"Dados Pessoais":</strong> conforme Art. 5º, I da LGPD.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">2. Aceitação dos Termos</h2>
                <p>Ao criar conta no Serviço, o Contratante declara que: (i) leu, entendeu e aceita integralmente estes Termos,
                a <a href="/privacy">Política de Privacidade</a> e o <a href="/dpa">Contrato de Operador (DPA)</a>;
                (ii) tem capacidade legal para contratar (maior de 18 anos ou representante legal de pessoa jurídica);
                (iii) as informações fornecidas no cadastro são verdadeiras e atualizadas.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">3. Descrição do Serviço</h2>
                <p>O atendente.online é uma plataforma SaaS multi-tenant que oferece:</p>
                <p>• Atendimento automatizado via WhatsApp, Instagram e Messenger com inteligência artificial (Claude/Anthropic);</p>
                <p>• Transcrição de áudios, análise de imagens e processamento de documentos PDF;</p>
                <p>• CRM com funil de vendas, campanhas/broadcast e quick replies;</p>
                <p>• E-commerce básico com integração ao Mercado Pago;</p>
                <p>• Agência Digital com IA para geração e aprovação de conteúdo;</p>
                <p>• Painel administrativo com 2FA, audit log e relatórios.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">4. Cadastro e Conta</h2>
                <p>4.1. O Contratante é responsável por manter senha e códigos de 2FA em sigilo.</p>
                <p>4.2. Não é permitido compartilhar conta com terceiros. Cada conta corresponde a uma única empresa/CNPJ
                ou CPF do contratante.</p>
                <p>4.3. A Contratada pode suspender ou encerrar contas que violem estes Termos, sem reembolso.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">5. Planos, Preços e Pagamento</h2>
                <p>5.1. Oferecemos quatro planos (Starter R$97, Profissional R$197, Business R$397, Agência R$997)
                com cobrança mensal recorrente via Mercado Pago.</p>
                <p>5.2. Período de teste gratuito de 7 dias está disponível para novos cadastros.
                Ao final do período, é necessário contratar plano pago para manter o acesso.</p>
                <p>5.3. <strong style="color:var(--text)">Direito de arrependimento (CDC Art. 49):</strong> o Contratante pessoa física pode rescindir o contrato
                em até 7 dias após a primeira contratação paga, com restituição integral.</p>
                <p>5.4. <strong style="color:var(--text)">Reajuste de preços:</strong> mediante aviso prévio de 30 dias por email. Se o Contratante não concordar,
                pode cancelar a assinatura antes da próxima cobrança.</p>
                <p>5.5. <strong style="color:var(--text)">Inadimplência:</strong> falha no pagamento suspende o serviço após 5 dias de tolerância.
                Após 30 dias, a conta é marcada como inativa e os dados ficam sujeitos à política de retenção.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">6. Limites por Plano</h2>
                <p>Cada plano possui limite mensal de mensagens processadas pela IA. Excedido o limite, mensagens
                adicionais não serão respondidas automaticamente até o início do próximo ciclo, ou o Contratante
                pode fazer upgrade para plano superior.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">7. Responsabilidades do Contratante</h2>
                <p>O Contratante se compromete a:</p>
                <p>7.1. <strong style="color:var(--text)">Cumprir as políticas da Meta:</strong> incluindo a
                <a href="https://www.whatsapp.com/legal/business-policy/" target="_blank" rel="noopener">WhatsApp Business Policy</a>,
                <a href="https://developers.facebook.com/policy" target="_blank" rel="noopener">Meta Platform Terms</a> e
                <a href="https://www.whatsapp.com/legal/messaging-commerce-policy" target="_blank" rel="noopener">Commerce Policy</a>.</p>
                <p>7.2. <strong style="color:var(--text)">Obter consentimento prévio:</strong> garantir que possui base legal
                (Art. 7º LGPD) para se comunicar com cada Cliente Final via WhatsApp.</p>
                <p>7.3. <strong style="color:var(--text)">Não enviar spam:</strong> não realizar campanhas em massa para listas adquiridas,
                contatos sem opt-in ou conteúdo enganoso.</p>
                <p>7.4. <strong style="color:var(--text)">Conteúdo lícito:</strong> não utilizar o Serviço para envio ou facilitação de
                conteúdo ilegal, discriminatório, fraudulento, de ódio, ou que viole direitos de terceiros.</p>
                <p>7.5. <strong style="color:var(--text)">Não burlar a IA:</strong> não tentar extrair prompts internos, contornar limites
                de uso, fazer engenharia reversa do sistema ou explorar vulnerabilidades sem autorização.</p>
                <p>7.6. <strong style="color:var(--text)">Atender solicitações de Clientes Finais:</strong> na qualidade de
                Controlador dos dados dos próprios clientes, responder solicitações LGPD (acesso, correção,
                exclusão) que receber.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">8. Conteúdo Gerado por IA</h2>
                <p>8.1. As respostas geradas pela IA podem conter imprecisões ou erros. O Contratante é responsável
                por revisar e corrigir respostas quando necessário (incluindo via funcionalidade "Assumir conversa").</p>
                <p>8.2. A Contratada não garante exatidão factual, completude ou adequação das respostas para
                propósitos específicos (médicos, jurídicos, financeiros).</p>
                <p>8.3. A Contratada se reserva o direito de bloquear prompts que violem políticas de uso da Anthropic.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">9. Propriedade Intelectual</h2>
                <p>9.1. O código-fonte, marca, design, fluxos e funcionalidades do Serviço são de propriedade
                exclusiva da Contratada e protegidos pela Lei nº 9.610/98 (Direitos Autorais) e Lei nº 9.279/96 (Propriedade Industrial).</p>
                <p>9.2. O Contratante mantém integralmente a propriedade dos dados e conteúdos que insere no Serviço
                (mensagens, base de conhecimento, produtos, etc.).</p>
                <p>9.3. O Contratante concede licença não-exclusiva para a Contratada processar esses dados estritamente
                para a prestação do Serviço.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">10. Disponibilidade e Suporte</h2>
                <p>10.1. O Serviço é fornecido "como está" (as-is), sem SLA contratual de uptime para os planos atuais.</p>
                <p>10.2. Manutenções programadas serão comunicadas com antecedência. Manutenções emergenciais
                podem ocorrer sem aviso prévio.</p>
                <p>10.3. Suporte é prestado por email (<a href="mailto:contato@atendente.online">contato@atendente.online</a>)
                em horário comercial (segunda a sexta, 9h às 18h, BRT).</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">11. Limitação de Responsabilidade</h2>
                <p>11.1. A Contratada não se responsabiliza por:</p>
                <p>• Indisponibilidade ou alterações em serviços de terceiros (Meta/WhatsApp, Anthropic, Groq, Mercado Pago, Railway);</p>
                <p>• Bloqueios, suspensões ou penalidades aplicadas pela Meta às contas do Contratante;</p>
                <p>• Perdas comerciais, lucros cessantes ou danos indiretos decorrentes de uso ou indisponibilidade do Serviço;</p>
                <p>• Conteúdo enviado ou recebido pelo Contratante ou seus Clientes Finais.</p>
                <p>11.2. A responsabilidade total da Contratada, em qualquer hipótese, fica limitada ao valor pago
                pelo Contratante nos 3 (três) meses anteriores ao evento que originou a reclamação.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">12. Cancelamento e Encerramento</h2>
                <p>12.1. O Contratante pode cancelar a qualquer momento pelo painel. O acesso continua até o final
                do período já pago.</p>
                <p>12.2. Não há reembolso proporcional de períodos não utilizados, exceto no direito de arrependimento (cl. 5.3).</p>
                <p>12.3. Após cancelamento, os dados ficam disponíveis para exportação por 30 dias e são anonimizados em seguida,
                conforme política de retenção (ver Política de Privacidade).</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">13. Alterações dos Termos</h2>
                <p>13.1. Estes Termos podem ser atualizados. A versão atual é <strong>{TERMS_OF_SERVICE_VERSION}</strong>.</p>
                <p>13.2. Alterações materiais serão notificadas por email com 30 dias de antecedência. O uso continuado
                após esse prazo constitui aceite das novas condições.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">14. Disposições Gerais</h2>
                <p>14.1. Estes Termos são regidos pelas leis da República Federativa do Brasil.</p>
                <p>14.2. Fica eleito o foro da Comarca de Quixadá-CE, com renúncia a qualquer outro.</p>
                <p>14.3. Se qualquer cláusula for considerada inválida, as demais permanecem em vigor.</p>

            </div>

            <div style="margin-top:32px;text-align:center;display:flex;gap:12px;justify-content:center;flex-wrap:wrap">
                <a href="/" class="btn btn-secondary">← Voltar ao início</a>
                <a href="/privacy" class="btn btn-secondary">Ver Política de Privacidade</a>
                <a href="/dpa" class="btn btn-secondary">Ver DPA</a>
            </div>
        </div>
    </div>"""
    return base_html("Termos de Serviço", content)


@app.route("/dpa")
def dpa_contract():
    """Data Processing Agreement / Contrato de Operador (LGPD Art. 39)."""
    content = f"""
    <div class="container" style="max-width:860px">
        <div class="card" style="margin-top:40px;padding:40px">
            <h1 style="font-size:28px;font-weight:700;margin-bottom:8px">Contrato de Operador de Dados (DPA)</h1>
            <p style="color:var(--text2);margin-bottom:24px;font-size:14px">
                <strong>Versão:</strong> {DPA_VERSION} &nbsp;|&nbsp;
                <strong>Última atualização:</strong> 18 de maio de 2026
            </p>

            <div style="background:var(--bg2);padding:16px;border-radius:8px;margin-bottom:24px;font-size:14px;line-height:1.7">
                <strong>📌 O que é este documento?</strong> Este é o Contrato de Operador de Dados (DPA — Data Processing Agreement),
                conforme exigido pela LGPD (Art. 39). Ele formaliza a relação entre você (Controlador) e o atendente.online
                (Operador) sobre o tratamento de dados pessoais dos seus clientes finais. Anexo aos Termos de Serviço.
            </div>

            <div style="color:var(--text2);font-size:15px;line-height:1.8">

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">Cláusula 1ª — Definições e Partes</h2>
                <p>São partes deste contrato:</p>
                <p>• <strong style="color:var(--text)">CONTROLADOR:</strong> o Contratante do serviço atendente.online,
                identificado nos dados de cadastro.</p>
                <p>• <strong style="color:var(--text)">OPERADOR:</strong> Clériston Almeida Capistrano, operador da plataforma atendente.online.</p>
                <p>Os termos "Dados Pessoais", "Titular", "Tratamento", "Controlador" e "Operador" seguem as definições
                do Art. 5º da Lei nº 13.709/2018 (LGPD).</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">Cláusula 2ª — Objeto</h2>
                <p>O OPERADOR realiza tratamento de Dados Pessoais dos clientes finais do CONTROLADOR
                (pessoas que interagem via WhatsApp, Instagram e Messenger) exclusivamente em nome e por conta deste,
                conforme as instruções aqui estabelecidas e seguindo os Termos de Serviço.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">Cláusula 3ª — Dados Tratados</h2>
                <p>O tratamento envolve as seguintes categorias de dados:</p>
                <p>• Identificadores: nome de perfil, número de telefone;</p>
                <p>• Conteúdo: mensagens em texto, áudio, imagem e documentos;</p>
                <p>• Metadados: data/hora, status de leitura, canal de origem.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">Cláusula 4ª — Finalidades</h2>
                <p>Os Dados serão tratados exclusivamente para:</p>
                <p>• Atendimento automatizado via IA;</p>
                <p>• Armazenamento de histórico de conversas;</p>
                <p>• Geração de métricas agregadas para o CONTROLADOR;</p>
                <p>• Processamento de pedidos (e-commerce, se habilitado);</p>
                <p>• Quaisquer finalidades acessórias necessárias à execução do contrato.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">Cláusula 5ª — Obrigações do OPERADOR</h2>
                <p>O OPERADOR se compromete a:</p>
                <p>5.1. Tratar os Dados apenas conforme instruções do CONTROLADOR e desta avença;</p>
                <p>5.2. Adotar medidas técnicas e organizacionais para proteção dos Dados (criptografia, controle de acesso,
                logs de auditoria, backups, conforme detalhado na Política de Privacidade);</p>
                <p>5.3. Comunicar ao CONTROLADOR, sem demora injustificada (em até 48 horas após ciência), qualquer incidente
                de segurança que possa afetar Dados sob seu tratamento (LGPD Art. 48);</p>
                <p>5.4. Não compartilhar Dados com terceiros, exceto sub-operadores listados na Política de Privacidade
                (Meta, Anthropic, Groq, Mercado Pago, Railway) e mediante garantias equivalentes;</p>
                <p>5.5. Manter sigilo profissional sobre os Dados, inclusive após o término deste contrato;</p>
                <p>5.6. Auxiliar o CONTROLADOR no atendimento de solicitações de Titulares e da ANPD;</p>
                <p>5.7. Eliminar ou devolver Dados ao final do contrato, exceto se obrigação legal exigir retenção (Art. 16 LGPD);</p>
                <p>5.8. Demonstrar conformidade quando solicitado pelo CONTROLADOR, mediante prazo razoável.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">Cláusula 6ª — Obrigações do CONTROLADOR</h2>
                <p>O CONTROLADOR se compromete a:</p>
                <p>6.1. Possuir base legal válida (LGPD Art. 7º) para o tratamento dos Dados que insere ou processa via o Serviço;</p>
                <p>6.2. Informar adequadamente seus próprios titulares sobre o tratamento de dados, conforme Art. 9º LGPD;</p>
                <p>6.3. Atender, em primeira instância, solicitações de seus próprios titulares (acesso, correção, exclusão);</p>
                <p>6.4. Não inserir no Serviço dados de menores de 18 anos sem consentimento específico dos responsáveis,
                nem dados sensíveis (saúde, religião, orientação sexual, biometria) sem base legal adequada (Art. 11 LGPD);</p>
                <p>6.5. Comunicar imediatamente ao OPERADOR qualquer incidente, dúvida ou risco identificado.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">Cláusula 7ª — Sub-operadores Autorizados</h2>
                <p>O CONTROLADOR autoriza expressamente o uso dos seguintes sub-operadores (lista completa na Política de Privacidade):</p>
                <p>• Meta Platforms (WhatsApp/Instagram/Messenger API);</p>
                <p>• Anthropic (Claude — geração de respostas com IA);</p>
                <p>• Groq (transcrição de áudio);</p>
                <p>• Mercado Pago (pagamentos);</p>
                <p>• Railway (hospedagem).</p>
                <p>Mudanças nesta lista serão comunicadas com 30 dias de antecedência, e o CONTROLADOR pode se opor
                à inclusão de novos sub-operadores; havendo objeção, o contrato pode ser rescindido sem multa.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">Cláusula 8ª — Transferência Internacional</h2>
                <p>Parte dos sub-operadores opera nos Estados Unidos. Tais transferências observam o Art. 33 da LGPD,
                em especial cláusulas contratuais padrão e padrões de proteção compatíveis com o ordenamento jurídico brasileiro.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">Cláusula 9ª — Responsabilidade e Indenização</h2>
                <p>9.1. Cada parte é responsável por suas obrigações nos termos da LGPD.</p>
                <p>9.2. O CONTROLADOR isenta o OPERADOR de responsabilidade por dados inseridos em violação às suas obrigações
                (cláusula 6ª) e por reclamações de seus próprios titulares baseadas em ausência de base legal do CONTROLADOR.</p>
                <p>9.3. A limitação de responsabilidade estabelecida nos Termos de Serviço aplica-se também a este DPA.</p>

                <h2 style="color:var(--text);font-size:20px;margin:28px 0 12px">Cláusula 10ª — Vigência e Disposições Finais</h2>
                <p>10.1. Este DPA tem vigência a partir do aceite dos Termos de Serviço e perdura enquanto houver
                tratamento de Dados pelo OPERADOR em nome do CONTROLADOR.</p>
                <p>10.2. O DPA é parte integrante dos Termos de Serviço. Em caso de conflito específico sobre
                proteção de dados, prevalece este DPA.</p>
                <p>10.3. Aplicam-se as leis brasileiras. Foro: Comarca de Quixadá-CE.</p>

                <p style="margin-top:24px;padding:12px;background:var(--bg2);border-radius:6px;font-size:13px">
                    <strong>Encarregado pelo Tratamento de Dados (DPO):</strong> Clériston Almeida Capistrano —
                    <a href="mailto:contato@atendente.online">contato@atendente.online</a>
                </p>

            </div>

            <div style="margin-top:32px;text-align:center;display:flex;gap:12px;justify-content:center;flex-wrap:wrap">
                <a href="/" class="btn btn-secondary">← Voltar ao início</a>
                <a href="/privacy" class="btn btn-secondary">Política de Privacidade</a>
                <a href="/terms" class="btn btn-secondary">Termos de Serviço</a>
                <a href="/dpo" class="btn btn-secondary">Contato do DPO</a>
            </div>
        </div>
    </div>"""
    return base_html("Contrato de Operador (DPA)", content)


@app.route("/dpo")
def dpo_contact():
    """Página oficial do Encarregado pelo Tratamento de Dados (LGPD Art. 41)."""
    content = """
    <div class="container" style="max-width:720px">
        <div class="card" style="margin-top:40px;padding:40px">
            <h1 style="font-size:28px;font-weight:700;margin-bottom:8px">Encarregado pelo Tratamento de Dados</h1>
            <p style="color:var(--text2);margin-bottom:24px;font-size:14px">
                <strong>DPO — Data Protection Officer</strong> · LGPD Art. 41
            </p>

            <div style="background:var(--bg2);padding:24px;border-radius:8px;margin-bottom:24px">
                <h2 style="color:var(--text);font-size:18px;margin-bottom:16px">Identificação do Encarregado</h2>
                <p style="color:var(--text2);line-height:1.9;font-size:15px">
                    <strong style="color:var(--text)">Nome:</strong> Clériston Almeida Capistrano<br>
                    <strong style="color:var(--text)">Função:</strong> Encarregado pelo Tratamento de Dados Pessoais (DPO)<br>
                    <strong style="color:var(--text)">E-mail:</strong> <a href="mailto:contato@atendente.online">contato@atendente.online</a><br>
                    <strong style="color:var(--text)">Localização:</strong> Quixadá-CE, Brasil<br>
                    <strong style="color:var(--text)">Prazo de resposta:</strong> até 15 dias úteis
                </p>
            </div>

            <div style="color:var(--text2);font-size:15px;line-height:1.8">
                <h2 style="color:var(--text);font-size:18px;margin:24px 0 12px">Atribuições do DPO (LGPD Art. 41, §2º)</h2>
                <p>• Aceitar reclamações e comunicações dos titulares, prestar esclarecimentos e adotar providências;</p>
                <p>• Receber comunicações da Autoridade Nacional de Proteção de Dados (ANPD) e adotar providências;</p>
                <p>• Orientar funcionários e contratados sobre práticas a serem tomadas em relação à proteção de dados pessoais;</p>
                <p>• Executar as demais atribuições determinadas pelo controlador ou estabelecidas em normas complementares.</p>

                <h2 style="color:var(--text);font-size:18px;margin:24px 0 12px">Como contatar</h2>
                <p>Para exercer seus direitos previstos na LGPD (Art. 18) — acesso, correção, exclusão, portabilidade,
                revogação de consentimento etc. — você pode:</p>
                <p>• <strong style="color:var(--text)">Usuários logados:</strong> acessar diretamente a página
                <a href="/conta/meus-dados">Meus Dados</a> no painel;</p>
                <p>• <strong style="color:var(--text)">Por email:</strong> escrever para
                <a href="mailto:contato@atendente.online">contato@atendente.online</a> com:</p>
                <p style="margin-left:24px">– Assunto: "Solicitação LGPD"</p>
                <p style="margin-left:24px">– Descrição do direito que deseja exercer</p>
                <p style="margin-left:24px">– Dados para verificação da identidade (email da conta)</p>

                <h2 style="color:var(--text);font-size:18px;margin:24px 0 12px">Autoridade Nacional</h2>
                <p>Caso não esteja satisfeito com nossa resposta, você pode apresentar reclamação à ANPD
                em <a href="https://www.gov.br/anpd" target="_blank" rel="noopener">gov.br/anpd</a>.</p>
            </div>

            <div style="margin-top:32px;text-align:center;display:flex;gap:12px;justify-content:center;flex-wrap:wrap">
                <a href="/" class="btn btn-secondary">← Voltar ao início</a>
                <a href="/privacy" class="btn btn-secondary">Política de Privacidade</a>
            </div>
        </div>
    </div>"""
    return base_html("Encarregado de Dados (DPO)", content)


# ════════════════════════════════════════════════════════════════
#  ROTAS LGPD — Direitos do Titular (Art. 18 da Lei 13.709/2018)
# ════════════════════════════════════════════════════════════════

@app.route("/conta/meus-dados")
@login_required
def lgpd_my_data():
    """LGPD Art. 18, II — Acesso aos dados pessoais.
    Mostra ao titular tudo que está armazenado sobre ele."""
    user = g.user
    db = get_db()

    # Resumo numérico (sem carregar dados pesados na tela)
    counts = {}
    try:
        counts["contatos"] = db.execute("SELECT COUNT(*) FROM contacts WHERE user_id=?", (user["id"],)).fetchone()[0]
        counts["conversas"] = db.execute("SELECT COUNT(*) FROM conversations WHERE user_id=?", (user["id"],)).fetchone()[0]
        counts["mensagens"] = db.execute(
            """SELECT COUNT(*) FROM messages m JOIN conversations c ON c.id=m.conversation_id
               WHERE c.user_id=?""", (user["id"],)
        ).fetchone()[0]
        counts["pedidos"] = db.execute("SELECT COUNT(*) FROM orders WHERE user_id=?", (user["id"],)).fetchone()[0]
        counts["pagamentos"] = db.execute("SELECT COUNT(*) FROM payments WHERE user_id=?", (user["id"],)).fetchone()[0]
        counts["base_conhecimento"] = db.execute("SELECT COUNT(*) FROM knowledge_base WHERE user_id=?", (user["id"],)).fetchone()[0]
        counts["produtos"] = db.execute("SELECT COUNT(*) FROM product_gallery WHERE user_id=?", (user["id"],)).fetchone()[0]
        counts["campanhas"] = db.execute("SELECT COUNT(*) FROM campaigns WHERE user_id=?", (user["id"],)).fetchone()[0]
    except Exception as e:
        safe_log(f"[LGPD MY_DATA] Erro contagem: {e}", level="ERROR")
        counts = {k: 0 for k in ["contatos","conversas","mensagens","pedidos","pagamentos","base_conhecimento","produtos","campanhas"]}

    # Plano e retenção
    plan_label = PLANS.get(user["plan"], PLANS["starter"])["name"]
    retention_days = {"starter": 90, "pro": 180, "business": 365, "agency": 730}.get(user["plan"], 90)
    retention_label = {"starter": "3 meses", "pro": "6 meses", "business": "12 meses", "agency": "24 meses"}.get(user["plan"], "3 meses")

    # Última atualização de consentimento
    last_consent = db.execute(
        """SELECT consent_version, created_at FROM consent_log
           WHERE user_id=? AND consent_type='privacy_policy'
           ORDER BY created_at DESC LIMIT 1""", (user["id"],)
    ).fetchone()
    last_consent_str = f"Versão {last_consent['consent_version']} em {last_consent['created_at']}" if last_consent else "Nunca aceito explicitamente (cliente anterior à LGPD)"

    content = f"""
    <div class="container" style="max-width:980px;padding:24px">
        <div style="margin-bottom:24px">
            <h1 style="font-size:26px;font-weight:700;margin-bottom:8px">🛡️ Meus Dados</h1>
            <p style="color:var(--text2);font-size:14px">
                Esta página garante seus direitos previstos no Art. 18 da LGPD.
                Aqui você pode <strong>visualizar, exportar, corrigir ou solicitar a exclusão</strong> dos seus dados pessoais.
            </p>
        </div>

        <div class="card" style="padding:24px;margin-bottom:20px">
            <h2 style="font-size:18px;margin-bottom:16px">📋 Dados de Cadastro</h2>
            <table style="width:100%;font-size:14px;line-height:1.9">
                <tr><td style="color:var(--text2);width:35%">Nome</td><td><strong>{(user['name'] or '—')}</strong></td></tr>
                <tr><td style="color:var(--text2)">Email</td><td><strong>{user['email']}</strong></td></tr>
                <tr><td style="color:var(--text2)">Empresa</td><td>{user.get('company') or '—'}</td></tr>
                <tr><td style="color:var(--text2)">Telefone</td><td>{user.get('phone') or '—'}</td></tr>
                <tr><td style="color:var(--text2)">Plano contratado</td><td><strong>{plan_label}</strong> ({user['plan_status']})</td></tr>
                <tr><td style="color:var(--text2)">Conta criada em</td><td>{user.get('created_at') or '—'}</td></tr>
                <tr><td style="color:var(--text2)">Último login</td><td>{user.get('last_login') or '—'}</td></tr>
                <tr><td style="color:var(--text2)">Mensagens usadas</td><td>{user.get('msgs_used',0)} / {user.get('msgs_limit',0)}</td></tr>
            </table>
        </div>

        <div class="card" style="padding:24px;margin-bottom:20px">
            <h2 style="font-size:18px;margin-bottom:16px">📊 Volume de Dados Armazenados</h2>
            <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px">
                <div style="background:var(--bg2);padding:14px;border-radius:8px;text-align:center">
                    <div style="font-size:24px;font-weight:700;color:var(--accent2)">{counts['contatos']}</div>
                    <div style="font-size:12px;color:var(--text2);margin-top:4px">Contatos</div>
                </div>
                <div style="background:var(--bg2);padding:14px;border-radius:8px;text-align:center">
                    <div style="font-size:24px;font-weight:700;color:var(--accent2)">{counts['conversas']}</div>
                    <div style="font-size:12px;color:var(--text2);margin-top:4px">Conversas</div>
                </div>
                <div style="background:var(--bg2);padding:14px;border-radius:8px;text-align:center">
                    <div style="font-size:24px;font-weight:700;color:var(--accent2)">{counts['mensagens']}</div>
                    <div style="font-size:12px;color:var(--text2);margin-top:4px">Mensagens</div>
                </div>
                <div style="background:var(--bg2);padding:14px;border-radius:8px;text-align:center">
                    <div style="font-size:24px;font-weight:700;color:var(--accent2)">{counts['pedidos']}</div>
                    <div style="font-size:12px;color:var(--text2);margin-top:4px">Pedidos</div>
                </div>
                <div style="background:var(--bg2);padding:14px;border-radius:8px;text-align:center">
                    <div style="font-size:24px;font-weight:700;color:var(--accent2)">{counts['pagamentos']}</div>
                    <div style="font-size:12px;color:var(--text2);margin-top:4px">Pagamentos</div>
                </div>
                <div style="background:var(--bg2);padding:14px;border-radius:8px;text-align:center">
                    <div style="font-size:24px;font-weight:700;color:var(--accent2)">{counts['base_conhecimento']}</div>
                    <div style="font-size:12px;color:var(--text2);margin-top:4px">Base de conhecimento</div>
                </div>
                <div style="background:var(--bg2);padding:14px;border-radius:8px;text-align:center">
                    <div style="font-size:24px;font-weight:700;color:var(--accent2)">{counts['produtos']}</div>
                    <div style="font-size:12px;color:var(--text2);margin-top:4px">Produtos</div>
                </div>
                <div style="background:var(--bg2);padding:14px;border-radius:8px;text-align:center">
                    <div style="font-size:24px;font-weight:700;color:var(--accent2)">{counts['campanhas']}</div>
                    <div style="font-size:12px;color:var(--text2);margin-top:4px">Campanhas</div>
                </div>
            </div>
            <p style="font-size:13px;color:var(--text2);margin-top:16px;background:var(--bg2);padding:10px;border-radius:6px">
                ℹ️ <strong>Política de retenção do seu plano ({plan_label}):</strong> mensagens são automaticamente
                excluídas após <strong>{retention_label}</strong>. Para retenção maior, faça upgrade de plano.
            </p>
        </div>

        <div class="card" style="padding:24px;margin-bottom:20px">
            <h2 style="font-size:18px;margin-bottom:16px">✅ Status dos Consentimentos</h2>
            <p style="color:var(--text2);font-size:14px;margin-bottom:12px">
                Último aceite registrado: <strong>{last_consent_str}</strong>
            </p>
            <a href="/conta/consentimentos" class="btn btn-secondary btn-sm">📜 Ver histórico completo de consentimentos →</a>
        </div>

        <div class="card" style="padding:24px;margin-bottom:20px">
            <h2 style="font-size:18px;margin-bottom:16px">🔐 Segurança da Conta</h2>
            <p style="color:var(--text2);font-size:14px;margin-bottom:12px">
                {'✅ <strong style="color:var(--green, #16a34a)">2FA ativado</strong> — sua conta tem proteção em duas camadas.' if user_2fa_enabled(user) else '⚠️ <strong>2FA não ativado</strong> — adicione uma segunda camada de proteção ao seu login.'}
            </p>
            <a href="/conta/2fa" class="btn btn-secondary btn-sm">🔐 {'Gerenciar 2FA' if user_2fa_enabled(user) else 'Ativar 2FA'} →</a>
        </div>

        <div class="card" style="padding:24px;margin-bottom:20px">
            <h2 style="font-size:18px;margin-bottom:16px">🎯 Exercer seus direitos (LGPD Art. 18)</h2>
            <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px">
                <a href="/conta/exportar" class="btn btn-primary" style="text-align:center;padding:14px">
                    📦 Exportar meus dados<br><span style="font-size:11px;font-weight:400;opacity:0.85">Portabilidade (Art. 18, V)</span>
                </a>
                <a href="/profile" class="btn btn-secondary" style="text-align:center;padding:14px">
                    ✏️ Corrigir cadastro<br><span style="font-size:11px;font-weight:400;opacity:0.85">Correção (Art. 18, III)</span>
                </a>
                <a href="/conta/consentimentos" class="btn btn-secondary" style="text-align:center;padding:14px">
                    🔧 Gerenciar consentimentos<br><span style="font-size:11px;font-weight:400;opacity:0.85">Revogação (Art. 18, IX)</span>
                </a>
                <a href="/conta/excluir" class="btn" style="background:#fef2f2;color:#dc2626;border:1px solid #fecaca;text-align:center;padding:14px">
                    🗑️ Excluir minha conta<br><span style="font-size:11px;font-weight:400;opacity:0.85">Eliminação (Art. 18, VI)</span>
                </a>
            </div>
            <p style="font-size:12px;color:var(--text2);margin-top:16px">
                Dúvidas? Contate o Encarregado (DPO) em <a href="mailto:contato@atendente.online">contato@atendente.online</a>.
                Prazo de resposta: até 15 dias úteis.
            </p>
        </div>
    </div>
    """
    return base_html("Meus Dados — LGPD", content, dict(user))


@app.route("/conta/exportar")
@login_required
def lgpd_export_data():
    """LGPD Art. 18, V — Portabilidade.
    Gera download JSON com todos os dados do titular."""
    user = g.user
    data = export_user_data_as_json(user["id"])
    if data is None:
        return "Erro ao gerar exportação. Tente novamente ou contate contato@atendente.online", 500

    # Log na auditoria
    try:
        register_consent(user["id"], user["email"], "data_processing", PRIVACY_POLICY_VERSION, accepted=True,
                         details="Exportação de dados realizada (LGPD Art. 18, V)")
    except Exception:
        pass

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"atendente_online_meus_dados_{user['id']}_{timestamp}.json"
    response = make_response(json.dumps(data, ensure_ascii=False, indent=2, default=str))
    response.headers["Content-Type"] = "application/json; charset=utf-8"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@app.route("/conta/consentimentos", methods=["GET","POST"])
@login_required
def lgpd_consents():
    """LGPD Art. 18, IX — Revogação de consentimento.
    Mostra histórico e permite revogar/conceder opt-ins voluntários."""
    user = g.user
    success = ""
    error = ""

    if request.method == "POST":
        action = request.form.get("action", "")
        consent_type = request.form.get("consent_type", "")
        if action not in ("grant", "revoke") or consent_type not in CONSENT_TYPES:
            error = "Solicitação inválida."
        else:
            # Tipos obrigatórios não podem ser revogados (somente excluindo a conta)
            if action == "revoke" and consent_type in ("privacy_policy", "terms_of_service", "dpa", "data_processing"):
                error = "Este consentimento é base para uso do serviço. Para revogá-lo, exclua sua conta."
            else:
                accepted = (action == "grant")
                ok = register_consent(
                    user["id"], user["email"], consent_type, PRIVACY_POLICY_VERSION,
                    accepted=accepted,
                    details=f"{'Concedido' if accepted else 'Revogado'} via página de consentimentos"
                )
                if ok:
                    success = f"Consentimento '{CONSENT_TYPES[consent_type]}' {'concedido' if accepted else 'revogado'} com sucesso."
                else:
                    error = "Erro ao registrar. Tente novamente."

    consents = get_user_consents(user["id"])
    # Status atual de cada tipo: pega o registro mais recente
    current_status = {}
    for c in consents:
        if c["consent_type"] not in current_status:
            current_status[c["consent_type"]] = c["accepted"]

    # Render do histórico
    if consents:
        history_rows = ""
        for c in consents[:50]:  # últimos 50
            badge_color = "#10b981" if c["accepted"] else "#dc2626"
            badge_text = "Aceito" if c["accepted"] else "Revogado"
            type_label = CONSENT_TYPES.get(c["consent_type"], c["consent_type"])
            history_rows += f"""
            <tr>
                <td style="padding:10px;border-bottom:1px solid var(--border);font-size:13px">{c['created_at']}</td>
                <td style="padding:10px;border-bottom:1px solid var(--border);font-size:13px">{type_label}</td>
                <td style="padding:10px;border-bottom:1px solid var(--border);font-size:13px">v{c['consent_version']}</td>
                <td style="padding:10px;border-bottom:1px solid var(--border)">
                    <span style="background:{badge_color};color:white;padding:2px 8px;border-radius:4px;font-size:11px">{badge_text}</span>
                </td>
            </tr>"""
        history_table = f"""
            <table style="width:100%;border-collapse:collapse;font-size:14px">
                <thead><tr style="background:var(--bg2)">
                    <th style="padding:10px;text-align:left">Data</th>
                    <th style="padding:10px;text-align:left">Tipo</th>
                    <th style="padding:10px;text-align:left">Versão</th>
                    <th style="padding:10px;text-align:left">Status</th>
                </tr></thead>
                <tbody>{history_rows}</tbody>
            </table>"""
    else:
        history_table = '<p style="color:var(--text2)">Nenhum registro de consentimento encontrado.</p>'

    marketing_status = current_status.get("marketing_email", 0)
    marketing_button = (
        '<form method="POST" style="display:inline">' + csrf_field() +
        '<input type="hidden" name="action" value="revoke">' +
        '<input type="hidden" name="consent_type" value="marketing_email">' +
        '<button type="submit" class="btn btn-secondary btn-sm">🚫 Revogar opt-in de marketing</button></form>'
        if marketing_status else
        '<form method="POST" style="display:inline">' + csrf_field() +
        '<input type="hidden" name="action" value="grant">' +
        '<input type="hidden" name="consent_type" value="marketing_email">' +
        '<button type="submit" class="btn btn-primary btn-sm">✅ Aceitar comunicações de marketing</button></form>'
    )

    alert = ""
    if success: alert = f'<div class="alert alert-success">{success}</div>'
    if error: alert = f'<div class="alert alert-error">{error}</div>'

    content = f"""
    <div class="container" style="max-width:880px;padding:24px">
        <a href="/conta/meus-dados" style="color:var(--text2);font-size:13px;text-decoration:none">← Voltar para Meus Dados</a>
        <h1 style="font-size:26px;font-weight:700;margin:8px 0 16px">📜 Consentimentos</h1>
        {alert}

        <div class="card" style="padding:24px;margin-bottom:20px">
            <h2 style="font-size:18px;margin-bottom:12px">Consentimentos atuais</h2>
            <table style="width:100%;font-size:14px">
                <tr><td style="padding:8px 0;color:var(--text2)">Política de Privacidade (obrigatório)</td>
                    <td style="text-align:right"><span style="background:#10b981;color:white;padding:2px 8px;border-radius:4px;font-size:11px">Aceito</span></td></tr>
                <tr><td style="padding:8px 0;color:var(--text2)">Termos de Serviço (obrigatório)</td>
                    <td style="text-align:right"><span style="background:#10b981;color:white;padding:2px 8px;border-radius:4px;font-size:11px">Aceito</span></td></tr>
                <tr><td style="padding:8px 0;color:var(--text2)">DPA — Contrato de Operador (obrigatório)</td>
                    <td style="text-align:right"><span style="background:#10b981;color:white;padding:2px 8px;border-radius:4px;font-size:11px">Aceito</span></td></tr>
                <tr><td style="padding:8px 0">Comunicações de marketing (opcional)</td>
                    <td style="text-align:right">{marketing_button}</td></tr>
            </table>
            <p style="font-size:12px;color:var(--text2);margin-top:14px">
                Os 3 consentimentos obrigatórios são base contratual e legal para uso do serviço.
                Para revogá-los, é necessário <a href="/conta/excluir">excluir sua conta</a>.
            </p>
        </div>

        <div class="card" style="padding:24px">
            <h2 style="font-size:18px;margin-bottom:12px">Histórico completo (últimos 50 registros)</h2>
            {history_table}
            <p style="font-size:11px;color:var(--text2);margin-top:12px">
                💾 Estes registros são mantidos por 5 anos para comprovação perante a ANPD, conforme LGPD.
            </p>
        </div>
    </div>
    """
    return base_html("Consentimentos — LGPD", content, dict(user))


@app.route("/conta/2fa", methods=["GET", "POST"])
@login_required
def user_2fa_setup():
    """Setup/gerenciamento de 2FA (TOTP) do usuário comum — opt-in.
    Fluxo:
    1. GET sem 2FA → tela de início com botão "Ativar 2FA"
    2. GET ?start=1 → gera secret pendente, mostra QR code para escanear
    3. POST action=verify_setup → valida código, ativa 2FA e gera backup codes
    4. GET ?setup=success → mostra backup codes uma única vez
    5. POST action=disable → exige senha de confirmação para desativar
    6. POST action=regenerate_backup → exige senha, gera novos backup codes
    """
    user = g.user
    db = get_db()
    msg = ""
    is_enabled = user_2fa_enabled(user)

    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "disable" and is_enabled:
            confirm_pw = request.form.get("confirm_password", "")
            raw = db.execute("SELECT password_hash FROM users WHERE id=?", (user["id"],)).fetchone()
            if not raw or not check_password(confirm_pw, raw["password_hash"]):
                msg = '<div class="alert alert-error">Senha incorreta. 2FA não foi desativado.</div>'
            else:
                db.execute("UPDATE users SET totp_secret='', totp_enabled=0, totp_backup_codes='' WHERE id=?", (user["id"],))
                db.commit()
                safe_log(f"[2FA] Desativado pelo próprio usuário (user_id={user['id']})", level="WARN")
                is_enabled = False
                msg = '<div class="alert alert-warning">⚠️ 2FA desativado. Sua conta está agora protegida apenas por senha. Recomendamos reativar.</div>'

        elif action == "regenerate_backup" and is_enabled:
            confirm_pw = request.form.get("confirm_password", "")
            raw = db.execute("SELECT password_hash FROM users WHERE id=?", (user["id"],)).fetchone()
            if not raw or not check_password(confirm_pw, raw["password_hash"]):
                msg = '<div class="alert alert-error">Senha incorreta. Backup codes não foram regenerados.</div>'
            else:
                new_codes = generate_backup_codes(8)
                enc = _encrypt_value(",".join(new_codes))
                db.execute("UPDATE users SET totp_backup_codes=? WHERE id=?", (enc, user["id"]))
                db.commit()
                session["2fa_backup_codes_shown"] = new_codes
                safe_log(f"[2FA] Backup codes regenerados (user_id={user['id']})")
                return redirect("/conta/2fa?setup=success")

        elif action == "verify_setup":
            pending_secret = session.get("pending_user_totp_secret", "")
            code = request.form.get("totp_code", "").strip()
            if not pending_secret:
                msg = '<div class="alert alert-error">Sessão expirada. Inicie o setup novamente.</div>'
            elif verify_totp_code(pending_secret, code):
                backup_codes = generate_backup_codes(8)
                enc_secret = _encrypt_value(pending_secret)
                enc_backup = _encrypt_value(",".join(backup_codes))
                db.execute(
                    "UPDATE users SET totp_secret=?, totp_enabled=1, totp_backup_codes=? WHERE id=?",
                    (enc_secret, enc_backup, user["id"])
                )
                db.commit()
                session.pop("pending_user_totp_secret", None)
                session["2fa_backup_codes_shown"] = backup_codes
                safe_log(f"[2FA] Ativado pelo usuário (user_id={user['id']})")
                return redirect("/conta/2fa?setup=success")
            else:
                msg = '<div class="alert alert-error">Código inválido. Tente novamente.</div>'

    # Página de sucesso (mostra backup codes uma única vez)
    if request.args.get("setup") == "success":
        codes = session.pop("2fa_backup_codes_shown", [])
        if not codes:
            return redirect("/conta/2fa")
        codes_html = "".join([f'<li style="font-family:monospace;font-size:16px;padding:8px 0;border-bottom:1px solid var(--border)">{c}</li>' for c in codes])
        content = f"""
        <div class="container" style="max-width:680px;padding:32px 24px">
            <h1 style="font-size:24px;margin-bottom:8px">✅ 2FA Ativado!</h1>
            <p style="color:var(--text2);margin-bottom:24px">Anote seus backup codes em local seguro — eles só aparecem agora.</p>
            <div class="alert alert-warning" style="margin-bottom:16px">
                ⚠️ <strong>IMPORTANTE:</strong> Cada backup code funciona <strong>uma vez</strong>. Use se perder o celular.
                Esses códigos <strong>não serão mostrados novamente</strong>.
            </div>
            <div class="card" style="padding:24px">
                <h2 style="font-size:16px;margin-bottom:12px">🔑 Seus 8 Backup Codes</h2>
                <ul style="list-style:none;padding:16px 20px;background:var(--bg2);border-radius:8px;margin:0">{codes_html}</ul>
                <div style="margin-top:20px;display:flex;gap:10px">
                    <a href="/conta/2fa" class="btn btn-primary">Já anotei, continuar</a>
                    <a href="/conta/meus-dados" class="btn btn-secondary">Voltar para Meus Dados</a>
                </div>
            </div>
        </div>
        """
        return base_html("Backup Codes 2FA", content, dict(user))

    # Iniciar novo setup (gera secret temporário)
    if not is_enabled and request.args.get("start") == "1":
        new_secret = generate_totp_secret()
        if not new_secret:
            content = '<div class="container" style="padding:32px"><div class="alert alert-error">❌ Biblioteca pyotp não instalada no servidor. Contate o suporte.</div></div>'
            return base_html("Erro 2FA", content, dict(user))

        session["pending_user_totp_secret"] = new_secret
        uri = generate_totp_uri(new_secret, issuer="atendente.online", account=user["email"])
        qr_data = generate_totp_qr_base64(uri)
        qr_html = (f'<img src="{qr_data}" alt="QR Code 2FA" style="width:220px;height:220px;background:white;padding:12px;border-radius:8px;display:block;margin:0 auto">'
                   if qr_data else '<p style="color:var(--red)">Erro ao gerar QR code. Cole o código manualmente abaixo.</p>')

        content = f"""
        <div class="container" style="max-width:880px;padding:32px 24px">
            <h1 style="font-size:24px;margin-bottom:8px">🔐 Configurar Autenticação em Dois Fatores</h1>
            <p style="color:var(--text2);margin-bottom:24px">Adicione uma segunda camada de proteção à sua conta usando um app autenticador.</p>
            {msg}
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px">
                <div class="card" style="padding:20px">
                    <h2 style="font-size:16px;margin-bottom:12px">📱 Passo 1: Escaneie o QR code</h2>
                    <p style="color:var(--text2);font-size:13px;margin-bottom:16px">Use Google Authenticator, Authy, Microsoft Authenticator, 1Password ou Bitwarden.</p>
                    {qr_html}
                    <p style="color:var(--text3);font-size:12px;text-align:center;margin-top:12px">Ou cole este código manualmente:</p>
                    <code style="display:block;padding:10px;background:var(--bg2);border-radius:6px;font-size:11px;word-break:break-all;text-align:center;border:1px solid var(--border)">{new_secret}</code>
                </div>
                <div class="card" style="padding:20px">
                    <h2 style="font-size:16px;margin-bottom:12px">✅ Passo 2: Confirme o código</h2>
                    <p style="color:var(--text2);font-size:13px;margin-bottom:16px">Digite o código de 6 dígitos que aparece no seu app:</p>
                    <form method="POST">{csrf_field()}
                        <input type="hidden" name="action" value="verify_setup">
                        <div class="form-group">
                            <input type="text" name="totp_code" class="form-input" required maxlength="6" autofocus
                                   placeholder="000000" autocomplete="off"
                                   style="font-size:22px;letter-spacing:6px;text-align:center;font-family:monospace">
                        </div>
                        <button type="submit" class="btn btn-primary btn-block">Ativar 2FA</button>
                        <a href="/conta/2fa" style="display:block;text-align:center;margin-top:10px;color:var(--text3);font-size:12px">Cancelar</a>
                    </form>
                </div>
            </div>
        </div>
        """
        return base_html("Configurar 2FA", content, dict(user))

    # Tela principal: status atual + ativar / desativar / regenerar
    if is_enabled:
        body = f"""
        <div class="card" style="padding:24px;margin-bottom:20px">
            <div style="display:flex;align-items:center;gap:12px;margin-bottom:12px">
                <div style="font-size:32px">🛡️</div>
                <div>
                    <h2 style="font-size:18px;margin:0">2FA Ativado</h2>
                    <p style="color:var(--text2);font-size:13px;margin:4px 0 0">Sua conta está protegida por autenticação em dois fatores.</p>
                </div>
            </div>
        </div>

        <div class="card" style="padding:24px;margin-bottom:20px">
            <h2 style="font-size:16px;margin-bottom:12px">🔄 Regerar backup codes</h2>
            <p style="color:var(--text2);font-size:13px;margin-bottom:16px">Use se você usou ou perdeu seus códigos antigos. <strong>Os códigos antigos deixam de funcionar.</strong></p>
            <form method="POST" onsubmit="return confirm('Tem certeza? Os backup codes antigos deixarão de funcionar.');">
                {csrf_field()}
                <input type="hidden" name="action" value="regenerate_backup">
                <div class="form-group">
                    <label class="form-label">Confirme sua senha</label>
                    <input type="password" name="confirm_password" class="form-input" required>
                </div>
                <button type="submit" class="btn btn-secondary">Gerar novos backup codes</button>
            </form>
        </div>

        <div class="card" style="padding:24px;border-left:3px solid var(--red, #dc2626)">
            <h2 style="font-size:16px;margin-bottom:12px;color:var(--red, #dc2626)">⚠️ Desativar 2FA</h2>
            <p style="color:var(--text2);font-size:13px;margin-bottom:16px">Sua conta voltará a depender apenas de senha — <strong>não recomendado</strong>.</p>
            <form method="POST" onsubmit="return confirm('Tem certeza que quer desativar 2FA?');">
                {csrf_field()}
                <input type="hidden" name="action" value="disable">
                <div class="form-group">
                    <label class="form-label">Confirme sua senha</label>
                    <input type="password" name="confirm_password" class="form-input" required>
                </div>
                <button type="submit" class="btn" style="background:#fef2f2;color:#dc2626;border:1px solid #fecaca">Desativar 2FA</button>
            </form>
        </div>
        """
    else:
        body = f"""
        <div class="card" style="padding:32px;text-align:center">
            <div style="font-size:42px;margin-bottom:12px">🔓</div>
            <h2 style="font-size:18px;margin-bottom:8px">2FA não está ativado</h2>
            <p style="color:var(--text2);font-size:14px;margin-bottom:24px;max-width:520px;margin-left:auto;margin-right:auto">
                A autenticação em dois fatores adiciona uma camada extra de proteção: além da senha,
                o login passa a exigir um código de 6 dígitos gerado por um app no seu celular.
                Recomendado se você processa pagamentos ou guarda dados de clientes.
            </p>
            <a href="/conta/2fa?start=1" class="btn btn-primary btn-lg">🔐 Ativar 2FA agora</a>
        </div>
        """

    content = f"""
    <div class="container" style="max-width:720px;padding:24px">
        <div style="margin-bottom:24px">
            <a href="/conta/meus-dados" style="color:var(--text2);font-size:13px;text-decoration:none">← Voltar para Meus Dados</a>
            <h1 style="font-size:26px;font-weight:700;margin:8px 0">🔐 Segurança da Conta</h1>
            <p style="color:var(--text2);font-size:14px">Gerencie a autenticação em dois fatores (2FA) da sua conta.</p>
        </div>
        {msg}
        {body}
    </div>
    """
    return base_html("Segurança — 2FA", content, dict(user))


@app.route("/conta/excluir", methods=["GET","POST"])
@login_required
def lgpd_delete_account():
    """LGPD Art. 18, VI — Exclusão de dados (anonimização híbrida).
    Fluxo:
    1. GET: tela de aviso + checkbox "entendo que é irreversível"
    2. POST step=request_code: envia código por email
    3. POST step=confirm: valida código + senha + anonimiza conta
    """
    user = g.user
    error = ""
    success = ""
    step = request.args.get("step", "warning")  # warning | confirm

    if request.method == "POST":
        action = request.form.get("action", "")

        # ── Etapa 1: solicitar código por email ──
        if action == "request_code":
            understand = request.form.get("understand") == "on"
            if not understand:
                error = "Você precisa confirmar que entende que esta ação é irreversível."
            else:
                code = generate_deletion_code(user["id"], user["email"])
                if not code:
                    error = "Erro ao gerar código. Tente novamente em alguns minutos."
                else:
                    # Enviar email com o código
                    try:
                        html_body = f"""
                        <div style="font-family:Arial,sans-serif;max-width:520px;margin:auto;padding:20px;background:#fff">
                            <h2 style="color:#dc2626">⚠️ Confirmação de exclusão de conta</h2>
                            <p>Olá, {esc(user.get('name') or user['email'])}!</p>
                            <p>Recebemos uma solicitação para excluir sua conta no <strong>atendente.online</strong>.</p>
                            <p>Para confirmar, digite o código abaixo na tela de confirmação:</p>
                            <div style="background:#f3f4f6;padding:20px;text-align:center;font-size:32px;letter-spacing:6px;font-weight:700;border-radius:8px;margin:20px 0">
                                {code}
                            </div>
                            <p style="font-size:13px;color:#6b7280">
                                <strong>Este código expira em 15 minutos.</strong><br>
                                Se você não solicitou esta ação, ignore este email e altere sua senha imediatamente.
                            </p>
                            <hr style="margin:24px 0;border:none;border-top:1px solid #e5e7eb">
                            <p style="font-size:12px;color:#9ca3af">
                                Esta ação removerá: conversas, mensagens, contatos, base de conhecimento, campanhas e configurações.<br>
                                Dados de pagamento serão mantidos anonimizados por 5 anos por obrigação fiscal.
                            </p>
                        </div>"""
                        sent = send_email(user["email"], "🗑️ Confirmação de exclusão de conta — atendente.online", html_body)
                        if sent:
                            session["deletion_pending"] = True
                            return redirect("/conta/excluir?step=confirm")
                        else:
                            error = "Não foi possível enviar o email. Tente novamente ou contate o suporte."
                    except Exception as e:
                        safe_log(f"[DELETE] Erro envio email: {e}", level="ERROR")
                        error = "Erro ao enviar email de confirmação."

        # ── Etapa 2: confirmar com código + senha ──
        elif action == "confirm_delete":
            if not session.get("deletion_pending"):
                error = "Sessão de exclusão expirada. Reinicie o processo."
            else:
                code = request.form.get("code", "").strip()
                password = request.form.get("password", "")
                if not code or len(code) != 6:
                    error = "Digite o código de 6 dígitos enviado por email."
                elif not password:
                    error = "Digite sua senha para confirmar."
                elif not check_password(password, user["password_hash"]):
                    error = "Senha incorreta."
                elif not verify_deletion_code(user["email"], code):
                    error = "Código inválido ou expirado."
                else:
                    # Executar anonimização
                    ok = anonymize_user_account(user["id"], reason="user_self_service")
                    if ok:
                        session.clear()
                        # Render página final de despedida
                        farewell = """
                        <div class="container" style="max-width:560px;padding:60px 24px;text-align:center">
                            <div style="font-size:48px;margin-bottom:16px">👋</div>
                            <h1 style="font-size:24px;margin-bottom:12px">Sua conta foi excluída</h1>
                            <p style="color:var(--text2);line-height:1.7;margin-bottom:24px">
                                Suas conversas, mensagens, contatos e configurações foram removidos.<br>
                                Dados de pagamento permanecem anonimizados por 5 anos (obrigação fiscal).<br>
                                Obrigado por ter usado o atendente.online.
                            </p>
                            <a href="/" class="btn btn-primary">← Voltar para o início</a>
                        </div>
                        """
                        return base_html("Conta excluída", farewell)
                    else:
                        error = "Erro ao processar exclusão. Contate o suporte: contato@atendente.online"

    alert = ""
    if error: alert = f'<div class="alert alert-error">{error}</div>'
    if success: alert = f'<div class="alert alert-success">{success}</div>'

    # ── Render da tela conforme o step ──
    if step == "confirm" and session.get("deletion_pending"):
        masked = mask_email(user["email"])
        content = f"""
        <div class="container" style="max-width:560px;padding:24px">
            <a href="/conta/meus-dados" style="color:var(--text2);font-size:13px;text-decoration:none">← Cancelar e voltar</a>
            <div class="card" style="padding:32px;margin-top:12px;border:2px solid #fecaca">
                <h1 style="font-size:22px;color:#dc2626;margin-bottom:12px">🗑️ Confirmar exclusão</h1>
                <p style="color:var(--text2);margin-bottom:20px">
                    Enviamos um código de 6 dígitos para <strong>{masked}</strong>.
                    Digite-o abaixo junto com sua senha para confirmar a exclusão.
                </p>
                {alert}
                <form method="POST">
                    {csrf_field()}
                    <input type="hidden" name="action" value="confirm_delete">
                    <div class="form-group">
                        <label class="form-label">Código recebido por email</label>
                        <input type="text" name="code" class="form-input" maxlength="6"
                            placeholder="000000" required autofocus
                            style="text-align:center;font-size:24px;letter-spacing:6px;font-weight:700">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Sua senha atual</label>
                        <input type="password" name="password" class="form-input" required>
                    </div>
                    <button type="submit" class="btn btn-block" style="background:#dc2626;color:white;padding:12px">
                        ⚠️ Excluir minha conta DEFINITIVAMENTE
                    </button>
                </form>
                <p style="font-size:12px;color:var(--text2);margin-top:16px;text-align:center">
                    O código expira em 15 minutos. <a href="/conta/excluir">Reiniciar processo</a>
                </p>
            </div>
        </div>
        """
    else:
        # Tela inicial de aviso
        content = f"""
        <div class="container" style="max-width:680px;padding:24px">
            <a href="/conta/meus-dados" style="color:var(--text2);font-size:13px;text-decoration:none">← Voltar para Meus Dados</a>
            <div class="card" style="padding:32px;margin-top:12px;border:2px solid #fecaca">
                <h1 style="font-size:24px;color:#dc2626;margin-bottom:12px">⚠️ Excluir minha conta</h1>
                <p style="color:var(--text2);line-height:1.7;margin-bottom:16px">
                    Esta é uma ação <strong>irreversível</strong>. Ao confirmar, executaremos:
                </p>

                <h3 style="font-size:15px;color:#dc2626;margin:16px 0 8px">🗑️ Será permanentemente apagado:</h3>
                <ul style="color:var(--text2);font-size:14px;line-height:1.9;padding-left:20px">
                    <li>Todas as suas conversas, mensagens e contatos</li>
                    <li>Base de conhecimento treinada</li>
                    <li>Campanhas e listas de envio</li>
                    <li>Pipeline CRM (cards e etapas)</li>
                    <li>Produtos cadastrados e galeria social</li>
                    <li>Tokens da Meta (Instagram/Messenger/WhatsApp), Telegram, Mercado Pago</li>
                    <li>Configurações de 2FA</li>
                </ul>

                <h3 style="font-size:15px;color:#92400e;margin:20px 0 8px">📦 Será mantido (anonimizado) por 5 anos:</h3>
                <ul style="color:var(--text2);font-size:14px;line-height:1.9;padding-left:20px">
                    <li>Histórico de pagamentos (obrigação fiscal — Lei nº 5.172/66, CTN Art. 195)</li>
                    <li>Pedidos com valor financeiro</li>
                    <li>Logs de auditoria com seu ID anonimizado</li>
                    <li>Registros de consentimento (prova perante ANPD)</li>
                </ul>

                <div style="background:var(--bg2);padding:14px;border-radius:8px;margin:20px 0;font-size:13px;color:var(--text2);line-height:1.6">
                    💡 <strong>Quer apenas pausar?</strong> Você pode <a href="/dashboard/settings">cancelar seu plano</a>
                    sem excluir a conta — assim você pode retornar depois com seus dados intactos.
                </div>

                {alert}

                <form method="POST">
                    {csrf_field()}
                    <input type="hidden" name="action" value="request_code">
                    <label style="display:flex;align-items:flex-start;gap:10px;cursor:pointer;font-size:14px;color:var(--text2);background:#fef2f2;padding:14px;border-radius:6px;margin-bottom:16px">
                        <input type="checkbox" name="understand" required style="margin-top:3px;flex-shrink:0;width:16px;height:16px">
                        <span>Entendo que esta ação é <strong>irreversível</strong>, que perderei acesso ao sistema imediatamente,
                        e que dados de pagamento serão mantidos por obrigação fiscal.</span>
                    </label>
                    <button type="submit" class="btn btn-block" style="background:#dc2626;color:white;padding:12px">
                        📧 Enviar código de confirmação por email
                    </button>
                </form>
            </div>
        </div>
        """
    return base_html("Excluir conta", content, dict(user))


@app.route("/conta/aceitar-termos", methods=["GET","POST"])
@login_required
def lgpd_reaccept_terms():
    """Tela exibida quando o usuário precisa aceitar termos novos ou atualizados.
    Bloqueia o resto do sistema até aceitar (exceto whitelist em login_required)."""
    user = g.user
    error = ""

    # Se o usuário já está em dia, redireciona para o dashboard
    if not user_needs_to_reaccept_terms(user["id"]):
        return redirect("/dashboard")

    if request.method == "POST":
        accept = request.form.get("accept_terms") == "on"
        if not accept:
            error = "Você precisa aceitar os documentos para continuar usando o serviço."
        else:
            register_consent(user["id"], user["email"], "privacy_policy", PRIVACY_POLICY_VERSION,
                             accepted=True, details="Re-aceite após atualização dos termos")
            register_consent(user["id"], user["email"], "terms_of_service", TERMS_OF_SERVICE_VERSION,
                             accepted=True, details="Re-aceite após atualização dos termos")
            register_consent(user["id"], user["email"], "dpa", DPA_VERSION,
                             accepted=True, details="Re-aceite após atualização dos termos")
            register_consent(user["id"], user["email"], "data_processing", PRIVACY_POLICY_VERSION,
                             accepted=True, details="Base legal: execução de contrato (LGPD Art. 7º, V) — re-aceite")
            return redirect("/dashboard")

    alert = f'<div class="alert alert-error">{error}</div>' if error else ""

    content = f"""
    <div class="container" style="max-width:680px;padding:24px">
        <div class="card" style="padding:32px;margin-top:40px">
            <div style="font-size:48px;text-align:center;margin-bottom:12px">📜</div>
            <h1 style="font-size:24px;font-weight:700;text-align:center;margin-bottom:8px">Atualização dos Termos</h1>
            <p style="color:var(--text2);text-align:center;margin-bottom:24px;font-size:14px;line-height:1.6">
                Atualizamos nossa <strong>Política de Privacidade</strong>, <strong>Termos de Serviço</strong> e adicionamos o
                <strong>Contrato de Operador (DPA)</strong> em conformidade com a LGPD.<br>
                Para continuar usando o atendente.online, por favor revise e aceite os documentos atualizados.
            </p>

            <div style="background:var(--bg2);padding:16px;border-radius:8px;margin-bottom:20px">
                <h3 style="font-size:14px;margin-bottom:10px">📋 O que mudou:</h3>
                <ul style="color:var(--text2);font-size:13px;line-height:1.8;padding-left:20px">
                    <li>Política de Privacidade reescrita com detalhamento LGPD</li>
                    <li>Termos de Serviço atualizados com cláusulas profissionais</li>
                    <li>Novo: Contrato de Operador (DPA) anexo aos termos</li>
                    <li>Designamos formalmente um Encarregado de Dados (DPO)</li>
                    <li>Você agora pode exercer todos os direitos do Art. 18 da LGPD</li>
                </ul>
            </div>

            <div style="text-align:center;margin-bottom:20px">
                <a href="/privacy" target="_blank" class="btn btn-secondary btn-sm" style="margin:4px">📄 Política de Privacidade</a>
                <a href="/terms" target="_blank" class="btn btn-secondary btn-sm" style="margin:4px">📄 Termos de Serviço</a>
                <a href="/dpa" target="_blank" class="btn btn-secondary btn-sm" style="margin:4px">📄 DPA</a>
            </div>

            {alert}

            <form method="POST">
                {csrf_field()}
                <div style="background:var(--bg2);padding:14px;border-radius:8px;margin-bottom:16px">
                    <label style="display:flex;align-items:flex-start;gap:10px;cursor:pointer;font-size:14px;color:var(--text2);line-height:1.5">
                        <input type="checkbox" name="accept_terms" required style="margin-top:3px;flex-shrink:0;width:16px;height:16px">
                        <span>Li e aceito a <a href="/privacy" target="_blank" style="color:var(--accent2)">Política de Privacidade</a>,
                        os <a href="/terms" target="_blank" style="color:var(--accent2)">Termos de Serviço</a> e o
                        <a href="/dpa" target="_blank" style="color:var(--accent2)">Contrato de Operador (DPA)</a> atualizados.</span>
                    </label>
                </div>
                <button type="submit" class="btn btn-primary btn-block btn-lg">✅ Aceitar e continuar</button>
            </form>

            <p style="font-size:12px;color:var(--text2);text-align:center;margin-top:16px">
                Se não concorda, você pode <a href="/conta/excluir">excluir sua conta</a> ou
                <a href="/logout">sair</a>.
            </p>
        </div>
    </div>
    """
    return base_html("Aceitar Termos Atualizados", content)


@app.route("/")
def landing():
    if "user_id" in session: return redirect("/dashboard")
    nav_logo = LOGO_NAV_B64
    content = f"""
    <nav class="nav-main"><div class="nav-inner">
        <a href="/"><img src="data:image/png;base64,{nav_logo}" alt="atendente.online" class="nav-logo-img"></a>
        <div class="nav-links">
            <a href="#features" class="nav-link">Recursos</a>
            <a href="#pricing" class="nav-link">Planos</a>
            <a href="/login" class="nav-link">Entrar</a>
            <a href="/register" class="btn btn-primary btn-sm" style="margin-left:8px">Começar grátis</a>
        </div>
    </div></nav>

    <div class="hero fade-in">
        <h1>Atendente + Agência Digital<br><span class="gradient">com inteligência artificial</span></h1>
        <p>Automatize seu WhatsApp com IA treinável E crie conteúdo diário automaticamente para suas redes sociais. Tudo em um só lugar.</p>
        <div class="hero-badges">
            <span class="hero-badge">✓ WhatsApp Business API</span>
            <span class="hero-badge">✓ Posts automáticos com IA</span>
            <span class="hero-badge">✓ Aprovação por Telegram</span>
            <span class="hero-badge">✓ 7 dias grátis</span>
        </div>
        <div style="display:flex;gap:12px;justify-content:center;flex-wrap:wrap">
            <a href="/register" class="btn btn-primary btn-lg">Começar grátis →</a>
            <a href="/login" class="btn btn-secondary btn-lg">Já tenho conta</a>
        </div>
    </div>

    <div id="features" class="features-grid">
        <div class="feature-card fade-in fade-in-1"><div class="feature-icon">🤖</div><h3>IA Treinável</h3><p>Ensine sobre seus produtos, preços e jeito de atender. A IA aprende o DNA do seu negócio.</p></div>
        <div class="feature-card fade-in fade-in-2"><div class="feature-icon">🎤</div><h3>Entende e Responde Áudio</h3><p>Transcreve áudios e responde por voz automaticamente. Seu cliente fala, a IA fala de volta.</p></div>
        <div class="feature-card fade-in fade-in-3"><div class="feature-icon">📷</div><h3>Analisa Imagens</h3><p>Entende fotos de produtos, comprovantes e documentos enviados.</p></div>
        <div class="feature-card fade-in fade-in-1"><div class="feature-icon">📄</div><h3>Lê PDFs</h3><p>Extrai e processa texto de documentos. Orçamentos, contratos e mais.</p></div>
        <div class="feature-card fade-in fade-in-2"><div class="feature-icon">📊</div><h3>Painel Completo</h3><p>Conversas em tempo real, métricas de atendimento e controle total.</p></div>
        <div class="feature-card fade-in fade-in-3"><div class="feature-icon">⚡</div><h3>Respostas Rápidas</h3><p>Atalhos para mensagens frequentes. Atenda em segundos.</p></div>
    </div>

    <div style="padding:80px 24px 40px;background:linear-gradient(135deg,rgba(99,102,241,0.08),rgba(168,85,247,0.08));border-radius:24px;margin:40px 24px;max-width:1200px;margin-left:auto;margin-right:auto">
        <div style="text-align:center;max-width:800px;margin:0 auto">
            <p style="color:var(--accent2);font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:2px;margin-bottom:12px">🚀 Novidade</p>
            <h2 style="font-size:36px;margin-bottom:16px;letter-spacing:-0.5px">Agência Digital com IA</h2>
            <p style="color:var(--text2);font-size:17px;margin-bottom:40px;line-height:1.6">
                Economize R$ 2.000-5.000/mês que pagaria para uma agência.<br>
                Nossa IA cria posts diários para suas redes sociais e te envia pelo Telegram para aprovar.
            </p>
        </div>
        <div class="grid-3" style="max-width:1000px;margin:0 auto;gap:24px">
            <div class="card fade-in fade-in-1" style="padding:28px;text-align:left">
                <div style="font-size:32px;margin-bottom:12px">🖼️</div>
                <h3 style="margin-bottom:8px">Biblioteca de mídia</h3>
                <p style="color:var(--text2);font-size:14px;line-height:1.6">Envie fotos e vídeos do seu negócio. A IA escolhe a melhor para cada dia e tema (produto, bastidor, motivacional).</p>
            </div>
            <div class="card fade-in fade-in-2" style="padding:28px;text-align:left">
                <div style="font-size:32px;margin-bottom:12px">✨</div>
                <h3 style="margin-bottom:8px">Legendas geradas por IA</h3>
                <p style="color:var(--text2);font-size:14px;line-height:1.6">Claude analisa a imagem, entende seu negócio e cria legendas envolventes com hashtags relevantes. Tom personalizado.</p>
            </div>
            <div class="card fade-in fade-in-3" style="padding:28px;text-align:left">
                <div style="font-size:32px;margin-bottom:12px">📱</div>
                <h3 style="margin-bottom:8px">Aprovação por Telegram</h3>
                <p style="color:var(--text2);font-size:14px;line-height:1.6">Receba a sugestão de post direto no seu Telegram. Aprove com 1 clique ou rejeite se preferir algo diferente.</p>
            </div>
        </div>
        <div style="text-align:center;margin-top:40px">
            <a href="/register" class="btn btn-primary btn-lg">Experimente grátis por 7 dias →</a>
        </div>
    </div>

    <div id="channels" style="padding:80px 24px 40px;max-width:1100px;margin:0 auto;text-align:center">
        <p style="color:var(--accent2);font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:2px;margin-bottom:12px">Multicanal</p>
        <h2 style="font-size:32px;margin-bottom:12px;letter-spacing:-0.5px">Um atendente, todos os canais</h2>
        <p style="color:var(--text2);font-size:16px;margin-bottom:40px">Seu cliente fala por qualquer canal, sua IA responde do mesmo jeito.</p>
        <div class="grid-3" style="gap:20px">
            <div class="card fade-in fade-in-1" style="padding:24px">
                <div style="font-size:36px;margin-bottom:12px">🟢</div>
                <h3 style="margin-bottom:8px">WhatsApp</h3>
                <p style="color:var(--text2);font-size:13px">Canal principal. Texto, áudio, imagens, PDFs, localização.</p>
            </div>
            <div class="card fade-in fade-in-2" style="padding:24px">
                <div style="font-size:36px;margin-bottom:12px">📷</div>
                <h3 style="margin-bottom:8px">Instagram Direct</h3>
                <p style="color:var(--text2);font-size:13px">Atende mensagens diretas do Instagram na mesma interface.</p>
            </div>
            <div class="card fade-in fade-in-3" style="padding:24px">
                <div style="font-size:36px;margin-bottom:12px">💬</div>
                <h3 style="margin-bottom:8px">Messenger</h3>
                <p style="color:var(--text2);font-size:13px">Mensagens da sua página do Facebook respondidas automaticamente.</p>
            </div>
        </div>
    </div>

    <div style="text-align:center;padding:40px 24px 20px">
        <p style="color:var(--accent2);font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:2px;margin-bottom:12px">Como funciona</p>
        <h2 style="font-size:32px;margin-bottom:48px;letter-spacing:-0.5px">Simples como 1, 2, 3</h2>
        <div class="grid-3" style="max-width:900px;margin:0 auto 60px;text-align:center">
            <div class="card fade-in fade-in-1" style="text-align:center;padding:32px">
                <div style="font-size:36px;font-weight:800;color:var(--accent);margin-bottom:12px">1</div>
                <h3 style="margin-bottom:8px">Conecte seu WhatsApp</h3>
                <p style="color:var(--text2);font-size:14px">Vincule seu número do WhatsApp Business em poucos cliques.</p></div>
            <div class="card fade-in fade-in-2" style="text-align:center;padding:32px">
                <div style="font-size:36px;font-weight:800;color:var(--accent);margin-bottom:12px">2</div>
                <h3 style="margin-bottom:8px">Treine a IA</h3>
                <p style="color:var(--text2);font-size:14px">Cadastre produtos, preços, FAQ e o tom de voz da sua empresa.</p></div>
            <div class="card fade-in fade-in-3" style="text-align:center;padding:32px">
                <div style="font-size:36px;font-weight:800;color:var(--accent);margin-bottom:12px">3</div>
                <h3 style="margin-bottom:8px">Venda no automático</h3>
                <p style="color:var(--text2);font-size:14px">A IA atende seus clientes 24/7 enquanto você foca no que importa.</p></div>
        </div>
    </div>

    <div id="pricing" style="text-align:center;padding:60px 24px 80px;max-width:1400px;margin:0 auto">
        <p style="color:var(--accent2);font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:2px;margin-bottom:12px">Planos e preços</p>
        <h2 style="font-size:36px;margin-bottom:12px;letter-spacing:-0.5px">Escolha o plano ideal para você</h2>
        <p style="color:var(--text2);margin-bottom:48px;font-size:16px">7 dias grátis em todos os planos · Cancele quando quiser · Sem fidelidade</p>

        <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(270px,1fr));gap:20px;max-width:1300px;margin:0 auto;text-align:left">

            <!-- STARTER -->
            <div class="plan-card fade-in fade-in-1" style="background:rgba(255,255,255,0.02);border:1px solid rgba(255,255,255,0.06);border-radius:16px;padding:32px 24px;display:flex;flex-direction:column">
                <div style="text-align:center;margin-bottom:20px">
                    <div style="font-size:14px;color:var(--text3);font-weight:600;text-transform:uppercase;letter-spacing:1px">STARTER</div>
                    <div style="font-size:12px;color:var(--text3);margin-top:4px">Para quem está começando</div>
                </div>
                <div style="text-align:center;margin-bottom:24px">
                    <div style="font-size:42px;font-weight:700;color:var(--text)">R$ 97<span style="font-size:16px;color:var(--text3);font-weight:400">/mês</span></div>
                    <div style="color:var(--text3);font-size:13px;margin-top:4px">500 mensagens incluídas</div>
                </div>
                <ul style="list-style:none;padding:0;margin:0 0 24px 0;flex:1">
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">✅ <strong>IA Claude</strong> (modelo top)</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">🎤 Entende áudio</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">🔊 Responde por áudio</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">📷 Analisa imagens</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">📄 Lê PDFs</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">📚 Base de conhecimento</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">📸 Galeria de produtos</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px">💬 Suporte por email</li>
                </ul>
                <a href="/register?plan=starter" class="btn btn-secondary btn-block" style="text-align:center">Começar grátis</a>
            </div>

            <!-- PRO (highlighted) -->
            <div class="plan-card fade-in fade-in-2" style="background:linear-gradient(135deg,rgba(99,102,241,0.08),rgba(168,85,247,0.08));border:2px solid var(--accent2);border-radius:16px;padding:32px 24px;display:flex;flex-direction:column;position:relative;transform:scale(1.02)">
                <div style="position:absolute;top:-12px;left:50%;transform:translateX(-50%);background:linear-gradient(135deg,var(--accent),var(--accent2));padding:4px 16px;border-radius:20px;font-size:11px;font-weight:700;color:white;letter-spacing:1px">⭐ MAIS POPULAR</div>
                <div style="text-align:center;margin-bottom:20px">
                    <div style="font-size:14px;color:var(--accent2);font-weight:600;text-transform:uppercase;letter-spacing:1px">PROFISSIONAL</div>
                    <div style="font-size:12px;color:var(--text3);margin-top:4px">Atendimento + Agência Digital</div>
                </div>
                <div style="text-align:center;margin-bottom:24px">
                    <div style="font-size:42px;font-weight:700;color:var(--text)">R$ 197<span style="font-size:16px;color:var(--text3);font-weight:400">/mês</span></div>
                    <div style="color:var(--text3);font-size:13px;margin-top:4px">2.000 mensagens incluídas</div>
                </div>
                <ul style="list-style:none;padding:0;margin:0 0 24px 0;flex:1">
                    <li style="padding:8px 0;color:var(--text);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)"><strong>Tudo do Starter +</strong></li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">📸 <strong>Agência Digital com IA</strong></li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">✨ Posts automáticos</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">📱 Aprovação via Telegram</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">🗓️ Agenda de publicações</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">⚡ Respostas rápidas</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">📥 Exportação de conversas</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px">🎯 Suporte prioritário</li>
                </ul>
                <a href="/register?plan=pro" class="btn btn-primary btn-block" style="text-align:center">Começar teste grátis</a>
            </div>

            <!-- BUSINESS -->
            <div class="plan-card fade-in fade-in-3" style="background:rgba(255,255,255,0.02);border:1px solid rgba(255,255,255,0.06);border-radius:16px;padding:32px 24px;display:flex;flex-direction:column;position:relative">
                <div style="position:absolute;top:-12px;right:20px;background:#10b981;padding:4px 12px;border-radius:12px;font-size:10px;font-weight:700;color:white;letter-spacing:1px">💰 VENDEDOR</div>
                <div style="text-align:center;margin-bottom:20px">
                    <div style="font-size:14px;color:var(--text3);font-weight:600;text-transform:uppercase;letter-spacing:1px">BUSINESS</div>
                    <div style="font-size:12px;color:var(--text3);margin-top:4px">Vendas + Campanhas + Agência</div>
                </div>
                <div style="text-align:center;margin-bottom:24px">
                    <div style="font-size:42px;font-weight:700;color:var(--text)">R$ 397<span style="font-size:16px;color:var(--text3);font-weight:400">/mês</span></div>
                    <div style="color:var(--text3);font-size:13px;margin-top:4px">10.000 mensagens incluídas</div>
                </div>
                <ul style="list-style:none;padding:0;margin:0 0 24px 0;flex:1">
                    <li style="padding:8px 0;color:var(--text);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)"><strong>Tudo do Profissional +</strong></li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">🛒 <strong>Venda direto no WhatsApp</strong></li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">💳 PIX + Cartão via Mercado Pago</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">📢 <strong>Campanhas em massa</strong></li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">🎯 CRM com funil de vendas</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">📊 Analytics avançado</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">🏷️ Tags e segmentação</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px">💬 Suporte via WhatsApp</li>
                </ul>
                <a href="/register?plan=business" class="btn btn-secondary btn-block" style="text-align:center">Começar teste grátis</a>
            </div>

            <!-- AGÊNCIA -->
            <div class="plan-card fade-in fade-in-3" style="background:linear-gradient(135deg,rgba(251,191,36,0.05),rgba(168,85,247,0.05));border:1px solid rgba(251,191,36,0.2);border-radius:16px;padding:32px 24px;display:flex;flex-direction:column;position:relative">
                <div style="position:absolute;top:-12px;right:20px;background:linear-gradient(135deg,#fbbf24,#a855f7);padding:4px 12px;border-radius:12px;font-size:10px;font-weight:700;color:white;letter-spacing:1px">👑 PREMIUM</div>
                <div style="text-align:center;margin-bottom:20px">
                    <div style="font-size:14px;color:#fbbf24;font-weight:600;text-transform:uppercase;letter-spacing:1px">AGÊNCIA</div>
                    <div style="font-size:12px;color:var(--text3);margin-top:4px">Para revender como serviço</div>
                </div>
                <div style="text-align:center;margin-bottom:24px">
                    <div style="font-size:42px;font-weight:700;color:var(--text)">R$ 997<span style="font-size:16px;color:var(--text3);font-weight:400">/mês</span></div>
                    <div style="color:var(--text3);font-size:13px;margin-top:4px">50.000 mensagens + multi-conta</div>
                </div>
                <ul style="list-style:none;padding:0;margin:0 0 24px 0;flex:1">
                    <li style="padding:8px 0;color:var(--text);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)"><strong>Tudo do Business +</strong></li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">🏢 <strong>Até 10 contas de cliente</strong></li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">👥 Sub-usuários atendentes</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">♾️ Campanhas ilimitadas</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">♾️ Posts automáticos ilimitados</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">🔌 API dedicada</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px;border-bottom:1px solid rgba(255,255,255,0.04)">📞 Suporte telefônico</li>
                    <li style="padding:8px 0;color:var(--text2);font-size:14px">🚀 Onboarding personalizado</li>
                </ul>
                <a href="/register?plan=agency" class="btn btn-secondary btn-block" style="text-align:center">Começar teste grátis</a>
            </div>

        </div>

        <div style="margin-top:48px;padding:24px;background:rgba(99,102,241,0.04);border:1px solid rgba(99,102,241,0.15);border-radius:12px;max-width:900px;margin-left:auto;margin-right:auto;text-align:left">
            <div style="display:flex;gap:16px;align-items:center;flex-wrap:wrap">
                <div style="font-size:32px">💡</div>
                <div style="flex:1;min-width:250px">
                    <h3 style="color:var(--text);margin:0 0 4px 0;font-size:16px">Por que pagar múltiplas ferramentas?</h3>
                    <p style="color:var(--text2);font-size:13px;margin:0;line-height:1.6">
                        Chatbot comum: R$ 597 · Ferramenta de campanhas: R$ 400 · Agência de conteúdo: R$ 2.000+<br>
                        <strong style="color:var(--accent2)">Total: R$ 2.997/mês.</strong>
                        Com atendente.online, tudo incluso por R$ 397/mês. <strong style="color:#10b981">Economize 87%.</strong>
                    </p>
                </div>
            </div>
        </div>

        <div style="margin-top:32px;color:var(--text3);font-size:12px;max-width:800px;margin-left:auto;margin-right:auto">
            💳 Aceitamos Cartão, PIX e Boleto · 🔒 Pagamento seguro via Mercado Pago · ❌ Cancele quando quiser, sem multa
        </div>
    </div>

    <footer style="text-align:center;padding:40px 24px;border-top:1px solid rgba(255,255,255,0.06);color:var(--text3);font-size:13px">
        <p>© 2026 atendente.online — Todos os direitos reservados</p>
        <p style="margin-top:8px"><a href="/privacy">Política de Privacidade</a> · <a href="/terms">Termos de Serviço</a></p>
        <p style="margin-top:12px;color:var(--text3);font-size:12px">Desenvolvido por <strong style="color:var(--text2)">Clériston Almeida Capistrano</strong> · <a href="mailto:contato@atendente.online" style="color:var(--accent2)">contato@atendente.online</a></p>
    </footer>"""
    return base_html("Atendente IA para WhatsApp", content)


@app.route("/register", methods=["GET","POST"])
def register():
    error = ""
    if request.method == "POST":
        name = request.form.get("name","").strip()
        email = request.form.get("email","").strip().lower()
        password = request.form.get("password","")
        company = request.form.get("company","").strip()
        plan = request.form.get("plan","starter")
        # LGPD: aceite obrigatório de Política, Termos e DPA
        accept_terms = request.form.get("accept_terms") == "on"
        accept_marketing = request.form.get("accept_marketing") == "on"  # opcional
        if not name or not email or not password:
            error = "Preencha todos os campos obrigatórios."
        elif not accept_terms:
            error = "Você precisa aceitar a Política de Privacidade, os Termos de Serviço e o DPA para criar uma conta."
        else:
            pw_ok, pw_msg = validate_password_strength(password)
            if not pw_ok:
                error = pw_msg
            else:
                db = get_db()
                existing = db.execute("SELECT id, email_verified FROM users WHERE email=?", (email,)).fetchone()
                if existing and existing["email_verified"]:
                    error = "Este email já está cadastrado."
                else:
                    if existing and not existing["email_verified"]:
                        db.execute("DELETE FROM users WHERE id=?", (existing["id"],))
                    trial_end = (datetime.now() + timedelta(days=7)).isoformat()
                    msgs_limit = PLANS.get(plan, PLANS["starter"])["msgs"]
                    db.execute("INSERT INTO users (email,password_hash,name,company,plan,plan_status,msgs_limit,trial_ends_at,email_verified) VALUES (?,?,?,?,?,?,?,?,0)",
                        (email, hash_password(password), name, company, plan, "trial", msgs_limit, trial_end))
                    db.commit()
                    # LGPD: registrar consentimentos formais para comprovação perante ANPD
                    new_user = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
                    new_user_id = new_user["id"] if new_user else None
                    register_consent(new_user_id, email, "privacy_policy", PRIVACY_POLICY_VERSION, accepted=True,
                                     details=f"Aceite no signup. Plano: {plan}")
                    register_consent(new_user_id, email, "terms_of_service", TERMS_OF_SERVICE_VERSION, accepted=True,
                                     details=f"Aceite no signup. Plano: {plan}")
                    register_consent(new_user_id, email, "dpa", DPA_VERSION, accepted=True,
                                     details=f"Aceite no signup. Plano: {plan}")
                    register_consent(new_user_id, email, "data_processing", PRIVACY_POLICY_VERSION, accepted=True,
                                     details="Base legal: execução de contrato (LGPD Art. 7º, V)")
                    if accept_marketing:
                        register_consent(new_user_id, email, "marketing_email", PRIVACY_POLICY_VERSION, accepted=True,
                                         details="Opt-in voluntário no signup")
                    send_verification_code(email)
                    session["pending_email"] = email
                    return redirect("/verify-email")
    plan = request.args.get("plan","starter")
    alert = f'<div class="alert alert-error">{error}</div>' if error else ""
    content = f"""<div class="auth-container"><div class="auth-card">
        <a href="/" style="display:block;text-align:center;margin-bottom:24px"><img src="data:image/png;base64,{LOGO_NAV_B64}" alt="atendente.online" style="height:56px"></a><h2>Criar conta grátis</h2>{alert}
        <form method="POST">{csrf_field()}<input type="hidden" name="plan" value="{plan}">
        <div class="form-group"><label class="form-label">Seu nome *</label><input type="text" name="name" class="form-input" required></div>
        <div class="form-group"><label class="form-label">Email *</label><input type="email" name="email" class="form-input" required></div>
        <div class="form-group"><label class="form-label">Empresa</label><input type="text" name="company" class="form-input"></div>
        <div class="form-group"><label class="form-label">Senha *</label><input type="password" name="password" class="form-input" required></div>

        <div class="form-group" style="background:var(--bg2);padding:14px;border-radius:8px;margin-top:20px">
            <label style="display:flex;align-items:flex-start;gap:10px;cursor:pointer;font-size:13px;line-height:1.5;color:var(--text2)">
                <input type="checkbox" name="accept_terms" required style="margin-top:3px;flex-shrink:0;width:16px;height:16px">
                <span>Li e aceito a <a href="/privacy" target="_blank" style="color:var(--accent2)">Política de Privacidade</a>,
                os <a href="/terms" target="_blank" style="color:var(--accent2)">Termos de Serviço</a> e o
                <a href="/dpa" target="_blank" style="color:var(--accent2)">Contrato de Operador (DPA)</a>. *</span>
            </label>
        </div>

        <div class="form-group" style="padding:0 14px">
            <label style="display:flex;align-items:flex-start;gap:10px;cursor:pointer;font-size:13px;line-height:1.5;color:var(--text2)">
                <input type="checkbox" name="accept_marketing" style="margin-top:3px;flex-shrink:0;width:16px;height:16px">
                <span>Aceito receber novidades, dicas e ofertas por email (opcional, pode ser cancelado a qualquer momento).</span>
            </label>
        </div>

        <button type="submit" class="btn btn-primary btn-block btn-lg" style="margin-top:8px">Criar conta →</button></form>
        <div class="auth-divider">Já tem conta? <a href="/login">Entrar</a></div></div></div>"""
    return base_html("Criar Conta", content)


@app.route("/verify-email", methods=["GET","POST"])
def verify_email():
    email = session.get("pending_email", "")
    if not email:
        return redirect("/register")

    error = request.args.get("error", "")
    success = ""
    if request.method == "POST":
        code = request.form.get("code","").strip()
        if not code or len(code) != 6:
            error = "Digite o código de 6 dígitos."
        elif verify_code(email, code):
            db = get_db()
            user = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
            if user:
                session.pop("pending_email", None)
                session["user_id"] = user["id"]
                return redirect("/dashboard")
            error = "Erro ao ativar conta. Tente novamente."
        else:
            error = "Código inválido ou expirado."

    alert = f'<div class="alert alert-error">{error}</div>' if error else ""
    masked = email[:3] + "***" + email[email.index("@"):] if "@" in email else email
    content = f"""<div class="auth-container"><div class="auth-card" style="text-align:center">
        <a href="/" style="display:block;margin-bottom:24px"><img src="data:image/png;base64,{LOGO_NAV_B64}" alt="atendente.online" style="height:56px"></a>
        <div style="font-size:48px;margin-bottom:16px">📧</div>
        <h2>Verifique seu email</h2>
        <p style="color:var(--text2);margin-bottom:24px">Enviamos um código de 6 dígitos para<br><strong style="color:var(--accent2)">{masked}</strong></p>
        {alert}
        <form method="POST">{csrf_field()}
        <div class="form-group"><input type="text" name="code" class="form-input" placeholder="000000" maxlength="6"
            style="text-align:center;font-size:28px;letter-spacing:8px;font-weight:700" required autofocus></div>
        <button type="submit" class="btn btn-primary btn-block btn-lg">Verificar →</button></form>
        <div style="margin-top:20px">
            <a href="/resend-code" style="color:var(--text2);font-size:13px">Não recebeu? Reenviar código</a>
        </div></div></div>"""
    return base_html("Verificar Email", content)


@app.route("/resend-code")
def resend_code():
    """Reenvia código de verificação por email. Cooldown 60s + rate limit por IP."""
    email = session.get("pending_email", "")
    if not email:
        return redirect("/register")

    client_ip = request.remote_addr or "unknown"
    # Rate limit por IP: 5 reenvios em 1h (impede flood mesmo com vários emails)
    if not check_rate_limit(client_ip, max_attempts=5, window=3600):
        safe_log(f"[RESEND] Bloqueado por rate limit: {mask_email(email)} de {client_ip}", level="WARN")
        return redirect("/verify-email?error=Muitas+tentativas.+Aguarde+1+hora.")

    # Cooldown 60s por email: evita flood do mesmo destinatário
    try:
        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row
        recent = db_conn.execute(
            """SELECT created_at FROM verification_codes
               WHERE email=? ORDER BY created_at DESC LIMIT 1""", (email,)
        ).fetchone()
        db_conn.close()
        if recent and recent["created_at"]:
            try:
                last_time = datetime.fromisoformat(recent["created_at"])
                elapsed = (datetime.now() - last_time).total_seconds()
                if elapsed < 60:
                    wait = int(60 - elapsed)
                    safe_log(f"[RESEND] Cooldown ativo para {mask_email(email)} (faltam {wait}s)", level="WARN")
                    return redirect(f"/verify-email?error=Aguarde+{wait}s+antes+de+reenviar.")
            except Exception:
                pass
    except Exception as e:
        safe_log(f"[RESEND] Erro verificando cooldown: {e}", level="WARN")

    # Conta no rate limiter (mesmo que o envio falhe)
    record_login_attempt(client_ip, window=3600)
    send_verification_code(email)
    return redirect("/verify-email")


@app.route("/login", methods=["GET","POST"])
def login():
    error = request.args.get("error", "")
    client_ip = request.remote_addr or "unknown"
    # Etapa 2FA pendente?
    pending_2fa_user_id = session.get("pending_2fa_user_id")
    step_2fa = bool(pending_2fa_user_id)

    if request.method == "POST":
        # LGPD/Segurança: 3 tentativas em 15 min (rigoroso anti brute-force)
        if not check_rate_limit(client_ip, max_attempts=3, window=900):
            error = "Muitas tentativas de login. Aguarde 15 minutos antes de tentar novamente."
        elif step_2fa:
            # Etapa 2: valida código TOTP ou backup code do usuário
            db = get_db()
            user_row = db.execute("SELECT * FROM users WHERE id=?", (pending_2fa_user_id,)).fetchone()
            if not user_row:
                session.pop("pending_2fa_user_id", None)
                error = "Sessão de login expirada. Faça login novamente."
            else:
                user_dec = decrypt_user_row(user_row)
                code = request.form.get("totp_code", "").strip()
                code_ok = verify_totp_code(user_dec.get("totp_secret", ""), code)
                # Tenta backup code se TOTP falhar
                if not code_ok and code:
                    backup_raw = user_dec.get("totp_backup_codes", "") or ""
                    if backup_raw:
                        codes = [c.strip() for c in backup_raw.split(",") if c.strip()]
                        code_upper = code.upper().replace(" ", "").replace("-", "")
                        # Aceita com ou sem hífen
                        normalized_codes = [c.replace("-", "") for c in codes]
                        if code_upper in normalized_codes:
                            idx = normalized_codes.index(code_upper)
                            codes.pop(idx)
                            new_backup = ",".join(codes)
                            # Re-criptografa antes de salvar
                            enc_backup = _encrypt_value(new_backup) if new_backup else ""
                            db.execute("UPDATE users SET totp_backup_codes=? WHERE id=?", (enc_backup, pending_2fa_user_id))
                            db.commit()
                            code_ok = True
                            safe_log(f"[2FA] Backup code consumido (user_id={pending_2fa_user_id}, restantes={len(codes)})")
                if code_ok:
                    reset_login_attempts(client_ip)
                    session.pop("pending_2fa_user_id", None)
                    session.clear()
                    session.permanent = True
                    session["user_id"] = pending_2fa_user_id
                    db.execute("UPDATE users SET last_login=datetime('now') WHERE id=?", (pending_2fa_user_id,))
                    db.commit()
                    return redirect("/dashboard")
                else:
                    record_login_attempt(client_ip, window=900)
                    error = "Código inválido. Tente novamente ou use um backup code."
        else:
            email = request.form.get("email","").strip().lower()
            password = request.form.get("password","")
            db = get_db()
            user = db.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
            if user and check_password(password, user["password_hash"]):
                maybe_upgrade_password_hash(user["id"], password, user["password_hash"])
                user_dec = decrypt_user_row(user)
                if not user_dec.get("is_active", 1):
                    error = "Conta desativada. Entre em contato com o suporte."
                elif not user_dec.get("email_verified", 0):
                    session["pending_email"] = email
                    send_verification_code(email)
                    return redirect("/verify-email")
                elif user_2fa_enabled(user_dec):
                    # Senha OK + 2FA ativo → segue para etapa de código
                    session["pending_2fa_user_id"] = user["id"]
                    step_2fa = True
                    safe_log(f"[2FA] Senha OK, aguardando código (user_id={user['id']})")
                else:
                    reset_login_attempts(client_ip)
                    session.clear()
                    session.permanent = True
                    session["user_id"] = user["id"]
                    db.execute("UPDATE users SET last_login=datetime('now') WHERE id=?", (user["id"],))
                    db.commit()
                    return redirect("/dashboard")
            else:
                # Registra com a mesma janela (15 min) usada no check
                record_login_attempt(client_ip, window=900)
                error = "Email ou senha incorretos."

    # Cancelar 2FA pendente
    if request.args.get("cancel_2fa") == "1":
        session.pop("pending_2fa_user_id", None)
        return redirect("/login")

    alert = f'<div class="alert alert-error">{error}</div>' if error else ""

    if step_2fa:
        info_2fa = '<div class="alert alert-info" style="margin-bottom:16px">🔐 Digite o código de 6 dígitos do seu app autenticador (Google Authenticator, Authy, etc).</div>'
        content = f"""<div class="auth-container"><div class="auth-card">
            <a href="/" style="display:block;text-align:center;margin-bottom:24px"><img src="data:image/png;base64,{LOGO_NAV_B64}" alt="atendente.online" style="height:56px"></a>
            <h2>Verificação 2FA</h2>{alert}{info_2fa}
            <form method="POST">{csrf_field()}
                <div class="form-group">
                    <label class="form-label">Código (6 dígitos) ou Backup Code</label>
                    <input type="text" name="totp_code" class="form-input" required maxlength="12" autocomplete="off" autofocus
                           placeholder="000000" style="font-size:22px;letter-spacing:6px;text-align:center;font-family:monospace">
                    <small style="color:var(--text3);font-size:12px;margin-top:6px;display:block">Backup codes têm formato XXXX-XXXX (use um se perdeu o celular).</small>
                </div>
                <button type="submit" class="btn btn-primary btn-block btn-lg">Verificar</button>
            </form>
            <div class="auth-divider"><a href="/login?cancel_2fa=1">Cancelar e voltar</a></div>
        </div></div>"""
        return base_html("Verificação 2FA", content)

    content = f"""<div class="auth-container"><div class="auth-card">
        <a href="/" style="display:block;text-align:center;margin-bottom:24px"><img src="data:image/png;base64,{LOGO_NAV_B64}" alt="atendente.online" style="height:56px"></a><h2>Entrar</h2>{alert}
        <form method="POST">{csrf_field()}
        <div class="form-group"><label class="form-label">Email</label><input type="email" name="email" class="form-input" required></div>
        <div class="form-group"><label class="form-label">Senha</label><input type="password" name="password" class="form-input" required></div>
        <button type="submit" class="btn btn-primary btn-block btn-lg">Entrar</button></form>
        <div class="auth-divider">Não tem conta? <a href="/register">Criar conta grátis</a></div></div></div>"""
    return base_html("Login", content)

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


# ─── DASHBOARD ─────────────────────────────────────────────────
@app.route("/dashboard")
@login_required
def dashboard():
    user = g.user
    stats = get_user_stats(user["id"])
    plan = PLANS.get(user["plan"], PLANS["starter"])
    usage_pct = min(100, int((user["msgs_used"] / max(user["msgs_limit"],1)) * 100))
    usage_color = "var(--green)" if usage_pct < 70 else "var(--orange)" if usage_pct < 90 else "var(--red)"
    plan_badge = '<span class="badge badge-green">ATIVO</span>' if user["plan_status"]=="active" else '<span class="badge badge-orange">TRIAL</span>' if user["plan_status"]=="trial" else '<span class="badge badge-red">INATIVO</span>'

    db = get_db()
    recent = db.execute("""SELECT c.*, (SELECT content FROM messages WHERE conversation_id=c.id ORDER BY created_at DESC LIMIT 1) as last_msg
        FROM conversations c WHERE c.user_id=? ORDER BY c.last_message_at DESC LIMIT 5""", (user["id"],)).fetchall()
    convos_html = ""
    if recent:
        rows = "".join(f"""<tr><td><strong>{esc(c['customer_phone'])}</strong><br><span style="color:var(--text3);font-size:12px">{esc(c['customer_name'] or 'Sem nome')}</span></td>
            <td style="max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{esc((c['last_msg'] or '—')[:60])}</td>
            <td>{'<span class="badge badge-green">Ativa</span>' if c['status']=='active' else '<span class="badge badge-orange">Finalizada</span>'}{' <span class="badge badge-purple">Humano</span>' if c['is_human_takeover'] else ''}</td>
            <td style="color:var(--text3);font-size:12px">{esc((c['last_message_at'] or '')[:16])}</td></tr>""" for c in recent)
        convos_html = f'<div class="card"><div class="card-header"><span class="card-title">Conversas recentes</span><a href="/dashboard/conversations" class="btn btn-secondary btn-sm">Ver todas →</a></div><div class="table-wrap"><table><thead><tr><th>Cliente</th><th>Última msg</th><th>Status</th><th>Hora</th></tr></thead><tbody>{rows}</tbody></table></div></div>'
    else:
        convos_html = '<div class="card"><div class="empty-state"><div class="icon">💬</div><h3>Nenhuma conversa ainda</h3><p>Configure seu WhatsApp para começar.</p><a href="/dashboard/settings" class="btn btn-primary" style="margin-top:16px">Configurar →</a></div></div>'

    content = f"""<div class="container">
        <div class="page-header fade-in"><h1>Olá, {esc(user['name'].split()[0])}! 👋</h1><p>Plano {plan['name']} {plan_badge}</p></div>
        <div class="grid-4">
            <div class="stat-card fade-in fade-in-1"><div class="stat-icon stat-icon-green">💬</div><div class="stat-value">{stats['conversations']}</div><div class="stat-label">Conversas totais</div></div>
            <div class="stat-card fade-in fade-in-2"><div class="stat-icon stat-icon-blue">📨</div><div class="stat-value">{stats['today_messages']}</div><div class="stat-label">Mensagens hoje</div></div>
            <div class="stat-card fade-in fade-in-3"><div class="stat-icon stat-icon-purple">🧠</div><div class="stat-value">{stats['knowledge_items']}</div><div class="stat-label">Base de conhecimento</div></div>
            <div class="stat-card fade-in fade-in-4"><div class="stat-icon stat-icon-orange">📊</div><div class="stat-value">{usage_pct}%</div><div class="stat-label">Uso ({user['msgs_used']}/{user['msgs_limit']})</div>
                <div class="usage-bar-bg" style="margin-top:8px"><div class="usage-bar-fill" style="width:{usage_pct}%;background:{usage_color}"></div></div></div>
        </div>{convos_html}</div>"""
    return base_html("Dashboard", content, dict(user))


# ─── CONVERSATIONS ─────────────────────────────────────────────
@app.route("/dashboard/conversations")
@login_required
def conversations():
    db = get_db()
    user = g.user
    convos = db.execute("""SELECT c.*,(SELECT content FROM messages WHERE conversation_id=c.id ORDER BY created_at DESC LIMIT 1) as last_msg,
        (SELECT COUNT(*) FROM messages WHERE conversation_id=c.id) as msg_count FROM conversations c WHERE c.user_id=? ORDER BY c.last_message_at DESC""", (user["id"],)).fetchall()

    sidebar_items = ""
    first_id = None
    msgs_html = ""
    channel_icons = {"whatsapp": "🟢", "instagram": "📷", "messenger": "💬"}
    for c in convos:
        if not first_id: first_id = c["id"]
        active = "active" if c["id"] == first_id else ""
        name = esc(c["customer_name"] or c["customer_phone"])
        preview = esc((c["last_msg"] or "Sem mensagens")[:50])
        date = esc(to_br_date(c["last_message_at"]))
        # Canal: whatsapp por padrão (para conversas antigas)
        try:
            channel = c["channel"] or "whatsapp"
        except (KeyError, IndexError):
            channel = "whatsapp"
        icon = channel_icons.get(channel, "🟢")
        sidebar_items += f'<div class="chat-item {active}" data-conv-id="{int(c["id"])}"><span class="chat-item-time">{icon} {date}</span><div class="chat-item-name">{name}</div><div class="chat-item-preview">{preview}</div></div>'

    if first_id:
        messages = db.execute("SELECT * FROM messages WHERE conversation_id=? ORDER BY created_at", (first_id,)).fetchall()
        for m in messages:
            cls = "msg-bot" if m["sender"]=="bot" else "msg-customer"
            content = esc(m["content"])
            msg_type = esc(m["msg_type"])
            media_tag = f'<div class="msg-media">{msg_type}</div>' if m["msg_type"] not in ("text","") else ""
            time_str = esc(to_br_time(m["created_at"]))
            msgs_html += f'<div class="msg {cls}">{content}{media_tag}<div class="msg-time">{time_str}</div></div>'

    if not convos:
        return base_html("Conversas", '<div class="container"><div class="card"><div class="empty-state"><div class="icon">💬</div><h3>Nenhuma conversa ainda</h3><p>As conversas aparecerão aqui quando clientes enviarem mensagens.</p></div></div></div>', dict(user))

    first_name = esc(convos[0]['customer_name'] or convos[0]['customer_phone'])
    first_phone = esc(convos[0]['customer_phone'])
    content = f"""<div class="container"><div class="page-header"><h1>Conversas <span style="display:inline-flex;align-items:center;gap:6px;font-size:13px;color:var(--green2);font-weight:500;background:rgba(0,200,150,0.1);padding:4px 12px;border-radius:20px;vertical-align:middle"><span style="width:8px;height:8px;border-radius:50%;background:var(--green2);display:inline-block;animation:pulse 2s infinite"></span> ao vivo</span></h1><p>{len(convos)} conversas <a href="/dashboard/conversations/export" class="btn btn-sm btn-secondary" style="margin-left:12px">📥 Exportar CSV</a></p></div>
    <style>@keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:0.3}}}}</style>
        <div class="chat-container"><div class="chat-sidebar"><div class="chat-sidebar-header">
            <input type="text" class="form-input" id="search-chats" placeholder="Buscar..." style="font-size:13px;padding:8px 12px">
            </div><div id="chat-list">{sidebar_items}</div></div>
            <div class="chat-main"><div class="chat-header"><div><strong id="chat-name">{first_name}</strong>
                <span style="color:var(--text3);font-size:12px" id="chat-phone">{first_phone}</span></div>
                <div><a href="/dashboard/conversations/{first_id}/export" class="btn btn-sm btn-secondary" style="margin-right:8px">📥 Exportar</a>
                <button class="btn btn-secondary btn-sm" id="assume-btn">🙋 Assumir</button></div></div>
                <div class="chat-messages" id="chat-messages">{msgs_html}</div>
                <div style="padding:16px 24px;border-top:1px solid rgba(255,255,255,0.06);display:flex;gap:8px">
                    <input type="text" class="form-input" id="msg-input" placeholder="Digite..." style="flex:1">
                    <button class="btn btn-primary" id="send-btn">Enviar</button></div></div></div></div>
    <script nonce="{g.csp_nonce}">
    let activeConvId = {first_id or 0};
    let lastMsgCount = 0;
    const CSRF_TOKEN = '{generate_csrf_token()}';

    function formatBrTime(utcStr){{
        // Converte 'YYYY-MM-DD HH:MM:SS' (UTC) para HH:MM de Brasília
        if(!utcStr) return '';
        try {{
            // SQLite retorna sem timezone — adiciona 'Z' para marcar como UTC
            const iso = utcStr.replace(' ', 'T') + 'Z';
            const d = new Date(iso);
            if(isNaN(d)) return utcStr.substring(11,16);
            return d.toLocaleTimeString('pt-BR', {{hour:'2-digit', minute:'2-digit', timeZone:'America/Sao_Paulo'}});
        }} catch(e) {{
            return utcStr.substring(11,16);
        }}
    }}

    function formatBrDate(utcStr){{
        if(!utcStr) return '';
        try {{
            const iso = utcStr.replace(' ', 'T') + 'Z';
            const d = new Date(iso);
            if(isNaN(d)) return utcStr.substring(0,10);
            return d.toLocaleDateString('pt-BR', {{timeZone:'America/Sao_Paulo'}});
        }} catch(e) {{
            return utcStr.substring(0,10);
        }}
    }}

    function renderMessages(box, messages){{
        box.innerHTML='';
        messages.forEach(m=>{{
            const div=document.createElement('div');
            // bot e human aparecem do lado direito (saída); customer do lado esquerdo (entrada)
            const isFromUs = m.sender==='bot' || m.sender==='human';
            div.className='msg '+(isFromUs?'msg-bot':'msg-customer');
            div.textContent=m.content;
            const t=document.createElement('div');
            t.className='msg-time';
            // Indica visualmente quando foi humano vs bot
            const senderLabel = m.sender==='human' ? '👤 ' : (m.sender==='bot' ? '🤖 ' : '');
            t.textContent=senderLabel+formatBrTime(m.created_at);
            div.appendChild(t);
            box.appendChild(div);
        }});
        box.scrollTop=box.scrollHeight;
    }}

    function loadConversation(id,el){{
        activeConvId = id;
        document.querySelectorAll('.chat-item').forEach(i=>i.classList.remove('active'));
        if(el) el.classList.add('active');
        fetch('/api/conversations/'+id+'/messages').then(r=>r.json()).then(data=>{{
            renderMessages(document.getElementById('chat-messages'), data.messages);
            lastMsgCount = data.messages.length;
            document.getElementById('chat-name').textContent=data.customer_name||data.customer_phone;
            document.getElementById('chat-phone').textContent=data.customer_phone;
            updateAssumeBtn(data.is_human_takeover === 1);
        }});
    }}

    function refreshMessages(){{
        if(!activeConvId) return;
        fetch('/api/conversations/'+activeConvId+'/messages').then(r=>r.json()).then(data=>{{
            if(data.messages.length !== lastMsgCount){{
                renderMessages(document.getElementById('chat-messages'), data.messages);
                lastMsgCount = data.messages.length;
            }}
        }}).catch(()=>{{}});
    }}

    function refreshSidebar(){{
        fetch('/api/conversations').then(r=>r.json()).then(data=>{{
            const list = document.getElementById('chat-list');
            if(!data.conversations) return;
            list.innerHTML='';
            data.conversations.forEach(c=>{{
                const div=document.createElement('div');
                div.className='chat-item '+(c.id===activeConvId?'active':'');
                div.dataset.convId = c.id;
                const time=document.createElement('span');
                time.className='chat-item-time';
                time.textContent=formatBrDate(c.last_message_at);
                const name=document.createElement('div');
                name.className='chat-item-name';
                name.textContent=c.customer_name||c.customer_phone;
                const preview=document.createElement('div');
                preview.className='chat-item-preview';
                preview.textContent=(c.last_msg||'Sem mensagens').substring(0,50);
                div.appendChild(time);div.appendChild(name);div.appendChild(preview);
                list.appendChild(div);
            }});
        }}).catch(()=>{{}});
    }}

    setInterval(refreshMessages, 3000);
    setInterval(refreshSidebar, 10000);

    function sendMsg(){{
        if(!activeConvId){{ alert('Selecione uma conversa primeiro'); return; }}
        const i = document.getElementById('msg-input');
        const msg = i.value.trim();
        if(!msg) return;

        // Desabilita botão durante envio
        const sendBtn = document.getElementById('send-btn');
        if(sendBtn) sendBtn.disabled = true;

        // Adiciona mensagem na tela com indicador "enviando..."
        const b = document.getElementById('chat-messages');
        const div = document.createElement('div');
        div.className = 'msg msg-bot';
        div.style.opacity = '0.6';
        div.textContent = msg;
        const t = document.createElement('div');
        t.className = 'msg-time';
        t.textContent = 'enviando...';
        div.appendChild(t);
        b.appendChild(div);
        b.scrollTop = b.scrollHeight;
        i.value = '';

        // Envia para o backend que envia via WhatsApp API
        fetch('/api/conversations/' + activeConvId + '/send', {{
            method: 'POST',
            headers: {{
                'Content-Type': 'application/json',
                'X-CSRF-Token': CSRF_TOKEN
            }},
            body: JSON.stringify({{message: msg}})
        }})
        .then(r => r.json().then(d => ({{ok: r.ok, data: d}})))
        .then(({{ok, data}}) => {{
            if(sendBtn) sendBtn.disabled = false;
            if(ok && data.success){{
                div.style.opacity = '1';
                t.textContent = 'enviado ✓';
                // Recarrega mensagens para sincronizar
                setTimeout(refreshMessages, 500);
            }} else {{
                div.style.background = 'rgba(239,68,68,0.2)';
                div.style.borderLeft = '3px solid #ef4444';
                t.textContent = '❌ ' + (data.error || 'falha ao enviar');
                t.style.color = '#ef4444';
            }}
        }})
        .catch(err => {{
            if(sendBtn) sendBtn.disabled = false;
            div.style.background = 'rgba(239,68,68,0.2)';
            t.textContent = '❌ erro de conexão';
            t.style.color = '#ef4444';
            console.error('Erro ao enviar:', err);
        }});
    }}
    function filterChats(q){{document.querySelectorAll('.chat-item').forEach(i=>{{i.style.display=i.textContent.toLowerCase().includes(q.toLowerCase())?'':'none'}})}}
    function updateAssumeBtn(isHuman){{
        const btn = document.getElementById('assume-btn');
        if(!btn) return;
        if(isHuman){{
            btn.textContent = '🤖 Devolver para IA';
            btn.className = 'btn btn-primary btn-sm';
            btn.dataset.takeover = '1';
        }} else {{
            btn.textContent = '🙋 Assumir';
            btn.className = 'btn btn-secondary btn-sm';
            btn.dataset.takeover = '0';
        }}
    }}

    function toggleHuman(){{
        if(!activeConvId){{ alert('Selecione uma conversa primeiro'); return; }}
        const btn = document.getElementById('assume-btn');
        if(btn) btn.disabled = true;
        fetch('/api/conversations/' + activeConvId + '/toggle-human', {{
            method: 'POST',
            headers: {{
                'Content-Type': 'application/json',
                'X-CSRF-Token': CSRF_TOKEN
            }}
        }})
        .then(r => r.json())
        .then(data => {{
            if(btn) btn.disabled = false;
            if(data.success){{
                updateAssumeBtn(data.is_human_takeover === 1);
                if(data.is_human_takeover === 1){{
                    alert('✅ Você assumiu. A IA não responderá mais nesta conversa.');
                }} else {{
                    alert('🤖 IA retomou. Próximas mensagens serão respondidas automaticamente.');
                }}
            }} else {{
                alert('❌ Erro: ' + (data.error || 'falha ao alterar'));
            }}
        }})
        .catch(err => {{
            if(btn) btn.disabled = false;
            alert('❌ Erro de conexão: ' + err.message);
        }});
    }}

    const box=document.getElementById('chat-messages');
    if(box) box.scrollTop=box.scrollHeight;

    // ─── EVENT LISTENERS (CSP-compliant) ─────────────────────
    // Substitui onclick="..." inline por addEventListener,
    // que é permitido pelo CSP nonce.
    const sendBtn = document.getElementById('send-btn');
    if(sendBtn){{
        sendBtn.addEventListener('click', sendMsg);
    }}

    const msgInput = document.getElementById('msg-input');
    if(msgInput){{
        msgInput.addEventListener('keydown', function(e){{
            if(e.key === 'Enter'){{
                e.preventDefault();
                sendMsg();
            }}
        }});
    }}

    const assumeBtn = document.getElementById('assume-btn');
    if(assumeBtn){{
        assumeBtn.addEventListener('click', toggleHuman);
    }}

    const searchInput = document.getElementById('search-chats');
    if(searchInput){{
        searchInput.addEventListener('input', function(e){{
            filterChats(e.target.value);
        }});
    }}

    // Delegação de evento para chat-items (que podem ser recriados dinamicamente)
    const chatList = document.getElementById('chat-list');
    if(chatList){{
        chatList.addEventListener('click', function(e){{
            const item = e.target.closest('.chat-item');
            if(item && item.dataset.convId){{
                loadConversation(parseInt(item.dataset.convId), item);
            }}
        }});
    }}
    </script>"""
    return base_html("Conversas", content, dict(user))


# ─── TRAINING ──────────────────────────────────────────────────
@app.route("/dashboard/training", methods=["GET","POST"])
@login_required
def training():
    user = g.user; db = get_db(); msg = ""
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add_knowledge":
            title = request.form.get("title","").strip()
            ct = request.form.get("content","").strip()
            cat = request.form.get("category","geral")
            if title and ct:
                db.execute("INSERT INTO knowledge_base (user_id,title,content,category) VALUES (?,?,?,?)", (user["id"],title,ct,cat))
                db.commit(); msg = '<div class="alert alert-success">Item adicionado!</div>'
        elif action == "update_prompt":
            db.execute("UPDATE users SET ai_system_prompt=?,ai_tone=?,ai_greeting=? WHERE id=?",
                (request.form.get("system_prompt",""), request.form.get("tone","profissional"), request.form.get("greeting",""), user["id"]))
            db.commit(); msg = '<div class="alert alert-success">IA atualizada!</div>'
            user = db.execute("SELECT * FROM users WHERE id=?", (user["id"],)).fetchone()
        elif action == "delete_kb":
            db.execute("DELETE FROM knowledge_base WHERE id=? AND user_id=?", (request.form.get("kb_id"), user["id"]))
            db.commit(); msg = '<div class="alert alert-success">Removido!</div>'
        elif action == "import_file":
            # Processa arquivo enviado (PDF, XLSX, CSV, TXT)
            if "file" not in request.files:
                msg = '<div class="alert alert-error">Nenhum arquivo enviado</div>'
            else:
                file = request.files["file"]
                if file.filename == "":
                    msg = '<div class="alert alert-error">Arquivo vazio</div>'
                else:
                    file_title = request.form.get("file_title", "").strip() or file.filename
                    file_category = request.form.get("file_category", "geral")
                    # Lê conteúdo baseado no tipo
                    extracted_text = ""
                    try:
                        fname_lower = file.filename.lower()
                        if fname_lower.endswith(".pdf"):
                            # Salva temporariamente e extrai
                            import tempfile
                            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                                tmp.write(file.read())
                                tmp_path = tmp.name
                            extracted_text = extract_pdf_text(tmp_path)
                            try: os.remove(tmp_path)
                            except: pass
                        elif fname_lower.endswith(".txt"):
                            extracted_text = file.read().decode("utf-8", errors="ignore")
                        elif fname_lower.endswith(".csv"):
                            import csv as csv_mod, io
                            content_str = file.read().decode("utf-8", errors="ignore")
                            reader = csv_mod.reader(io.StringIO(content_str))
                            rows = list(reader)
                            # Formata como tabela legível
                            extracted_text = "\n".join([" | ".join(r) for r in rows])
                        elif fname_lower.endswith((".xlsx", ".xls")):
                            extracted_text = extract_spreadsheet_text(file)
                        else:
                            msg = '<div class="alert alert-error">Formato não suportado. Use: PDF, XLSX, CSV ou TXT</div>'

                        if extracted_text and len(extracted_text.strip()) > 10:
                            # Limita tamanho (máx ~50k caracteres para não estourar contexto)
                            if len(extracted_text) > 50000:
                                extracted_text = extracted_text[:50000] + "\n\n[...conteúdo truncado em 50k chars]"
                            db.execute(
                                "INSERT INTO knowledge_base (user_id,title,content,category) VALUES (?,?,?,?)",
                                (user["id"], file_title, extracted_text, file_category)
                            )
                            db.commit()
                            chars = len(extracted_text)
                            msg = f'<div class="alert alert-success">✅ Arquivo importado! ({chars} caracteres extraídos)</div>'
                        elif not msg:
                            msg = '<div class="alert alert-error">Não foi possível extrair texto do arquivo</div>'
                    except Exception as e:
                        msg = f'<div class="alert alert-error">Erro ao processar arquivo: {esc(str(e))}</div>'

    kb = db.execute("SELECT * FROM knowledge_base WHERE user_id=? ORDER BY created_at DESC", (user["id"],)).fetchall()
    kb_rows = "".join(f'<tr><td><strong>{esc(i["title"])}</strong></td><td><span class="badge badge-purple">{esc(i["category"])}</span></td><td style="max-width:300px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--text2)">{esc(i["content"][:100])}</td><td><form method="POST">{csrf_field()}<span style="display:inline"><input type="hidden" name="action" value="delete_kb"><input type="hidden" name="kb_id" value="{i["id"]}"><button type="submit" class="btn btn-danger btn-sm">✕</button></form></td></tr>' for i in kb)

    content = f"""<div class="container"><div class="page-header fade-in"><h1>Treinamento da IA 🧠</h1><p>Configure personalidade e base de conhecimento.</p></div>{msg}
        <div class="grid-2"><div class="card fade-in fade-in-1"><div class="card-header"><span class="card-title">Personalidade da IA</span></div>
            <form method="POST">{csrf_field()}<input type="hidden" name="action" value="update_prompt">
            <div class="form-group"><label class="form-label">System prompt</label><textarea name="system_prompt" class="form-input" rows="6">{user['ai_system_prompt']}</textarea></div>
            <div class="form-group"><label class="form-label">Tom de voz</label><select name="tone" class="form-input">
                <option value="profissional" {'selected' if user['ai_tone']=='profissional' else ''}>Profissional</option>
                <option value="descontraido" {'selected' if user['ai_tone']=='descontraido' else ''}>Descontraído</option>
                <option value="formal" {'selected' if user['ai_tone']=='formal' else ''}>Formal</option>
                <option value="amigavel" {'selected' if user['ai_tone']=='amigavel' else ''}>Amigável</option></select></div>
            <div class="form-group"><label class="form-label">Saudação</label><input type="text" name="greeting" class="form-input" value="{user['ai_greeting']}"></div>
            <button type="submit" class="btn btn-primary">Salvar</button></form></div>
        <div class="card fade-in fade-in-2"><div class="card-header"><span class="card-title">Adicionar conhecimento</span></div>
            <form method="POST">{csrf_field()}<input type="hidden" name="action" value="add_knowledge">
            <div class="form-group"><label class="form-label">Título</label><input type="text" name="title" class="form-input" placeholder="Ex: Tabela de preços" required></div>
            <div class="form-group"><label class="form-label">Categoria</label><select name="category" class="form-input">
                <option value="produtos">Produtos</option><option value="precos">Preços</option><option value="faq">FAQ</option><option value="politicas">Políticas</option><option value="geral">Geral</option></select></div>
            <div class="form-group"><label class="form-label">Conteúdo</label><textarea name="content" class="form-input" rows="6" placeholder="Informações que a IA deve saber..." required></textarea></div>
            <button type="submit" class="btn btn-success">+ Adicionar</button></form></div></div>

        <div class="card fade-in fade-in-3" style="margin-bottom:24px">
            <div class="card-header"><span class="card-title">📄 Importar arquivo (PDF, Excel, CSV)</span></div>
            <p style="color:var(--text3);font-size:13px;margin-bottom:16px">
                Envie um arquivo e o sistema extrai o conteúdo automaticamente. Útil para tabela de preços,
                catálogos, manuais ou planilhas de produtos.
            </p>
            <form method="POST" enctype="multipart/form-data">{csrf_field()}<input type="hidden" name="action" value="import_file">
                <div class="grid-2">
                    <div class="form-group">
                        <label class="form-label">Título (opcional)</label>
                        <input type="text" name="file_title" class="form-input" placeholder="Ex: Tabela de preços 2026" maxlength="100">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Categoria</label>
                        <select name="file_category" class="form-input">
                            <option value="precos">Preços</option>
                            <option value="produtos">Produtos</option>
                            <option value="faq">FAQ</option>
                            <option value="politicas">Políticas</option>
                            <option value="geral">Geral</option>
                        </select>
                    </div>
                </div>
                <div class="form-group">
                    <label class="form-label">Arquivo (PDF, XLSX, CSV ou TXT — máx 10MB)</label>
                    <input type="file" name="file" accept=".pdf,.xlsx,.xls,.csv,.txt,application/pdf,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/csv,text/plain" required
                        style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.08);padding:10px;border-radius:8px;color:var(--text);width:100%">
                </div>
                <button type="submit" class="btn btn-primary">📤 Importar e treinar IA</button>
            </form>
            <p style="color:var(--text3);font-size:12px;margin-top:12px">
                ℹ️ O conteúdo extraído será salvo como um item na base de conhecimento e a IA usará essas informações nas respostas.
            </p>
        </div>
        <div class="card fade-in fade-in-3"><div class="card-header"><span class="card-title">Base de conhecimento ({len(kb)} itens)</span></div>
            {'<div class="table-wrap"><table><thead><tr><th>Título</th><th>Categoria</th><th>Conteúdo</th><th></th></tr></thead><tbody>'+kb_rows+'</tbody></table></div>' if kb else '<div class="empty-state"><div class="icon">📚</div><h3>Base vazia</h3><p>Adicione informações sobre seus produtos.</p></div>'}</div></div>"""
    return base_html("Treinamento", content, dict(user))


# ─── QUICK REPLIES (NOVO) ─────────────────────────────────────
@app.route("/dashboard/quick-replies", methods=["GET","POST"])
@login_required
def quick_replies():
    user = g.user; db = get_db(); msg = ""
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            shortcut = request.form.get("shortcut","").strip()
            content_text = request.form.get("content","").strip()
            if shortcut and content_text:
                db.execute("INSERT INTO quick_replies (user_id,shortcut,content) VALUES (?,?,?)", (user["id"],shortcut,content_text))
                db.commit(); msg = '<div class="alert alert-success">Resposta rápida adicionada!</div>'
        elif action == "delete":
            db.execute("DELETE FROM quick_replies WHERE id=? AND user_id=?", (request.form.get("qr_id"), user["id"]))
            db.commit(); msg = '<div class="alert alert-success">Removida!</div>'

    qrs = db.execute("SELECT * FROM quick_replies WHERE user_id=? ORDER BY times_used DESC", (user["id"],)).fetchall()
    rows = "".join(f'<tr><td><code style="color:var(--accent2)">/{esc(q["shortcut"])}</code></td><td>{esc(q["content"][:80])}</td><td>{q["times_used"]}</td><td><form method="POST">{csrf_field()}<span style="display:inline"><input type="hidden" name="action" value="delete"><input type="hidden" name="qr_id" value="{q["id"]}"><button type="submit" class="btn btn-danger btn-sm">✕</button></form></td></tr>' for q in qrs)

    content = f"""<div class="container"><div class="page-header"><h1>Respostas Rápidas ⚡</h1><p>Atalhos para mensagens que você usa com frequência.</p></div>{msg}
        <div class="grid-2"><div class="card"><div class="card-header"><span class="card-title">Nova resposta rápida</span></div>
            <form method="POST">{csrf_field()}<input type="hidden" name="action" value="add">
            <div class="form-group"><label class="form-label">Atalho (ex: preco, horario)</label><input type="text" name="shortcut" class="form-input" placeholder="preco" required></div>
            <div class="form-group"><label class="form-label">Mensagem</label><textarea name="content" class="form-input" rows="4" placeholder="Nossos preços começam a partir de..." required></textarea></div>
            <button type="submit" class="btn btn-success">+ Adicionar</button></form></div>
        <div class="card"><div class="card-header"><span class="card-title">Como funciona</span></div>
            <div style="color:var(--text2);font-size:14px;line-height:1.8;padding:8px 0">
                <p>Quando você está atendendo manualmente no painel, digite <code style="color:var(--accent2)">/atalho</code> para inserir a mensagem rapidamente.</p>
                <p style="margin-top:12px">A IA também pode usar essas respostas como referência para responder perguntas frequentes de forma consistente.</p>
                <p style="margin-top:12px"><strong>Exemplos úteis:</strong></p>
                <p><code style="color:var(--accent2)">/preco</code> → Tabela de preços</p>
                <p><code style="color:var(--accent2)">/horario</code> → Horário de funcionamento</p>
                <p><code style="color:var(--accent2)">/pix</code> → Chave PIX e instruções</p>
                <p><code style="color:var(--accent2)">/frete</code> → Informações de entrega</p>
            </div></div></div>
        <div class="card"><div class="card-header"><span class="card-title">Respostas cadastradas ({len(qrs)})</span></div>
            {'<div class="table-wrap"><table><thead><tr><th>Atalho</th><th>Mensagem</th><th>Usos</th><th></th></tr></thead><tbody>'+rows+'</tbody></table></div>' if qrs else '<div class="empty-state"><div class="icon">⚡</div><h3>Nenhuma resposta rápida</h3><p>Crie atalhos para agilizar seu atendimento.</p></div>'}</div></div>"""
    return base_html("Respostas Rápidas", content, dict(user))


# ─── SETTINGS ──────────────────────────────────────────────────
def upload_whatsapp_profile_photo(phone_id, token, image_bytes, mime_type="image/jpeg"):
    """Atualiza a foto de perfil do WhatsApp Business via API da Meta"""
    if not phone_id or not token:
        return False, "Phone ID ou Token não configurados"

    try:
        import requests as req

        # Passo 1: Iniciar sessão de upload resumível
        session_url = "https://graph.facebook.com/v18.0/app/uploads"
        session_headers = {
            "Authorization": f"OAuth {token}",
        }
        session_params = {
            "file_length": str(len(image_bytes)),
            "file_type": mime_type,
            "access_token": token,
        }
        session_resp = req.post(session_url, headers=session_headers, params=session_params, timeout=15)

        if session_resp.status_code != 200:
            # Método alternativo: upload direto via /{phone-number-ID}
            safe_log(f"[PHOTO] Sessão falhou ({session_resp.status_code}), tentando upload direto...", level="ERROR")
            # Upload direto da foto usando profile_photo_handle
            files = {"file": ("photo.jpg", image_bytes, mime_type)}
            data = {
                "messaging_product": "whatsapp",
            }
            headers_direct = {"Authorization": f"Bearer {token}"}

            # Tenta fazer upload via /media
            media_url = f"https://graph.facebook.com/v18.0/{phone_id}/media"
            media_resp = req.post(media_url, headers=headers_direct, files=files, data=data, timeout=30)
            if media_resp.status_code != 200:
                return False, f"Upload falhou: {media_resp.status_code} {_short_resp_text(media_resp)}"
            media_id = media_resp.json().get("id", "")

            # Atualiza perfil com media_id
            profile_url = f"https://graph.facebook.com/v18.0/{phone_id}/whatsapp_business_profile"
            profile_headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
            profile_data = {
                "messaging_product": "whatsapp",
                "profile_picture_handle": media_id,
            }
            profile_resp = req.post(profile_url, headers=profile_headers, json=profile_data, timeout=15)
            if profile_resp.status_code == 200:
                return True, "Foto atualizada com sucesso!"
            return False, f"Erro ao definir perfil: {profile_resp.status_code} {_short_resp_text(profile_resp)}"

        # Sessão criada - continua upload
        upload_session_id = session_resp.json().get("id", "")

        # Passo 2: Upload dos bytes
        upload_url = f"https://graph.facebook.com/v18.0/{upload_session_id}"
        upload_headers = {
            "Authorization": f"OAuth {token}",
            "file_offset": "0",
        }
        upload_resp = req.post(upload_url, headers=upload_headers, data=image_bytes, timeout=30)

        if upload_resp.status_code != 200:
            return False, f"Upload falhou: {upload_resp.status_code}"

        upload_handle = upload_resp.json().get("h", "")

        # Passo 3: Atualizar perfil do WhatsApp com o handle
        profile_url = f"https://graph.facebook.com/v18.0/{phone_id}/whatsapp_business_profile"
        profile_headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        profile_data = {
            "messaging_product": "whatsapp",
            "profile_picture_handle": upload_handle,
        }
        profile_resp = req.post(profile_url, headers=profile_headers, json=profile_data, timeout=15)

        if profile_resp.status_code == 200:
            return True, "Foto de perfil atualizada com sucesso!"
        return False, f"Erro ao definir perfil: {_short_resp_text(profile_resp)}"

    except Exception as e:
        return False, f"Exceção: {str(e)}"


@app.route("/dashboard/settings/upload-photo", methods=["POST"])
@login_required
def upload_profile_photo():
    """Endpoint para upload de foto de perfil do WhatsApp.
    Valida imagem com Pillow e remove metadados EXIF."""
    user = g.user
    if not user["whatsapp_phone_id"] or not user["whatsapp_token"]:
        return redirect("/dashboard/settings?photo_error=" + "Configure%20Phone%20ID%20e%20Token%20primeiro")

    if "photo" not in request.files:
        return redirect("/dashboard/settings?photo_error=" + "Nenhum%20arquivo%20enviado")

    file = request.files["photo"]
    if file.filename == "":
        return redirect("/dashboard/settings?photo_error=" + "Arquivo%20vazio")

    # Lê bytes
    image_bytes = file.read()
    if len(image_bytes) > 5 * 1024 * 1024:
        return redirect("/dashboard/settings?photo_error=" + "M%C3%A1ximo%205MB")

    if len(image_bytes) < 100:
        return redirect("/dashboard/settings?photo_error=" + "Arquivo%20muito%20pequeno")

    # VALIDAÇÃO REAL com Pillow — rejeita arquivos maliciosos disfarçados
    validated_bytes, real_content_type = validate_and_normalize_image(image_bytes)
    if validated_bytes is None:
        return redirect("/dashboard/settings?photo_error=" + "Arquivo%20inv%C3%A1lido%20ou%20corrompido.%20Envie%20JPG%20ou%20PNG%20real.")

    # Upload para o Meta com bytes normalizados
    success, message = upload_whatsapp_profile_photo(
        user["whatsapp_phone_id"],
        user["whatsapp_token"],
        validated_bytes,
        real_content_type
    )

    if success:
        return redirect("/dashboard/settings?photo_ok=" + "Foto%20atualizada!")
    else:
        import urllib.parse
        return redirect("/dashboard/settings?photo_error=" + urllib.parse.quote(message))


@app.route("/dashboard/settings", methods=["GET","POST"])
@login_required
def settings():
    user = g.user; db = get_db(); msg = ""
    if request.method == "POST":
        # "Manter atual se vazio": se o usuário enviou campo vazio mas já tinha token, mantém o que estava
        wa_token_input = request.form.get("whatsapp_token","").strip()
        ig_token_input = request.form.get("instagram_token","").strip()
        msg_token_input = request.form.get("messenger_token","").strip()

        # Se campo veio vazio, lê o valor ATUAL do banco (criptografado) e mantém
        raw_current = db.execute("SELECT whatsapp_token, instagram_token, messenger_token FROM users WHERE id=?", (user["id"],)).fetchone()

        # Se usuário enviou novo token, criptografa; se vazio, mantém atual (já criptografado)
        wa_token_final = _encrypt_value(wa_token_input) if wa_token_input else (raw_current["whatsapp_token"] or "")
        ig_token_final = _encrypt_value(ig_token_input) if ig_token_input else (raw_current["instagram_token"] or "")
        msg_token_final = _encrypt_value(msg_token_input) if msg_token_input else (raw_current["messenger_token"] or "")

        db.execute("""UPDATE users SET 
            whatsapp_phone_id=?, whatsapp_token=?,
            instagram_page_id=?, instagram_token=?,
            messenger_page_id=?, messenger_token=?,
            business_hours=?, auto_reply_off_hours=?, 
            name=?, company=?, phone=? 
            WHERE id=?""",
            (request.form.get("whatsapp_phone_id","").strip(),
             wa_token_final,
             request.form.get("instagram_page_id","").strip(),
             ig_token_final,
             request.form.get("messenger_page_id","").strip(),
             msg_token_final,
             request.form.get("business_hours","08:00-18:00").strip(),
             request.form.get("auto_reply_off_hours","").strip(),
             request.form.get("name","").strip(),
             request.form.get("company","").strip(),
             request.form.get("phone","").strip(),
             user["id"]))
        db.commit(); msg = '<div class="alert alert-success">Configurações salvas!</div>'
        user = db.execute("SELECT * FROM users WHERE id=?", (user["id"],)).fetchone()

    # Mensagens de foto
    photo_ok = request.args.get("photo_ok", "")
    photo_error = request.args.get("photo_error", "")
    photo_msg = ""
    if photo_ok:
        photo_msg = f'<div class="alert alert-success">✅ {esc(photo_ok)}</div>'
    elif photo_error:
        photo_msg = f'<div class="alert alert-error">❌ {esc(photo_error)}</div>'

    base = get_setting("BASE_URL", BASE_URL)
    wa_verify = get_setting("WHATSAPP_VERIFY_TOKEN", WHATSAPP_VERIFY_TOKEN)
    webhook_url = f"{base}/webhook/whatsapp"
    # Escape de todos os valores do usuário para prevenir XSS
    u_name = esc(user['name'])
    u_company = esc(user['company'])
    u_phone = esc(user['phone'])
    u_hours = esc(user['business_hours'])
    u_reply = esc(user['auto_reply_off_hours'])
    u_wa_id = esc(user['whatsapp_phone_id'] or '')
    u_wa_token_mask = mask_secret(user['whatsapp_token'] or '')
    u_wa_token_configured = bool(user['whatsapp_token'])
    u_ig_token_mask = mask_secret(user['instagram_token'] or '')
    u_ig_token_configured = bool(user['instagram_token'])
    u_msg_token_mask = mask_secret(user['messenger_token'] or '')
    u_msg_token_configured = bool(user['messenger_token'])
    e_webhook = esc(webhook_url)
    e_verify = esc(wa_verify)
    # Placeholder específico: se já configurado, mostra máscara; senão, mostra instrução
    wa_placeholder = u_wa_token_mask if u_wa_token_configured else "Cole aqui o Access Token"
    wa_help = '<small style="color:var(--green2);font-size:11px">✓ Token configurado. Deixe em branco para manter o atual ou cole um novo para substituir.</small>' if u_wa_token_configured else ''
    content = f"""<div class="container"><div class="page-header fade-in"><h1>Configurações ⚙️</h1></div>{msg}{photo_msg}
        <div class="grid-2"><div class="card fade-in fade-in-1"><div class="card-header"><span class="card-title">Perfil e WhatsApp</span></div>
            <form method="POST">{csrf_field()}
            <div class="form-group"><label class="form-label">Nome</label><input type="text" name="name" class="form-input" value="{u_name}"></div>
            <div class="form-group"><label class="form-label">Empresa</label><input type="text" name="company" class="form-input" value="{u_company}"></div>
            <div class="form-group"><label class="form-label">Telefone</label><input type="text" name="phone" class="form-input" value="{u_phone}"></div>
            <div class="form-group"><label class="form-label">Horário de atendimento</label><input type="text" name="business_hours" class="form-input" value="{u_hours}"></div>
            <div class="form-group"><label class="form-label">Resposta fora do horário</label><textarea name="auto_reply_off_hours" class="form-input" rows="3">{u_reply}</textarea></div>
            <div class="form-group"><label class="form-label">WhatsApp Phone ID</label><input type="text" id="wp_phone_id" name="whatsapp_phone_id" class="form-input" value="{u_wa_id}" placeholder="Cole aqui o Phone Number ID" autocomplete="off" style="background:#2a2a3a;border:2px solid #00c896;color:#fff;cursor:text"></div>
            <div class="form-group"><label class="form-label">WhatsApp Token 🔒</label><input type="password" id="wp_token" name="whatsapp_token" class="form-input" value="" placeholder="{esc(wa_placeholder)}" autocomplete="off" style="background:#2a2a3a;border:2px solid #00c896;color:#fff;cursor:text">{wa_help}</div>
            <button type="submit" class="btn btn-primary">Salvar</button></form>

            <div style="margin-top:24px;padding-top:24px;border-top:1px solid rgba(255,255,255,0.06)">
                <h3 style="font-size:16px;margin-bottom:12px;color:var(--text)">📷 Foto do WhatsApp</h3>
                <p style="color:var(--text3);font-size:13px;margin-bottom:12px">Envie uma foto JPG ou PNG (máx 5MB) para aparecer como foto de perfil do atendente no WhatsApp.</p>
                <form method="POST" action="/dashboard/settings/upload-photo" enctype="multipart/form-data">{csrf_field()}
                    <div class="form-group">
                        <input type="file" name="photo" accept="image/jpeg,image/png,image/jpg" required
                            style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.08);padding:10px;border-radius:8px;color:var(--text);width:100%">
                    </div>
                    <button type="submit" class="btn btn-primary" style="background:#00c896">📤 Enviar foto</button>
                </form>
                <p style="color:var(--text3);font-size:12px;margin-top:8px">⚠️ Requer Phone ID e Token configurados acima.</p>
            </div>

            <script nonce="{g.csp_nonce}">
            document.addEventListener('DOMContentLoaded', function() {{
                ['wp_phone_id','wp_token'].forEach(function(id) {{
                    var el = document.getElementById(id);
                    if(el) {{
                        el.removeAttribute('readonly');
                        el.removeAttribute('disabled');
                        el.style.pointerEvents = 'auto';
                        el.style.userSelect = 'text';
                        el.addEventListener('click', function() {{ this.focus(); this.select(); }});
                    }}
                }});
            }});
            </script></div>
        <div><div class="card fade-in fade-in-2" style="margin-bottom:24px"><div class="card-header"><span class="card-title">Webhook URL</span></div>
            <p style="color:var(--text2);font-size:14px;margin-bottom:12px">Configure no Meta Business:</p>
            <p style="color:var(--text3);font-size:12px;margin-bottom:4px">📱 WhatsApp:</p>
            <div style="background:var(--bg4);padding:10px 14px;border-radius:var(--radius-sm);font-family:var(--mono);font-size:12px;word-break:break-all;color:var(--accent2);margin-bottom:8px">{e_webhook}</div>
            <p style="color:var(--text3);font-size:12px;margin-bottom:4px">📷 Instagram:</p>
            <div style="background:var(--bg4);padding:10px 14px;border-radius:var(--radius-sm);font-family:var(--mono);font-size:12px;word-break:break-all;color:var(--accent2);margin-bottom:8px">{esc(base)}/webhook/instagram/{user['id']}</div>
            <p style="color:var(--text3);font-size:12px;margin-bottom:4px">💬 Messenger:</p>
            <div style="background:var(--bg4);padding:10px 14px;border-radius:var(--radius-sm);font-family:var(--mono);font-size:12px;word-break:break-all;color:var(--accent2);margin-bottom:8px">{esc(base)}/webhook/messenger/{user['id']}</div>
            <p style="color:var(--text3);font-size:12px;margin-top:8px">Token: <code style="color:var(--accent2)">{e_verify}</code></p></div>

        <div class="card fade-in fade-in-3" style="margin-bottom:24px">
            <div class="card-header"><span class="card-title">📷 Instagram Direct</span></div>
            <p style="color:var(--text3);font-size:13px;margin-bottom:12px">Conecte sua conta Instagram Business/Creator para responder mensagens Direct automaticamente.</p>
            <form method="POST">{csrf_field()}
                <input type="hidden" name="name" value="{u_name}">
                <input type="hidden" name="company" value="{u_company}">
                <input type="hidden" name="phone" value="{u_phone}">
                <input type="hidden" name="business_hours" value="{u_hours}">
                <input type="hidden" name="auto_reply_off_hours" value="{u_reply}">
                <input type="hidden" name="whatsapp_phone_id" value="{u_wa_id}">
                <input type="hidden" name="messenger_page_id" value="{esc(user['messenger_page_id'] or '')}">
                <div class="form-group">
                    <label class="form-label">Instagram Page ID</label>
                    <input type="text" name="instagram_page_id" class="form-input" value="{esc(user['instagram_page_id'] or '')}" placeholder="ID da página conectada ao Instagram" autocomplete="off" style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.08)">
                </div>
                <div class="form-group">
                    <label class="form-label">Instagram Access Token 🔒</label>
                    <input type="password" name="instagram_token" class="form-input" value="" placeholder="{esc(mask_secret(user['instagram_token'] or '')) if user['instagram_token'] else 'Token da Page com permissões Instagram'}" autocomplete="off" style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.08)">
                    {'<small style="color:var(--green2);font-size:11px">✓ Token configurado. Deixe em branco para manter.</small>' if user['instagram_token'] else ''}
                </div>
                <button type="submit" class="btn btn-primary">Salvar Instagram</button>
            </form>
        </div>

        <div class="card fade-in fade-in-4" style="margin-bottom:24px">
            <div class="card-header"><span class="card-title">💬 Facebook Messenger</span></div>
            <p style="color:var(--text3);font-size:13px;margin-bottom:12px">Conecte sua página do Facebook para responder mensagens do Messenger automaticamente.</p>
            <form method="POST">{csrf_field()}
                <input type="hidden" name="name" value="{u_name}">
                <input type="hidden" name="company" value="{u_company}">
                <input type="hidden" name="phone" value="{u_phone}">
                <input type="hidden" name="business_hours" value="{u_hours}">
                <input type="hidden" name="auto_reply_off_hours" value="{u_reply}">
                <input type="hidden" name="whatsapp_phone_id" value="{u_wa_id}">
                <input type="hidden" name="instagram_page_id" value="{esc(user['instagram_page_id'] or '')}">
                <div class="form-group">
                    <label class="form-label">Page ID do Facebook</label>
                    <input type="text" name="messenger_page_id" class="form-input" value="{esc(user['messenger_page_id'] or '')}" placeholder="ID da página do Facebook" autocomplete="off" style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.08)">
                </div>
                <div class="form-group">
                    <label class="form-label">Page Access Token 🔒</label>
                    <input type="password" name="messenger_token" class="form-input" value="" placeholder="{esc(mask_secret(user['messenger_token'] or '')) if user['messenger_token'] else 'Token da Page com permissão pages_messaging'}" autocomplete="off" style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.08)">
                    {'<small style="color:var(--green2);font-size:11px">✓ Token configurado. Deixe em branco para manter.</small>' if user['messenger_token'] else ''}
                </div>
                <button type="submit" class="btn btn-primary">Salvar Messenger</button>
            </form>
        </div>

        <div class="card fade-in fade-in-5"><div class="card-header"><span class="card-title">Mídias suportadas</span></div>
            <div style="color:var(--text2);font-size:14px;line-height:1.8">
                <p>✅ <strong style="color:var(--text)">Texto</strong> — lê e responde normalmente</p>
                <p>✅ <strong style="color:var(--text)">Áudio</strong> — transcreve com Groq/Whisper e responde por voz</p>
                <p>✅ <strong style="color:var(--text)">Imagens</strong> — analisa com Claude Vision</p>
                <p>✅ <strong style="color:var(--text)">PDFs</strong> — extrai texto e interpreta</p>
                <p>✅ <strong style="color:var(--text)">Localização</strong> — recebe e processa</p>
                <p>✅ <strong style="color:var(--text)">Contatos</strong> — recebe dados do contato</p>
                <p>✅ <strong style="color:var(--text)">Stickers/Reações</strong> — registra</p>
            </div></div></div></div></div>"""
    return base_html("Configurações", content, dict(user))


# ─── BILLING ───────────────────────────────────────────────────
@app.route("/dashboard/billing")
@login_required
def billing():
    user = g.user; db = get_db()
    plan = PLANS.get(user["plan"], PLANS["starter"])
    payments = db.execute("SELECT * FROM payments WHERE user_id=? ORDER BY created_at DESC LIMIT 10", (user["id"],)).fetchall()
    payment_rows = ""
    for p in payments:
        p_date = (p["created_at"] or "")[:10]
        p_plan = PLANS.get(p["plan"], {}).get("name", p["plan"])
        p_cls = "badge-green" if p["status"]=="approved" else "badge-orange" if p["status"]=="pending" else "badge-red"
        p_label = {"approved":"Aprovado","pending":"Pendente","rejected":"Rejeitado"}.get(p["status"], p["status"])
        payment_rows += f'<tr><td>{p_date}</td><td>R$ {p["amount"]:.2f}</td><td>{p_plan}</td><td><span class="badge {p_cls}">{p_label}</span></td></tr>'

    # Gera cards dos 4 planos com features completas
    plans_html = ""
    plan_order = ["starter", "pro", "business", "agency"]
    for key in plan_order:
        p = PLANS.get(key)
        if not p:
            continue
        is_current = key == user["plan"]
        is_popular = key == "pro"
        is_premium = key == "agency"

        border_style = ""
        badge = ""
        if is_popular:
            border_style = "border:2px solid var(--accent2);transform:scale(1.02)"
            badge = '<div style="position:absolute;top:-12px;left:50%;transform:translateX(-50%);background:linear-gradient(135deg,var(--accent),var(--accent2));padding:4px 14px;border-radius:20px;font-size:11px;font-weight:700;color:white;letter-spacing:1px">⭐ MAIS POPULAR</div>'
        elif is_premium:
            border_style = "border:1px solid rgba(251,191,36,0.3);background:linear-gradient(135deg,rgba(251,191,36,0.04),rgba(168,85,247,0.04))"
            badge = '<div style="position:absolute;top:-12px;right:20px;background:linear-gradient(135deg,#fbbf24,#a855f7);padding:4px 12px;border-radius:12px;font-size:10px;font-weight:700;color:white">👑 PREMIUM</div>'
        elif key == "business":
            badge = '<div style="position:absolute;top:-12px;right:20px;background:#10b981;padding:4px 12px;border-radius:12px;font-size:10px;font-weight:700;color:white">💰 VENDEDOR</div>'

        feats_html = "".join([
            f'<li style="padding:8px 0;color:var(--text2);font-size:13px;border-bottom:1px solid rgba(255,255,255,0.04)">✓ {f}</li>'
            for f in p.get("features", [])
        ])

        if is_current:
            btn = '<button class="btn" style="width:100%;background:rgba(16,185,129,0.15);color:#10b981;border:1px solid rgba(16,185,129,0.3);pointer-events:none">✓ Plano atual</button>'
        else:
            cta = p.get("cta", "Assinar")
            btn_class = "btn-primary" if is_popular else "btn-secondary"
            btn = f'<a href="/api/mercadopago/create-preference?plan={key}" class="btn {btn_class}" style="width:100%;text-align:center;text-decoration:none">{cta} →</a>'

        plans_html += f"""
        <div style="background:rgba(255,255,255,0.02);border:1px solid rgba(255,255,255,0.06);{border_style};border-radius:16px;padding:28px 20px;position:relative;display:flex;flex-direction:column">
            {badge}
            <div style="text-align:center;margin-bottom:16px">
                <div style="font-size:13px;color:var(--text3);font-weight:600;text-transform:uppercase;letter-spacing:1px">{p['name']}</div>
                <div style="font-size:11px;color:var(--text3);margin-top:4px">{p.get('tagline', p['desc'])}</div>
            </div>
            <div style="text-align:center;margin-bottom:20px">
                <div style="font-size:36px;font-weight:700;color:var(--text)">R$ {p['price']:.0f}<span style="font-size:14px;color:var(--text3);font-weight:400">/mês</span></div>
                <div style="color:var(--text3);font-size:12px;margin-top:4px">{p['msgs']:,} mensagens/mês</div>
            </div>
            <ul style="list-style:none;padding:0;margin:0 0 20px 0;flex:1">
                {feats_html}
            </ul>
            {btn}
        </div>
        """

    status_map = {"active":"Ativo","trial":"Período de teste","inactive":"Inativo","cancelled":"Cancelado"}
    cls_map = {"active":"badge-green","trial":"badge-orange","inactive":"badge-red","cancelled":"badge-red"}

    content = f"""<div class="container">
        <div class="page-header fade-in"><h1>Plano e Pagamento 💳</h1><p>Escolha o plano ideal para o tamanho do seu negócio</p></div>

        <div class="card fade-in" style="margin-bottom:32px">
            <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:16px">
                <div>
                    <div style="font-size:13px;color:var(--text3);text-transform:uppercase;letter-spacing:0.5px">Plano atual</div>
                    <div style="font-size:24px;font-weight:700;margin-top:4px">{plan['name']} <span class="badge {cls_map.get(user['plan_status'],'badge-orange')}">{status_map.get(user['plan_status'],user['plan_status'])}</span></div>
                    <div style="color:var(--text2);margin-top:4px">R$ {plan['price']:.0f}/mês · {user['msgs_used']}/{user['msgs_limit']} mensagens usadas</div>
                </div>
                <a href="#plans" class="btn btn-primary">Alterar plano ↓</a>
            </div>
        </div>

        <div id="plans" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:20px;margin-top:40px;margin-bottom:32px">
            {plans_html}
        </div>

        <div class="alert alert-info" style="margin-bottom:32px">
            💡 <strong>Dúvidas?</strong> Todos os planos incluem 7 dias grátis, cancelamento a qualquer momento e suporte em português.
            Pagamento via Cartão, PIX ou Boleto pelo Mercado Pago.
        </div>

        <div class="card fade-in fade-in-2">
            <div class="card-header"><span class="card-title">📋 Histórico de pagamentos</span></div>
            {'<div class="table-wrap"><table><thead><tr><th>Data</th><th>Valor</th><th>Plano</th><th>Status</th></tr></thead><tbody>'+payment_rows+'</tbody></table></div>' if payments else '<div style="text-align:center;padding:40px;color:var(--text3)"><div style="font-size:36px;margin-bottom:8px">📋</div><h3 style="margin:0;font-size:16px">Nenhum pagamento ainda</h3><p style="font-size:13px;margin-top:4px">Seus pagamentos aparecerão aqui</p></div>'}
        </div>
    </div>"""
    return base_html("Pagamento", content, dict(user))


# ─── MERCADO PAGO ──────────────────────────────────────────────
@app.route("/api/mercadopago/create-preference")
@login_required
def mp_create_preference():
    plan_key = request.args.get("plan","starter")
    plan = PLANS.get(plan_key)
    if not plan: return jsonify({"error":"Plano inválido"}), 400
    user = g.user
    mp_token = get_setting("MERCADOPAGO_ACCESS_TOKEN", MERCADOPAGO_ACCESS_TOKEN)
    base = get_setting("BASE_URL", BASE_URL)
    try:
        import mercadopago
        sdk = mercadopago.SDK(mp_token)
        pref = sdk.preference().create({"items":[{"title":f"atendente.online — {plan['name']}","quantity":1,"unit_price":plan["price"],"currency_id":"BRL"}],
            "payer":{"email":user["email"],"name":user["name"]},
            "back_urls":{"success":f"{base}/api/mercadopago/callback?status=success&plan={plan_key}","failure":f"{base}/api/mercadopago/callback?status=failure&plan={plan_key}","pending":f"{base}/api/mercadopago/callback?status=pending&plan={plan_key}"},
            "auto_return":"approved","notification_url":f"{base}/api/mercadopago/webhook","external_reference":f"user_{user['id']}_plan_{plan_key}_{int(time.time())}","statement_descriptor":"ATENDENTE.ONLINE"})
        checkout_url = pref["response"].get("init_point", pref["response"].get("sandbox_init_point",""))
        if checkout_url: return redirect(checkout_url)
        return redirect("/dashboard/billing?error=Erro ao criar pagamento")
    except ImportError:
        # Se SDK não instalado, só mostra tela de simulação em DEV
        is_dev = os.getenv("FLASK_ENV", "").lower() == "development"
        if not is_dev:
            return redirect("/dashboard/billing?error=Sistema de pagamento indisponível")
        return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><link rel="icon" type="image/png" href="/favicon.ico"><link rel="apple-touch-icon" href="/apple-touch-icon.png"><meta name="theme-color" content="#6366f1"><title>Checkout Simulado (DEV)</title>
        <style>body{{font-family:'DM Sans',sans-serif;background:#0a0e14;color:#f0f4f8;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0}}
        .box{{background:#111827;padding:40px;border-radius:16px;max-width:400px;text-align:center;border:1px solid rgba(255,255,255,0.1)}}
        .price{{font-size:36px;font-weight:700;color:#34d399;margin:16px 0}}.btn{{display:inline-block;padding:14px 32px;background:#00c896;color:white;border-radius:8px;text-decoration:none;font-weight:600}}
        .warn{{background:#7f1d1d;color:#fff;padding:8px;border-radius:6px;margin-bottom:16px;font-size:12px}}</style></head>
        <body><div class="box"><div class="warn">⚠️ MODO DESENVOLVIMENTO</div><h2>Checkout Simulado</h2><p style="color:#94a3b8">Plano {plan['name']}</p>
        <div class="price">R$ {plan['price']:.0f}<small style="font-size:14px;color:#94a3b8">/mês</small></div>
        <p style="color:#94a3b8;font-size:13px;margin-bottom:24px">SDK não instalado. Simulação de teste.</p>
        <a href="{base}/api/mercadopago/callback?status=success&plan={plan_key}&simulated=1" class="btn">Simular aprovação ✓</a><br><br>
        <a href="/dashboard/billing" style="color:#94a3b8;font-size:13px">← Voltar</a></div></body></html>"""
    except Exception as e:
        err_id = secrets.token_hex(6)
        safe_log(f"[MP] Erro checkout id={err_id}: {e}", level="ERROR")
        return redirect(f"/dashboard/billing?error=Não foi possível iniciar o checkout. Tente novamente (cod {err_id}).")

@app.route("/api/mercadopago/callback")
@login_required
def mp_callback():
    status = request.args.get("status",""); plan_key = request.args.get("plan","starter")
    plan = PLANS.get(plan_key, PLANS["starter"]); user = g.user; db = get_db()
    simulated = request.args.get("simulated","")

    # Simulação SÓ funciona em modo desenvolvimento explícito
    # (variável de ambiente FLASK_ENV=development)
    is_dev = os.getenv("FLASK_ENV", "").lower() == "development"

    if status == "success" and simulated == "1":
        if not is_dev:
            # Em produção, simulação está DESABILITADA — registra tentativa suspeita
            try:
                user_email = user["email"]
            except:
                user_email = ""
            safe_log(f"[SECURITY] Tentativa de bypass com simulated=1 por user {user['id']} ({user_email})", level="WARN")
            return redirect("/dashboard/billing?error=Operação não permitida")
        # Em dev, permite simular
        pid = f"sim_{int(time.time())}"
        db.execute("UPDATE users SET plan=?,plan_status='active',msgs_limit=?,msgs_used=0 WHERE id=?", (plan_key, plan["msgs"], user["id"]))
        db.execute("INSERT INTO payments (user_id,mp_payment_id,amount,status,plan) VALUES (?,?,?,?,?)", (user["id"],pid,plan["price"],"approved",plan_key))
        db.commit()
    elif status == "success":
        # Pagamento real — registra como pendente, webhook ativa depois
        pid = request.args.get("payment_id", "")
        db.execute("INSERT INTO payments (user_id,mp_payment_id,amount,status,plan) VALUES (?,?,?,?,?)", (user["id"],pid,plan["price"],"pending",plan_key))
        db.commit()
    return redirect("/dashboard/billing")

@app.route("/api/mercadopago/webhook", methods=["POST"])
def mp_webhook():
    # Valida assinatura — proteção contra fraude
    if not validate_mp_signature(request):
        log_webhook_error("mercadopago", None, "InvalidSignature", "Assinatura MP inválida", None)
        return jsonify({"status": "invalid_signature"}), 401

    data = request.json or {}
    if data.get("type") == "payment":
        pid = data.get("data",{}).get("id")
        if pid:
            try:
                event_key = f"payment:{pid}"
                if not register_processed_webhook_event("mercadopago_subscription", event_key, None, {"id": pid, "type": "payment"}):
                    return jsonify({"status": "duplicate"}), 200

                import mercadopago
                mp_token = get_setting("MERCADOPAGO_ACCESS_TOKEN", MERCADOPAGO_ACCESS_TOKEN)
                sdk = mercadopago.SDK(mp_token)
                payment = sdk.payment().get(pid)["response"]
                ext = payment.get("external_reference", "")
                ref = parse_mp_subscription_reference(ext)
                if not ref:
                    return jsonify({"status":"invalid_reference"}), 200

                uid = ref["user_id"]
                pk = ref["plan_key"]
                plan = PLANS.get(pk)
                amount = float(payment.get("transaction_amount") or 0)
                currency = payment.get("currency_id", "")
                if not plan or currency != "BRL" or abs(amount - float(plan["price"])) > 0.01:
                    log_webhook_error("mercadopago", uid, "InvalidPaymentData", f"amount={amount} currency={currency} ref={ext}", {"payment_id": pid})
                    return jsonify({"status":"invalid_payment_data"}), 200

                db_c = sqlite3.connect(DATABASE)
                db_c.row_factory = sqlite3.Row
                user_row = db_c.execute("SELECT id FROM users WHERE id=?", (uid,)).fetchone()
                if not user_row:
                    db_c.close()
                    return jsonify({"status":"user_not_found"}), 200

                if payment.get("status") == "approved":
                    db_c.execute("UPDATE users SET plan=?,plan_status='active',msgs_limit=?,msgs_used=0 WHERE id=?", (pk, plan["msgs"], uid))
                    db_c.execute("UPDATE payments SET status='approved' WHERE mp_payment_id=? AND user_id=?", (str(pid), uid))
                    db_c.execute("INSERT OR IGNORE INTO payments (user_id,mp_payment_id,amount,status,plan) VALUES (?,?,?,?,?)", (uid, str(pid), amount, "approved", pk))
                elif payment.get("status") == "rejected":
                    db_c.execute("UPDATE payments SET status='rejected' WHERE mp_payment_id=? AND user_id=?", (str(pid), uid))
                else:
                    db_c.execute("INSERT OR IGNORE INTO payments (user_id,mp_payment_id,amount,status,plan) VALUES (?,?,?,?,?)", (uid, str(pid), amount, payment.get("status", "pending"), pk))
                db_c.commit()
                db_c.close()
                safe_log(f"[MP] Webhook: payment {pid} status={payment.get('status')} user={uid} plan={pk}")
            except Exception as e:
                safe_log(f"[MP] Webhook error: {e}", level="ERROR")
    return jsonify({"status":"ok"}), 200


# ─── API ───────────────────────────────────────────────────────
@app.route("/api/conversations/<int:conv_id>/messages")
@login_required
def api_conv_messages(conv_id):
    db = get_db()
    conv = db.execute("SELECT * FROM conversations WHERE id=? AND user_id=?", (conv_id, g.user["id"])).fetchone()
    if not conv: return jsonify({"error":"Não encontrada"}), 404
    messages = db.execute("SELECT * FROM messages WHERE conversation_id=? ORDER BY created_at", (conv_id,)).fetchall()
    return jsonify({"customer_phone":conv["customer_phone"],"customer_name":conv["customer_name"],"is_human_takeover":conv["is_human_takeover"],"messages":[dict(m) for m in messages]})


@app.route("/api/conversations/<int:conv_id>/toggle-human", methods=["POST"])
@login_required
def api_conv_toggle_human(conv_id):
    """Alterna entre IA respondendo automaticamente vs atendente humano.
    is_human_takeover=1 -> IA NAO responde mensagens dessa conversa.
    is_human_takeover=0 -> IA volta a responder automaticamente.
    CSRF eh validado automaticamente pelo @app.before_request csrf_protect."""
    try:
        db = get_db()
        conv = db.execute(
            "SELECT * FROM conversations WHERE id=? AND user_id=?",
            (conv_id, g.user["id"])
        ).fetchone()
        if not conv:
            return jsonify({"error": "Conversa nao encontrada"}), 404

        new_status = 0 if conv["is_human_takeover"] else 1
        db.execute(
            "UPDATE conversations SET is_human_takeover=? WHERE id=?",
            (new_status, conv_id)
        )
        db.commit()

        safe_log(f"[CONV {conv_id}] Takeover alterado: {'humano' if new_status else 'IA'} (user {g.user['id']})")
        return jsonify({
            "success": True,
            "is_human_takeover": new_status,
            "message": "Atendimento humano ativo" if new_status else "IA retomada"
        })
    except Exception as e:
        err_id = secrets.token_hex(6)
        safe_log(f"[TOGGLE HUMAN] Erro id={err_id}: {e}", level="ERROR")
        return jsonify({"error": "Erro interno ao alternar atendimento", "err_id": err_id}), 500


@app.route("/api/conversations/<int:conv_id>/send", methods=["POST"])
@login_required
def api_conv_send_message(conv_id):
    """Envia mensagem manual do atendente para o cliente.
    O atendente digita no painel e o sistema envia via WhatsApp Business API."""
    try:
        data = request.get_json(silent=True) or {}
        message = (data.get("message") or "").strip()

        if not message:
            return jsonify({"error": "Mensagem vazia"}), 400

        if len(message) > 4096:
            return jsonify({"error": "Mensagem muito longa (máx 4096 caracteres)"}), 400

        db = get_db()

        # Valida que a conversa pertence ao usuário
        conv = db.execute(
            "SELECT * FROM conversations WHERE id=? AND user_id=?",
            (conv_id, g.user["id"])
        ).fetchone()
        if not conv:
            return jsonify({"error": "Conversa não encontrada"}), 404

        # Pega tokens do usuário (já descriptografados pelo login_required)
        user = g.user
        phone_id = user.get("whatsapp_phone_id", "")
        token = user.get("whatsapp_token", "")

        if not phone_id or not token:
            return jsonify({
                "error": "WhatsApp não configurado. Vá em Configurações."
            }), 400

        # Envia via WhatsApp API
        result = send_whatsapp_message(phone_id, token, conv["customer_phone"], message)

        if not result or not result.get("success"):
            error_msg = result.get("error", "Falha desconhecida") if result else "Sem resposta da API"
            return jsonify({
                "error": f"Falha ao enviar: {error_msg}"
            }), 500

        # Registra no banco como mensagem do atendente humano
        db.execute(
            """INSERT INTO messages (conversation_id, sender, content, msg_type)
               VALUES (?, ?, ?, ?)""",
            (conv_id, "human", message, "text")
        )

        # Atualiza timestamp da conversa
        db.execute(
            "UPDATE conversations SET last_message_at=datetime('now') WHERE id=?",
            (conv_id,)
        )
        db.commit()

        return jsonify({
            "success": True,
            "message": "Mensagem enviada"
        })

    except Exception as e:
        err_id = secrets.token_hex(6)
        safe_log(f"[SEND MANUAL] Erro id={err_id}: {e}", level="ERROR")
        return jsonify({"error": "Erro interno ao enviar mensagem", "err_id": err_id}), 500


@app.route("/api/conversations")
@login_required
def api_conversations_list():
    db = get_db()
    convos = db.execute("""SELECT c.id, c.customer_phone, c.customer_name, c.last_message_at,
        (SELECT content FROM messages WHERE conversation_id=c.id ORDER BY created_at DESC LIMIT 1) as last_msg,
        (SELECT COUNT(*) FROM messages WHERE conversation_id=c.id) as msg_count
        FROM conversations c WHERE c.user_id=? ORDER BY c.last_message_at DESC""", (g.user["id"],)).fetchall()
    return jsonify({"conversations": [dict(c) for c in convos]})


@app.route("/dashboard/conversations/export")
@login_required
def export_all_conversations():
    """Exporta todas as conversas do usuário em CSV"""
    db = get_db()
    rows = db.execute("""SELECT c.id as conv_id, c.customer_phone, c.customer_name, c.last_message_at,
        m.sender, m.content, m.msg_type, m.created_at
        FROM conversations c
        LEFT JOIN messages m ON m.conversation_id = c.id
        WHERE c.user_id=?
        ORDER BY c.id, m.created_at""", (g.user["id"],)).fetchall()

    lines = ["ID,Telefone,Nome,Remetente,Mensagem,Tipo,Data"]
    for r in rows:
        lines.append(",".join([
            csv_safe(r["conv_id"]),
            csv_safe(r["customer_phone"]),
            csv_safe(r["customer_name"]),
            csv_safe(r["sender"]),
            csv_safe(r["content"]),
            csv_safe(r["msg_type"]),
            csv_safe(r["created_at"])
        ]))

    csv_content = "\n".join(lines)
    # UTF-8 BOM para Excel abrir acentos corretamente
    response = make_response("\ufeff" + csv_content)
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = f'attachment; filename="conversas_{datetime.now().strftime("%Y%m%d")}.csv"'
    return response


@app.route("/dashboard/conversations/<int:conv_id>/export")
@login_required
def export_single_conversation(conv_id):
    """Exporta uma conversa específica em CSV"""
    db = get_db()
    conv = db.execute("SELECT * FROM conversations WHERE id=? AND user_id=?", (conv_id, g.user["id"])).fetchone()
    if not conv:
        return "Conversa não encontrada", 404

    messages = db.execute("SELECT * FROM messages WHERE conversation_id=? ORDER BY created_at", (conv_id,)).fetchall()

    lines = ["Data,Remetente,Tipo,Mensagem"]
    for m in messages:
        lines.append(",".join([
            csv_safe(m["created_at"]),
            csv_safe("Bot" if m["sender"] == "bot" else (conv["customer_name"] or conv["customer_phone"])),
            csv_safe(m["msg_type"]),
            csv_safe(m["content"])
        ]))

    csv_content = "\n".join(lines)
    response = make_response("\ufeff" + csv_content)
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    safe_phone = re.sub(r'[^0-9]', '', conv["customer_phone"] or "contato")
    response.headers["Content-Disposition"] = f'attachment; filename="conversa_{safe_phone}_{datetime.now().strftime("%Y%m%d")}.csv"'
    return response


@app.route("/dashboard/conversations/<int:conv_id>/print")
@login_required
def print_conversation(conv_id):
    """Versão imprimível de uma conversa"""
    db = get_db()
    conv = db.execute("SELECT * FROM conversations WHERE id=? AND user_id=?", (conv_id, g.user["id"])).fetchone()
    if not conv:
        return "Conversa não encontrada", 404

    messages = db.execute("SELECT * FROM messages WHERE conversation_id=? ORDER BY created_at", (conv_id,)).fetchall()
    user = g.user

    msgs_html = ""
    for m in messages:
        sender = "Bot" if m["sender"] == "bot" else esc(conv["customer_name"] or conv["customer_phone"])
        bg = "#e8f5e9" if m["sender"] == "bot" else "#f5f5f5"
        date = esc(m["created_at"] or "")
        content = esc(m["content"])
        msgs_html += f'<div style="background:{bg};padding:12px;margin-bottom:8px;border-radius:8px"><div style="font-size:12px;color:#666;margin-bottom:4px"><strong>{sender}</strong> — {date}</div><div>{content}</div></div>'

    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><link rel="icon" type="image/png" href="/favicon.ico"><link rel="apple-touch-icon" href="/apple-touch-icon.png"><meta name="theme-color" content="#6366f1">
<title>Conversa — {esc(conv['customer_name'] or conv['customer_phone'])}</title>
<style>
body{{font-family:Arial,sans-serif;max-width:800px;margin:0 auto;padding:24px;background:#fff;color:#000}}
h1{{border-bottom:2px solid #00c896;padding-bottom:8px}}
.info{{background:#f9f9f9;padding:12px;margin-bottom:20px;border-radius:8px;font-size:14px}}
@media print{{.no-print{{display:none}}}}
</style></head><body>
<h1>Conversa com {esc(conv['customer_name'] or conv['customer_phone'])}</h1>
<div class="info">
<strong>Telefone:</strong> {esc(conv['customer_phone'])}<br>
<strong>Última mensagem:</strong> {esc(conv['last_message_at'])}<br>
<strong>Total de mensagens:</strong> {len(messages)}<br>
<strong>Empresa:</strong> {esc(user['company'] or user['name'])}
</div>
<div class="no-print" style="margin-bottom:20px">
<button id="print-conversation-btn" style="padding:10px 20px;background:#00c896;color:white;border:none;border-radius:6px;cursor:pointer;font-size:14px">🖨️ Imprimir</button>
<a href="/dashboard/conversations/{conv_id}/export" style="margin-left:8px;padding:10px 20px;background:#0ea5e9;color:white;text-decoration:none;border-radius:6px;font-size:14px">📥 Baixar CSV</a>
<a href="/dashboard/conversations" style="margin-left:8px;padding:10px 20px;background:#666;color:white;text-decoration:none;border-radius:6px;font-size:14px">← Voltar</a>
</div>
{msgs_html}
<div style="margin-top:40px;font-size:12px;color:#999;text-align:center">
Exportado em {datetime.now().strftime("%d/%m/%Y %H:%M")} — atendente.online
</div>
<script nonce="{getattr(g, 'csp_nonce', '')}">
document.getElementById('print-conversation-btn')?.addEventListener('click', function(){{ window.print(); }});
</script>
</body></html>"""


# ─── WHATSAPP WEBHOOK ─────────────────────────────────────────
def verify_whatsapp_signature(request_data, signature_header, app_secret):
    """Valida assinatura X-Hub-Signature-256 da Meta"""
    if not signature_header or not app_secret:
        return False
    if not signature_header.startswith("sha256="):
        return False
    import hmac
    expected = hmac.new(app_secret.encode(), request_data, hashlib.sha256).hexdigest()
    received = signature_header.replace("sha256=", "")
    return hmac.compare_digest(expected, received)


@app.route("/webhook/whatsapp", methods=["GET","POST"])
@app.route("/webhook/whatsapp/<int:user_id>", methods=["GET","POST"])
def whatsapp_webhook(user_id=None):
    if request.method == "GET":
        wa_verify = get_setting("WHATSAPP_VERIFY_TOKEN", WHATSAPP_VERIFY_TOKEN)
        if request.args.get("hub.mode") == "subscribe" and request.args.get("hub.verify_token") == wa_verify:
            return request.args.get("hub.challenge",""), 200
        return "Forbidden", 403

    # Valida assinatura da Meta — OBRIGATÓRIO em produção
    app_secret = get_setting("WHATSAPP_APP_SECRET", "")
    is_dev = os.getenv("FLASK_ENV", "").lower() == "development"

    if app_secret:
        # APP_SECRET configurado → valida assinatura
        signature = request.headers.get("X-Hub-Signature-256", "")
        if not verify_whatsapp_signature(request.get_data(), signature, app_secret):
            safe_log(f"[WEBHOOK] Assinatura inválida do WhatsApp/Meta", level="WARN")
            return jsonify({"status":"invalid signature"}), 403
    elif not is_dev:
        # Produção sem APP_SECRET → recusa (força configuração)
        safe_log(f"[WEBHOOK] REJEITADO: WHATSAPP_APP_SECRET não configurado em produção", level="ERROR")
        return jsonify({"status":"webhook not configured - APP_SECRET required"}), 503

    data = request.json or {}
    try:
        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row

        for entry in data.get("entry", []):
            for change in entry.get("changes", []):
                value = change.get("value", {})
                metadata = value.get("metadata", {}) or {}
                phone_number_id = metadata.get("phone_number_id", "")
                raw_user = resolve_user_by_whatsapp_phone_id(db_conn, phone_number_id)
                if not raw_user:
                    log_webhook_error("whatsapp", None, "UserNotMapped", f"phone_number_id={phone_number_id}", {"metadata": metadata})
                    continue

                user = decrypt_user_row(raw_user)
                resolved_user_id = user["id"]

                for msg in value.get("messages", []):
                    wamid = msg.get("id", "")
                    if wamid and not register_processed_webhook_event("whatsapp", wamid, resolved_user_id, msg):
                        continue

                    sender_phone = msg.get("from", "")

                    blocked = db_conn.execute("SELECT id FROM blocked_contacts WHERE user_id=? AND phone=?", (resolved_user_id, sender_phone)).fetchone()
                    if blocked:
                        continue

                    media_result = process_whatsapp_media(msg, user["whatsapp_token"])
                    if not media_result["content"]:
                        continue

                    conv = db_conn.execute(
                        "SELECT * FROM conversations WHERE user_id=? AND customer_phone=? AND status='active' AND channel='whatsapp'",
                        (resolved_user_id, sender_phone)
                    ).fetchone()
                    if not conv:
                        contact_name = ""
                        contacts = value.get("contacts", [])
                        if contacts:
                            contact_name = contacts[0].get("profile", {}).get("name", "")
                        db_conn.execute(
                            "INSERT INTO conversations (user_id,customer_phone,customer_name,channel) VALUES (?,?,?,?)",
                            (resolved_user_id, sender_phone, contact_name, "whatsapp")
                        )
                        db_conn.commit()
                        conv = db_conn.execute(
                            "SELECT * FROM conversations WHERE user_id=? AND customer_phone=? AND status='active' AND channel='whatsapp'",
                            (resolved_user_id, sender_phone)
                        ).fetchone()

                    db_conn.execute(
                        "INSERT INTO messages (conversation_id,sender,content,msg_type,media_url,external_message_id) VALUES (?,?,?,?,?,?)",
                        (conv["id"], "customer", media_result["content"], media_result["type"], media_result.get("media_path", ""), wamid)
                    )
                    db_conn.execute("UPDATE conversations SET last_message_at=datetime('now') WHERE id=?", (conv["id"],))

                    if conv["is_human_takeover"]:
                        db_conn.commit()
                        continue

                    can_process, access_code, access_message = validate_user_messaging_access(user)
                    if not can_process:
                        db_conn.execute(
                            "INSERT INTO messages (conversation_id,sender,content,msg_type) VALUES (?,?,?,?)",
                            (conv["id"], "system", f"[BLOQUEIO AUTOMATICO: {access_code}] {access_message}", "system")
                        )
                        db_conn.commit()
                        continue

                    ai_input = media_result.get("description", media_result["content"])
                    if media_result["type"] == "audio":
                        ai_input = f"[MENSAGEM DE AUDIO DO CLIENTE]: {ai_input}"

                    commerce_triggered = False
                    if user["commerce_enabled"] and user["mp_access_token"] and media_result["type"] == "text":
                        products = db_conn.execute(
                            "SELECT * FROM product_gallery WHERE user_id=? AND active=1 AND price > 0",
                            (resolved_user_id,)
                        ).fetchall()
                        products_list = [dict(p) for p in products]

                        if products_list:
                            intent = detect_purchase_intent(ai_input, products_list)
                            if intent and intent.get("is_purchase") and intent.get("confidence") in ("high", "medium"):
                                db_conn.close()
                                order = create_order_from_intent(user, conv["id"], sender_phone, intent)
                                if order:
                                    commerce_msg = format_order_message(order)
                                    send_whatsapp_message(
                                        user["whatsapp_phone_id"],
                                        user["whatsapp_token"],
                                        sender_phone,
                                        commerce_msg
                                    )
                                    db_conn = sqlite3.connect(DATABASE)
                                    db_conn.row_factory = sqlite3.Row
                                    db_conn.execute(
                                        "INSERT INTO messages (conversation_id,sender,content,msg_type) VALUES (?,?,?,?)",
                                        (conv["id"], "bot", commerce_msg, "text")
                                    )
                                    db_conn.execute("UPDATE users SET msgs_used=msgs_used+1 WHERE id=?", (resolved_user_id,))
                                    db_conn.commit()
                                    commerce_triggered = True
                                else:
                                    db_conn = sqlite3.connect(DATABASE)
                                    db_conn.row_factory = sqlite3.Row

                    if commerce_triggered:
                        continue

                    ai_response = generate_ai_response(user, conv["id"], ai_input, db_conn)

                    db_conn.execute("INSERT INTO messages (conversation_id,sender,content,msg_type) VALUES (?,?,?,?)", (conv["id"], "bot", ai_response, "text"))
                    db_conn.execute("UPDATE users SET msgs_used=msgs_used+1 WHERE id=?", (resolved_user_id,))
                    db_conn.commit()

                    product = find_matching_product(resolved_user_id, ai_input)
                    if product and os.path.exists(product["file_path"]):
                        safe_log(f"[GALLERY] Enviando foto: {product['name']}")
                        caption = product["description"] or product["name"]
                        send_whatsapp_image(
                            user["whatsapp_phone_id"],
                            user["whatsapp_token"],
                            sender_phone,
                            product["file_path"],
                            caption=caption
                        )
                        db_conn.execute(
                            "INSERT INTO messages (conversation_id,sender,content,msg_type) VALUES (?,?,?,?)",
                            (conv["id"], "bot", f"[Foto: {product['name']}]", "image")
                        )
                        db_conn.commit()

                    audio_sent = False
                    if media_result["type"] == "audio":
                        safe_log(f"[VOICE] Cliente enviou audio, gerando resposta por voz...")
                        audio_path = text_to_audio(ai_response)
                        if audio_path:
                            audio_sent = send_whatsapp_audio(user["whatsapp_phone_id"], user["whatsapp_token"], sender_phone, audio_path)
                            try:
                                os.remove(audio_path)
                            except Exception:
                                pass

                    if not audio_sent:
                        send_whatsapp_message(user["whatsapp_phone_id"], user["whatsapp_token"], sender_phone, ai_response)
        db_conn.close()
    except sqlite3.OperationalError as e:
        # Erro transitório de DB — Meta deve retentar
        safe_log(f"[WA WEBHOOK] DB error (retentativa necessária): {e}", level="ERROR")
        log_webhook_error("whatsapp", user_id, "DBOperationalError", str(e), data)
        return jsonify({"status": "temporary_error"}), 500
    except (ConnectionError, TimeoutError) as e:
        # Erro de conexão com API externa — retentar
        safe_log(f"[WA WEBHOOK] Erro de conexão: {e}", level="ERROR")
        log_webhook_error("whatsapp", user_id, type(e).__name__, str(e), data)
        return jsonify({"status": "temporary_error"}), 500
    except Exception as e:
        # Erro permanente — logamos mas respondemos 200 para Meta não retentar infinitamente
        safe_log(f"[WA WEBHOOK] Erro permanente: {e}", level="ERROR")
        import traceback
        traceback.print_exc()
        log_webhook_error("whatsapp", user_id, type(e).__name__, str(e), data)
    return jsonify({"status":"ok"}), 200


# ═══════════════════════════════════════════════════════════════
#  INSTAGRAM DIRECT — Webhook e envio de mensagens
# ═══════════════════════════════════════════════════════════════

def send_instagram_message(page_id, token, recipient_id, message):
    """Envia mensagem via Instagram Direct"""
    if not page_id or not token:
        log_debug("[IG SEND] Página ID ou token vazio")
        return False
    try:
        import requests as req
        url = f"https://graph.facebook.com/v18.0/{page_id}/messages"
        params = {"access_token": token}
        payload = {
            "recipient": {"id": recipient_id},
            "message": {"text": message[:1000]},  # Instagram limit
            "messaging_type": "RESPONSE"
        }
        resp = req.post(url, params=params, json=payload, timeout=15)
        if resp.status_code == 200:
            safe_log(f"[IG SEND] ✓ Enviado para {recipient_id}")
            return True
        safe_log(f"[IG SEND] Erro {resp.status_code}: {_short_resp_text(resp)}", level="ERROR")
        return False
    except Exception as e:
        safe_log(f"[IG SEND] Exceção: {e}", level="ERROR")
        return False


@app.route("/webhook/instagram/<int:user_id>", methods=["GET", "POST"])
def webhook_instagram(user_id):
    """Webhook do Instagram Direct"""
    # Verificação inicial (quando Meta conecta)
    if request.method == "GET":
        mode = request.args.get("hub.mode")
        token = request.args.get("hub.verify_token")
        challenge = request.args.get("hub.challenge")
        verify_token = get_setting("WHATSAPP_VERIFY_TOKEN", WHATSAPP_VERIFY_TOKEN)
        if mode == "subscribe" and token == verify_token:
            return challenge, 200
        return "Unauthorized", 403

    # Valida assinatura em produção
    app_secret = get_setting("WHATSAPP_APP_SECRET", "")
    is_dev = os.getenv("FLASK_ENV", "").lower() == "development"
    if app_secret:
        signature = request.headers.get("X-Hub-Signature-256", "")
        if not verify_whatsapp_signature(request.get_data(), signature, app_secret):
            safe_log(f"[IG WEBHOOK] Assinatura inválida")
            return jsonify({"status":"invalid signature"}), 403
    elif not is_dev:
        return jsonify({"status":"APP_SECRET required"}), 503

    data = request.json or {}
    try:
        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row
        raw_user = db_conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
        if not raw_user:
            db_conn.close()
            return jsonify({"status":"user not found"}), 404
        user = decrypt_user_row(raw_user)

        # Instagram envia messaging events
        for entry in data.get("entry", []):
            for event in entry.get("messaging", []):
                sender_id = event.get("sender", {}).get("id", "")
                if not sender_id or sender_id == user["instagram_page_id"]:
                    continue  # Ignora echo da própria página

                msg = event.get("message", {})
                if not msg or msg.get("is_echo"):
                    continue
                event_id = msg.get("mid") or event.get("timestamp") or f"{sender_id}:{msg.get('text','')[:40]}"
                if not register_processed_webhook_event("instagram", event_id, user_id, {"sender": sender_id, "mid": msg.get("mid", "")}):
                    continue

                message_text = msg.get("text", "")
                if not message_text:
                    # Mensagem não é texto (imagem, sticker, etc) — ignora por ora
                    continue

                # Busca ou cria conversa
                conv = db_conn.execute(
                    "SELECT * FROM conversations WHERE user_id=? AND customer_phone=? AND channel='instagram'",
                    (user_id, sender_id)
                ).fetchone()

                if not conv:
                    db_conn.execute(
                        "INSERT INTO conversations (user_id, customer_phone, customer_name, channel, last_message_at) VALUES (?,?,?,?,datetime('now'))",
                        (user_id, sender_id, f"IG: {sender_id[:10]}", "instagram")
                    )
                    db_conn.commit()
                    conv = db_conn.execute(
                        "SELECT * FROM conversations WHERE user_id=? AND customer_phone=? AND channel='instagram'",
                        (user_id, sender_id)
                    ).fetchone()

                # Salva mensagem do cliente
                db_conn.execute(
                    "INSERT INTO messages (conversation_id, sender, content, msg_type) VALUES (?,?,?,?)",
                    (conv["id"], "customer", message_text, "text")
                )
                db_conn.execute(
                    "UPDATE conversations SET last_message_at=datetime('now') WHERE id=?",
                    (conv["id"],)
                )
                db_conn.commit()

                if conv["is_human_takeover"]:
                    continue

                can_process, access_code, access_message = validate_user_messaging_access(user)
                if not can_process:
                    db_conn.execute(
                        "INSERT INTO messages (conversation_id, sender, content, msg_type) VALUES (?,?,?,?)",
                        (conv["id"], "system", f"[BLOQUEIO AUTOMATICO: {access_code}] {access_message}", "system")
                    )
                    db_conn.commit()
                    continue

                # Gera resposta com IA
                ai_response = generate_ai_response(user, conv["id"], message_text, db_conn)

                # Salva resposta do bot
                db_conn.execute(
                    "INSERT INTO messages (conversation_id, sender, content, msg_type) VALUES (?,?,?,?)",
                    (conv["id"], "bot", ai_response, "text")
                )
                db_conn.execute(
                    "UPDATE users SET msgs_used=msgs_used+1 WHERE id=?",
                    (user_id,)
                )
                db_conn.commit()

                # Envia resposta no Instagram
                send_instagram_message(
                    user["instagram_page_id"],
                    user["instagram_token"],
                    sender_id,
                    ai_response
                )
        db_conn.close()
    except sqlite3.OperationalError as e:
        safe_log(f"[IG WEBHOOK] DB error: {e}", level="ERROR")
        log_webhook_error("instagram", user_id, "DBOperationalError", str(e), data)
        return jsonify({"status": "temporary_error"}), 500
    except (ConnectionError, TimeoutError) as e:
        safe_log(f"[IG WEBHOOK] Erro de conexão: {e}", level="ERROR")
        log_webhook_error("instagram", user_id, type(e).__name__, str(e), data)
        return jsonify({"status": "temporary_error"}), 500
    except Exception as e:
        safe_log(f"[IG WEBHOOK] Erro permanente: {e}", level="ERROR")
        import traceback
        traceback.print_exc()
        log_webhook_error("instagram", user_id, type(e).__name__, str(e), data)
    return jsonify({"status":"ok"}), 200


# ═══════════════════════════════════════════════════════════════
#  MESSENGER (Facebook) — Webhook e envio de mensagens
# ═══════════════════════════════════════════════════════════════

def send_messenger_message(page_id, token, recipient_id, message):
    """Envia mensagem via Facebook Messenger"""
    if not page_id or not token:
        return False
    try:
        import requests as req
        url = f"https://graph.facebook.com/v18.0/{page_id}/messages"
        params = {"access_token": token}
        payload = {
            "recipient": {"id": recipient_id},
            "message": {"text": message[:2000]},
            "messaging_type": "RESPONSE"
        }
        resp = req.post(url, params=params, json=payload, timeout=15)
        if resp.status_code == 200:
            safe_log(f"[MSG SEND] ✓ Enviado para {recipient_id}")
            return True
        safe_log(f"[MSG SEND] Erro {resp.status_code}: {_short_resp_text(resp)}", level="ERROR")
        return False
    except Exception as e:
        safe_log(f"[MSG SEND] Exceção: {e}", level="ERROR")
        return False


@app.route("/webhook/messenger/<int:user_id>", methods=["GET", "POST"])
def webhook_messenger(user_id):
    """Webhook do Facebook Messenger"""
    if request.method == "GET":
        mode = request.args.get("hub.mode")
        token = request.args.get("hub.verify_token")
        challenge = request.args.get("hub.challenge")
        verify_token = get_setting("WHATSAPP_VERIFY_TOKEN", WHATSAPP_VERIFY_TOKEN)
        if mode == "subscribe" and token == verify_token:
            return challenge, 200
        return "Unauthorized", 403

    # Valida assinatura
    app_secret = get_setting("WHATSAPP_APP_SECRET", "")
    is_dev = os.getenv("FLASK_ENV", "").lower() == "development"
    if app_secret:
        signature = request.headers.get("X-Hub-Signature-256", "")
        if not verify_whatsapp_signature(request.get_data(), signature, app_secret):
            return jsonify({"status":"invalid signature"}), 403
    elif not is_dev:
        return jsonify({"status":"APP_SECRET required"}), 503

    data = request.json or {}
    try:
        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row
        raw_user = db_conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
        if not raw_user:
            db_conn.close()
            return jsonify({"status":"user not found"}), 404
        user = decrypt_user_row(raw_user)

        for entry in data.get("entry", []):
            for event in entry.get("messaging", []):
                sender_id = event.get("sender", {}).get("id", "")
                if not sender_id or sender_id == user["messenger_page_id"]:
                    continue

                msg = event.get("message", {})
                if not msg or msg.get("is_echo"):
                    continue
                event_id = msg.get("mid") or event.get("timestamp") or f"{sender_id}:{msg.get('text','')[:40]}"
                if not register_processed_webhook_event("messenger", event_id, user_id, {"sender": sender_id, "mid": msg.get("mid", "")}):
                    continue

                message_text = msg.get("text", "")
                if not message_text:
                    continue

                # Busca ou cria conversa
                conv = db_conn.execute(
                    "SELECT * FROM conversations WHERE user_id=? AND customer_phone=? AND channel='messenger'",
                    (user_id, sender_id)
                ).fetchone()

                if not conv:
                    db_conn.execute(
                        "INSERT INTO conversations (user_id, customer_phone, customer_name, channel, last_message_at) VALUES (?,?,?,?,datetime('now'))",
                        (user_id, sender_id, f"FB: {sender_id[:10]}", "messenger")
                    )
                    db_conn.commit()
                    conv = db_conn.execute(
                        "SELECT * FROM conversations WHERE user_id=? AND customer_phone=? AND channel='messenger'",
                        (user_id, sender_id)
                    ).fetchone()

                db_conn.execute(
                    "INSERT INTO messages (conversation_id, sender, content, msg_type) VALUES (?,?,?,?)",
                    (conv["id"], "customer", message_text, "text")
                )
                db_conn.execute(
                    "UPDATE conversations SET last_message_at=datetime('now') WHERE id=?",
                    (conv["id"],)
                )
                db_conn.commit()

                if conv["is_human_takeover"]:
                    continue

                can_process, access_code, access_message = validate_user_messaging_access(user)
                if not can_process:
                    db_conn.execute(
                        "INSERT INTO messages (conversation_id, sender, content, msg_type) VALUES (?,?,?,?)",
                        (conv["id"], "system", f"[BLOQUEIO AUTOMATICO: {access_code}] {access_message}", "system")
                    )
                    db_conn.commit()
                    continue

                ai_response = generate_ai_response(user, conv["id"], message_text, db_conn)

                db_conn.execute(
                    "INSERT INTO messages (conversation_id, sender, content, msg_type) VALUES (?,?,?,?)",
                    (conv["id"], "bot", ai_response, "text")
                )
                db_conn.execute(
                    "UPDATE users SET msgs_used=msgs_used+1 WHERE id=?",
                    (user_id,)
                )
                db_conn.commit()

                send_messenger_message(
                    user["messenger_page_id"],
                    user["messenger_token"],
                    sender_id,
                    ai_response
                )
        db_conn.close()
    except sqlite3.OperationalError as e:
        safe_log(f"[MSG WEBHOOK] DB error: {e}", level="ERROR")
        log_webhook_error("messenger", user_id, "DBOperationalError", str(e), data)
        return jsonify({"status": "temporary_error"}), 500
    except (ConnectionError, TimeoutError) as e:
        safe_log(f"[MSG WEBHOOK] Erro de conexão: {e}", level="ERROR")
        log_webhook_error("messenger", user_id, type(e).__name__, str(e), data)
        return jsonify({"status": "temporary_error"}), 500
    except Exception as e:
        safe_log(f"[MSG WEBHOOK] Erro permanente: {e}", level="ERROR")
        import traceback
        traceback.print_exc()
        log_webhook_error("messenger", user_id, type(e).__name__, str(e), data)
    return jsonify({"status":"ok"}), 200


# ═══════════════════════════════════════════════════════════════
#  AGÊNCIA DIGITAL — Biblioteca de mídia + IA + Telegram
# ═══════════════════════════════════════════════════════════════

def send_telegram_message(bot_token, chat_id, text, parse_mode="HTML"):
    """Envia mensagem de texto via Telegram Bot"""
    if not bot_token or not chat_id:
        return None
    try:
        import requests as req
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        payload = {
            "chat_id": chat_id,
            "text": text[:4000],
            "parse_mode": parse_mode
        }
        resp = req.post(url, json=payload, timeout=15)
        if resp.status_code == 200:
            return resp.json().get("result", {}).get("message_id")
        safe_log(f"[TG] Erro {resp.status_code}: {_short_resp_text(resp)}", level="ERROR")
        return None
    except Exception as e:
        safe_log(f"[TG] Exceção: {e}", level="ERROR")
        return None


def send_telegram_photo(bot_token, chat_id, photo_path, caption="", reply_markup=None):
    """Envia foto via Telegram Bot com botões de aprovação"""
    if not bot_token or not chat_id:
        return None
    try:
        import requests as req, json as json_mod
        url = f"https://api.telegram.org/bot{bot_token}/sendPhoto"
        with open(photo_path, "rb") as f:
            files = {"photo": f}
            data = {
                "chat_id": chat_id,
                "caption": caption[:1024],
                "parse_mode": "HTML"
            }
            if reply_markup:
                data["reply_markup"] = json_mod.dumps(reply_markup)
            resp = req.post(url, files=files, data=data, timeout=30)
        if resp.status_code == 200:
            return resp.json().get("result", {}).get("message_id")
        safe_log(f"[TG PHOTO] Erro {resp.status_code}: {_short_resp_text(resp)}", level="ERROR")
        return None
    except Exception as e:
        safe_log(f"[TG PHOTO] Exceção: {e}", level="ERROR")
        return None


def generate_social_caption(user, media_description, theme="geral"):
    """Usa Claude para gerar legenda + hashtags para post"""
    try:
        import requests as req
        api_key = get_setting("ANTHROPIC_API_KEY", "")
        if not api_key:
            return None

        business_context = user["social_business_context"] or user["ai_system_prompt"] or ""
        tone = user["social_post_tone"] or "profissional"

        prompt = f"""Você é um especialista em redes sociais. Crie uma legenda para post no Instagram/Facebook.

CONTEXTO DO NEGÓCIO:
{business_context}

TEMA DA IMAGEM: {theme}
DESCRIÇÃO DA IMAGEM: {media_description or 'Imagem promocional'}

TOM: {tone}

REGRAS:
- Legenda envolvente, máximo 150 palavras
- Comece com um gancho (pergunta, fato, emoji)
- Inclua call-to-action natural no final
- 8 a 12 hashtags relevantes ao final
- Use emojis com moderação (2-4 no total)
- Linguagem brasileira, natural

FORMATO DA RESPOSTA (exato):
LEGENDA: [texto da legenda aqui]
HASHTAGS: #tag1 #tag2 #tag3 ...

Responda apenas com a LEGENDA e HASHTAGS, nada mais."""

        resp = req.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={
                "model": "claude-sonnet-4-6",
                "max_tokens": 500,
                "messages": [{"role": "user", "content": prompt}]
            },
            timeout=30
        )
        if resp.status_code != 200:
            safe_log(f"[SOCIAL AI] Erro: {resp.status_code}", level="ERROR")
            return None

        text = resp.json()["content"][0]["text"]
        # Parse da resposta
        caption = ""
        hashtags = ""
        for line in text.split("\n"):
            if line.startswith("LEGENDA:"):
                caption = line.replace("LEGENDA:", "").strip()
            elif line.startswith("HASHTAGS:"):
                hashtags = line.replace("HASHTAGS:", "").strip()
            elif caption and not line.startswith("HASHTAGS:") and not hashtags:
                # Continua a legenda se está em múltiplas linhas
                caption += "\n" + line.strip()

        # Fallback: se não conseguir parsear, usa tudo como caption
        if not caption:
            caption = text[:500]

        return {"caption": caption.strip(), "hashtags": hashtags.strip()}
    except Exception as e:
        safe_log(f"[SOCIAL AI] Exceção: {e}", level="ERROR")
        return None


def describe_image_for_caption(image_path):
    """Usa Claude Vision para descrever a imagem (ajuda a gerar legenda contextual)"""
    try:
        import base64 as b64, requests as req
        api_key = get_setting("ANTHROPIC_API_KEY", "")
        if not api_key:
            return ""

        with open(image_path, "rb") as f:
            image_data = b64.b64encode(f.read()).decode("utf-8")

        ext = image_path.lower().split(".")[-1]
        media_type = "image/jpeg" if ext in ("jpg", "jpeg") else f"image/{ext}"

        resp = req.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={
                "model": "claude-sonnet-4-6",
                "max_tokens": 200,
                "messages": [{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_data}},
                    {"type": "text", "text": "Descreva esta imagem em uma frase curta em português, focando em elementos que podem inspirar uma legenda de rede social. Máximo 30 palavras."}
                ]}]
            },
            timeout=30
        )
        if resp.status_code == 200:
            return resp.json()["content"][0]["text"]
        return ""
    except Exception as e:
        safe_log(f"[VISION] Erro: {e}", level="ERROR")
        return ""


def create_social_post(user_id, media_id=None):
    """Cria um post: escolhe mídia, gera legenda, envia para Telegram aprovar"""
    try:
        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row

        raw_user = db_conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
        if not raw_user:
            db_conn.close()
            return None
        user = decrypt_user_row(raw_user)

        # Se não especificou media_id, escolhe a menos usada
        if media_id:
            media = db_conn.execute(
                "SELECT * FROM social_media_library WHERE id=? AND user_id=?",
                (media_id, user_id)
            ).fetchone()
        else:
            media = db_conn.execute(
                """SELECT * FROM social_media_library
                   WHERE user_id=?
                   ORDER BY times_used ASC, last_used_at ASC
                   LIMIT 1""",
                (user_id,)
            ).fetchone()

        if not media:
            db_conn.close()
            safe_log(f"[SOCIAL] Usuário {user_id} sem mídia na biblioteca")
            return None

        # Descreve a imagem com IA
        description = describe_image_for_caption(media["file_path"])

        # Gera legenda + hashtags
        ai_result = generate_social_caption(user, description, media["theme"])
        if not ai_result:
            db_conn.close()
            return None

        caption = ai_result["caption"]
        hashtags = ai_result["hashtags"]

        # Cria registro do post
        cur = db_conn.execute(
            """INSERT INTO scheduled_posts
               (user_id, media_id, caption, hashtags, status, scheduled_for)
               VALUES (?,?,?,?,?,?)""",
            (user_id, media["id"], caption, hashtags, "pending",
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        post_id = cur.lastrowid
        db_conn.commit()

        # Envia para aprovação via Telegram
        if user["telegram_bot_token"] and user["telegram_chat_id"]:
            full_text = f"{caption}\n\n{hashtags}"
            message = (
                f"📸 <b>Sugestão de post</b>\n\n"
                f"{full_text}\n\n"
                f"<i>Post ID: {post_id}</i>\n"
                f"Aprove pela interface web ou responda aqui."
            )
            msg_id = send_telegram_photo(
                user["telegram_bot_token"],
                user["telegram_chat_id"],
                media["file_path"],
                caption=full_text[:1000],
                reply_markup={
                    "inline_keyboard": [[
                        {"text": "✅ Aprovar", "callback_data": f"approve_{post_id}"},
                        {"text": "❌ Rejeitar", "callback_data": f"reject_{post_id}"}
                    ]]
                }
            )
            if msg_id:
                db_conn.execute(
                    "UPDATE scheduled_posts SET telegram_message_id=? WHERE id=?",
                    (str(msg_id), post_id)
                )
                db_conn.commit()

        db_conn.close()
        return post_id
    except Exception as e:
        safe_log(f"[SOCIAL] Erro create_social_post: {e}", level="ERROR")
        return None


# ─── ROTAS: BIBLIOTECA DE MÍDIA ──────────────────────────────
@app.route("/dashboard/social")
@login_required
def social_dashboard():
    """Dashboard principal da agência digital"""
    user = g.user
    db = get_db()

    # Estatísticas
    total_media = db.execute(
        "SELECT COUNT(*) as c FROM social_media_library WHERE user_id=?", (user["id"],)
    ).fetchone()["c"]
    total_posts = db.execute(
        "SELECT COUNT(*) as c FROM scheduled_posts WHERE user_id=?", (user["id"],)
    ).fetchone()["c"]
    pending = db.execute(
        "SELECT COUNT(*) as c FROM scheduled_posts WHERE user_id=? AND status='pending'", (user["id"],)
    ).fetchone()["c"]
    approved = db.execute(
        "SELECT COUNT(*) as c FROM scheduled_posts WHERE user_id=? AND status='approved'", (user["id"],)
    ).fetchone()["c"]

    # Mídia recente
    media_items = db.execute(
        "SELECT * FROM social_media_library WHERE user_id=? ORDER BY created_at DESC LIMIT 8",
        (user["id"],)
    ).fetchall()

    media_html = ""
    if media_items:
        for m in media_items:
            img_url = f"/media/social/{m['id']}"
            media_html += f"""
            <div class="card" style="padding:12px">
                <img src="{img_url}" style="width:100%;height:140px;object-fit:cover;border-radius:8px;margin-bottom:8px" alt="">
                <p style="font-size:12px;color:var(--text3);margin:0">{esc(m['theme'])} · usada {m['times_used']}x</p>
            </div>
            """
    else:
        media_html = '<p style="color:var(--text3);grid-column:1/-1;text-align:center;padding:40px">Nenhuma mídia cadastrada ainda.</p>'

    # Posts recentes
    posts = db.execute(
        """SELECT sp.*, sml.file_path FROM scheduled_posts sp
           LEFT JOIN social_media_library sml ON sp.media_id=sml.id
           WHERE sp.user_id=? ORDER BY sp.created_at DESC LIMIT 10""",
        (user["id"],)
    ).fetchall()

    posts_html = ""
    if posts:
        for p in posts:
            status_cls = {
                "pending": "badge-orange",
                "approved": "badge-green",
                "rejected": "badge-red",
                "posted": "badge-purple"
            }.get(p["status"], "badge-orange")
            status_label = {
                "pending": "Aguardando aprovação",
                "approved": "Aprovado",
                "rejected": "Rejeitado",
                "posted": "Publicado"
            }.get(p["status"], p["status"])

            img_tag = ""
            if p["media_id"]:
                img_tag = f'<img src="/media/social/{p["media_id"]}" style="width:80px;height:80px;object-fit:cover;border-radius:6px;margin-right:12px">'

            posts_html += f"""
            <div style="display:flex;padding:12px;border-bottom:1px solid rgba(255,255,255,0.06);align-items:flex-start">
                {img_tag}
                <div style="flex:1">
                    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px">
                        <span class="badge {status_cls}">{status_label}</span>
                        <span style="color:var(--text3);font-size:11px">{to_br_datetime(p['created_at'])}</span>
                    </div>
                    <p style="color:var(--text2);font-size:13px;margin:0 0 4px 0;line-height:1.4">{esc(p['caption'][:160])}{'...' if len(p['caption'] or '') > 160 else ''}</p>
                    <p style="color:var(--accent2);font-size:11px;margin:0">{esc(p['hashtags'][:120])}</p>
                    <div style="margin-top:8px;display:flex;gap:8px">
                        {'<form method="POST" action="/dashboard/social/posts/' + str(p['id']) + '/approve" style="display:inline">' + csrf_field() + '<button type="submit" class="btn btn-success btn-sm">✅ Aprovar</button></form>' if p['status']=='pending' else ''}
                        {'<form method="POST" action="/dashboard/social/posts/' + str(p['id']) + '/reject" style="display:inline">' + csrf_field() + '<button type="submit" class="btn btn-danger btn-sm">❌ Rejeitar</button></form>' if p['status']=='pending' else ''}
                    </div>
                </div>
            </div>
            """
    else:
        posts_html = '<p style="color:var(--text3);text-align:center;padding:40px">Nenhum post gerado ainda. Cadastre mídias e clique em "Gerar post agora".</p>'

    content = f"""<div class="container">
        <div class="page-header fade-in">
            <h1>📸 Agência Digital</h1>
            <p>Sistema automatizado de criação de conteúdo com IA + aprovação por Telegram</p>
        </div>

        <div class="grid-4 fade-in fade-in-1">
            <div class="metric-card"><div style="font-size:24px">🖼️</div><div class="metric-value">{total_media}</div><div class="metric-label">Mídias cadastradas</div></div>
            <div class="metric-card"><div style="font-size:24px">📝</div><div class="metric-value">{total_posts}</div><div class="metric-label">Posts gerados</div></div>
            <div class="metric-card"><div style="font-size:24px">⏳</div><div class="metric-value" style="color:var(--orange)">{pending}</div><div class="metric-label">Aguardando aprovação</div></div>
            <div class="metric-card"><div style="font-size:24px">✅</div><div class="metric-value" style="color:var(--green2)">{approved}</div><div class="metric-label">Aprovados</div></div>
        </div>

        <div class="grid-2 fade-in fade-in-2">
            <div class="card">
                <div class="card-header"><span class="card-title">🚀 Ações rápidas</span></div>
                <form method="POST" action="/dashboard/social/generate" style="margin-bottom:12px">{csrf_field()}
                    <button type="submit" class="btn btn-primary" style="width:100%">✨ Gerar novo post agora</button>
                </form>
                <a href="/dashboard/social/library" class="btn" style="width:100%;background:rgba(255,255,255,0.05);margin-bottom:12px;display:block;text-align:center;text-decoration:none">📤 Gerenciar biblioteca de mídia</a>
                <a href="/dashboard/social/settings" class="btn" style="width:100%;background:rgba(255,255,255,0.05);display:block;text-align:center;text-decoration:none">⚙️ Configurações e Telegram</a>
            </div>
            <div class="card">
                <div class="card-header"><span class="card-title">📚 Mídias recentes</span></div>
                <div class="grid-4" style="gap:8px">
                    {media_html}
                </div>
            </div>
        </div>

        <div class="card fade-in fade-in-3" style="margin-top:24px">
            <div class="card-header"><span class="card-title">📋 Histórico de posts</span></div>
            <div>{posts_html}</div>
        </div>
    </div>"""
    return base_html("Agência Digital", content, dict(user))


@app.route("/dashboard/social/library", methods=["GET", "POST"])
@login_required
def social_library():
    """Gerenciar biblioteca de mídia"""
    user = g.user
    db = get_db()
    msg = ""

    if request.method == "POST":
        if "photo" in request.files:
            file = request.files["photo"]
            if file.filename:
                image_bytes = file.read()
                if len(image_bytes) > 10 * 1024 * 1024:
                    msg = '<div class="alert alert-error">Arquivo maior que 10MB</div>'
                else:
                    # Validação real com Pillow
                    validated_bytes, real_ct = validate_and_normalize_image(image_bytes)
                    if validated_bytes is None:
                        msg = '<div class="alert alert-error">Arquivo inválido. Envie JPG ou PNG real.</div>'
                    else:
                        lib_dir = os.path.join(MEDIA_FOLDER, "social")
                        os.makedirs(lib_dir, exist_ok=True)
                        ext = "jpg" if "jpeg" in real_ct else "png"
                        ts = int(time.time() * 1000)
                        fname = f"user{user['id']}_{ts}.{ext}"
                        fpath = os.path.join(lib_dir, fname)
                        with open(fpath, "wb") as f:
                            f.write(validated_bytes)
                        theme = request.form.get("theme", "geral")
                        description = request.form.get("description", "")
                        db.execute(
                            """INSERT INTO social_media_library
                               (user_id, file_path, file_type, theme, description)
                               VALUES (?,?,?,?,?)""",
                            (user["id"], fpath, real_ct, theme, description)
                        )
                        db.commit()
                        msg = '<div class="alert alert-success">✅ Mídia adicionada!</div>'

    items = db.execute(
        "SELECT * FROM social_media_library WHERE user_id=? ORDER BY created_at DESC",
        (user["id"],)
    ).fetchall()

    items_html = ""
    if items:
        for m in items:
            img_url = f"/media/social/{m['id']}"
            items_html += f"""
            <div class="card" style="padding:12px">
                <img src="{img_url}" style="width:100%;height:160px;object-fit:cover;border-radius:8px;margin-bottom:8px" alt="">
                <p style="font-size:13px;color:var(--text);margin:0 0 4px 0;font-weight:500">{esc(m['theme'])}</p>
                <p style="font-size:11px;color:var(--text3);margin:0 0 8px 0">Usada {m['times_used']}x</p>
                {('<p style="font-size:11px;color:var(--text2);margin:0 0 8px 0">' + esc(m['description'][:80]) + '</p>') if m['description'] else ''}
                <form method="POST" action="/dashboard/social/library/{m['id']}/delete" style="margin:0">{csrf_field()}
                    <button type="submit" class="btn btn-sm" style="background:rgba(239,68,68,0.2);color:#ef4444;width:100%;font-size:12px">🗑️ Excluir</button>
                </form>
            </div>
            """
    else:
        items_html = '<p style="color:var(--text3);grid-column:1/-1;text-align:center;padding:40px">Nenhuma mídia cadastrada. Use o formulário acima para enviar.</p>'

    content = f"""<div class="container">
        <div class="page-header fade-in">
            <h1>🖼️ Biblioteca de Mídia</h1>
            <p>Cadastre fotos que a IA usará para gerar posts automaticamente</p>
        </div>
        {msg}

        <div class="card fade-in fade-in-1" style="margin-bottom:24px">
            <div class="card-header"><span class="card-title">📤 Adicionar nova mídia</span></div>
            <form method="POST" enctype="multipart/form-data">{csrf_field()}
                <div class="grid-2">
                    <div class="form-group">
                        <label class="form-label">Tema da imagem *</label>
                        <select name="theme" class="form-input" required>
                            <option value="produto">Produto / Serviço</option>
                            <option value="motivacional">Motivacional</option>
                            <option value="bastidores">Bastidores</option>
                            <option value="dica">Dica / Educacional</option>
                            <option value="depoimento">Depoimento / Cliente</option>
                            <option value="promocao">Promoção</option>
                            <option value="evento">Evento / Data comemorativa</option>
                            <option value="geral">Geral</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label class="form-label">Imagem (JPG/PNG, máx 10MB) *</label>
                        <input type="file" name="photo" accept="image/jpeg,image/png" required
                            style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.08);padding:10px;border-radius:8px;color:var(--text);width:100%">
                    </div>
                </div>
                <div class="form-group">
                    <label class="form-label">Descrição (opcional — ajuda a IA gerar legenda melhor)</label>
                    <textarea name="description" class="form-input" rows="2" placeholder="Ex: Mesa posta para jantar romântico, detalhes em dourado"></textarea>
                </div>
                <button type="submit" class="btn btn-primary">📤 Adicionar à biblioteca</button>
            </form>
        </div>

        <div class="card fade-in fade-in-2">
            <div class="card-header"><span class="card-title">📸 Mídias cadastradas ({len(items)})</span></div>
            <div class="grid-4" style="gap:16px">{items_html}</div>
        </div>
    </div>"""
    return base_html("Biblioteca", content, dict(user))


@app.route("/dashboard/social/library/<int:media_id>/delete", methods=["POST"])
@login_required
def social_library_delete(media_id):
    user = g.user
    db = get_db()
    item = db.execute(
        "SELECT * FROM social_media_library WHERE id=? AND user_id=?",
        (media_id, user["id"])
    ).fetchone()
    if item:
        try:
            if os.path.exists(item["file_path"]):
                os.remove(item["file_path"])
        except Exception:
            pass
        db.execute("DELETE FROM social_media_library WHERE id=?", (media_id,))
        db.commit()
    return redirect("/dashboard/social/library")


@app.route("/media/social/<int:media_id>")
@login_required
def serve_social_media(media_id):
    user = g.user
    db = get_db()
    m = db.execute(
        "SELECT * FROM social_media_library WHERE id=? AND user_id=?",
        (media_id, user["id"])
    ).fetchone()
    if not m or not os.path.exists(m["file_path"]):
        return "Não encontrado", 404
    return send_file(m["file_path"], mimetype=m["file_type"])


@app.route("/dashboard/social/generate", methods=["POST"])
@login_required
def social_generate():
    """Gera um post novo agora"""
    user = g.user
    post_id = create_social_post(user["id"])
    if post_id:
        return redirect(f"/dashboard/social?ok=1")
    return redirect("/dashboard/social?err=no_media")


@app.route("/dashboard/social/posts/<int:post_id>/approve", methods=["POST"])
@login_required
def social_approve_post(post_id):
    user = g.user
    db = get_db()
    post = db.execute(
        "SELECT * FROM scheduled_posts WHERE id=? AND user_id=?",
        (post_id, user["id"])
    ).fetchone()
    if post:
        db.execute(
            "UPDATE scheduled_posts SET status='approved', approved_at=datetime('now') WHERE id=?",
            (post_id,)
        )
        # Incrementa uso da mídia
        if post["media_id"]:
            db.execute(
                "UPDATE social_media_library SET times_used=times_used+1, last_used_at=datetime('now') WHERE id=?",
                (post["media_id"],)
            )
        db.commit()
        # Notifica via Telegram se configurado
        if user["telegram_bot_token"] and user["telegram_chat_id"]:
            send_telegram_message(
                user["telegram_bot_token"],
                user["telegram_chat_id"],
                f"✅ <b>Post #{post_id} aprovado!</b>\n\nCaption e imagem prontas para publicação."
            )
    return redirect("/dashboard/social")


@app.route("/dashboard/social/posts/<int:post_id>/reject", methods=["POST"])
@login_required
def social_reject_post(post_id):
    user = g.user
    db = get_db()
    db.execute(
        "UPDATE scheduled_posts SET status='rejected' WHERE id=? AND user_id=?",
        (post_id, user["id"])
    )
    db.commit()
    return redirect("/dashboard/social")


@app.route("/dashboard/social/settings", methods=["GET", "POST"])
@login_required
def social_settings():
    user = g.user
    db = get_db()
    msg = ""

    if request.method == "POST":
        # Dias da semana selecionados (1=Seg, 2=Ter, ..., 7=Dom)
        days_selected = request.form.getlist("post_days")
        days_csv = ",".join(sorted(days_selected)) if days_selected else ""

        # Múltiplos horários (pode ser "09:00,14:00,18:00")
        times_raw = request.form.get("social_post_times", "09:00").strip()
        # Valida formato HH:MM separados por vírgula
        import re as re_mod
        times_list = []
        for t in times_raw.split(","):
            t = t.strip()
            if re_mod.match(r'^\d{1,2}:\d{2}$', t):
                times_list.append(t)
        times_csv = ",".join(times_list) if times_list else "09:00"

        tg_token_input = request.form.get("telegram_bot_token", "").strip()

        # Se veio vazio, mantém token atual
        if tg_token_input:
            tg_token_final = _encrypt_value(tg_token_input)
        else:
            raw = db.execute("SELECT telegram_bot_token FROM users WHERE id=?", (user["id"],)).fetchone()
            tg_token_final = raw["telegram_bot_token"] or ""

        db.execute(
            """UPDATE users SET
                telegram_bot_token=?,
                telegram_chat_id=?,
                social_post_time=?,
                social_post_times=?,
                social_post_days=?,
                social_auto_enabled=?,
                social_post_tone=?,
                social_business_context=?
               WHERE id=?""",
            (
                tg_token_final,
                request.form.get("telegram_chat_id", "").strip(),
                times_list[0] if times_list else "09:00",
                times_csv,
                days_csv,
                1 if request.form.get("social_auto_enabled") else 0,
                request.form.get("social_post_tone", "profissional"),
                request.form.get("social_business_context", "").strip(),
                user["id"]
            )
        )
        db.commit()
        raw_user = db.execute("SELECT * FROM users WHERE id=?", (user["id"],)).fetchone()
        user = decrypt_user_row(raw_user)
        msg = '<div class="alert alert-success">✅ Configurações salvas!</div>'

    # Testa Telegram se configurado
    if request.args.get("test_telegram") == "1":
        if user["telegram_bot_token"] and user["telegram_chat_id"]:
            result = send_telegram_message(
                user["telegram_bot_token"],
                user["telegram_chat_id"],
                "🎉 <b>Telegram conectado!</b>\n\nSeu atendente.online está pronto para enviar posts para aprovação aqui."
            )
            if result:
                msg = '<div class="alert alert-success">✅ Mensagem de teste enviada com sucesso!</div>'
            else:
                msg = '<div class="alert alert-error">❌ Não consegui enviar. Verifique Token e Chat ID.</div>'

    # Monta os checkboxes dos dias da semana
    current_days = (user["social_post_days"] or "1,2,3,4,5").split(",")
    days_labels = [
        ("1", "Seg"), ("2", "Ter"), ("3", "Qua"), ("4", "Qui"),
        ("5", "Sex"), ("6", "Sáb"), ("7", "Dom")
    ]
    days_checkboxes = ""
    for day_num, day_label in days_labels:
        checked = "checked" if day_num in current_days else ""
        days_checkboxes += f"""
        <label style="display:inline-flex;align-items:center;gap:6px;background:rgba(255,255,255,0.04);padding:8px 14px;border-radius:8px;cursor:pointer;border:1px solid rgba(255,255,255,0.06);font-size:13px">
            <input type="checkbox" name="post_days" value="{day_num}" {checked} style="margin:0">
            {day_label}
        </label>
        """

    current_times = user["social_post_times"] or "09:00"

    # Calcula próximos agendamentos para mostrar preview
    next_schedule_html = ""
    if user["social_auto_enabled"] and current_days[0]:
        from datetime import datetime as dt_mod, timedelta as td_mod
        now = dt_mod.now()
        upcoming = []
        times_arr = [t.strip() for t in current_times.split(",") if t.strip()]
        day_names_pt = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]

        for i in range(14):  # próximos 14 dias
            check_date = now + td_mod(days=i)
            weekday = check_date.isoweekday()  # 1=Seg, 7=Dom
            if str(weekday) in current_days:
                for t in times_arr:
                    try:
                        h, m = t.split(":")
                        scheduled = check_date.replace(hour=int(h), minute=int(m), second=0, microsecond=0)
                        if scheduled > now:
                            upcoming.append((scheduled, day_names_pt[weekday-1]))
                    except:
                        pass
            if len(upcoming) >= 5:
                break
        upcoming = sorted(upcoming)[:5]

        if upcoming:
            items = ""
            for dt, day_name in upcoming:
                items += f'<li style="padding:6px 0;border-bottom:1px solid rgba(255,255,255,0.05);color:var(--text2);font-size:13px">📅 {day_name}, {dt.strftime("%d/%m")} às {dt.strftime("%H:%M")}</li>'
            next_schedule_html = f"""
            <div class="card" style="margin-top:16px;background:rgba(99,102,241,0.05);border:1px solid rgba(99,102,241,0.2)">
                <div class="card-header"><span class="card-title" style="color:var(--accent2)">📆 Próximos 5 agendamentos</span></div>
                <ul style="list-style:none;padding:0;margin:0">{items}</ul>
            </div>
            """

    content = f"""<div class="container">
        <div class="page-header fade-in">
            <h1>⚙️ Configurações da Agência</h1>
            <p>Configure Telegram, agenda de publicações e personalidade da IA</p>
        </div>
        {msg}

        <form method="POST">{csrf_field()}
            <div class="grid-2" style="margin-bottom:24px">
                <div class="card fade-in fade-in-1">
                    <div class="card-header"><span class="card-title">📱 Telegram Bot</span></div>
                    <p style="color:var(--text3);font-size:13px;margin-bottom:16px">
                        Posts serão enviados para o Telegram para sua aprovação.
                        <br><a href="https://telegram.me/BotFather" target="_blank" style="color:var(--accent2)">Como criar um bot no Telegram →</a>
                    </p>
                    <div class="form-group">
                        <label class="form-label">Bot Token 🔒</label>
                        <input type="password" name="telegram_bot_token" class="form-input" value="" placeholder="{esc(mask_secret(user['telegram_bot_token'] or '')) if user['telegram_bot_token'] else '123456:ABC-DEF...'}" autocomplete="off">
                        {f'<small style="color:var(--green2);font-size:11px">✓ Configurado. Deixe em branco para manter.</small>' if user['telegram_bot_token'] else ''}
                    </div>
                    <div class="form-group">
                        <label class="form-label">Chat ID (seu)</label>
                        <input type="text" name="telegram_chat_id" class="form-input" value="{esc(user['telegram_chat_id'] or '')}" placeholder="123456789">
                        <small style="color:var(--text3)">Acesse <a href="https://telegram.me/userinfobot" target="_blank" style="color:var(--accent2)">@userinfobot</a> para descobrir seu chat ID</small>
                    </div>
                </div>

                <div class="card fade-in fade-in-2">
                    <div class="card-header"><span class="card-title">🤖 Personalidade da IA</span></div>
                    <div class="form-group">
                        <label class="form-label">Tom dos posts</label>
                        <select name="social_post_tone" class="form-input">
                            <option value="profissional" {'selected' if user['social_post_tone']=='profissional' else ''}>Profissional</option>
                            <option value="descontraido" {'selected' if user['social_post_tone']=='descontraido' else ''}>Descontraído</option>
                            <option value="inspirador" {'selected' if user['social_post_tone']=='inspirador' else ''}>Inspirador</option>
                            <option value="humoristico" {'selected' if user['social_post_tone']=='humoristico' else ''}>Humorístico</option>
                            <option value="elegante" {'selected' if user['social_post_tone']=='elegante' else ''}>Elegante / Sofisticado</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label class="form-label">Contexto do negócio para a IA</label>
                        <textarea name="social_business_context" class="form-input" rows="6" placeholder="Ex: Somos uma pizzaria no Centro de Fortaleza...">{esc(user['social_business_context'] or '')}</textarea>
                        <small style="color:var(--text3)">A IA usará isso para criar legendas alinhadas com seu negócio</small>
                    </div>
                </div>
            </div>

            <div class="card fade-in fade-in-3" style="margin-bottom:24px">
                <div class="card-header"><span class="card-title">📆 Agenda de publicações</span></div>
                <p style="color:var(--text3);font-size:13px;margin-bottom:20px">
                    Configure quais dias e horários a IA deve gerar posts automaticamente para sua aprovação.
                </p>

                <div class="form-group">
                    <label class="form-label" style="margin-bottom:12px">Dias da semana</label>
                    <div style="display:flex;gap:8px;flex-wrap:wrap">
                        {days_checkboxes}
                    </div>
                    <small style="color:var(--text3);display:block;margin-top:8px">Selecione os dias em que quer gerar posts.</small>
                </div>

                <div class="form-group">
                    <label class="form-label">Horários (separados por vírgula, formato HH:MM)</label>
                    <input type="text" name="social_post_times" class="form-input"
                           value="{esc(current_times)}"
                           placeholder="09:00, 14:00, 18:00">
                    <small style="color:var(--text3);display:block;margin-top:4px">
                        Exemplo: <code style="background:rgba(255,255,255,0.05);padding:2px 6px;border-radius:4px">09:00, 14:00, 18:00</code>
                        (gera 3 posts por dia). Deixe só um horário para 1 post por dia.
                    </small>
                </div>

                <div class="form-group">
                    <label class="form-label" style="display:inline-flex;align-items:center;gap:10px;cursor:pointer;padding:12px;background:rgba(0,200,150,0.05);border:1px solid rgba(0,200,150,0.2);border-radius:8px;width:100%">
                        <input type="checkbox" name="social_auto_enabled" value="1" {'checked' if user['social_auto_enabled'] else ''} style="width:18px;height:18px">
                        <div>
                            <strong>Ativar geração automática de posts</strong>
                            <p style="margin:4px 0 0;font-size:12px;color:var(--text3)">Quando ativado, a IA gera posts nos dias e horários selecionados acima.</p>
                        </div>
                    </label>
                </div>

                {next_schedule_html}
            </div>

            <div style="display:flex;gap:12px;justify-content:flex-end">
                <a href="/dashboard/social/settings?test_telegram=1" class="btn" style="background:rgba(255,255,255,0.05)">🧪 Testar Telegram</a>
                <button type="submit" class="btn btn-primary">💾 Salvar configurações</button>
            </div>
        </form>
    </div>"""
    return base_html("Configurações Agência", content, dict(user))


def run_social_scheduler():
    """Executa o scheduler de posts automáticos (chamado a cada minuto).
    Usa UPDATE condicional como lock atômico — garante que apenas 1 worker dispara
    o post mesmo com múltiplas instâncias Gunicorn."""
    try:
        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row

        now = datetime.now()
        current_time = now.strftime("%H:%M")
        current_weekday = str(now.isoweekday())  # 1=Seg, 7=Dom
        current_minute = now.strftime("%Y-%m-%d %H:%M")

        # Busca usuários com auto ativado
        users = db_conn.execute(
            """SELECT * FROM users
               WHERE social_auto_enabled=1
               AND is_active=1
               AND social_post_days!=''
               AND social_post_times!=''"""
        ).fetchall()

        for user in users:
            # Verifica se é dia certo
            post_days = (user["social_post_days"] or "").split(",")
            if current_weekday not in post_days:
                continue

            # Verifica se é horário certo
            post_times = (user["social_post_times"] or "").split(",")
            post_times_clean = [t.strip() for t in post_times]
            if current_time not in post_times_clean:
                continue

            # Lock atômico: só dispara se conseguir atualizar (WHERE social_last_run != current_minute)
            # Se outro worker já atualizou, o UPDATE retorna rowcount=0 e pulamos
            cursor = db_conn.execute(
                """UPDATE users SET social_last_run=?
                   WHERE id=? AND (social_last_run IS NULL OR social_last_run != ?)""",
                (current_minute, user["id"], current_minute)
            )
            db_conn.commit()

            if cursor.rowcount == 0:
                # Outro worker já pegou essa execução
                continue

            safe_log(f"[SCHEDULER] Gerando post para user {user['id']} ({user['email']}) às {current_time}")

            # Gera o post (chama em thread separada para não bloquear)
            import threading
            threading.Thread(target=create_social_post, args=(user["id"],), daemon=True).start()

        db_conn.close()
    except Exception as e:
        safe_log(f"[SCHEDULER] Erro: {e}", level="ERROR")


def start_social_scheduler():
    """Inicia o scheduler em background (executa a cada minuto)"""
    import threading
    def loop():
        import time as t_mod
        while True:
            try:
                run_social_scheduler()
            except Exception as e:
                safe_log(f"[SCHEDULER LOOP] Erro: {e}", level="ERROR")
            # Espera até o próximo minuto
            now = datetime.now()
            sleep_seconds = 60 - now.second
            t_mod.sleep(sleep_seconds)

    scheduler_thread = threading.Thread(target=loop, daemon=True)
    scheduler_thread.start()
    safe_log("[SCHEDULER] Scheduler de posts iniciado (roda a cada minuto)")


# ═══════════════════════════════════════════════════════════════
#  COMÉRCIO — Pagamentos via PIX/Cartão direto no WhatsApp
# ═══════════════════════════════════════════════════════════════

def mp_create_pix_payment(access_token, amount, description, payer_phone, external_reference):
    """Cria cobrança PIX via Mercado Pago"""
    if not access_token:
        return None
    try:
        import requests as req, uuid as uuid_mod
        url = "https://api.mercadopago.com/v1/payments"
        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
            "X-Idempotency-Key": str(uuid_mod.uuid4())
        }
        expires_at = (datetime.now() + timedelta(hours=24)).strftime("%Y-%m-%dT%H:%M:%S.000-03:00")
        payload = {
            "transaction_amount": float(amount),
            "description": description[:150],
            "payment_method_id": "pix",
            "external_reference": external_reference,
            "date_of_expiration": expires_at,
            "payer": {
                "email": f"customer_{payer_phone[-9:] if len(payer_phone) >= 9 else payer_phone}@atendente.online",
                "first_name": "Cliente",
                "last_name": "WhatsApp"
            }
        }
        resp = req.post(url, headers=headers, json=payload, timeout=15)
        if resp.status_code == 201:
            data = resp.json()
            poi = data.get("point_of_interaction", {}).get("transaction_data", {})
            return {
                "id": str(data.get("id", "")),
                "qr_code": poi.get("qr_code", ""),
                "qr_code_base64": poi.get("qr_code_base64", ""),
                "ticket_url": poi.get("ticket_url", ""),
                "copy_paste": poi.get("qr_code", ""),
                "status": data.get("status", "pending")
            }
        safe_log(f"[MP PIX] Erro {resp.status_code}", level="ERROR")
        return None
    except Exception as e:
        safe_log(f"[MP PIX] Exceção: {e}", level="ERROR")
        return None


def mp_create_checkout_preference(access_token, items, payer_phone, external_reference, notification_url):
    """Cria link de checkout Mercado Pago com TODAS as opções habilitadas:
    PIX, Cartão de Crédito (até 12x), Cartão de Débito e Boleto.

    Para PIX aparecer corretamente, a conta MP precisa:
    1. Ter CPF/CNPJ verificado
    2. Ter PIX habilitado em 'Ferramentas > Meios de pagamento'
    3. Usar access token de produção (não test)
    """
    if not access_token:
        return None
    try:
        import requests as req, uuid as uuid_mod

        url = "https://api.mercadopago.com/checkout/preferences"
        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
            "X-Idempotency-Key": str(uuid_mod.uuid4())
        }

        # Gera email de placeholder do comprador (MP exige para PIX aparecer)
        phone_clean = ''.join(c for c in payer_phone if c.isdigit())
        payer_email = f"customer_{phone_clean[-9:] if len(phone_clean) >= 9 else phone_clean}@atendente.online"

        # Data de expiração: 24h
        expires_at = (datetime.now() + timedelta(hours=24)).strftime("%Y-%m-%dT%H:%M:%S.000-03:00")

        payload = {
            "items": items,
            "external_reference": external_reference,
            "notification_url": notification_url,
            "payer": {
                "email": payer_email,
                "name": "Cliente",
                "surname": "WhatsApp",
                "phone": {
                    "area_code": phone_clean[2:4] if len(phone_clean) >= 4 else "",
                    "number": phone_clean[4:] if len(phone_clean) >= 4 else phone_clean
                }
            },
            "back_urls": {
                "success": "https://atendente.online/payment/success",
                "pending": "https://atendente.online/payment/pending",
                "failure": "https://atendente.online/payment/failure"
            },
            "auto_return": "approved",
            "payment_methods": {
                # NÃO excluir nenhum método — permite PIX, Cartão, Débito, Boleto
                "excluded_payment_methods": [],
                "excluded_payment_types": [],
                "installments": 12,
                "default_installments": 1
            },
            "expires": True,
            "expiration_date_from": datetime.now().strftime("%Y-%m-%dT%H:%M:%S.000-03:00"),
            "expiration_date_to": expires_at,
            "statement_descriptor": "ATENDENTEONLINE",
            "binary_mode": False  # False = aceita pagamentos pendentes (PIX, Boleto); True = só cartão instantâneo
        }

        resp = req.post(url, headers=headers, json=payload, timeout=15)
        if resp.status_code in (200, 201):
            data = resp.json()
            return {
                "id": data.get("id", ""),
                "checkout_url": data.get("init_point", ""),
                "sandbox_url": data.get("sandbox_init_point", "")
            }

        # Erro — log detalhado
        error_detail = resp.text[:500]
        safe_log(f"[MP CHECKOUT] Erro {resp.status_code}", level="ERROR")

        # Retorna informação do erro para o admin poder ver
        return {
            "error": True,
            "status_code": resp.status_code,
            "detail": error_detail
        }
    except Exception as e:
        safe_log(f"[MP CHECKOUT] Exceção: {e}", level="ERROR")
        return {
            "error": True,
            "detail": str(e)
        }


def detect_purchase_intent(message, products):
    """Usa IA para detectar se mensagem do cliente é intenção de compra"""
    try:
        import requests as req, json as json_mod
        api_key = get_setting("ANTHROPIC_API_KEY", "")
        if not api_key or not products:
            return None

        def _get(p, key, default=None):
            try:
                if hasattr(p, 'get') and not hasattr(p, 'keys'):
                    return p.get(key, default)
                val = p[key]
                return val if val is not None else default
            except (KeyError, IndexError, TypeError):
                return default

        products_text_lines = []
        for p in products:
            active = _get(p, 'active', 1)
            price = _get(p, 'price', 0) or 0
            if not active or price <= 0:
                continue
            stock = _get(p, 'stock', -1)
            line = f"- ID:{p['id']} | {p['name']} | R$ {price:.2f}"
            if stock is not None and stock >= 0:
                line += f" (estoque: {stock})"
            products_text_lines.append(line)

        products_text = "\n".join(products_text_lines)

        if not products_text:
            return None

        prompt = f"""Você analisa mensagens de clientes no WhatsApp e identifica intenção de compra.

CATÁLOGO DISPONÍVEL:
{products_text}

MENSAGEM DO CLIENTE:
"{message}"

Analise e responda APENAS em JSON válido (sem ```json, sem explicação):
{{"is_purchase": true/false, "items": [{{"product_id": N, "quantity": N}}], "confidence": "high/medium/low"}}

REGRAS:
- is_purchase = true apenas se o cliente está claramente querendo comprar
- Se não menciona produto específico: is_purchase = false
- Se só está perguntando preço: is_purchase = false
- Se pede "1 pizza", "2 camisetas", etc: is_purchase = true
- quantidade padrão = 1 se não especificado

Responda apenas com o JSON:"""

        resp = req.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 300,
                "messages": [{"role": "user", "content": prompt}]
            },
            timeout=15
        )
        if resp.status_code != 200:
            return None

        text = resp.json()["content"][0]["text"].strip()
        # Remove markdown se tiver
        text = text.replace("```json", "").replace("```", "").strip()
        result = json_mod.loads(text)

        if not result.get("is_purchase"):
            return None

        return result
    except Exception as e:
        safe_log(f"[PURCHASE AI] Erro: {e}", level="ERROR")
        return None


def create_order_from_intent(user, conversation_id, customer_phone, purchase_data):
    """Cria pedido e gera cobrança PIX automaticamente.
    VALIDAÇÃO RIGOROSA:
    - Preço SEMPRE do banco (nunca do que a IA retornou)
    - Quantidade máx 100 por item
    - Valor total entre R$ 1 e R$ 10.000 por pedido
    - Verifica estoque se configurado
    - Rejeita pedidos sem itens válidos
    - Não cria pedidos duplicados (mesmo telefone + mesmos itens em < 2 min)
    """
    try:
        # Limites de segurança (evita abuso via prompt injection)
        MAX_QTY_PER_ITEM = 100
        MAX_ORDER_TOTAL = 10000.00
        MIN_ORDER_TOTAL = 1.00
        MAX_ITEMS_PER_ORDER = 20

        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row

        # Anti-duplicidade: checa se criou pedido similar recentemente
        recent_order = db_conn.execute(
            """SELECT id FROM orders
               WHERE user_id=? AND customer_phone=?
               AND datetime(created_at) > datetime('now', '-2 minutes')
               AND payment_status='pending'
               ORDER BY id DESC LIMIT 1""",
            (user["id"], customer_phone)
        ).fetchone()
        if recent_order:
            safe_log(f"[ORDER] Bloqueado pedido duplicado (order #{recent_order['id']} recente)")
            db_conn.close()
            return None

        items_detail = []
        total = 0
        items_json = []
        rejected_items = []

        # Valida input
        raw_items = purchase_data.get("items", [])
        if not raw_items or not isinstance(raw_items, list):
            safe_log("[ORDER] Rejeitado: sem items válidos", level="ERROR")
            db_conn.close()
            return None

        # Limite de itens por pedido
        if len(raw_items) > MAX_ITEMS_PER_ORDER:
            safe_log(f"[ORDER] Rejeitado: mais de {MAX_ITEMS_PER_ORDER} itens", level="ERROR")
            log_webhook_error("commerce_ai", user["id"], "TooManyItems",
                            f"Tentativa de pedido com {len(raw_items)} itens", purchase_data)
            db_conn.close()
            return None

        for item in raw_items:
            try:
                pid = int(item.get("product_id", 0))
            except (ValueError, TypeError):
                rejected_items.append(f"product_id inválido: {item}")
                continue

            if pid <= 0:
                continue

            # SEMPRE busca do banco — NUNCA confia no que a IA disser sobre preço
            product = db_conn.execute(
                "SELECT * FROM product_gallery WHERE id=? AND user_id=? AND active=1",
                (pid, user["id"])
            ).fetchone()
            if not product:
                rejected_items.append(f"produto {pid} não encontrado ou inativo")
                continue

            price_from_db = float(product["price"] or 0)
            if price_from_db <= 0:
                rejected_items.append(f"{product['name']}: sem preço configurado")
                continue

            # Valida quantidade
            try:
                qty = int(item.get("quantity", 1))
            except (ValueError, TypeError):
                qty = 1
            if qty < 1:
                qty = 1
            if qty > MAX_QTY_PER_ITEM:
                rejected_items.append(f"{product['name']}: quantidade {qty} excede limite {MAX_QTY_PER_ITEM}")
                qty = MAX_QTY_PER_ITEM
                log_webhook_error("commerce_ai", user["id"], "QuantityCapped",
                                f"Qty capped from {item.get('quantity')} to {MAX_QTY_PER_ITEM}", purchase_data)

            # Verifica estoque se configurado (stock >= 0)
            try:
                stock = int(product["stock"]) if product["stock"] is not None else -1
            except (ValueError, TypeError):
                stock = -1
            if stock >= 0 and qty > stock:
                rejected_items.append(f"{product['name']}: estoque insuficiente (disponível: {stock})")
                if stock == 0:
                    continue
                qty = stock  # Limita à quantidade disponível

            subtotal = price_from_db * qty
            total += subtotal
            items_detail.append({
                "name": product["name"],
                "quantity": qty,
                "price": price_from_db,
                "subtotal": subtotal
            })
            items_json.append({
                "id": str(product["id"]),
                "title": product["name"][:250],
                "quantity": qty,
                "unit_price": price_from_db,
                "currency_id": "BRL"
            })

        # Valida total final
        if total < MIN_ORDER_TOTAL:
            safe_log(f"[ORDER] Rejeitado: total R$ {total:.2f} abaixo do mínimo R$ {MIN_ORDER_TOTAL:.2f}", level="ERROR")
            db_conn.close()
            return None

        if total > MAX_ORDER_TOTAL:
            safe_log(f"[ORDER] Rejeitado: total R$ {total:.2f} acima do máximo R$ {MAX_ORDER_TOTAL:.2f}", level="ERROR")
            log_webhook_error("commerce_ai", user["id"], "TotalExceedsLimit",
                            f"Total R$ {total:.2f} > limite R$ {MAX_ORDER_TOTAL:.2f}", purchase_data)
            db_conn.close()
            return None

        if not items_detail:
            safe_log(f"[ORDER] Rejeitado: sem itens válidos. Rejeitados: {rejected_items}", level="ERROR")
            db_conn.close()
            return None

        # Cria registro do pedido
        import json as json_mod
        notes_text = "; ".join(rejected_items[:3]) if rejected_items else ""
        cur = db_conn.execute(
            """INSERT INTO orders
               (user_id, conversation_id, customer_phone, items, total, payment_status, expires_at, notes)
               VALUES (?,?,?,?,?,?,?,?)""",
            (user["id"], conversation_id, customer_phone,
             json_mod.dumps(items_detail, ensure_ascii=False),
             total, "pending",
             (datetime.now() + timedelta(hours=24)).strftime("%Y-%m-%d %H:%M:%S"),
             notes_text)
        )
        order_id = cur.lastrowid
        db_conn.commit()

        # Gera APENAS o link de checkout do Mercado Pago
        # Esse link já oferece PIX + Cartão Crédito + Cartão Débito + Boleto
        # O cliente escolhe direto na página do Mercado Pago
        items_names = ", ".join([f"{i['quantity']}x {i['name']}" for i in items_detail])

        checkout = mp_create_checkout_preference(
            user["mp_access_token"],
            items_json,
            customer_phone,
            f"order_{order_id}",
            f"https://atendente.online/webhook/mp-commerce/{user['id']}"
        )

        # Verifica se é resposta de erro
        is_valid = checkout and not checkout.get("error") and checkout.get("checkout_url")

        if is_valid:
            db_conn.execute(
                "UPDATE orders SET mp_payment_id=?, mp_checkout_url=? WHERE id=?",
                (checkout["id"], checkout["checkout_url"], order_id)
            )
            db_conn.commit()

        db_conn.close()

        return {
            "order_id": order_id,
            "total": total,
            "items": items_detail,
            "checkout_url": checkout["checkout_url"] if is_valid else None
        }
    except Exception as e:
        safe_log(f"[ORDER] Erro: {e}", level="ERROR")
        return None


def format_order_message(order_data):
    """Formata mensagem do pedido com link único de pagamento.
    O link do Mercado Pago oferece todas as opções: PIX, Cartão Crédito,
    Cartão Débito e Boleto. O cliente escolhe na página segura do MP."""
    items_text = "\n".join([f"• {i['quantity']}x {i['name']} — R$ {i['subtotal']:.2f}"
                            for i in order_data["items"]])

    msg = f"""🛒 *Seu pedido #{order_data['order_id']}*

{items_text}

💰 *Total: R$ {order_data['total']:.2f}*

━━━━━━━━━━━━━━━━━━━━
💳 *COMO PAGAR*

Clique no link abaixo para finalizar o pagamento:
"""
    if order_data.get("checkout_url"):
        msg += f"""
👉 {order_data['checkout_url']}

Na página segura do Mercado Pago, você pode escolher:
✅ PIX (aprovação imediata)
✅ Cartão de Crédito (até 12x)
✅ Cartão de Débito
✅ Boleto Bancário
"""
    else:
        msg += "\n⚠️ Link de pagamento indisponível no momento. Por favor, entre em contato."

    msg += """
━━━━━━━━━━━━━━━━━━━━
✅ Após o pagamento, confirmaremos automaticamente seu pedido!
⏰ Link válido por 24 horas."""
    return msg


# ─── ROTA: Webhook do Mercado Pago para confirmação de pagamento ───
@app.route("/webhook/mp-commerce/<int:user_id>", methods=["POST"])
def mp_commerce_webhook(user_id):
    """Recebe notificação de pagamento do Mercado Pago.

    Valida assinatura com o webhook secret do tenant (coluna users.mp_webhook_secret).
    Se o tenant não tiver secret próprio, usa o global (MP_WEBHOOK_SECRET). Isso permite
    migração gradual: cada cliente pode mover seu webhook para uma URL com secret próprio
    sem forçar mudança em massa.
    """
    try:
        # Determina qual secret usar — tenant primeiro, global como fallback
        tenant_secret = ""
        try:
            db_lookup = sqlite3.connect(DATABASE)
            db_lookup.row_factory = sqlite3.Row
            row = db_lookup.execute("SELECT mp_webhook_secret FROM users WHERE id=?", (user_id,)).fetchone()
            db_lookup.close()
            if row and row["mp_webhook_secret"]:
                tenant_secret = _decrypt_value(row["mp_webhook_secret"]) or ""
        except Exception as e:
            safe_log(f"[MP WEBHOOK] Erro buscando secret do tenant {user_id}: {e}", level="ERROR")

        # Valida assinatura MP — impede fraude de pagamento.
        # Se tenant_secret estiver vazio, validate_mp_signature cai no global automaticamente.
        if not validate_mp_signature(request, webhook_secret=tenant_secret or None):
            log_webhook_error("mercadopago", user_id, "InvalidSignature", "Assinatura MP inválida", None)
            return jsonify({"status": "invalid_signature"}), 401

        data = request.json or {}
        payment_id = data.get("data", {}).get("id") or data.get("id")

        if not payment_id:
            return jsonify({"status": "no_id"}), 200

        event_key = f"payment:{payment_id}"
        if not register_processed_webhook_event("mercadopago_commerce", event_key, user_id, {"id": payment_id}):
            return jsonify({"status": "duplicate"}), 200

        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row

        raw_user = db_conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
        if not raw_user or not raw_user["mp_access_token"]:
            db_conn.close()
            return jsonify({"status": "user_not_found"}), 404
        user = decrypt_user_row(raw_user)

        # Consulta status do pagamento no MP
        import requests as req
        resp = req.get(
            f"https://api.mercadopago.com/v1/payments/{payment_id}",
            headers={"Authorization": f"Bearer {user['mp_access_token']}"},
            timeout=10
        )
        if resp.status_code != 200:
            db_conn.close()
            return jsonify({"status": "mp_error"}), 200

        payment = resp.json()
        status = payment.get("status", "")
        external_ref = payment.get("external_reference", "")
        amount = float(payment.get("transaction_amount") or 0)
        currency = payment.get("currency_id", "")

        if not external_ref.startswith("order_"):
            db_conn.close()
            return jsonify({"status": "ok"}), 200

        order_id = int(external_ref.replace("order_", ""))
        order = db_conn.execute(
            "SELECT * FROM orders WHERE id=? AND user_id=?",
            (order_id, user_id)
        ).fetchone()

        if not order:
            db_conn.close()
            return jsonify({"status": "order_not_found"}), 200

        if currency != "BRL":
            db_conn.close()
            log_webhook_error("mercadopago", user_id, "InvalidCurrency", f"currency={currency}", {"payment_id": payment_id})
            return jsonify({"status": "invalid_currency"}), 200

        if abs(amount - float(order["total"] or 0)) > 0.01:
            db_conn.close()
            log_webhook_error("mercadopago", user_id, "InvalidAmount", f"amount={amount} expected={order['total']}", {"payment_id": payment_id})
            return jsonify({"status": "invalid_amount"}), 200

        if order["user_id"] != user_id:
            db_conn.close()
            return jsonify({"status": "user_mismatch"}), 200

        if status == "approved" and order["payment_status"] != "paid":
            db_conn.execute(
                "UPDATE orders SET payment_status='paid', paid_at=datetime('now'), mp_payment_id=? WHERE id=?",
                (str(payment_id), order_id)
            )
            db_conn.commit()

            # Envia confirmação via WhatsApp
            if user["whatsapp_phone_id"] and user["whatsapp_token"]:
                msg = f"""✅ *Pagamento confirmado!*

Pedido #{order_id} foi pago com sucesso.
Valor: R$ {order['total']:.2f}

Obrigado pela compra! 🎉
Em breve entraremos em contato para combinar a entrega/retirada."""
                send_whatsapp_message(
                    user["whatsapp_phone_id"],
                    user["whatsapp_token"],
                    order["customer_phone"],
                    msg
                )

            # Atualiza contato
            db_conn.execute(
                """UPDATE contacts SET
                    total_orders=total_orders+1,
                    total_spent=total_spent+?,
                    last_contact_at=datetime('now')
                   WHERE user_id=? AND phone=?""",
                (order["total"], user_id, order["customer_phone"])
            )
            db_conn.commit()

        db_conn.close()
        return jsonify({"status": "ok"}), 200
    except Exception as e:
        safe_log(f"[MP COMMERCE WEBHOOK] Erro: {e}", level="ERROR")
        return jsonify({"status": "error"}), 200


# ─── ROTAS: Dashboard de Comércio ──────────────────────────────
@app.route("/dashboard/commerce")
@login_required
def commerce_dashboard():
    """Dashboard de vendas pelo WhatsApp"""
    user = g.user
    db = get_db()

    # Estatísticas
    total_sales = db.execute(
        "SELECT COALESCE(SUM(total),0) as s FROM orders WHERE user_id=? AND payment_status='paid'",
        (user["id"],)
    ).fetchone()["s"]
    total_orders = db.execute(
        "SELECT COUNT(*) as c FROM orders WHERE user_id=? AND payment_status='paid'",
        (user["id"],)
    ).fetchone()["c"]
    pending_orders = db.execute(
        "SELECT COUNT(*) as c FROM orders WHERE user_id=? AND payment_status='pending'",
        (user["id"],)
    ).fetchone()["c"]

    # Vendas do mês
    month_sales = db.execute(
        """SELECT COALESCE(SUM(total),0) as s FROM orders
           WHERE user_id=? AND payment_status='paid'
           AND date(paid_at) >= date('now', 'start of month')""",
        (user["id"],)
    ).fetchone()["s"]

    avg_ticket = total_sales / total_orders if total_orders > 0 else 0

    # Últimos pedidos
    orders = db.execute(
        """SELECT o.* FROM orders o
           WHERE o.user_id=?
           ORDER BY o.created_at DESC LIMIT 20""",
        (user["id"],)
    ).fetchall()

    orders_html = ""
    if orders:
        import json as json_mod
        for o in orders:
            status_cls = {
                "paid": "badge-green",
                "pending": "badge-orange",
                "expired": "badge-red",
                "cancelled": "badge-red"
            }.get(o["payment_status"], "badge-orange")
            status_label = {
                "paid": "Pago",
                "pending": "Aguardando",
                "expired": "Expirado",
                "cancelled": "Cancelado"
            }.get(o["payment_status"], o["payment_status"])

            try:
                items = json_mod.loads(o["items"] or "[]")
                items_str = ", ".join([f"{i['quantity']}x {i['name'][:30]}" for i in items[:3]])
            except:
                items_str = "—"

            orders_html += f"""
            <tr>
                <td>#{o['id']}</td>
                <td>{esc(o['customer_phone'][:13] if o['customer_phone'] else '')}</td>
                <td style="font-size:12px">{esc(items_str)}</td>
                <td><strong>R$ {o['total']:.2f}</strong></td>
                <td><span class="badge {status_cls}">{status_label}</span></td>
                <td style="font-size:12px;color:var(--text3)">{to_br_datetime(o['created_at'])}</td>
            </tr>
            """
    else:
        orders_html = '<tr><td colspan="6" style="text-align:center;color:var(--text3);padding:40px">Nenhum pedido ainda. Configure seus produtos com preço e ative o comércio.</td></tr>'

    # Check MP configurado
    mp_configured = bool(user["mp_access_token"])
    commerce_active = bool(user["commerce_enabled"])

    alert_html = ""
    if not mp_configured:
        alert_html = '<div class="alert alert-error">⚠️ Configure seu Mercado Pago em <a href="/dashboard/commerce/settings" style="color:var(--accent2)">Configurações</a> para começar a vender.</div>'
    elif not commerce_active:
        alert_html = '<div class="alert alert-warning">⏸️ Comércio desativado. <a href="/dashboard/commerce/settings" style="color:var(--accent2)">Ative nas configurações</a> para que a IA gere cobranças automáticas.</div>'

    content = f"""<div class="container">
        <div class="page-header fade-in">
            <h1>🛒 Comércio no WhatsApp</h1>
            <p>Venda direto pelo WhatsApp com PIX e cartão</p>
        </div>
        {alert_html}

        <div class="grid-4 fade-in fade-in-1">
            <div class="metric-card"><div style="font-size:24px">💰</div><div class="metric-value">R$ {total_sales:.2f}</div><div class="metric-label">Total vendido</div></div>
            <div class="metric-card"><div style="font-size:24px">📅</div><div class="metric-value">R$ {month_sales:.2f}</div><div class="metric-label">Este mês</div></div>
            <div class="metric-card"><div style="font-size:24px">📦</div><div class="metric-value">{total_orders}</div><div class="metric-label">Pedidos pagos</div></div>
            <div class="metric-card"><div style="font-size:24px">⏳</div><div class="metric-value" style="color:var(--orange)">{pending_orders}</div><div class="metric-label">Aguardando pagamento</div></div>
        </div>

        <div class="grid-2 fade-in fade-in-2" style="margin-top:24px">
            <div class="card">
                <div class="card-header"><span class="card-title">🚀 Ações rápidas</span></div>
                <a href="/dashboard/gallery" class="btn btn-primary" style="width:100%;margin-bottom:8px;display:block;text-align:center;text-decoration:none">📦 Gerenciar produtos e preços</a>
                <a href="/dashboard/commerce/settings" class="btn" style="width:100%;background:rgba(255,255,255,0.05);margin-bottom:8px;display:block;text-align:center;text-decoration:none">⚙️ Configurar Mercado Pago</a>
                <a href="/dashboard/commerce/orders" class="btn" style="width:100%;background:rgba(255,255,255,0.05);display:block;text-align:center;text-decoration:none">📋 Ver todos os pedidos</a>
            </div>
            <div class="card">
                <div class="card-header"><span class="card-title">📊 Performance</span></div>
                <div style="padding:16px 0">
                    <div style="display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid rgba(255,255,255,0.05)">
                        <span style="color:var(--text2)">Ticket médio</span>
                        <strong>R$ {avg_ticket:.2f}</strong>
                    </div>
                    <div style="display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid rgba(255,255,255,0.05)">
                        <span style="color:var(--text2)">Taxa de conversão</span>
                        <strong>{(total_orders/(total_orders+pending_orders)*100 if (total_orders+pending_orders) > 0 else 0):.1f}%</strong>
                    </div>
                    <div style="display:flex;justify-content:space-between;padding:8px 0">
                        <span style="color:var(--text2)">Comércio ativo</span>
                        <strong style="color:{'var(--green2)' if commerce_active else 'var(--orange)'}">{'✅ Sim' if commerce_active else '⏸️ Não'}</strong>
                    </div>
                </div>
            </div>
        </div>

        <div class="card fade-in fade-in-3" style="margin-top:24px">
            <div class="card-header"><span class="card-title">📋 Pedidos recentes</span></div>
            <div class="table-wrap">
                <table>
                    <thead>
                        <tr><th>#</th><th>Cliente</th><th>Itens</th><th>Total</th><th>Status</th><th>Data</th></tr>
                    </thead>
                    <tbody>{orders_html}</tbody>
                </table>
            </div>
        </div>
    </div>"""
    return base_html("Comércio", content, dict(user))


@app.route("/dashboard/commerce/settings", methods=["GET", "POST"])
@login_required
def commerce_settings():
    """Configurações de comércio (Mercado Pago)"""
    user = g.user
    db = get_db()
    msg = ""

    if request.method == "POST":
        mp_token_input = request.form.get("mp_access_token", "").strip()
        mp_webhook_secret_input = request.form.get("mp_webhook_secret", "").strip()

        # Se veio vazio, mantém o atual
        if mp_token_input:
            mp_token_final = _encrypt_value(mp_token_input)
        else:
            raw = db.execute("SELECT mp_access_token FROM users WHERE id=?", (user["id"],)).fetchone()
            mp_token_final = raw["mp_access_token"] or ""

        if mp_webhook_secret_input:
            mp_webhook_secret_final = _encrypt_value(mp_webhook_secret_input)
        else:
            raw = db.execute("SELECT mp_webhook_secret FROM users WHERE id=?", (user["id"],)).fetchone()
            mp_webhook_secret_final = (raw["mp_webhook_secret"] if raw else "") or ""

        db.execute(
            """UPDATE users SET
                mp_access_token=?,
                mp_public_key=?,
                mp_webhook_secret=?,
                commerce_enabled=?,
                auto_payment_enabled=?
               WHERE id=?""",
            (
                mp_token_final,
                request.form.get("mp_public_key", "").strip(),
                mp_webhook_secret_final,
                1 if request.form.get("commerce_enabled") else 0,
                1 if request.form.get("auto_payment_enabled") else 0,
                user["id"]
            )
        )
        db.commit()
        # Re-carrega com descriptografia
        raw_user = db.execute("SELECT * FROM users WHERE id=?", (user["id"],)).fetchone()
        user = decrypt_user_row(raw_user)
        msg = '<div class="alert alert-success">✅ Configurações salvas!</div>'

    content = f"""<div class="container">
        <div class="page-header fade-in">
            <h1>⚙️ Configurações de Comércio</h1>
            <p>Conecte seu Mercado Pago para vender direto pelo WhatsApp</p>
        </div>
        {msg}

        <div class="card fade-in fade-in-1" style="margin-bottom:24px">
            <div class="card-header"><span class="card-title">💳 Credenciais Mercado Pago</span></div>
            <p style="color:var(--text3);font-size:13px;margin-bottom:20px">
                Obtenha suas credenciais em <a href="https://www.mercadopago.com.br/developers/panel/app" target="_blank" style="color:var(--accent2)">Mercado Pago Developers →</a>
                <br>Cadastre uma aplicação e copie o <strong>Access Token</strong> de produção.
            </p>
            <form method="POST">{csrf_field()}
                <div class="form-group">
                    <label class="form-label">Access Token (produção) 🔒</label>
                    <input type="password" name="mp_access_token" class="form-input"
                           value=""
                           placeholder="{esc(mask_secret(user['mp_access_token'] or '')) if user['mp_access_token'] else 'APP_USR-...'}"
                           autocomplete="off">
                    <small style="color:var(--text3)">{'✓ Token configurado. Deixe em branco para manter ou cole um novo para substituir.' if user['mp_access_token'] else 'Começa com APP_USR- (produção) ou TEST- (teste)'}</small>
                </div>
                <div class="form-group">
                    <label class="form-label">Public Key (opcional)</label>
                    <input type="text" name="mp_public_key" class="form-input"
                           value="{esc(user['mp_public_key'] or '')}"
                           placeholder="APP_USR-...">
                </div>

                <div class="form-group" style="background:rgba(99,102,241,0.04);border:1px solid rgba(99,102,241,0.18);border-radius:8px;padding:14px">
                    <label class="form-label">🔐 Webhook Secret (seu, por conta) <span style="font-size:11px;color:var(--text3);font-weight:400">— recomendado</span></label>
                    <input type="password" name="mp_webhook_secret" class="form-input"
                           value=""
                           placeholder="{'••••••••' if user.get('mp_webhook_secret') else 'Cole a chave secreta de assinatura do seu webhook MP'}"
                           autocomplete="off">
                    <small style="color:var(--text3);display:block;margin-top:6px;line-height:1.5">
                        {'✓ Secret configurado para sua conta. Deixe em branco para manter.' if user.get('mp_webhook_secret') else 'Se não configurar, é usado o secret global do sistema (menos seguro em ambiente multi-cliente).'}
                        <br>📍 <strong>URL do seu webhook:</strong>
                        <code style="display:inline-block;padding:2px 6px;background:var(--bg2);border-radius:4px;font-size:11px">{request.host_url.rstrip('/')}/webhook/mp-commerce/{user['id']}</code>
                        <br>Configure essa URL e a chave secreta em <a href="https://www.mercadopago.com.br/developers/panel/app" target="_blank" style="color:var(--accent2)">Mercado Pago Developers → Webhooks</a>.
                    </small>
                </div>

                <div class="form-group">
                    <label class="form-label" style="display:inline-flex;align-items:center;gap:10px;cursor:pointer;padding:14px;background:rgba(0,200,150,0.05);border:1px solid rgba(0,200,150,0.2);border-radius:8px;width:100%">
                        <input type="checkbox" name="commerce_enabled" value="1" {'checked' if user['commerce_enabled'] else ''} style="width:18px;height:18px">
                        <div>
                            <strong>Ativar comércio via WhatsApp</strong>
                            <p style="margin:4px 0 0;font-size:12px;color:var(--text3)">Quando o cliente pedir para comprar, a IA gera cobrança automática.</p>
                        </div>
                    </label>
                </div>

                <div class="form-group">
                    <label class="form-label" style="display:inline-flex;align-items:center;gap:10px;cursor:pointer;padding:14px;background:rgba(99,102,241,0.05);border:1px solid rgba(99,102,241,0.2);border-radius:8px;width:100%">
                        <input type="checkbox" name="auto_payment_enabled" value="1" {'checked' if user['auto_payment_enabled'] else ''} style="width:18px;height:18px">
                        <div>
                            <strong>Geração automática de PIX + link de cartão</strong>
                            <p style="margin:4px 0 0;font-size:12px;color:var(--text3)">A IA envia PIX e link de checkout juntos. Desative se quiser só PIX.</p>
                        </div>
                    </label>
                </div>

                <button type="submit" class="btn btn-primary">💾 Salvar</button>
            </form>
        </div>

        <div class="card fade-in fade-in-2">
            <div class="card-header"><span class="card-title">📖 Como funciona</span></div>
            <div style="color:var(--text2);font-size:14px;line-height:1.7">
                <p><strong>1.</strong> Cadastre produtos com preço na <a href="/dashboard/gallery" style="color:var(--accent2)">Galeria</a>.</p>
                <p><strong>2.</strong> Quando cliente disser "quero comprar X", a IA identifica automaticamente.</p>
                <p><strong>3.</strong> Sistema gera PIX + link de cartão via Mercado Pago.</p>
                <p><strong>4.</strong> Envia mensagem formatada com todas opções.</p>
                <p><strong>5.</strong> Quando cliente pagar, Mercado Pago avisa o sistema automaticamente.</p>
                <p><strong>6.</strong> Bot confirma pagamento no WhatsApp instantaneamente.</p>
            </div>
        </div>
    </div>"""
    return base_html("Comércio — Config", content, dict(user))


@app.route("/dashboard/commerce/orders")
@login_required
def commerce_orders():
    """Lista completa de pedidos"""
    user = g.user
    db = get_db()
    orders = db.execute(
        "SELECT * FROM orders WHERE user_id=? ORDER BY created_at DESC",
        (user["id"],)
    ).fetchall()

    rows = ""
    import json as json_mod
    for o in orders:
        status_cls = {"paid":"badge-green","pending":"badge-orange","expired":"badge-red"}.get(o["payment_status"],"badge-orange")
        status_label = {"paid":"Pago","pending":"Aguardando","expired":"Expirado"}.get(o["payment_status"],o["payment_status"])
        try:
            items = json_mod.loads(o["items"] or "[]")
            items_str = ", ".join([f"{i['quantity']}x {i['name'][:30]}" for i in items])
        except:
            items_str = "—"
        rows += f"""<tr>
            <td>#{o['id']}</td>
            <td>{esc(o['customer_phone'] or '')}</td>
            <td style="font-size:12px;max-width:300px">{esc(items_str)}</td>
            <td><strong>R$ {o['total']:.2f}</strong></td>
            <td><span class="badge {status_cls}">{status_label}</span></td>
            <td style="font-size:12px">{to_br_datetime(o['created_at'])}</td>
            <td style="font-size:12px">{to_br_datetime(o['paid_at']) if o['paid_at'] else '—'}</td>
        </tr>"""

    content = f"""<div class="container">
        <div class="page-header"><h1>Pedidos ({len(orders)})</h1></div>
        <div class="card"><div class="table-wrap"><table>
            <thead><tr><th>#</th><th>Cliente</th><th>Itens</th><th>Total</th><th>Status</th><th>Criado</th><th>Pago</th></tr></thead>
            <tbody>{rows or '<tr><td colspan="7" style="text-align:center;color:var(--text3);padding:40px">Nenhum pedido</td></tr>'}</tbody>
        </table></div></div>
    </div>"""
    return base_html("Pedidos", content, dict(user))


# ═══════════════════════════════════════════════════════════════
#  CAMPANHAS / BROADCAST — Envio em massa
# ═══════════════════════════════════════════════════════════════

def run_campaign(campaign_id):
    """Processa uma campanha enviando mensagens com rate limit"""
    try:
        import time as t_mod
        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row

        campaign = db_conn.execute("SELECT * FROM campaigns WHERE id=?", (campaign_id,)).fetchone()
        if not campaign:
            db_conn.close()
            return

        user = decrypt_user_row(db_conn.execute("SELECT * FROM users WHERE id=?", (campaign["user_id"],)).fetchone())
        if not user:
            db_conn.close()
            return

        # Atualiza status
        db_conn.execute(
            "UPDATE campaigns SET status='running', started_at=datetime('now') WHERE id=?",
            (campaign_id,)
        )
        db_conn.commit()

        contacts = db_conn.execute(
            "SELECT * FROM campaign_contacts WHERE campaign_id=? AND status='pending'",
            (campaign_id,)
        ).fetchall()

        sent = 0
        failed = 0

        for contact in contacts:
            allowed_window, window_error = validate_campaign_contact_window(db_conn, campaign, contact)
            if not allowed_window:
                db_conn.execute(
                    "UPDATE campaign_contacts SET status='failed', error=? WHERE id=?",
                    (window_error[:200], contact["id"])
                )
                failed += 1
                continue

            # Personaliza a mensagem com variáveis
            msg = campaign["message"]
            try:
                import json as json_mod
                variables = json_mod.loads(contact["variables"] or "{}")
                for k, v in variables.items():
                    msg = msg.replace(f"{{{k}}}", str(v))
                msg = msg.replace("{nome}", contact["name"] or "")
                msg = msg.replace("{telefone}", contact["phone"] or "")
            except:
                pass

            # Envia via WhatsApp
            try:
                result = send_whatsapp_message(
                    user["whatsapp_phone_id"],
                    user["whatsapp_token"],
                    contact["phone"],
                    msg
                )
                if result and result.get("success") is True:
                    db_conn.execute(
                        "UPDATE campaign_contacts SET status='sent', sent_at=datetime('now') WHERE id=?",
                        (contact["id"],)
                    )
                    sent += 1
                else:
                    db_conn.execute(
                        "UPDATE campaign_contacts SET status='failed', error=? WHERE id=?",
                        ((result or {}).get("error", "send_error")[:200], contact["id"])
                    )
                    failed += 1
            except Exception as e:
                db_conn.execute(
                    "UPDATE campaign_contacts SET status='failed', error=? WHERE id=?",
                    (str(e)[:200], contact["id"])
                )
                failed += 1

            # Atualiza contadores a cada 10 envios
            if (sent + failed) % 10 == 0:
                db_conn.execute(
                    "UPDATE campaigns SET sent_count=?, failed_count=? WHERE id=?",
                    (sent, failed, campaign_id)
                )
                db_conn.commit()

            # Rate limit: 1 mensagem a cada 2 segundos (30/min)
            t_mod.sleep(2)

        # Finaliza campanha
        db_conn.execute(
            "UPDATE campaigns SET sent_count=?, failed_count=?, status='completed', completed_at=datetime('now') WHERE id=?",
            (sent, failed, campaign_id)
        )
        db_conn.commit()
        db_conn.close()
        safe_log(f"[CAMPAIGN] #{campaign_id} completa: {sent} enviadas, {failed} falhas", level="ERROR")
    except Exception as e:
        safe_log(f"[CAMPAIGN] Erro: {e}", level="ERROR")


@app.route("/dashboard/campaigns")
@login_required
def campaigns_dashboard():
    """Dashboard de campanhas"""
    user = g.user
    db = get_db()

    campaigns = db.execute(
        "SELECT * FROM campaigns WHERE user_id=? ORDER BY created_at DESC LIMIT 30",
        (user["id"],)
    ).fetchall()

    total_sent = db.execute(
        "SELECT COALESCE(SUM(sent_count),0) as s FROM campaigns WHERE user_id=?",
        (user["id"],)
    ).fetchone()["s"]

    rows = ""
    for c in campaigns:
        status_cls = {
            "draft": "badge-gray",
            "scheduled": "badge-orange",
            "running": "badge-blue",
            "completed": "badge-green",
            "failed": "badge-red"
        }.get(c["status"], "badge-gray")
        status_label = {
            "draft": "Rascunho",
            "scheduled": "Agendada",
            "running": "Em execução",
            "completed": "Completa",
            "failed": "Falhou"
        }.get(c["status"], c["status"])

        progress = 0
        if c["total_contacts"] > 0:
            progress = int((c["sent_count"] / c["total_contacts"]) * 100)

        actions = ""
        if c["status"] == "draft":
            actions = f'<form method="POST" action="/dashboard/campaigns/{c["id"]}/start" style="display:inline">{csrf_field()}<button class="btn btn-success btn-sm">▶ Iniciar</button></form>'

        rows += f"""<tr>
            <td><strong>{esc(c['name'])}</strong></td>
            <td>{c['total_contacts']}</td>
            <td>{c['sent_count']}</td>
            <td>{c['failed_count']}</td>
            <td><div style="background:rgba(255,255,255,0.05);border-radius:6px;height:8px;overflow:hidden"><div style="background:var(--accent2);height:100%;width:{progress}%"></div></div><small style="color:var(--text3)">{progress}%</small></td>
            <td><span class="badge {status_cls}">{status_label}</span></td>
            <td style="font-size:12px">{to_br_datetime(c['created_at'])}</td>
            <td>{actions}</td>
        </tr>"""

    content = f"""<div class="container">
        <div class="page-header fade-in">
            <h1>📢 Campanhas</h1>
            <p>Envie mensagens em massa via WhatsApp (broadcast)</p>
        </div>

        <div class="grid-3 fade-in fade-in-1">
            <div class="metric-card"><div style="font-size:24px">📢</div><div class="metric-value">{len(campaigns)}</div><div class="metric-label">Campanhas</div></div>
            <div class="metric-card"><div style="font-size:24px">📤</div><div class="metric-value">{total_sent}</div><div class="metric-label">Mensagens enviadas</div></div>
            <div class="metric-card"><div style="font-size:24px">✉️</div><div class="metric-value">1/2s</div><div class="metric-label">Rate limit seguro</div></div>
        </div>

        <div style="margin:24px 0"><a href="/dashboard/campaigns/new" class="btn btn-primary">➕ Nova campanha</a></div>

        <div class="card fade-in fade-in-2">
            <div class="card-header"><span class="card-title">Todas as campanhas</span></div>
            <div class="table-wrap">
                <table>
                    <thead><tr><th>Nome</th><th>Contatos</th><th>Enviadas</th><th>Falhas</th><th>Progresso</th><th>Status</th><th>Criada</th><th>Ação</th></tr></thead>
                    <tbody>{rows or '<tr><td colspan="8" style="text-align:center;color:var(--text3);padding:40px">Nenhuma campanha criada</td></tr>'}</tbody>
                </table>
            </div>
        </div>

        <div class="alert alert-warning" style="margin-top:24px">
            ⚠️ <strong>Importante:</strong> Respeite as regras da Meta. Mensagens em massa para contatos sem interação em 24h precisam de <strong>template aprovado</strong>.
            Excesso de bloqueios pode pausar sua conta.
        </div>
    </div>"""
    return base_html("Campanhas", content, dict(user))


@app.route("/dashboard/campaigns/new", methods=["GET", "POST"])
@login_required
def campaign_new():
    """Criar nova campanha"""
    user = g.user
    db = get_db()
    msg = ""

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        message = request.form.get("message", "").strip()
        contacts_raw = request.form.get("contacts", "").strip()

        if not name or not message or not contacts_raw:
            msg = '<div class="alert alert-error">Preencha todos os campos</div>'
        else:
            # Processa lista de contatos (uma por linha: telefone,nome)
            contacts_list = []
            for line in contacts_raw.split("\n"):
                line = line.strip()
                if not line:
                    continue
                parts = [p.strip() for p in line.split(",")]
                phone = parts[0]
                name_c = parts[1] if len(parts) > 1 else ""
                # Limpa telefone (só dígitos)
                import re as re_mod
                phone = re_mod.sub(r'\D', '', phone)
                if len(phone) >= 10:
                    # Adiciona 55 se não tiver
                    if not phone.startswith("55"):
                        phone = "55" + phone
                    contacts_list.append({"phone": phone, "name": name_c})

            if not contacts_list:
                msg = '<div class="alert alert-error">Nenhum contato válido encontrado. Use formato: telefone,nome (um por linha)</div>'
            else:
                cur = db.execute(
                    """INSERT INTO campaigns (user_id, name, message, total_contacts, status)
                       VALUES (?,?,?,?,?)""",
                    (user["id"], name, message, len(contacts_list), "draft")
                )
                cid = cur.lastrowid
                for c in contacts_list:
                    db.execute(
                        "INSERT INTO campaign_contacts (campaign_id, phone, name, status) VALUES (?,?,?,?)",
                        (cid, c["phone"], c["name"], "pending")
                    )
                db.commit()
                return redirect(f"/dashboard/campaigns?created={cid}")

    content = f"""<div class="container">
        <div class="page-header"><h1>➕ Nova campanha</h1></div>
        {msg}
        <form method="POST" class="card">{csrf_field()}
            <div class="form-group">
                <label class="form-label">Nome da campanha *</label>
                <input type="text" name="name" class="form-input" required placeholder="Ex: Black Friday — Oferta Especial">
            </div>
            <div class="form-group">
                <label class="form-label">Mensagem *</label>
                <textarea name="message" class="form-input" rows="5" required placeholder="Olá {{nome}}! Temos uma oferta especial para você..."></textarea>
                <small style="color:var(--text3)">Use <code>{{nome}}</code> para personalizar com o nome do contato.</small>
            </div>
            <div class="form-group">
                <label class="form-label">Lista de contatos *</label>
                <textarea name="contacts" class="form-input" rows="10" required
                    placeholder="5585999999999,João&#10;5585988888888,Maria&#10;..."></textarea>
                <small style="color:var(--text3)">Um contato por linha. Formato: <code>telefone,nome</code> (nome opcional).</small>
            </div>
            <button type="submit" class="btn btn-primary">Criar campanha (rascunho)</button>
            <a href="/dashboard/campaigns" class="btn" style="background:rgba(255,255,255,0.05);margin-left:8px">Cancelar</a>
        </form>
    </div>"""
    return base_html("Nova campanha", content, dict(user))


@app.route("/dashboard/campaigns/<int:cid>/start", methods=["POST"])
@login_required
def campaign_start(cid):
    """Inicia envio da campanha em background"""
    user = g.user
    db = get_db()
    campaign = db.execute(
        "SELECT * FROM campaigns WHERE id=? AND user_id=?",
        (cid, user["id"])
    ).fetchone()
    if campaign and campaign["status"] == "draft":
        pending_contacts = db.execute(
            "SELECT * FROM campaign_contacts WHERE campaign_id=? AND status='pending' LIMIT 2000",
            (cid,)
        ).fetchall()
        for contact in pending_contacts:
            allowed_window, window_error = validate_campaign_contact_window(db, campaign, contact)
            if not allowed_window:
                db.execute(
                    "UPDATE campaigns SET status='blocked', completed_at=datetime('now') WHERE id=?",
                    (cid,)
                )
                db.commit()
                return redirect(f"/dashboard/campaigns?error={window_error}")
        import threading
        threading.Thread(target=run_campaign, args=(cid,), daemon=True).start()
    return redirect("/dashboard/campaigns")


# ═══════════════════════════════════════════════════════════════
#  CONTATOS / CRM — Gestão de contatos e segmentação
# ═══════════════════════════════════════════════════════════════

@app.route("/dashboard/contacts")
@login_required
def contacts_dashboard():
    """Lista de contatos"""
    user = g.user
    db = get_db()

    # Sincroniza contatos das conversas
    db.execute(
        """INSERT OR IGNORE INTO contacts (user_id, phone, name, created_at, last_contact_at)
           SELECT DISTINCT user_id, customer_phone, customer_name, datetime('now'), last_message_at
           FROM conversations WHERE user_id=? AND customer_phone NOT IN
           (SELECT phone FROM contacts WHERE user_id=?)""",
        (user["id"], user["id"])
    )
    db.commit()

    contacts = db.execute(
        "SELECT * FROM contacts WHERE user_id=? ORDER BY last_contact_at DESC LIMIT 200",
        (user["id"],)
    ).fetchall()

    rows = ""
    for c in contacts:
        stage_cls = {"lead":"badge-gray","customer":"badge-green","vip":"badge-purple","inactive":"badge-red"}.get(c["lifecycle_stage"],"badge-gray")
        rows += f"""<tr>
            <td><strong>{esc(c['name'] or 'Sem nome')}</strong></td>
            <td>{esc(c['phone'])}</td>
            <td><span class="badge {stage_cls}">{esc(c['lifecycle_stage'])}</span></td>
            <td style="font-size:12px">{esc(c['tags'][:50] if c['tags'] else '—')}</td>
            <td>{c['total_orders']}</td>
            <td><strong>R$ {c['total_spent']:.2f}</strong></td>
            <td style="font-size:12px">{to_br_datetime(c['last_contact_at']) if c['last_contact_at'] else '—'}</td>
        </tr>"""

    content = f"""<div class="container">
        <div class="page-header"><h1>👥 Contatos ({len(contacts)})</h1><p>Todos os contatos que conversaram com você</p></div>
        <div class="card"><div class="table-wrap"><table>
            <thead><tr><th>Nome</th><th>Telefone</th><th>Estágio</th><th>Tags</th><th>Pedidos</th><th>Gasto total</th><th>Último contato</th></tr></thead>
            <tbody>{rows or '<tr><td colspan="7" style="text-align:center;color:var(--text3);padding:40px">Nenhum contato</td></tr>'}</tbody>
        </table></div></div>
    </div>"""
    return base_html("Contatos", content, dict(user))


# ═══════════════════════════════════════════════════════════════
#  PIPELINE / FUNIL — Kanban de oportunidades
# ═══════════════════════════════════════════════════════════════

@app.route("/dashboard/pipeline")
@login_required
def pipeline_dashboard():
    """Kanban de funil de vendas"""
    user = g.user
    db = get_db()

    # Cria estágios padrão se não existir nenhum
    stages = db.execute("SELECT * FROM pipeline_stages WHERE user_id=? ORDER BY position", (user["id"],)).fetchall()
    if not stages:
        default_stages = [
            ("Novo lead", "#6366f1", 0),
            ("Em contato", "#f59e0b", 1),
            ("Negociando", "#3b82f6", 2),
            ("Fechado - Ganho", "#10b981", 3),
            ("Fechado - Perdido", "#ef4444", 4)
        ]
        for name, color, pos in default_stages:
            db.execute(
                "INSERT INTO pipeline_stages (user_id, name, color, position) VALUES (?,?,?,?)",
                (user["id"], name, color, pos)
            )
        db.commit()
        stages = db.execute("SELECT * FROM pipeline_stages WHERE user_id=? ORDER BY position", (user["id"],)).fetchall()

    # Monta colunas
    columns_html = ""
    total_value = 0
    for stage in stages:
        cards = db.execute(
            "SELECT * FROM pipeline_cards WHERE user_id=? AND stage_id=? ORDER BY position",
            (user["id"], stage["id"])
        ).fetchall()

        stage_value = sum(c["value"] for c in cards)
        total_value += stage_value if "Ganho" in stage["name"] else 0

        cards_html = ""
        for card in cards:
            cards_html += f"""
            <div class="card" style="padding:14px;margin-bottom:10px;background:rgba(255,255,255,0.03);border-left:3px solid {stage['color']};cursor:pointer">
                <strong style="font-size:14px;display:block;margin-bottom:4px">{esc(card['title'])}</strong>
                {f'<div style="color:var(--green2);font-size:13px;margin-bottom:4px">💰 R$ {card["value"]:.2f}</div>' if card['value'] > 0 else ''}
                {f'<div style="color:var(--text3);font-size:11px">{esc(card["notes"][:60])}</div>' if card['notes'] else ''}
            </div>
            """

        columns_html += f"""
        <div style="flex:1;min-width:250px;background:rgba(255,255,255,0.02);border-radius:12px;padding:16px;border-top:3px solid {stage['color']}">
            <div style="display:flex;justify-content:space-between;margin-bottom:12px">
                <strong>{esc(stage['name'])}</strong>
                <span style="background:rgba(255,255,255,0.05);padding:2px 8px;border-radius:10px;font-size:11px;color:var(--text3)">{len(cards)}</span>
            </div>
            {f'<div style="color:var(--text3);font-size:11px;margin-bottom:12px">R$ {stage_value:.2f} em oportunidades</div>' if stage_value > 0 else ''}
            <div>{cards_html or '<p style="color:var(--text3);font-size:12px;text-align:center;padding:20px">Arraste cards para cá</p>'}</div>
        </div>
        """

    content = f"""<div class="container">
        <div class="page-header"><h1>🎯 Funil de vendas</h1><p>Acompanhe suas oportunidades em tempo real — R$ {total_value:.2f} ganhos</p></div>
        <div style="display:flex;gap:16px;overflow-x:auto;padding-bottom:20px">
            {columns_html}
        </div>
        <div class="alert alert-info" style="margin-top:20px">
            💡 <strong>Em breve:</strong> Arrastar-e-soltar para mover cards entre etapas. Por ora, os cards são criados automaticamente quando IA identifica uma oportunidade.
        </div>
    </div>"""
    return base_html("Pipeline", content, dict(user))


def fetch_weather(message):
    """Busca previsão do tempo usando Open-Meteo (grátis, sem API key, confiável)"""
    try:
        import requests as req
        import re as re_mod

        msg_lower = message.lower()
        # Remove prefixo [MENSAGEM DE ÁUDIO DO CLIENTE]: se presente
        msg_lower = re_mod.sub(r'^\[.*?\]:\s*', '', msg_lower)
        # Remove pontuação
        msg_clean = re_mod.sub(r'[?!.,;:]', ' ', msg_lower)

        # Estratégia 1: pega o que vem DEPOIS de "em", "para", "de", "no", "na"
        # Procura nomes de cidade evitando palavras-chave de clima
        patterns = [
            r'(?:em|para|de|do|na|no)\s+(?!tempo|clima|previs|chuva|temp|calor|frio)([a-záàâãéèêíïóôõöúçñ\s]+?)(?:\s+(?:hoje|amanhã|agora|$))',
            r'(?:em|para|de|do|na|no)\s+(?!tempo|clima|previs|chuva|temp|calor|frio)([a-záàâãéèêíïóôõöúçñ\s]+?)$',
        ]
        city = ""
        for pat in patterns:
            match = re_mod.search(pat, msg_clean)
            if match:
                city = match.group(1).strip()
                # Remove palavras sobrando
                stop_words = {"hoje", "amanhã", "amanha", "agora", "aqui", "lá", "la", "aí", "ai"}
                city_words = [w for w in city.split() if w not in stop_words and len(w) > 2]
                city = " ".join(city_words).strip()
                if city:
                    break

        # Estratégia 2: pega a última palavra significativa (maiúscula no original = nome próprio)
        if not city:
            words_original = message.split()
            # Procura palavras que começam com maiúscula (sem contar início de frase)
            proper_nouns = []
            for i, w in enumerate(words_original):
                w_clean = re_mod.sub(r'[?!.,;:]', '', w)
                if w_clean and w_clean[0].isupper() and i > 0:
                    proper_nouns.append(w_clean.lower())
            if proper_nouns:
                city = proper_nouns[-1]  # última palavra própria

        # Estratégia 3: fallback com remoção de palavras comuns
        if not city:
            remove_words = {"como", "está", "esta", "qual", "tempo", "clima", "temperatura", "previsão",
                           "previsao", "chuva", "chover", "vai", "hoje", "amanhã", "amanha",
                           "em", "de", "do", "da", "no", "na", "para", "o", "a", "é", "e",
                           "frio", "calor", "quente", "agora", "aqui", "lá", "la", "tá", "ta",
                           "mensagem", "áudio", "audio", "cliente", "faz", "fazer",
                           "me", "diga", "fala", "saber", "dizer", "ver", "posso",
                           "qualé", "quale", "que", "tal", "eu", "você", "voce",
                           "queria", "quero", "gostaria", "amigo", "amiga", "por", "favor",
                           "obrigado", "obrigada", "oi", "olá"}
            words = msg_clean.split()
            city_words = [w for w in words if w not in remove_words and len(w) > 2]
            # Pega apenas a última palavra (geralmente é o nome da cidade)
            if city_words:
                city = city_words[-1]

        if not city:
            city = "Fortaleza"

        safe_log(f"[WEATHER] Cidade extraída: '{city}'")

        # 1. Geocodificação: converte nome → coordenadas
        geo_url = f"https://geocoding-api.open-meteo.com/v1/search?name={city}&count=1&language=pt&format=json"
        geo_resp = req.get(geo_url, timeout=10)
        if geo_resp.status_code != 200:
            safe_log(f"[WEATHER] Erro geocoding: {geo_resp.status_code}", level="ERROR")
            return ""

        geo_data = geo_resp.json()
        results = geo_data.get("results", [])
        if not results:
            safe_log(f"[WEATHER] Cidade não encontrada: {city}")
            return ""

        loc = results[0]
        lat = loc.get("latitude")
        lon = loc.get("longitude")
        city_name = loc.get("name", city)
        country = loc.get("country", "")
        admin1 = loc.get("admin1", "")

        weather_url = (
            f"https://api.open-meteo.com/v1/forecast"
            f"?latitude={lat}&longitude={lon}"
            f"&current=temperature_2m,apparent_temperature,relative_humidity_2m,wind_speed_10m,weather_code,is_day"
            f"&daily=temperature_2m_max,temperature_2m_min,precipitation_probability_max,weather_code"
            f"&timezone=America/Sao_Paulo&forecast_days=1"
        )
        weather_resp = req.get(weather_url, timeout=10)
        if weather_resp.status_code != 200:
            safe_log(f"[WEATHER] Erro clima: {weather_resp.status_code}", level="ERROR")
            return ""

        data = weather_resp.json()
        current = data.get("current", {})
        daily = data.get("daily", {})

        temp = current.get("temperature_2m", "?")
        feels = current.get("apparent_temperature", "?")
        humidity = current.get("relative_humidity_2m", "?")
        wind = current.get("wind_speed_10m", "?")
        wcode = current.get("weather_code", 0)
        is_day = current.get("is_day", 1)

        max_temp = daily.get("temperature_2m_max", [None])[0]
        min_temp = daily.get("temperature_2m_min", [None])[0]
        rain_prob = daily.get("precipitation_probability_max", [None])[0]

        desc = _weather_code_to_pt(wcode, is_day)

        location = f"{city_name}"
        if admin1 and admin1 != city_name:
            location += f", {admin1}"
        if country:
            location += f", {country}"

        parts = [
            f"Cidade: {location}",
            f"Condição: {desc}",
            f"Temperatura: {temp}°C (sensação {feels}°C)",
        ]
        if max_temp is not None and min_temp is not None:
            parts.append(f"Mínima: {min_temp}°C, Máxima: {max_temp}°C")
        if rain_prob is not None:
            parts.append(f"Chance de chuva: {rain_prob}%")
        parts.append(f"Umidade: {humidity}%")
        parts.append(f"Vento: {wind} km/h")

        weather_text = ". ".join(parts) + "."
        safe_log(f"[WEATHER] ✅ {weather_text}")
        return weather_text

    except Exception as e:
        safe_log(f"[WEATHER] Exceção: {e}", level="ERROR")
        return ""


def _weather_code_to_pt(code, is_day=1):
    """Converte código WMO (Open-Meteo) para descrição em português"""
    codes = {
        0: "céu limpo" if is_day else "noite limpa",
        1: "predominantemente ensolarado" if is_day else "predominantemente limpo",
        2: "parcialmente nublado",
        3: "nublado",
        45: "nevoeiro",
        48: "nevoeiro com geada",
        51: "garoa leve",
        53: "garoa moderada",
        55: "garoa intensa",
        61: "chuva leve",
        63: "chuva moderada",
        65: "chuva forte",
        71: "neve leve",
        73: "neve moderada",
        75: "neve forte",
        77: "granizo",
        80: "pancadas de chuva leves",
        81: "pancadas de chuva moderadas",
        82: "pancadas de chuva fortes",
        95: "tempestade com raios",
        96: "tempestade com granizo leve",
        99: "tempestade com granizo forte",
    }
    return codes.get(code, "condição desconhecida")


def generate_ai_response(user, conversation_id, message, db_conn):
    history = list(reversed(db_conn.execute("SELECT sender,content FROM messages WHERE conversation_id=? ORDER BY created_at DESC LIMIT 10", (conversation_id,)).fetchall()))
    kb_items = db_conn.execute("SELECT title,content FROM knowledge_base WHERE user_id=? LIMIT 20", (user["id"],)).fetchall()
    qr_items = db_conn.execute("SELECT shortcut,content FROM quick_replies WHERE user_id=? LIMIT 20", (user["id"],)).fetchall()
    gallery_items = db_conn.execute("SELECT name,keywords,description FROM product_gallery WHERE user_id=?", (user["id"],)).fetchall()

    kb_context = "\n".join([f"- {i['title']}: {i['content']}" for i in kb_items])
    qr_context = "\n".join([f"- /{q['shortcut']}: {q['content']}" for q in qr_items])
    gallery_context = "\n".join([f"- {g['name']}: {g['description'] or 'sem descrição'}" for g in gallery_items])

    tone_map = {"profissional":"Profissional mas acessível.","descontraido":"Descontraído, com emojis moderados.","formal":"Formal e respeitoso.","amigavel":"Amigável e caloroso."}

    # Verifica se a mensagem vai disparar envio automático de foto
    matched_product = find_matching_product(user["id"], message)
    gallery_note = ""
    if matched_product:
        gallery_note = f"\n\n⚠️ IMPORTANTE: O sistema JÁ VAI ENVIAR automaticamente a foto de '{matched_product['name']}' para o cliente junto com sua resposta. NÃO diga que não pode enviar foto. NÃO mencione que vai enviar a foto (ela já está sendo enviada). Apenas responda normalmente descrevendo o produto ou confirmando o pedido."

    # Info sobre comércio se estiver ativo
    commerce_note = ""
    if user["commerce_enabled"] and user["mp_access_token"]:
        products_with_price = db_conn.execute(
            "SELECT name, price FROM product_gallery WHERE user_id=? AND active=1 AND price > 0",
            (user["id"],)
        ).fetchall()
        if products_with_price:
            prices_text = "\n".join([f"- {p['name']}: R$ {p['price']:.2f}" for p in products_with_price])
            commerce_note = f"""

🛒 SISTEMA DE PAGAMENTO ATIVO:
Este sistema ACEITA pagamentos automaticamente! Quando o cliente pedir para comprar, o sistema gera:
- PIX instantâneo (copia e cola)
- Link de cartão de crédito/débito (Mercado Pago)
- Link de boleto

PRODUTOS COM PREÇO (aceitam compra direto no WhatsApp):
{prices_text}

REGRAS DE VENDA:
- NUNCA diga "não aceitamos pagamento pelo WhatsApp" — aceita sim!
- Quando o cliente disser "quero comprar X" ou pedir o produto, o sistema gera PIX e link automaticamente.
- Você apenas confirma o pedido naturalmente e explica que vai gerar os dados de pagamento.
- Exemplo: "Ótimo! Vou gerar seu PIX agora mesmo, só um instante..."
- Se o cliente perguntar se aceita cartão/PIX: SIM, aceitamos ambos."""

    system_prompt = f"""{user['ai_system_prompt']}

Tom: {tone_map.get(user['ai_tone'],'Profissional.')}

INFORMAÇÕES DO NEGÓCIO:
{kb_context or 'Nenhuma info cadastrada.'}

RESPOSTAS RÁPIDAS DISPONÍVEIS:
{qr_context or 'Nenhuma.'}

FOTOS DE PRODUTOS DISPONÍVEIS (enviadas automaticamente quando o cliente pedir):
{gallery_context or 'Nenhuma foto cadastrada.'}{commerce_note}

REGRAS:
- Responda de forma breve (máx 3 parágrafos curtos)
- Não invente informações sobre produtos ou preços
- Se não souber, diga que vai verificar
- Horário: {user['business_hours']}
- IMPORTANTE: O sistema CONSEGUE enviar fotos de produtos automaticamente através da galeria cadastrada acima. Se o cliente pedir foto de um produto que está na lista, NÃO diga que não pode enviar foto — o sistema envia automaticamente. Apenas descreva o produto ou fale sobre ele.
- Se o cliente enviar áudio, você receberá a transcrição marcada como [MENSAGEM DE ÁUDIO DO CLIENTE]. Responda de forma conversacional e natural, como se estivesse falando (a resposta será convertida em áudio). NÃO use formatação como asteriscos, bullets, listas numeradas ou markdown. Escreva em frases corridas e naturais. Seja breve, no máximo 3 frases.
- Se o cliente enviar imagem, você receberá a descrição da imagem
- Se o cliente enviar PDF, você receberá o texto extraído
- Se o cliente perguntar sobre tempo/clima, você receberá dados reais marcados como [DADOS DO CLIMA ATUALIZADOS]. Use esses dados para responder com precisão.{gallery_note}
"""

    api_messages = [{"role":"assistant" if h["sender"]=="bot" else "user","content":h["content"]} for h in history]

    # Detecta perguntas sobre clima/tempo e busca dados reais
    weather_data = ""
    weather_keywords = ["tempo", "clima", "temperatura", "previsão", "chuva", "chover", "frio", "calor", "quente", "ensolarado", "nublado"]
    msg_lower = message.lower()
    if any(kw in msg_lower for kw in weather_keywords):
        weather_data = fetch_weather(message)
        if weather_data:
            message = f"{message}\n\n[DADOS DO CLIMA ATUALIZADOS]: {weather_data}"

    api_messages.append({"role":"user","content":message})

    ai_engine = get_setting("AI_ENGINE", "claude")  # claude ou openai

    # Tenta Claude primeiro (se configurado e selecionado)
    api_key = get_setting("ANTHROPIC_API_KEY")
    if api_key and ai_engine == "claude":
        try:
            import requests as req
            resp = req.post("https://api.anthropic.com/v1/messages",
                headers={"x-api-key":api_key,"anthropic-version":"2023-06-01","content-type":"application/json"},
                json={"model":"claude-sonnet-4-6","max_tokens":500,"system":system_prompt,"messages":api_messages}, timeout=30)
            if resp.status_code == 200:
                result = resp.json()
                tokens_in = result.get("usage",{}).get("input_tokens",0)
                tokens_out = result.get("usage",{}).get("output_tokens",0)
                cost = (tokens_in * 3 / 1000000) + (tokens_out * 15 / 1000000)
                db_conn.execute("INSERT INTO api_usage_log (user_id,api_name,tokens_in,tokens_out,cost_estimate) VALUES (?,?,?,?,?)",
                    (user["id"],"anthropic",tokens_in,tokens_out,cost))
                return result["content"][0]["text"]
            else:
                safe_log(f"Claude error: {resp.status_code} {_short_resp_text(resp)}", level="ERROR")
        except Exception as e: safe_log(f"Claude error: {e}", level="ERROR")

    # Tenta OpenAI/ChatGPT (se configurado)
    openai_key = get_setting("OPENAI_API_KEY")
    if openai_key and (ai_engine == "openai" or not api_key):
        try:
            import requests as req
            openai_messages = [{"role": "system", "content": system_prompt}] + api_messages
            resp = req.post("https://api.openai.com/v1/chat/completions",
                headers={"Authorization": f"Bearer {openai_key}", "Content-Type": "application/json"},
                json={"model": "gpt-4o-mini", "max_tokens": 500, "messages": openai_messages}, timeout=30)
            if resp.status_code == 200:
                result = resp.json()
                usage = result.get("usage", {})
                tokens_in = usage.get("prompt_tokens", 0)
                tokens_out = usage.get("completion_tokens", 0)
                cost = (tokens_in * 0.15 / 1000000) + (tokens_out * 0.6 / 1000000)
                db_conn.execute("INSERT INTO api_usage_log (user_id,api_name,tokens_in,tokens_out,cost_estimate) VALUES (?,?,?,?,?)",
                    (user["id"],"openai",tokens_in,tokens_out,cost))
                return result["choices"][0]["message"]["content"]
            else:
                safe_log(f"OpenAI error: {resp.status_code} {_short_resp_text(resp)}", level="ERROR")
        except Exception as e: safe_log(f"OpenAI error: {e}", level="ERROR")

    return user["ai_greeting"] or "Olá! Obrigado por entrar em contato. Como posso ajudar?"


def send_whatsapp_message(phone_id, token, to, message):
    masked_to = mask_phone(to)
    safe_log(f"[WA SEND] Tentando enviar para {masked_to}...")
    if not phone_id or not token:
        safe_log("[WA SEND] ERRO: Phone ID ou Token vazio!", level="ERROR")
        return {"success": False, "error": "Phone ID ou Token não configurado"}
    try:
        import requests as req
        url = f"https://graph.facebook.com/v18.0/{phone_id}/messages"
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        payload = {"messaging_product": "whatsapp", "to": to, "type": "text", "text": {"body": message}}
        resp = req.post(url, headers=headers, json=payload, timeout=15)
        if resp.status_code != 200:
            safe_log(f"[WA SEND] ERRO! Status {resp.status_code}", level="ERROR")
            # Tenta extrair erro humanizado da resposta — sem logar o body inteiro (que pode ter PII)
            try:
                err_data = resp.json()
                err_msg = err_data.get("error", {}).get("message", "")
                err_code = err_data.get("error", {}).get("code", "")
                safe_log(f"[WA SEND] Meta error code={err_code} msg={err_msg[:200]}", level="ERROR")
                if not err_msg:
                    err_msg = "Erro desconhecido da API Meta"
            except Exception:
                err_msg = f"HTTP {resp.status_code}"
            return {"success": False, "error": err_msg, "status_code": resp.status_code}
        else:
            safe_log(f"[WA SEND] ✓ Mensagem enviada para {masked_to}")
            return {"success": True}
    except Exception as e:
        safe_log(f"[WA SEND] EXCEÇÃO: {e}", level="ERROR")
        return {"success": False, "error": str(e)}


def prepare_tts_text(text):
    """Prepara texto para TTS, melhorando pronúncia de palavras inglesas"""
    # Dicionário de palavras inglesas comuns → pronúncia fonética em PT-BR
    english_words = {
        "smart": "esmárt", "center": "cênter", "online": "onlái-ne",
        "delivery": "delíveri", "store": "estóre", "shop": "xóp",
        "shopping": "xóping", "drive": "dráive", "fitness": "fítness",
        "beauty": "biúti", "fashion": "féxon", "design": "dezáin",
        "designer": "dezáiner", "marketing": "márketing", "business": "bízness",
        "feedback": "fíd-béck", "insight": "ínsait", "startup": "estartâp",
        "software": "sóftuer", "hardware": "rárdiuer", "network": "nétuork",
        "meeting": "míting", "deadline": "déd-láine", "budget": "bâdjet",
        "target": "târguet", "performance": "perfórmance", "coach": "côutch",
        "coaching": "côutching", "premium": "prêmium", "express": "êx-préss",
        "service": "sêrvice", "self-service": "sêlf sêrvice",
        "pet": "pétt", "petshop": "pétt xóp", "coworking": "co-uôrking",
        "hub": "râb", "tech": "téc", "food": "fúd", "drink": "drínk",
        "coffee": "cófi", "burger": "bârguer", "happy hour": "répi áuer",
        "sale": "sêil", "off": "óff", "black friday": "bléck fráidei",
        "free": "frí", "clean": "clín", "house": "ráus", "home": "rôum",
        "personal": "persônau", "trainer": "trêiner", "check-up": "tchéc-âp",
        "check": "tchéc", "whatsapp": "uótsép", "instagram": "ínstagrem",
        "facebook": "fêis-búk", "google": "gúgou", "youtube": "iú-tiúb",
        "iphone": "ái-fone", "android": "êndróid", "bluetooth": "blú-tufe",
        "wifi": "uái-fái", "site": "sáite", "link": "línk", "click": "clíck",
        "like": "láik", "post": "pôust", "story": "estóri", "stories": "estóris",
        "live": "láive", "streaming": "estríming", "playlist": "plêi-líst",
    }

    # Remove emojis para TTS
    text = re.sub(r'[\U00010000-\U0010ffff]', '', text)
    text = re.sub(r'[👋🤖📷📄📊⚡🎤💬✅❌⚠️🔑]', '', text)

    # Remove formatação markdown
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)  # **bold** → bold
    text = re.sub(r'\*(.+?)\*', r'\1', text)       # *italic* → italic
    text = re.sub(r'__(.+?)__', r'\1', text)       # __underline__ → underline
    text = re.sub(r'~~(.+?)~~', r'\1', text)       # ~~strike~~ → strike
    text = re.sub(r'`(.+?)`', r'\1', text)         # `code` → code
    text = re.sub(r'^#+\s*', '', text, flags=re.MULTILINE)  # ### heading → heading
    text = re.sub(r'^\s*[-*•]\s+', '', text, flags=re.MULTILINE)  # - item → item
    text = re.sub(r'^\s*\d+\.\s+', '', text, flags=re.MULTILINE)  # 1. item → item
    text = re.sub(r'\[(.+?)\]\(.+?\)', r'\1', text)  # [link](url) → link
    text = text.replace('*', '').replace('#', '').replace('_', ' ')  # limpa restantes

    # Substitui palavras inglesas (case-insensitive, palavras inteiras apenas)
    for eng, phonetic in english_words.items():
        pattern = re.compile(r'\b' + re.escape(eng) + r'\b', re.IGNORECASE)
        text = pattern.sub(phonetic, text)

    # Remove múltiplos espaços
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def text_to_audio(text, output_path=None):
    """Converte texto em áudio usando Edge TTS (grátis)"""
    try:
        import asyncio
        import edge_tts

        if not output_path:
            output_path = os.path.join(MEDIA_FOLDER, f"tts_{secrets.token_hex(8)}.mp3")

        # Prepara texto para melhor pronúncia
        clean_text = prepare_tts_text(text)
        safe_log(f"[TTS] Texto preparado: {clean_text[:80]}...")

        async def _generate():
            voice = "pt-BR-AntonioNeural"
            communicate = edge_tts.Communicate(clean_text, voice, rate="+5%")
            await communicate.save(output_path)

        # Roda async no sync context
        try:
            loop = asyncio.get_event_loop()
            if loop.is_running():
                import concurrent.futures
                with concurrent.futures.ThreadPoolExecutor() as pool:
                    pool.submit(lambda: asyncio.run(_generate())).result(timeout=30)
            else:
                loop.run_until_complete(_generate())
        except RuntimeError:
            asyncio.run(_generate())

        if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
            safe_log(f"[TTS] Áudio gerado: {output_path} ({os.path.getsize(output_path)} bytes)")
            return output_path
        safe_log("[TTS] Arquivo gerado vazio")
        return None
    except Exception as e:
        safe_log(f"[TTS] Erro: {e}", level="ERROR")
        return None


def upload_whatsapp_media(phone_id, token, filepath, mime_type="audio/mpeg"):
    """Faz upload de mídia para o WhatsApp e retorna o media_id"""
    try:
        import requests as req
        url = f"https://graph.facebook.com/v18.0/{phone_id}/media"
        headers = {"Authorization": f"Bearer {token}"}
        with open(filepath, "rb") as f:
            files = {"file": (os.path.basename(filepath), f, mime_type)}
            data = {"messaging_product": "whatsapp", "type": mime_type}
            resp = req.post(url, headers=headers, files=files, data=data, timeout=30)
        if resp.status_code == 200:
            media_id = resp.json().get("id", "")
            safe_log(f"[WA UPLOAD] Mídia enviada: {media_id}")
            return media_id
        else:
            safe_log(f"[WA UPLOAD] Erro: {resp.status_code}", level="ERROR")
            return None
    except Exception as e:
        safe_log(f"[WA UPLOAD] Exceção: {e}", level="ERROR")
        return None


def send_whatsapp_image(phone_id, token, to, image_path, caption=""):
    """Envia imagem pelo WhatsApp"""
    masked_to = mask_phone(to)
    media_id = upload_whatsapp_media(phone_id, token, image_path, "image/jpeg")
    if not media_id:
        safe_log("[WA IMAGE] Falha no upload", level="ERROR")
        return False
    try:
        import requests as req
        url = f"https://graph.facebook.com/v18.0/{phone_id}/messages"
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        payload = {"messaging_product": "whatsapp", "to": to, "type": "image",
                   "image": {"id": media_id}}
        if caption:
            payload["image"]["caption"] = caption[:1024]
        resp = req.post(url, headers=headers, json=payload, timeout=15)
        if resp.status_code == 200:
            safe_log(f"[WA IMAGE] ✓ Imagem enviada para {masked_to}")
            return True
        # Extrai code/message do erro sem logar body inteiro
        try:
            err = resp.json().get("error", {})
            safe_log(f"[WA IMAGE] Erro code={err.get('code','')} msg={err.get('message','')[:200]}", level="ERROR")
        except Exception:
            safe_log(f"[WA IMAGE] Erro HTTP {resp.status_code}", level="ERROR")
        return False
    except Exception as e:
        safe_log(f"[WA IMAGE] Exceção: {e}", level="ERROR")
        return False


def find_matching_product(user_id, message):
    """Encontra produto na galeria que combina com a mensagem do cliente"""
    try:
        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row
        products = db_conn.execute(
            "SELECT * FROM product_gallery WHERE user_id=?", (user_id,)
        ).fetchall()
        db_conn.close()

        if not products:
            return None

        msg_lower = message.lower()
        # Remove acentos para busca mais robusta
        import unicodedata
        def strip_accents(s):
            return ''.join(c for c in unicodedata.normalize('NFD', s)
                          if unicodedata.category(c) != 'Mn')

        msg_norm = strip_accents(msg_lower)

        # Calcula score de cada produto baseado em matches de keywords/nome
        best_match = None
        best_score = 0
        for p in products:
            score = 0
            # Nome do produto
            name_norm = strip_accents(p["name"].lower())
            if name_norm in msg_norm:
                score += 10
            for word in name_norm.split():
                if len(word) > 3 and word in msg_norm:
                    score += 3

            # Keywords (separadas por vírgula)
            keywords = p["keywords"] or ""
            for kw in keywords.split(","):
                kw = kw.strip().lower()
                kw_norm = strip_accents(kw)
                if not kw_norm or len(kw_norm) < 3:
                    continue
                if kw_norm in msg_norm:
                    score += 5

            if score > best_score:
                best_score = score
                best_match = p

        # Só retorna se o score for suficientemente alto
        if best_score >= 3:
            return best_match
        return None
    except Exception as e:
        safe_log(f"[GALLERY] Erro busca: {e}", level="ERROR")
        return None


# ─── GALERIA DE PRODUTOS ──────────────────────────────────────
@app.route("/dashboard/gallery")
@login_required
def gallery():
    user = g.user
    db = get_db()
    products = db.execute(
        "SELECT * FROM product_gallery WHERE user_id=? ORDER BY created_at DESC",
        (user["id"],)
    ).fetchall()

    # Mensagens
    ok_msg = request.args.get("ok", "")
    err_msg = request.args.get("err", "")
    alert = ""
    if ok_msg:
        alert = f'<div class="alert alert-success">✅ {esc(ok_msg)}</div>'
    elif err_msg:
        alert = f'<div class="alert alert-error">❌ {esc(err_msg)}</div>'

    # Grid de produtos cadastrados
    products_html = ""
    if products:
        for p in products:
            img_url = f"/media/gallery/{p['id']}"
            # Helper para pegar coluna com fallback (sqlite3.Row não tem .get)
            def _col(row, key, default=None):
                try:
                    val = row[key]
                    return val if val is not None else default
                except (KeyError, IndexError):
                    return default

            p_price = _col(p, 'price', 0) or 0
            p_stock = _col(p, 'stock', -1)
            p_active = _col(p, 'active', 1)

            price_html = f'<div style="background:rgba(16,185,129,0.15);color:#10b981;padding:4px 10px;border-radius:6px;display:inline-block;font-weight:600;font-size:13px;margin-bottom:8px">💰 R$ {p_price:.2f}</div>' if p_price > 0 else '<div style="background:rgba(239,68,68,0.15);color:#ef4444;padding:4px 10px;border-radius:6px;display:inline-block;font-size:12px;margin-bottom:8px">⚠️ Sem preço</div>'
            stock_html = ''
            if p_stock is not None and p_stock >= 0:
                stock_html = f'<div style="color:var(--text3);font-size:11px;margin-bottom:8px">Estoque: <strong>{p_stock}</strong></div>'
            products_html += f"""
            <div class="card" style="padding:16px">
                <img src="{img_url}" style="width:100%;height:180px;object-fit:cover;border-radius:8px;margin-bottom:12px" alt="{esc(p['name'])}">
                <h3 style="font-size:16px;margin-bottom:6px;color:var(--text)">{esc(p['name'])}</h3>
                {price_html}
                {stock_html}
                <p style="font-size:13px;color:var(--text3);margin-bottom:8px">{esc(p['description'] or 'Sem descrição')}</p>
                <p style="font-size:12px;color:var(--accent2);margin-bottom:12px"><strong>Keywords:</strong> {esc(p['keywords'] or 'nenhuma')}</p>
                <div style="display:flex;gap:6px">
                    <form method="POST" action="/dashboard/gallery/{p['id']}/update-price" style="flex:1;margin:0">{csrf_field()}
                        <div style="display:flex;gap:4px">
                            <input type="number" step="0.01" name="price" value="{p_price}"
                                   style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.1);padding:6px 8px;border-radius:6px;color:var(--text);width:70%;font-size:12px"
                                   placeholder="Preço">
                            <button type="submit" class="btn btn-sm" style="background:var(--accent2);color:white;width:30%;font-size:11px;padding:6px">💾</button>
                        </div>
                    </form>
                    <form method="POST" action="/dashboard/gallery/delete" style="margin:0">{csrf_field()}
                        <input type="hidden" name="id" value="{int(p['id'])}">
                        <button type="submit" class="btn btn-sm" style="background:rgba(239,68,68,0.2);color:#ef4444;font-size:11px;padding:6px 10px">🗑️</button>
                    </form>
                </div>
            </div>
            """
    else:
        products_html = '<p style="color:var(--text3);grid-column:1/-1;text-align:center;padding:40px">Nenhuma foto cadastrada ainda. Adicione produtos acima para que a IA possa enviá-los automaticamente.</p>'

    content = f"""<div class="container">
        <div class="page-header fade-in">
            <h1>Galeria de Produtos 📸</h1>
            <p>Cadastre fotos de produtos/serviços com preço. A IA envia e gera cobrança automática!</p>
        </div>
        {alert}

        <div class="card fade-in fade-in-1" style="margin-bottom:24px">
            <div class="card-header"><span class="card-title">➕ Adicionar novo produto</span></div>
            <form method="POST" action="/dashboard/gallery/upload" enctype="multipart/form-data">{csrf_field()}
                <div class="grid-2">
                    <div class="form-group">
                        <label class="form-label">Nome do produto *</label>
                        <input type="text" name="name" class="form-input" placeholder="Ex: Pizza Calabresa" required maxlength="100">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Palavras-chave (separadas por vírgula) *</label>
                        <input type="text" name="keywords" class="form-input" placeholder="Ex: pizza, calabresa, calabreza, peperoni" required maxlength="300">
                    </div>
                </div>
                <div class="grid-3">
                    <div class="form-group">
                        <label class="form-label">💰 Preço (R$) *</label>
                        <input type="number" name="price" class="form-input" step="0.01" min="0" placeholder="49.90">
                        <small style="color:var(--text3)">Coloque 0 se não for para vender direto</small>
                    </div>
                    <div class="form-group">
                        <label class="form-label">📦 Estoque</label>
                        <input type="number" name="stock" class="form-input" placeholder="Deixe vazio p/ ilimitado" min="0">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Categoria</label>
                        <input type="text" name="category" class="form-input" placeholder="Ex: Pizzas">
                    </div>
                </div>
                <div class="form-group">
                    <label class="form-label">Descrição</label>
                    <textarea name="description" class="form-input" rows="2" placeholder="Ex: Pizza grande de calabresa com cebola, mussarela e orégano" maxlength="500"></textarea>
                </div>
                <div class="form-group">
                    <label class="form-label">Imagem (JPG ou PNG, máx 5MB) *</label>
                    <input type="file" name="photo" accept="image/jpeg,image/png,image/jpg" required
                        style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.08);padding:10px;border-radius:8px;color:var(--text);width:100%">
                </div>
                <button type="submit" class="btn btn-primary">📤 Cadastrar produto</button>
            </form>
        </div>

        <div class="card fade-in fade-in-2">
            <div class="card-header"><span class="card-title">📦 Produtos cadastrados ({len(products)})</span></div>
            <div class="grid-3" style="gap:20px">
                {products_html}
            </div>
        </div>
    </div>"""
    return base_html("Galeria", content, dict(user))


@app.route("/dashboard/gallery/<int:pid>/update-price", methods=["POST"])
@login_required
def gallery_update_price(pid):
    """Atualiza o preço de um produto"""
    user = g.user
    db = get_db()
    try:
        price = float(request.form.get("price", "0").replace(",", "."))
        db.execute(
            "UPDATE product_gallery SET price=? WHERE id=? AND user_id=?",
            (price, pid, user["id"])
        )
        db.commit()
    except:
        pass
    return redirect("/dashboard/gallery?ok=Preço%20atualizado")


@app.route("/dashboard/gallery/upload", methods=["POST"])
@login_required
def gallery_upload():
    user = g.user
    name = request.form.get("name", "").strip()
    keywords = request.form.get("keywords", "").strip()
    description = request.form.get("description", "").strip()

    if not name or not keywords:
        return redirect("/dashboard/gallery?err=" + "Nome%20e%20palavras-chave%20s%C3%A3o%20obrigat%C3%B3rios")

    if "photo" not in request.files:
        return redirect("/dashboard/gallery?err=" + "Nenhuma%20imagem%20enviada")

    file = request.files["photo"]
    if file.filename == "":
        return redirect("/dashboard/gallery?err=" + "Arquivo%20vazio")

    image_bytes = file.read()
    if len(image_bytes) > 5 * 1024 * 1024:
        return redirect("/dashboard/gallery?err=" + "M%C3%A1ximo%205MB")

    # VALIDAÇÃO REAL com Pillow — rejeita arquivos falsificados
    validated_bytes, real_content_type = validate_and_normalize_image(image_bytes)
    if validated_bytes is None:
        return redirect("/dashboard/gallery?err=" + "Arquivo%20inv%C3%A1lido%20ou%20corrompido.%20Envie%20JPG%20ou%20PNG%20real.")

    # Salva no diretório de mídia
    gallery_dir = os.path.join(MEDIA_FOLDER, "gallery")
    os.makedirs(gallery_dir, exist_ok=True)

    # Nome único do arquivo (usa content_type validado, não o do cliente)
    ext = "jpg" if "jpeg" in real_content_type else "png"
    timestamp = int(time.time() * 1000)
    filename = f"user{user['id']}_{timestamp}.{ext}"
    file_path = os.path.join(gallery_dir, filename)

    with open(file_path, "wb") as f:
        f.write(validated_bytes)

    # Atualiza content_type com o verificado
    file_content_type = real_content_type

    # Salva no banco
    db = get_db()
    try:
        price = float(request.form.get("price", "0").replace(",", "."))
    except:
        price = 0
    try:
        stock_raw = request.form.get("stock", "").strip()
        stock = int(stock_raw) if stock_raw else -1
    except:
        stock = -1
    category = request.form.get("category", "").strip()

    db.execute(
        """INSERT INTO product_gallery
           (user_id, name, keywords, description, file_path, file_type, price, stock, category, active)
           VALUES (?,?,?,?,?,?,?,?,?,1)""",
        (user["id"], name, keywords, description, file_path, file_content_type, price, stock, category)
    )
    db.commit()

    return redirect("/dashboard/gallery?ok=" + "Produto%20cadastrado!")


@app.route("/dashboard/gallery/delete", methods=["POST"])
@login_required
def gallery_delete():
    user = g.user
    product_id = request.form.get("id", "")
    try:
        product_id = int(product_id)
    except ValueError:
        return redirect("/dashboard/gallery?err=ID%20inv%C3%A1lido")

    db = get_db()
    product = db.execute(
        "SELECT * FROM product_gallery WHERE id=? AND user_id=?",
        (product_id, user["id"])
    ).fetchone()

    if not product:
        return redirect("/dashboard/gallery?err=Produto%20n%C3%A3o%20encontrado")

    # Remove arquivo físico
    try:
        if os.path.exists(product["file_path"]):
            os.remove(product["file_path"])
    except Exception as e:
        safe_log(f"[GALLERY] Erro ao remover arquivo: {e}", level="ERROR")

    # Remove do banco
    db.execute("DELETE FROM product_gallery WHERE id=?", (product_id,))
    db.commit()

    return redirect("/dashboard/gallery?ok=" + "Produto%20removido")


@app.route("/media/gallery/<int:product_id>")
@login_required
def serve_gallery_image(product_id):
    """Serve imagem da galeria (somente dono)"""
    user = g.user
    db = get_db()
    product = db.execute(
        "SELECT * FROM product_gallery WHERE id=? AND user_id=?",
        (product_id, user["id"])
    ).fetchone()

    if not product or not os.path.exists(product["file_path"]):
        return "Imagem não encontrada", 404

    return send_file(product["file_path"], mimetype=product["file_type"])


def send_whatsapp_audio(phone_id, token, to, audio_path):
    """Envia áudio pelo WhatsApp"""
    masked_to = mask_phone(to)
    media_id = upload_whatsapp_media(phone_id, token, audio_path, "audio/mpeg")
    if not media_id:
        safe_log("[WA AUDIO] Falha no upload, enviando como texto", level="WARN")
        return False
    try:
        import requests as req
        url = f"https://graph.facebook.com/v18.0/{phone_id}/messages"
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        payload = {"messaging_product": "whatsapp", "to": to, "type": "audio", "audio": {"id": media_id}}
        resp = req.post(url, headers=headers, json=payload, timeout=15)
        safe_log(f"[WA AUDIO] Status: {resp.status_code}")
        if resp.status_code == 200:
            safe_log(f"[WA AUDIO] Áudio enviado para {masked_to}")
            return True
        else:
            try:
                err = resp.json().get("error", {})
                safe_log(f"[WA AUDIO] Erro code={err.get('code','')} msg={err.get('message','')[:200]}", level="ERROR")
            except Exception:
                safe_log(f"[WA AUDIO] Erro HTTP {resp.status_code}", level="ERROR")
            return False
    except Exception as e:
        safe_log(f"[WA AUDIO] Exceção: {e}", level="ERROR")
        return False


# ═══════════════════════════════════════════════════════════════
#  PAINEL ADMINISTRATIVO (DONO DO SISTEMA)
# ═══════════════════════════════════════════════════════════════

@app.route("/admin/login", methods=["GET","POST"])
def admin_login():
    """Login admin com suporte a 2FA (TOTP).
    Fluxo:
    1. Primeiro POST: valida email+senha. Se 2FA ativo, pede código. Senão, entra.
    2. Segundo POST (com totp_code): valida código ou backup code.
    """
    error = ""
    client_ip = request.remote_addr or "unknown"
    step_2fa = session.get("admin_awaiting_2fa", False)

    if request.method == "POST":
        if not check_rate_limit(client_ip):
            error = "Muitas tentativas. Aguarde 5 minutos."
        elif not ADMIN_PASSWORD:
            error = "ADMIN_PASSWORD não configurada. Defina no Railway → Variables."
        elif step_2fa:
            # Etapa 2: validar código TOTP ou backup code
            code = request.form.get("totp_code", "").strip()
            totp_secret = get_setting("ADMIN_TOTP_SECRET", "")
            code_ok = verify_totp_code(totp_secret, code)

            # Se não for TOTP válido, tenta backup code
            if not code_ok and code:
                backup_codes_raw = get_setting("ADMIN_BACKUP_CODES", "")
                if backup_codes_raw:
                    codes = [c.strip() for c in backup_codes_raw.split(",") if c.strip()]
                    code_upper = code.upper().replace(" ", "")
                    if code_upper in codes:
                        # Consome o backup code (uso único)
                        codes.remove(code_upper)
                        set_setting("ADMIN_BACKUP_CODES", ",".join(codes))
                        code_ok = True
                        log_admin_action("backup_code_used", details=f"Restantes: {len(codes)}")

            if code_ok:
                reset_login_attempts(client_ip)
                session.pop("admin_awaiting_2fa", None)
                # Preserva CSRF token ao limpar sessao para evitar 403 na primeira acao admin
                _preserved_csrf = session.get("_csrf_token")
                session.clear()
                if _preserved_csrf:
                    session["_csrf_token"] = _preserved_csrf
                session.permanent = True
                session["is_admin"] = True
                log_admin_action("login_success", details="2FA ok")
                return redirect("/admin")
            else:
                record_login_attempt(client_ip)
                error = "Código inválido. Tente novamente."
        else:
            import hmac as hmac_mod
            email_ok = hmac_mod.compare_digest(request.form.get("email", ""), ADMIN_EMAIL)
            password_ok = hmac_mod.compare_digest(request.form.get("password", ""), ADMIN_PASSWORD)
            if email_ok and password_ok:
                # Etapa 1: senha OK. Precisa de 2FA?
                if is_admin_2fa_enabled():
                    session["admin_awaiting_2fa"] = True
                    step_2fa = True
                    log_admin_action("password_ok_awaiting_2fa")
                else:
                    # LGPD/Segurança: em produção, 2FA é OBRIGATÓRIO.
                    # Só permite login direto em dev local ou se ADMIN_ALLOW_NO_2FA=1 estiver setada
                    # (apenas para o bootstrap inicial, quando 2FA ainda não foi configurado).
                    is_dev = os.getenv("FLASK_ENV", "").lower() == "development"
                    allow_no_2fa = os.getenv("ADMIN_ALLOW_NO_2FA", "").lower() in ("1", "true", "yes")
                    if not (is_dev or allow_no_2fa):
                        record_login_attempt(client_ip)
                        log_admin_action("login_blocked_no_2fa")
                        error = ("⚠️ Acesso bloqueado: 2FA não está configurado. "
                                 "Por segurança LGPD, o painel admin exige 2FA em produção. "
                                 "Defina temporariamente a variável de ambiente ADMIN_ALLOW_NO_2FA=1 "
                                 "no Railway, faça login uma vez, configure 2FA em /admin/2fa, "
                                 "remova a variável e refaça login.")
                    else:
                        reset_login_attempts(client_ip)
                        # Preserva CSRF token ao limpar sessao para evitar 403 na primeira acao admin
                        _preserved_csrf = session.get("_csrf_token")
                        session.clear()
                        if _preserved_csrf:
                            session["_csrf_token"] = _preserved_csrf
                        session.permanent = True
                        session["is_admin"] = True
                        # Marca origem do bypass para auditar
                        session["bypassed_2fa"] = True
                        log_admin_action("login_success_bypass_2fa",
                                         details=f"FLASK_ENV={os.getenv('FLASK_ENV','')} ADMIN_ALLOW_NO_2FA={os.getenv('ADMIN_ALLOW_NO_2FA','')}")
                        return redirect("/admin/2fa?bootstrap=1")
            else:
                record_login_attempt(client_ip)
                error = "Credenciais inválidas."
                log_admin_action("login_failed", details=f"email={request.form.get('email','')[:50]}")

    alert = f'<div class="alert alert-error">{error}</div>' if error else ""

    # Se estiver aguardando 2FA, mostra tela diferente
    if step_2fa:
        warning_2fa = '<div class="alert alert-info" style="margin-bottom:16px">🔐 Digite o código de 6 dígitos do seu app autenticador (Google Authenticator, Authy, etc)</div>'
        form_body = f"""<form method="POST">{csrf_field()}
            <div class="form-group">
                <label class="form-label">Código 2FA (6 dígitos)</label>
                <input type="text" name="totp_code" class="form-input" required maxlength="9" autocomplete="off" autofocus
                       placeholder="000000" style="font-size:24px;letter-spacing:8px;text-align:center;font-family:monospace">
                <small style="color:var(--text3);font-size:12px;margin-top:6px;display:block">Ou use um dos seus <strong>backup codes</strong> (formato XXXX-XXXX)</small>
            </div>
            <button type="submit" class="btn btn-primary btn-block btn-lg" style="background:var(--red)">Verificar</button>
            <a href="/admin/login?reset=1" style="display:block;text-align:center;margin-top:12px;color:var(--text3);font-size:12px">Cancelar e voltar</a>
        </form>"""
    else:
        warning_2fa = ''
        if not is_admin_2fa_enabled():
            warning_2fa = '<div class="alert alert-warning" style="margin-bottom:16px">⚠️ 2FA não configurado. Após o login, ative em <strong>Configurações 2FA</strong> para mais segurança.</div>'
        form_body = f"""<form method="POST">{csrf_field()}
            <div class="form-group"><label class="form-label">Email admin</label><input type="email" name="email" class="form-input" required autofocus></div>
            <div class="form-group"><label class="form-label">Senha</label><input type="password" name="password" class="form-input" required></div>
            <button type="submit" class="btn btn-primary btn-block btn-lg" style="background:var(--red)">Entrar</button>
        </form>"""

    # Limpar sessão 2FA se user cancelar
    if request.args.get("reset") == "1":
        session.pop("admin_awaiting_2fa", None)
        return redirect("/admin/login")

    return f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"><link rel="icon" type="image/png" href="/favicon.ico"><link rel="apple-touch-icon" href="/apple-touch-icon.png"><meta name="theme-color" content="#6366f1"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Admin — atendente.online</title><link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>{GLOBAL_CSS}</style></head><body>
<div class="auth-container"><div class="auth-card" style="border-top:3px solid var(--red)">
    <div style="text-align:center;margin-bottom:24px"><img src="data:image/png;base64,{LOGO_NAV_B64}" alt="atendente.online" style="height:56px"><span class="admin-badge" style="margin-left:8px;vertical-align:middle">ADMIN</span></div>
    <h2>{'Verificação 2FA' if step_2fa else 'Painel Administrativo'}</h2>
    {alert}
    {warning_2fa}
    {form_body}
</div></div></body></html>"""

@app.route("/admin/logout")
def admin_logout():
    log_admin_action("logout")
    session.pop("is_admin", None)
    session.pop("admin_awaiting_2fa", None)
    return redirect("/admin/login")


# ─── ADMIN DASHBOARD ──────────────────────────────────────────
@app.route("/admin")
@admin_required
def admin_dashboard():
    s = get_admin_stats()
    profit = s["mrr"] - s["total_api_cost"]

    # Alertas de segurança
    security_alerts = []
    if _get_fernet() is None:
        security_alerts.append("🚨 <strong>CRÍTICO: criptografia indisponível</strong> — biblioteca 'cryptography' não instalada. Segredos salvos podem estar em texto puro.")
    if not get_setting("WHATSAPP_APP_SECRET", ""):
        security_alerts.append("🚨 <strong>WHATSAPP_APP_SECRET não configurado</strong> — webhook está recusando mensagens em produção. Configure em APIs.")
    if not os.getenv("SECRET_KEY"):
        security_alerts.append("⚠️ <strong>SECRET_KEY usando valor aleatório</strong> — sessions serão invalidadas a cada restart. Defina SECRET_KEY no Railway → Variables.")
    if not os.getenv("ADMIN_PASSWORD"):
        security_alerts.append("⚠️ <strong>ADMIN_PASSWORD não configurada</strong> — defina no Railway → Variables para uma senha forte.")
    # 2FA admin
    if not is_admin_2fa_enabled():
        security_alerts.append("🚨 <strong>2FA admin não configurado</strong> — configure imediatamente em <a href='/admin/2fa' style='color:var(--red);text-decoration:underline'>/admin/2fa</a>. Após configurar, remova ADMIN_ALLOW_NO_2FA das env vars do Railway.")
    if os.getenv("ADMIN_ALLOW_NO_2FA", "").lower() in ("1", "true", "yes"):
        security_alerts.append("🚨 <strong>ADMIN_ALLOW_NO_2FA ativa!</strong> Esta variável permite login admin sem 2FA — é apenas para bootstrap. <strong>REMOVA AGORA do Railway → Variables</strong> se o 2FA já estiver configurado.")

    alerts_html = ""
    if security_alerts:
        alerts_html = '<div class="card fade-in" style="border:2px solid var(--red);margin-bottom:24px;background:rgba(239,68,68,0.05)"><div class="card-header"><span class="card-title" style="color:var(--red)">⚠️ Alertas de Segurança</span></div>'
        for alert in security_alerts:
            alerts_html += f'<p style="margin:8px 0;color:var(--text2);font-size:14px">{alert}</p>'
        alerts_html += '</div>'

    content = f"""<div class="container">
        <div class="page-header fade-in"><h1>Dashboard Administrativo 🏢</h1><p>Visão geral do seu SaaS</p></div>
        {alerts_html}
        
        <div class="grid-5 fade-in fade-in-1">
            <div class="metric-card"><div style="font-size:24px">👥</div><div class="metric-value">{s['total_users']}</div><div class="metric-label">Clientes totais</div>
                <div class="metric-trend trend-up">+{s['new_users_today']} hoje</div></div>
            <div class="metric-card"><div style="font-size:24px">✅</div><div class="metric-value" style="color:var(--green2)">{s['active_users']}</div><div class="metric-label">Assinaturas ativas</div></div>
            <div class="metric-card"><div style="font-size:24px">⏳</div><div class="metric-value" style="color:var(--orange)">{s['trial_users']}</div><div class="metric-label">Em trial</div></div>
            <div class="metric-card"><div style="font-size:24px">💰</div><div class="metric-value" style="color:var(--green2)">R$ {s['mrr']:.0f}</div><div class="metric-label">MRR (receita mensal)</div></div>
            <div class="metric-card"><div style="font-size:24px">📊</div><div class="metric-value" style="color:var(--accent2)">R$ {s['total_revenue']:.0f}</div><div class="metric-label">Receita total</div></div>
        </div>

        <div class="grid-4 fade-in fade-in-2">
            <div class="stat-card"><div class="stat-icon stat-icon-green">💬</div><div class="stat-value">{s['total_conversations']}</div><div class="stat-label">Conversas totais</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-blue">📨</div><div class="stat-value">{s['total_messages']}</div><div class="stat-label">Mensagens totais</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-orange">📨</div><div class="stat-value">{s['msgs_today']}</div><div class="stat-label">Mensagens hoje</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-red">💸</div><div class="stat-value">US$ {s['total_api_cost']:.2f}</div><div class="stat-label">Custo total de API</div></div>
        </div>

        <div class="grid-2 fade-in fade-in-3">
            <div class="card">
                <div class="card-header"><span class="card-title">Distribuição por plano</span></div>
                <table><thead><tr><th>Plano</th><th>Ativos</th><th>Receita/mês</th></tr></thead><tbody>
                {''.join(f'<tr><td><span class="badge badge-purple">{PLANS[k]["name"]}</span></td><td>{s["by_plan"].get(k,0)}</td><td>R$ {s["by_plan"].get(k,0) * PLANS[k]["price"]:.0f}</td></tr>' for k in PLANS)}
                <tr style="font-weight:700"><td>Total</td><td>{s['active_users']}</td><td>R$ {s['mrr']:.0f}</td></tr>
                </tbody></table>
            </div>
            <div class="card">
                <div class="card-header"><span class="card-title">Saúde do negócio</span></div>
                <div style="padding:16px 0">
                    <div style="display:flex;justify-content:space-between;margin-bottom:16px"><span style="color:var(--text2)">Taxa de conversão trial→pago</span>
                        <strong>{(s['active_users']/max(s['total_users'],1)*100):.0f}%</strong></div>
                    <div style="display:flex;justify-content:space-between;margin-bottom:16px"><span style="color:var(--text2)">Ticket médio</span>
                        <strong>R$ {(s['mrr']/max(s['active_users'],1)):.0f}</strong></div>
                    <div style="display:flex;justify-content:space-between;margin-bottom:16px"><span style="color:var(--text2)">Custo API / cliente</span>
                        <strong>US$ {(s['total_api_cost']/max(s['active_users'],1)):.2f}</strong></div>
                    <div style="display:flex;justify-content:space-between;margin-bottom:16px"><span style="color:var(--text2)">Lucro estimado (MRR - API)</span>
                        <strong style="color:{'var(--green2)' if profit > 0 else 'var(--red)'}">R$ {profit:.0f}</strong></div>
                    <div style="display:flex;justify-content:space-between"><span style="color:var(--text2)">Clientes inativos</span>
                        <strong style="color:var(--red)">{s['inactive_users']}</strong></div>
                </div>
            </div>
        </div>
    </div>"""
    return admin_html("Admin Dashboard", content)


# ─── ADMIN: CLIENTES ──────────────────────────────────────────
@app.route("/admin/users")
@admin_required
def admin_users():
    db = get_db()
    users = db.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall()
    rows = ""
    for u in users:
        status_cls = {"active":"badge-green","trial":"badge-orange","inactive":"badge-red","cancelled":"badge-red"}.get(u["plan_status"],"badge-orange")
        status_txt = {"active":"Ativo","trial":"Trial","inactive":"Inativo","cancelled":"Cancelado"}.get(u["plan_status"],u["plan_status"])
        stats = get_user_stats(u["id"])
        plan_name = PLANS.get(u['plan'], {}).get('name', u['plan'])
        rows += f"""<tr>
            <td><strong>{esc(u['name'])}</strong><br><span style="color:var(--text3);font-size:12px">{esc(u['email'])}</span></td>
            <td>{esc(u['company'] or '—')}</td>
            <td><span class="badge badge-purple">{plan_name}</span></td>
            <td><span class="badge {status_cls}">{status_txt}</span></td>
            <td>{u['msgs_used']}/{u['msgs_limit']}</td>
            <td>{stats['conversations']}</td>
            <td style="font-size:12px;color:var(--text3)">{(u['created_at'] or '')[:10]}</td>
            <td style="font-size:12px;color:var(--text3)">{(u['last_login'] or 'Nunca')[:10]}</td>
            <td>
                <form method="POST" action="/admin/users/{u['id']}/change-plan" style="display:inline;margin-right:4px">
                    {csrf_field()}
                    <select name="plan" class="admin-plan-select" style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.1);color:var(--text);padding:4px 8px;border-radius:6px;font-size:12px">
                        <option value="starter" {'selected' if u['plan']=='starter' else ''}>Starter</option>
                        <option value="pro" {'selected' if u['plan']=='pro' else ''}>Pro</option>
                        <option value="business" {'selected' if u['plan']=='business' else ''}>Business</option>
                        <option value="agency" {'selected' if u['plan']=='agency' else ''}>Agência</option>
                    </select>
                </form>
                <form method="POST" action="/admin/users/{u['id']}/toggle" style="display:inline">
                    {csrf_field()}
                    <button type="submit" class="btn {'btn-danger' if u['is_active'] else 'btn-success'} btn-sm">{'Desativar' if u['is_active'] else 'Ativar'}</button>
                </form>
            </td></tr>"""

    content = f"""<div class="container"><div class="page-header"><h1>Clientes ({len(users)})</h1><p>Todos os clientes cadastrados no sistema</p></div>

        <div class="card" style="margin-bottom:24px">
            <div class="card-header"><span class="card-title">➕ Criar cliente manualmente (sem verificação de email)</span></div>
            <p style="color:var(--text3);font-size:13px;margin-bottom:16px">Use para criar contas de teste ou contornar problemas com SMTP. A conta é criada já verificada e ativa.</p>
            <form method="POST" action="/admin/users/create">{csrf_field()}
                <div class="grid-3">
                    <div class="form-group">
                        <label class="form-label">Nome *</label>
                        <input type="text" name="name" class="form-input" required placeholder="Nome completo">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Email *</label>
                        <input type="email" name="email" class="form-input" required placeholder="email@exemplo.com">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Senha *</label>
                        <input type="text" name="password" class="form-input" required placeholder="senha forte">
                    </div>
                </div>
                <div class="grid-3">
                    <div class="form-group">
                        <label class="form-label">Empresa</label>
                        <input type="text" name="company" class="form-input" placeholder="Nome da empresa">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Telefone</label>
                        <input type="text" name="phone" class="form-input" placeholder="+55 11 99999-9999">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Plano</label>
                        <select name="plan" class="form-input">
                            <option value="starter">Starter</option>
                            <option value="pro">Pro</option>
                            <option value="business" selected>Business</option>
                            <option value="agency">Agência</option>
                        </select>
                    </div>
                </div>
                <button type="submit" class="btn btn-success">✅ Criar cliente</button>
            </form>
        </div>

        <div class="card"><div class="table-wrap"><table><thead><tr><th>Cliente</th><th>Empresa</th><th>Plano</th><th>Status</th><th>Msgs</th><th>Conversas</th><th>Cadastro</th><th>Último login</th><th>Ação</th></tr></thead>
        <tbody>{rows}</tbody></table></div></div>
        <script nonce="{getattr(g, 'csp_nonce', '')}">
        document.querySelectorAll('.admin-plan-select').forEach(function(el) {{
            el.addEventListener('change', function() {{
                this.form.submit();
            }});
        }});
        </script>
        </div>"""
    return admin_html("Clientes", content)


@app.route("/admin/users/create", methods=["POST"])
@admin_required
def admin_create_user():
    """Cria usuário manualmente, já verificado e ativo"""
    db = get_db()
    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "").strip()
    company = request.form.get("company", "").strip()
    phone = request.form.get("phone", "").strip()
    plan = request.form.get("plan", "business")

    if not name or not email or not password:
        return redirect("/admin/users")

    # Verifica se já existe
    existing = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
    if existing:
        return redirect("/admin/users?error=existe")

    # Pega limite do plano
    plan_info = PLANS.get(plan, PLANS["starter"])
    msgs_limit = plan_info.get("msgs", 500)

    # Cria usuário já verificado e ativo
    trial_ends = (datetime.now() + timedelta(days=365)).strftime("%Y-%m-%d")
    db.execute("""INSERT INTO users
        (email, password_hash, name, company, phone, plan, plan_status, 
         msgs_limit, msgs_used, trial_ends_at, email_verified, is_active)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (email, hash_password(password), name, company, phone, plan,
         "active", msgs_limit, 0, trial_ends, 1, 1))
    db.commit()
    return redirect("/admin/users")


@app.route("/admin/users/<int:uid>/toggle", methods=["POST"])
@admin_required
def admin_toggle_user(uid):
    db = get_db()
    user = db.execute("SELECT is_active, email FROM users WHERE id=?", (uid,)).fetchone()
    if user:
        new_status = 0 if user["is_active"] else 1
        db.execute("UPDATE users SET is_active=? WHERE id=?", (new_status, uid))
        db.commit()
        log_admin_action(
            "user_activated" if new_status else "user_deactivated",
            target_type="user", target_id=uid,
            details=f"email={user['email']}"
        )
    return redirect("/admin/users")


@app.route("/admin/users/<int:uid>/change-plan", methods=["POST"])
@admin_required
def admin_change_plan(uid):
    """Muda o plano de um usuário e ativa a assinatura"""
    db = get_db()
    new_plan = request.form.get("plan", "starter")
    if new_plan not in PLANS:
        return redirect("/admin/users")

    plan_info = PLANS[new_plan]
    msgs_limit = plan_info.get("msgs", 500)

    old_user = db.execute("SELECT email, plan FROM users WHERE id=?", (uid,)).fetchone()
    db.execute(
        "UPDATE users SET plan=?, plan_status='active', msgs_limit=?, email_verified=1, is_active=1 WHERE id=?",
        (new_plan, msgs_limit, uid)
    )
    db.commit()

    if old_user:
        log_admin_action(
            "user_plan_changed",
            target_type="user", target_id=uid,
            details=f"email={old_user['email']} de {old_user['plan']} para {new_plan}"
        )
    return redirect("/admin/users")


# ─── ADMIN: 2FA (SETUP, DISABLE, AUDIT) ──────────────────────
# ─── ADMIN: DEBUG MERCADO PAGO ──────────────────────────────────
@app.route("/admin/mp-debug", methods=["GET", "POST"])
@admin_required
def admin_mp_debug():
    """Testa a geração de link de pagamento do Mercado Pago.
    Útil para diagnosticar por que PIX não aparece."""
    result_html = ""

    if request.method == "POST":
        try:
            import time as time_mod
            import json as json_mod

            # Pega dados do formulário
            user_id = int(request.form.get("user_id", 0))
            test_amount = float(request.form.get("amount", "10.00"))

            db = get_db()
            raw_user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
            if not raw_user:
                result_html = '<div class="alert alert-error">❌ Usuário não encontrado</div>'
            else:
                user = decrypt_user_row(raw_user)
                mp_token = user.get("mp_access_token", "")

                if not mp_token:
                    result_html = '<div class="alert alert-error">❌ Usuário não tem MP token configurado em /dashboard/commerce/settings</div>'
                else:
                    # Detecta se é token de teste
                    is_test = mp_token.startswith("TEST-")

                    # Cria preference de teste
                    items_test = [{
                        "title": "Pedido de Teste",
                        "quantity": 1,
                        "unit_price": test_amount,
                        "currency_id": "BRL"
                    }]

                    checkout = mp_create_checkout_preference(
                        mp_token,
                        items_test,
                        "5585999999999",
                        f"test_debug_{int(time_mod.time())}",
                        f"https://atendente.online/webhook/mp-commerce/{user_id}"
                    )

                    # Trata erro 403 e outros do MP
                    if checkout and checkout.get("error"):
                        status_code = checkout.get("status_code", "?")
                        detail = checkout.get("detail", "")

                        # Erro 403 = permissão / PolicyAgent
                        if status_code == 403 or "PolicyAgent" in detail or "UNAUTHORIZED" in detail:
                            result_html = f'''
                            <div class="alert alert-error">
                                ❌ <strong>Erro 403 — Token sem permissão</strong>
                            </div>
                            <div class="card" style="margin-top:16px;border-left:4px solid #ef4444">
                                <div class="card-header"><span class="card-title">🚨 Diagnóstico</span></div>
                                <div style="padding:16px">
                                    <p style="margin-bottom:12px">O Mercado Pago bloqueou a criação do link. Isso acontece quando:</p>
                                    <ul style="list-style:none;padding-left:0;line-height:2">
                                        <li><strong>1.</strong> Token foi gerado em aplicação que NÃO tem "Checkout Pro" habilitado</li>
                                        <li><strong>2.</strong> Conta MP não está verificada (CPF/CNPJ pendente)</li>
                                        <li><strong>3.</strong> Credenciais de produção ainda não foram liberadas</li>
                                        <li><strong>4.</strong> Token foi revogado/expirou</li>
                                    </ul>

                                    <div class="alert alert-info" style="margin-top:16px">
                                        <strong>💡 Solução rápida:</strong><br><br>
                                        1. Acesse <a href="https://www.mercadopago.com.br/developers/panel/app" target="_blank" style="color:#00c896">mercadopago.com.br/developers/panel/app</a><br>
                                        2. Crie uma NOVA aplicação marcando <strong>"Checkout Pro"</strong> nos produtos<br>
                                        3. Pegue o token em <strong>"Credenciais de teste"</strong> (começa com <code>TEST-</code>)<br>
                                        4. Cole no sistema em <a href="/dashboard/commerce/settings" style="color:#00c896">/dashboard/commerce/settings</a><br>
                                        5. Teste aqui novamente<br><br>
                                        Depois de homologar, use as <strong>"Credenciais de produção"</strong> (APP_USR-).
                                    </div>
                                </div>
                            </div>
                            <div class="card" style="margin-top:16px">
                                <div class="card-header"><span class="card-title">📋 Resposta bruta do MP</span></div>
                                <pre style="padding:12px;background:#0a0e14;color:#94a3b8;font-size:11px;overflow-x:auto;border-radius:6px">Status: {status_code}\\n{esc(detail)}</pre>
                            </div>
                            '''
                        else:
                            result_html = f'''
                            <div class="alert alert-error">
                                ❌ Erro {status_code}: {esc(detail[:200])}
                            </div>
                            '''
                    elif checkout and checkout.get("checkout_url"):
                        # Consulta os detalhes da preference criada
                        import requests as req
                        detail_resp = req.get(
                            f"https://api.mercadopago.com/checkout/preferences/{checkout['id']}",
                            headers={"Authorization": f"Bearer {mp_token}"},
                            timeout=10
                        )

                        detail = detail_resp.json() if detail_resp.status_code == 200 else {}

                        # Consulta os métodos de pagamento habilitados NA CONTA
                        methods_resp = req.get(
                            "https://api.mercadopago.com/v1/payment_methods",
                            headers={"Authorization": f"Bearer {mp_token}"},
                            timeout=10
                        )
                        methods = methods_resp.json() if methods_resp.status_code == 200 else []

                        # Separa por tipo
                        pix_methods = [m for m in methods if m.get("id") == "pix" or "pix" in m.get("name", "").lower()]
                        credit_methods = [m for m in methods if m.get("payment_type_id") == "credit_card"]
                        debit_methods = [m for m in methods if m.get("payment_type_id") == "debit_card"]
                        ticket_methods = [m for m in methods if m.get("payment_type_id") == "ticket"]

                        pix_status = "✅ HABILITADO" if pix_methods else "❌ NÃO HABILITADO"
                        pix_color = "#10b981" if pix_methods else "#ef4444"

                        result_html = f"""
                        <div class="alert alert-{'warning' if is_test else 'success'}">
                            {'⚠️ Token de TESTE (sandbox). Em produção use o token real.' if is_test else '✅ Token de PRODUÇÃO'}
                        </div>

                        <div class="card" style="margin-top:16px">
                            <div class="card-header"><span class="card-title">🔗 Link gerado</span></div>
                            <div style="padding:16px">
                                <p><strong>Preference ID:</strong> <code>{esc(checkout['id'])}</code></p>
                                <p style="margin-top:12px"><strong>URL:</strong></p>
                                <a href="{esc(checkout['checkout_url'])}" target="_blank" style="color:#00c896;word-break:break-all">{esc(checkout['checkout_url'])}</a>
                                <br><br>
                                <a href="{esc(checkout['checkout_url'])}" target="_blank" class="btn btn-primary">🧪 Abrir link em nova aba</a>
                            </div>
                        </div>

                        <div class="card" style="margin-top:16px;border-left:4px solid {pix_color}">
                            <div class="card-header"><span class="card-title">💰 PIX nesta conta MP: {pix_status}</span></div>
                            <div style="padding:16px">
                                {'<p style="color:#10b981">PIX aparece nos métodos disponíveis da conta.</p>' if pix_methods else ''}
                                {'''<div class="alert alert-error" style="margin:12px 0">
                                    <strong>⚠️ Para habilitar PIX nesta conta:</strong><br>
                                    1. Acesse <a href="https://www.mercadopago.com.br/settings/account/pix" target="_blank" style="color:#00c896">mercadopago.com.br/settings/account/pix</a><br>
                                    2. Cadastre uma chave PIX (CPF, CNPJ, email ou aleatória)<br>
                                    3. Verifique se sua conta está validada (CPF/CNPJ comprovado)<br>
                                    4. Aguarde aprovação do MP (pode levar minutos)<br>
                                    5. Gere um novo link de pagamento
                                </div>''' if not pix_methods else ''}
                            </div>
                        </div>

                        <div class="grid-4" style="margin-top:16px;gap:8px">
                            <div class="stat-card"><div class="stat-value">{len(credit_methods)}</div><div class="stat-label">Cartões Crédito</div></div>
                            <div class="stat-card"><div class="stat-value">{len(debit_methods)}</div><div class="stat-label">Cartões Débito</div></div>
                            <div class="stat-card"><div class="stat-value">{len(ticket_methods)}</div><div class="stat-label">Boletos</div></div>
                            <div class="stat-card"><div class="stat-value">{len(pix_methods)}</div><div class="stat-label">PIX</div></div>
                        </div>

                        <div class="card" style="margin-top:16px">
                            <div class="card-header"><span class="card-title">🔍 Detalhes da Preference</span></div>
                            <pre style="padding:12px;background:#0a0e14;color:#94a3b8;font-size:11px;overflow-x:auto;border-radius:6px;max-height:300px">{esc(json_mod.dumps(detail, indent=2, ensure_ascii=False)[:3000])}</pre>
                        </div>
                        """
                    else:
                        result_html = '<div class="alert alert-error">❌ Falha ao criar preference. Verifique o log no console.</div>'

        except Exception as e:
            result_html = f'<div class="alert alert-error">❌ Erro: {esc(str(e))}</div>'

    # Listar usuários com MP configurado
    db = get_db()
    users_raw = db.execute("SELECT id, email FROM users WHERE mp_access_token != '' AND mp_access_token IS NOT NULL ORDER BY id DESC LIMIT 50").fetchall()
    users_options = ""
    for u in users_raw:
        users_options += f'<option value="{u["id"]}">#{u["id"]} — {esc(u["email"])}</option>'

    if not users_options:
        users_options = '<option value="">Nenhum usuário com MP configurado</option>'

    content = f"""<div class="container">
        <div class="page-header">
            <h1>🧪 Debug Mercado Pago</h1>
            <p>Testa a geração de link e diagnostica por que PIX não aparece</p>
        </div>

        <div class="card">
            <div class="card-header"><span class="card-title">Testar geração de link</span></div>
            <form method="POST" style="padding:16px">{csrf_field()}
                <div class="form-group">
                    <label class="form-label">Usuário (com MP configurado)</label>
                    <select name="user_id" class="form-input" required>{users_options}</select>
                </div>
                <div class="form-group">
                    <label class="form-label">Valor do teste (R$)</label>
                    <input type="number" name="amount" class="form-input" value="10.00" step="0.01" min="1" max="1000">
                    <small style="color:var(--text3)">Recomendado: valor baixo (R$ 1-10) para não gerar cobrança real em teste</small>
                </div>
                <button type="submit" class="btn btn-primary">🧪 Gerar link de teste</button>
            </form>
        </div>

        {result_html}
    </div>"""
    return admin_html("Debug Mercado Pago", content)


@app.route("/admin/2fa", methods=["GET", "POST"])
@admin_required
def admin_2fa_setup():
    """Setup/gerenciamento de 2FA do admin"""
    msg = ""
    is_enabled = is_admin_2fa_enabled()

    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "disable" and is_enabled:
            # Exige senha de confirmação
            confirm_pw = request.form.get("confirm_password", "")
            import hmac as hmac_mod
            if not hmac_mod.compare_digest(confirm_pw, ADMIN_PASSWORD):
                msg = '<div class="alert alert-error">Senha incorreta. 2FA não foi desativado.</div>'
            else:
                set_setting("ADMIN_TOTP_SECRET", "")
                set_setting("ADMIN_BACKUP_CODES", "")
                log_admin_action("2fa_disabled")
                is_enabled = False
                msg = '<div class="alert alert-warning">⚠️ 2FA desativado. Reconfigure para manter segurança.</div>'

        elif action == "verify_setup":
            # Usuário acabou de escanear o QR, está validando
            pending_secret = session.get("pending_totp_secret", "")
            code = request.form.get("totp_code", "").strip()
            if not pending_secret:
                msg = '<div class="alert alert-error">Sessão expirada. Inicie o setup novamente.</div>'
            elif verify_totp_code(pending_secret, code):
                # Salva definitivamente
                backup_codes = generate_backup_codes(8)
                set_setting("ADMIN_TOTP_SECRET", pending_secret)
                set_setting("ADMIN_BACKUP_CODES", ",".join(backup_codes))
                session.pop("pending_totp_secret", None)
                log_admin_action("2fa_enabled")
                session["2fa_backup_codes_shown"] = backup_codes  # Mostra uma vez só
                return redirect("/admin/2fa?setup=success")
            else:
                msg = '<div class="alert alert-error">Código inválido. Tente novamente.</div>'

    # Página de sucesso mostra backup codes (única vez)
    if request.args.get("setup") == "success":
        codes = session.pop("2fa_backup_codes_shown", [])
        codes_html = "".join([f'<li style="font-family:monospace;font-size:16px;padding:6px 0;border-bottom:1px solid rgba(255,255,255,0.05)">{c}</li>' for c in codes])
        content = f"""<div class="container">
            <div class="page-header"><h1>✅ 2FA Ativado!</h1><p>Guarde seus backup codes em local seguro</p></div>
            <div class="alert alert-success">🎉 2FA está funcionando. A partir de agora, todo login exigirá código do seu app autenticador.</div>
            <div class="card" style="margin-top:16px">
                <div class="card-header"><span class="card-title">🔑 Seus Backup Codes</span></div>
                <p style="color:var(--text2);margin-bottom:16px">Use um destes códigos <strong>se perder seu celular</strong>. Cada código funciona apenas <strong>uma vez</strong>.</p>
                <div class="alert alert-warning">⚠️ <strong>IMPORTANTE:</strong> Anote ou salve em gerenciador de senhas. Esses códigos não serão mostrados novamente!</div>
                <ul style="list-style:none;padding:20px;background:rgba(0,0,0,0.3);border-radius:8px;margin-top:16px">{codes_html}</ul>
                <a href="/admin/2fa" class="btn btn-primary" style="margin-top:16px">Já anotei, continuar</a>
            </div>
        </div>"""
        return admin_html("2FA Ativado", content)

    # Iniciar novo setup (gera secret temporário)
    if not is_enabled and request.args.get("start") == "1":
        new_secret = generate_totp_secret()
        if not new_secret:
            content = '<div class="container"><div class="alert alert-error">❌ pyotp não instalado. Execute: pip install pyotp qrcode</div></div>'
            return admin_html("Erro 2FA", content)

        session["pending_totp_secret"] = new_secret
        uri = generate_totp_uri(new_secret, "atendente.online", ADMIN_EMAIL)
        qr_data = generate_totp_qr_base64(uri)

        qr_html = f'<img src="{qr_data}" alt="QR Code 2FA" style="width:220px;height:220px;background:white;padding:12px;border-radius:8px;display:block;margin:0 auto">' if qr_data else '<p style="color:var(--red)">Erro ao gerar QR code</p>'

        content = f"""<div class="container">
            <div class="page-header"><h1>🔐 Configurar 2FA</h1><p>Passo 1: Escaneie o QR code no seu app</p></div>
            <div class="grid-2">
                <div class="card">
                    <div class="card-header"><span class="card-title">📱 Escaneie com:</span></div>
                    <ul style="list-style:none;padding:0;color:var(--text2);font-size:14px">
                        <li style="padding:6px 0">• Google Authenticator</li>
                        <li style="padding:6px 0">• Authy</li>
                        <li style="padding:6px 0">• Microsoft Authenticator</li>
                        <li style="padding:6px 0">• 1Password, Bitwarden, etc</li>
                    </ul>
                    <div style="margin-top:20px">
                        {qr_html}
                    </div>
                    <p style="color:var(--text3);font-size:12px;text-align:center;margin-top:12px">Ou cole este código manualmente:</p>
                    <code style="display:block;padding:10px;background:rgba(0,0,0,0.3);border-radius:6px;font-size:11px;word-break:break-all;text-align:center">{new_secret}</code>
                </div>
                <div class="card">
                    <div class="card-header"><span class="card-title">✅ Passo 2: Confirme o código</span></div>
                    <p style="color:var(--text2);margin-bottom:16px">Depois de escanear, digite o código de 6 dígitos que aparece no app:</p>
                    <form method="POST">{csrf_field()}
                        <input type="hidden" name="action" value="verify_setup">
                        <div class="form-group">
                            <input type="text" name="totp_code" class="form-input" required maxlength="6" autocomplete="off" autofocus
                                   placeholder="000000" style="font-size:24px;letter-spacing:8px;text-align:center;font-family:monospace">
                        </div>
                        <button type="submit" class="btn btn-primary btn-block">Ativar 2FA →</button>
                    </form>
                    {msg}
                </div>
            </div>
            <a href="/admin/2fa" style="display:block;text-align:center;margin-top:20px;color:var(--text3)">← Cancelar</a>
        </div>"""
        return admin_html("Setup 2FA", content)

    # Tela principal: status
    status_card = ""
    if is_enabled:
        backup_count = len([c for c in (get_setting("ADMIN_BACKUP_CODES", "") or "").split(",") if c.strip()])
        status_card = f"""
        <div class="card" style="border-left:4px solid #10b981">
            <div class="card-header"><span class="card-title">✅ 2FA Ativado</span></div>
            <p style="color:var(--text2);margin-bottom:16px">Seu login requer código do app autenticador. Backup codes restantes: <strong>{backup_count}</strong></p>
            <form method="POST" style="margin-top:16px">{csrf_field()}
                <input type="hidden" name="action" value="disable">
                <p style="color:var(--text3);font-size:13px;margin-bottom:8px">Para desativar, digite sua senha admin:</p>
                <div class="form-group">
                    <input type="password" name="confirm_password" class="form-input" placeholder="Senha admin" required>
                </div>
                <button type="submit" id="disable-2fa-btn" class="btn" style="background:rgba(239,68,68,0.2);color:#ef4444">🚫 Desativar 2FA</button>
            </form>
        </div>
        """
    else:
        status_card = """
        <div class="card" style="border-left:4px solid #f59e0b">
            <div class="card-header"><span class="card-title">⚠️ 2FA Desativado</span></div>
            <p style="color:var(--text2);margin-bottom:20px">Sem 2FA, qualquer pessoa com sua senha pode acessar o painel admin. <strong>Recomendamos fortemente ativar.</strong></p>
            <ul style="color:var(--text3);font-size:13px;margin-bottom:20px;padding-left:20px;line-height:1.8">
                <li>Leva menos de 2 minutos para configurar</li>
                <li>Funciona com Google Authenticator, Authy, etc</li>
                <li>Gera 8 backup codes para emergência</li>
            </ul>
            <a href="/admin/2fa?start=1" class="btn btn-primary">🔐 Configurar 2FA agora</a>
        </div>
        """

    content = f"""<div class="container">
        <div class="page-header"><h1>🔐 Autenticação de 2 Fatores</h1><p>Proteção extra para sua conta admin</p></div>
        {msg}
        {status_card}
    </div>
    <script nonce="{getattr(g, 'csp_nonce', '')}">
    document.getElementById('disable-2fa-btn')?.addEventListener('click', function(event) {{
        if (!window.confirm('Desativar 2FA? Isso reduz a segurança da conta admin.')) {{
            event.preventDefault();
        }}
    }});
    </script>"""
    return admin_html("2FA Admin", content)


@app.route("/admin/audit-log")
@admin_required
def admin_audit_log_view():
    """Histórico de ações admin (auditoria)"""
    db = get_db()
    logs = db.execute(
        "SELECT * FROM admin_audit_log ORDER BY created_at DESC LIMIT 500"
    ).fetchall()

    total = db.execute("SELECT COUNT(*) as c FROM admin_audit_log").fetchone()["c"]

    # Últimos 7 dias
    recent = db.execute(
        "SELECT COUNT(*) as c FROM admin_audit_log WHERE datetime(created_at) > datetime('now', '-7 days')"
    ).fetchone()["c"]

    rows = ""
    for log in logs:
        action_cls = "badge-green"
        if "fail" in log["action"] or "error" in log["action"]:
            action_cls = "badge-red"
        elif "disabled" in log["action"] or "delete" in log["action"]:
            action_cls = "badge-orange"
        elif "enabled" in log["action"] or "created" in log["action"]:
            action_cls = "badge-blue"

        rows += f"""<tr>
            <td style="font-size:12px;color:var(--text3);white-space:nowrap">{to_br_datetime(log['created_at'])}</td>
            <td><span class="badge {action_cls}">{esc(log['action'])}</span></td>
            <td style="font-size:12px">{esc(log['target_type'] or '—')} {(': ' + esc(log['target_id'])) if log['target_id'] else ''}</td>
            <td style="font-size:12px;color:var(--text3);font-family:monospace">{esc(log['ip_address'] or '—')}</td>
            <td style="font-size:11px;color:var(--text3);max-width:250px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{esc(log['details'] or '')}</td>
        </tr>"""

    content = f"""<div class="container">
        <div class="page-header"><h1>📋 Auditoria Admin</h1><p>Histórico de todas as ações realizadas no painel</p></div>
        <div class="grid-3" style="margin-bottom:24px">
            <div class="stat-card"><div class="stat-icon stat-icon-blue">📋</div><div class="stat-value">{total}</div><div class="stat-label">Total de ações</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-green">📅</div><div class="stat-value">{recent}</div><div class="stat-label">Últimos 7 dias</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-orange">🕐</div><div class="stat-value">500</div><div class="stat-label">Exibindo últimas</div></div>
        </div>
        <div class="card">
            <div class="card-header"><span class="card-title">Histórico detalhado</span></div>
            <div class="table-wrap">
                <table>
                    <thead><tr><th>Quando</th><th>Ação</th><th>Alvo</th><th>IP</th><th>Detalhes</th></tr></thead>
                    <tbody>{rows or '<tr><td colspan="5" style="text-align:center;padding:40px;color:var(--text3)">Nenhuma ação registrada ainda</td></tr>'}</tbody>
                </table>
            </div>
        </div>
    </div>"""
    return admin_html("Auditoria Admin", content)


# ─── ADMIN: ERROS DE WEBHOOK ──────────────────────────────────
@app.route("/admin/webhook-errors")
@admin_required
def admin_webhook_errors():
    """Dashboard de erros de webhook para diagnosticar falhas"""
    db = get_db()
    errors = db.execute(
        """SELECT we.*, u.email FROM webhook_errors we
           LEFT JOIN users u ON we.user_id=u.id
           ORDER BY we.created_at DESC LIMIT 200"""
    ).fetchall()

    unresolved = db.execute(
        "SELECT COUNT(*) as c FROM webhook_errors WHERE resolved=0"
    ).fetchone()["c"]
    total = db.execute("SELECT COUNT(*) as c FROM webhook_errors").fetchone()["c"]

    rows = ""
    for e in errors:
        source_cls = {
            "whatsapp": "badge-green",
            "instagram": "badge-purple",
            "messenger": "badge-blue",
            "mercadopago": "badge-orange"
        }.get(e["source"], "badge-gray")
        resolved_label = '<span class="badge badge-green">Resolvido</span>' if e["resolved"] else '<span class="badge badge-red">Aberto</span>'
        rows += f"""<tr>
            <td style="font-size:12px;color:var(--text3)">{to_br_datetime(e['created_at'])}</td>
            <td><span class="badge {source_cls}">{esc(e['source'])}</span></td>
            <td><strong>{esc(e['error_type'] or '—')}</strong></td>
            <td style="font-size:12px;max-width:400px;color:var(--text2)">{esc((e['error_message'] or '')[:200])}</td>
            <td style="font-size:11px;color:var(--text3)">{esc(e['email'] or (f'user_{e["user_id"]}' if e['user_id'] else '—'))}</td>
            <td>{resolved_label}</td>
            <td>
                {'<form method="POST" action="/admin/webhook-errors/' + str(e['id']) + '/resolve" style="display:inline">' + csrf_field() + '<button class="btn btn-sm btn-success">Resolver</button></form>' if not e['resolved'] else ''}
            </td>
        </tr>"""

    content = f"""<div class="container">
        <div class="page-header">
            <h1>🚨 Erros de Webhook</h1>
            <p>Falhas detectadas nos webhooks para revisão</p>
        </div>

        <div class="grid-3" style="margin-bottom:24px">
            <div class="stat-card"><div class="stat-icon stat-icon-red">🚨</div><div class="stat-value">{unresolved}</div><div class="stat-label">Erros abertos</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-blue">📋</div><div class="stat-value">{total}</div><div class="stat-label">Total de erros</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-green">✅</div><div class="stat-value">{total - unresolved}</div><div class="stat-label">Resolvidos</div></div>
        </div>

        <div class="card">
            <div class="card-header"><span class="card-title">Últimos 200 erros</span></div>
            <div class="table-wrap">
                <table>
                    <thead><tr><th>Quando</th><th>Fonte</th><th>Tipo</th><th>Mensagem</th><th>Usuário</th><th>Status</th><th>Ação</th></tr></thead>
                    <tbody>{rows or '<tr><td colspan="7" style="text-align:center;color:var(--text3);padding:40px">✅ Nenhum erro registrado. Tudo funcionando!</td></tr>'}</tbody>
                </table>
            </div>
        </div>
    </div>"""
    return admin_html("Erros Webhook", content)


@app.route("/admin/webhook-errors/<int:eid>/resolve", methods=["POST"])
@admin_required
def admin_resolve_webhook_error(eid):
    db = get_db()
    db.execute("UPDATE webhook_errors SET resolved=1 WHERE id=?", (eid,))
    db.commit()
    return redirect("/admin/webhook-errors")


# ─── ADMIN: PAGAMENTOS ────────────────────────────────────────
@app.route("/admin/payments")
@admin_required
def admin_payments():
    db = get_db()
    payments = db.execute("""SELECT p.*, u.name, u.email FROM payments p 
        JOIN users u ON p.user_id=u.id ORDER BY p.created_at DESC LIMIT 100""").fetchall()
    
    total_approved = db.execute("SELECT COALESCE(SUM(amount),0) as s FROM payments WHERE status='approved'").fetchone()["s"]
    total_pending = db.execute("SELECT COALESCE(SUM(amount),0) as s FROM payments WHERE status='pending'").fetchone()["s"]
    count_approved = db.execute("SELECT COUNT(*) as c FROM payments WHERE status='approved'").fetchone()["c"]

    rows = ""
    for p in payments:
        p_date = (p['created_at'] or '')[:16]
        p_plan = PLANS.get(p['plan'], {}).get('name', p['plan'])
        p_cls = 'badge-green' if p['status']=='approved' else 'badge-orange' if p['status']=='pending' else 'badge-red'
        p_label = {"approved":"Aprovado","pending":"Pendente","rejected":"Rejeitado"}.get(p['status'], p['status'])
        rows += f"""<tr><td>{p_date}</td><td><strong>{esc(p['name'])}</strong><br><span style="color:var(--text3);font-size:12px">{esc(p['email'])}</span></td>
        <td><span class="badge badge-purple">{p_plan}</span></td><td>R$ {p['amount']:.2f}</td>
        <td><span class="badge {p_cls}">{p_label}</span></td>
        <td style="font-size:12px;color:var(--text3)">{esc(p['mp_payment_id'] or '—')}</td></tr>"""

    content = f"""<div class="container"><div class="page-header"><h1>Pagamentos 💰</h1><p>Histórico de todas as transações</p></div>
        <div class="grid-4" style="margin-bottom:32px">
            <div class="stat-card"><div class="stat-icon stat-icon-green">✅</div><div class="stat-value">R$ {total_approved:.0f}</div><div class="stat-label">Total recebido</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-orange">⏳</div><div class="stat-value">R$ {total_pending:.0f}</div><div class="stat-label">Pendente</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-blue">🧾</div><div class="stat-value">{count_approved}</div><div class="stat-label">Pagamentos aprovados</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-purple">💳</div><div class="stat-value">{len(payments)}</div><div class="stat-label">Total transações</div></div>
        </div>
        <div class="card"><div class="card-header"><span class="card-title">Todas as transações</span></div>
        <div class="table-wrap"><table><thead><tr><th>Data</th><th>Cliente</th><th>Plano</th><th>Valor</th><th>Status</th><th>ID MP</th></tr></thead>
        <tbody>{rows}</tbody></table></div></div></div>"""
    return admin_html("Pagamentos", content)


# ─── ADMIN: USO DE API ────────────────────────────────────────
@app.route("/admin/usage")
@admin_required
def admin_usage():
    db = get_db()
    # Per-user usage
    usage = db.execute("""SELECT u.name, u.email, u.plan, u.msgs_used, u.msgs_limit,
        COALESCE(SUM(a.tokens_in),0) as total_tokens_in, COALESCE(SUM(a.tokens_out),0) as total_tokens_out,
        COALESCE(SUM(a.cost_estimate),0) as total_cost, COUNT(a.id) as api_calls
        FROM users u LEFT JOIN api_usage_log a ON u.id = a.user_id
        GROUP BY u.id ORDER BY total_cost DESC""").fetchall()

    rows = ""
    for u in usage:
        plan_name = PLANS.get(u['plan'], {}).get('name', u['plan'])
        plan_price = PLANS.get(u['plan'], {}).get('price', 0)
        is_healthy = plan_price > u['total_cost'] * 5.5
        health_color = 'var(--green2)' if is_healthy else 'var(--red)'
        health_label = 'Saudável' if is_healthy else 'Atenção'
        rows += f"""<tr><td><strong>{esc(u['name'])}</strong><br><span style="color:var(--text3);font-size:12px">{esc(u['email'])}</span></td>
        <td><span class="badge badge-purple">{plan_name}</span></td>
        <td>{u['msgs_used']}/{u['msgs_limit']}</td><td>{u['api_calls']}</td>
        <td>{u['total_tokens_in']:,}</td><td>{u['total_tokens_out']:,}</td>
        <td><strong>US$ {u['total_cost']:.4f}</strong></td>
        <td style="color:{health_color}">{health_label}</td></tr>"""

    total_cost = sum(u['total_cost'] for u in usage)
    total_calls = sum(u['api_calls'] for u in usage)

    content = f"""<div class="container"><div class="page-header"><h1>Uso de API 📊</h1><p>Monitoramento de custos por cliente</p></div>
        <div class="grid-4" style="margin-bottom:32px">
            <div class="stat-card"><div class="stat-icon stat-icon-red">💸</div><div class="stat-value">US$ {total_cost:.2f}</div><div class="stat-label">Custo total de API</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-blue">🔄</div><div class="stat-value">{total_calls}</div><div class="stat-label">Chamadas de API</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-purple">📉</div><div class="stat-value">US$ {(total_cost/max(len(usage),1)):.3f}</div><div class="stat-label">Custo médio/cliente</div></div>
            <div class="stat-card"><div class="stat-icon stat-icon-green">💰</div><div class="stat-value">US$ {(total_cost/max(total_calls,1)):.4f}</div><div class="stat-label">Custo médio/chamada</div></div>
        </div>
        <div class="card"><div class="card-header"><span class="card-title">Uso por cliente</span></div>
        <div class="table-wrap"><table><thead><tr><th>Cliente</th><th>Plano</th><th>Msgs</th><th>Chamadas API</th><th>Tokens in</th><th>Tokens out</th><th>Custo</th><th>Saúde</th></tr></thead>
        <tbody>{rows}</tbody></table></div></div></div>"""
    return admin_html("Uso de API", content)


# ─── ADMIN: LOGS ──────────────────────────────────────────────
@app.route("/admin/logs")
@admin_required
def admin_logs():
    db = get_db()
    recent_users = db.execute("SELECT name,email,created_at,plan FROM users ORDER BY created_at DESC LIMIT 20").fetchall()
    recent_payments = db.execute("SELECT p.*,u.name FROM payments p JOIN users u ON p.user_id=u.id ORDER BY p.created_at DESC LIMIT 20").fetchall()
    
    user_rows = ""
    for u in recent_users:
        pn = PLANS.get(u["plan"], {}).get("name", u["plan"])
        user_rows += f'<tr><td style="color:var(--text3);font-size:12px">{(u["created_at"] or "")[:16]}</td><td>👤 Novo cadastro</td><td><strong>{esc(u["name"])}</strong> ({esc(u["email"])}) — Plano {esc(pn)}</td></tr>'
    pay_rows = "".join(f'<tr><td style="color:var(--text3);font-size:12px">{(p["created_at"] or "")[:16]}</td><td>💳 Pagamento</td><td><strong>{esc(p["name"])}</strong> — R$ {p["amount"]:.2f} ({p["status"]})</td></tr>' for p in recent_payments)

    content = f"""<div class="container"><div class="page-header"><h1>Logs do Sistema 📋</h1><p>Atividade recente</p></div>
        <div class="grid-2">
            <div class="card"><div class="card-header"><span class="card-title">Últimos cadastros</span></div>
            <div class="table-wrap"><table><thead><tr><th>Data</th><th>Evento</th><th>Detalhes</th></tr></thead><tbody>{user_rows}</tbody></table></div></div>
            <div class="card"><div class="card-header"><span class="card-title">Últimos pagamentos</span></div>
            <div class="table-wrap"><table><thead><tr><th>Data</th><th>Evento</th><th>Detalhes</th></tr></thead><tbody>{pay_rows}</tbody></table></div></div>
        </div></div>"""
    return admin_html("Logs", content)


# ─── ADMIN: BACKUPS E MANUTENÇÃO (LGPD) ─────────────────────────
@app.route("/admin/emergency-backup-download", methods=["GET", "POST"])
@admin_required
def admin_emergency_backup_download():
    """Baixa cópia consistente do atendeia.db pelo navegador, CRIPTOGRAFADA com passphrase.

    Fluxo:
    1. GET → mostra formulário pedindo passphrase (mínimo 12 caracteres).
    2. POST → gera backup com sqlite3.backup() (consistente mesmo com writers ativos),
       criptografa com Fernet usando chave derivada da passphrase via PBKDF2-SHA256 (600k iters),
       prefixa 16 bytes de salt no arquivo final, envia como download e LIMPA o tempfile
       depois que a resposta foi entregue.

    Formato do .enc:
        [16 bytes de salt][token Fernet (base64 ASCII)]

    Para decriptar offline (Python ≥3.8 com `cryptography`):
        from cryptography.fernet import Fernet
        import hashlib, base64
        blob = open("atendeia-backup.db.enc", "rb").read()
        salt, ct = blob[:16], blob[16:]
        key = base64.urlsafe_b64encode(
            hashlib.pbkdf2_hmac("sha256", PASSPHRASE.encode(), salt, 600000)[:32]
        )
        open("atendeia-backup.db", "wb").write(Fernet(key).decrypt(ct))
    """
    import tempfile
    from flask import send_file, after_this_request

    if request.method == "GET":
        content = f"""
        <div class="container" style="max-width:680px;padding:32px">
            <h1 style="font-size:22px;margin-bottom:8px">🔒 Backup de Emergência (criptografado)</h1>
            <p style="color:var(--text2);font-size:14px;margin-bottom:20px">
                Gera uma cópia consistente do banco e <strong>criptografa antes do download</strong>.
                A passphrase NÃO é armazenada — guarde-a em local seguro, pois sem ela o backup é inútil.
            </p>
            <div class="alert alert-warning" style="margin-bottom:20px">
                ⚠️ <strong>Importante:</strong> use uma passphrase forte —
                <strong>16+ caracteres</strong> OU <strong>12+ caracteres com pelo menos 3 classes</strong>
                (minúscula, MAIÚSCULA, dígito, símbolo). Se perder a passphrase,
                <strong>o backup não pode ser restaurado</strong>.
            </div>
            <div class="card" style="padding:24px">
                <form method="POST">{csrf_field()}
                    <div class="form-group">
                        <label class="form-label">Passphrase de criptografia (mín. 12 caracteres)</label>
                        <input type="password" name="passphrase" class="form-input" required minlength="12" autofocus
                               autocomplete="new-password" placeholder="••••••••••••">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Confirme a passphrase</label>
                        <input type="password" name="passphrase_confirm" class="form-input" required minlength="12"
                               autocomplete="new-password" placeholder="••••••••••••">
                    </div>
                    <button type="submit" class="btn btn-primary btn-block">🔐 Gerar e baixar backup criptografado</button>
                </form>
            </div>
            <div class="card" style="padding:20px;margin-top:16px;background:var(--bg2)">
                <h3 style="font-size:14px;margin-bottom:8px">Como restaurar</h3>
                <pre style="font-size:11px;background:rgba(0,0,0,0.3);padding:12px;border-radius:6px;overflow-x:auto;color:#9ca3af">from cryptography.fernet import Fernet
import hashlib, base64
blob = open("atendeia-backup.db.enc", "rb").read()
salt, ct = blob[:16], blob[16:]
key = base64.urlsafe_b64encode(
    hashlib.pbkdf2_hmac("sha256", PASSPHRASE.encode(), salt, 600000)[:32]
)
open("atendeia-backup.db", "wb").write(Fernet(key).decrypt(ct))</pre>
            </div>
        </div>
        """
        return admin_html("Backup de Emergência", content)

    # POST: gera e criptografa
    passphrase = request.form.get("passphrase", "")
    passphrase_confirm = request.form.get("passphrase_confirm", "")

    # Validação de força: 16+ chars OU 12+ chars com mistura de classes
    # (minúscula, MAIÚSCULA, dígito, símbolo — exige pelo menos 3 das 4 classes).
    def _passphrase_strong(p):
        if not p:
            return False, "Passphrase obrigatória."
        if len(p) >= 16:
            return True, ""
        if len(p) < 12:
            return False, "Passphrase muito curta — mínimo 12 caracteres (ou 16 sem mistura de classes)."
        classes = 0
        if any(c.islower() for c in p): classes += 1
        if any(c.isupper() for c in p): classes += 1
        if any(c.isdigit() for c in p): classes += 1
        if any(not c.isalnum() for c in p): classes += 1
        if classes < 3:
            return False, "Passphrase fraca — use 16+ caracteres OU 12+ com pelo menos 3 destas classes: minúscula, MAIÚSCULA, dígito, símbolo."
        return True, ""

    ok, err = _passphrase_strong(passphrase)
    if not ok:
        return err, 400
    if passphrase != passphrase_confirm:
        return "Passphrases não conferem.", 400

    tmp_path = None
    try:
        from cryptography.fernet import Fernet
        import base64 as b64_mod, hashlib as hashlib_mod

        tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tmp_path = tmp.name
        tmp.close()
        src = sqlite3.connect(DATABASE)
        dst = sqlite3.connect(tmp_path)
        with dst:
            src.backup(dst)
        src.close()
        dst.close()

        with open(tmp_path, "rb") as f:
            db_bytes = f.read()

        # Apaga o .db em claro o quanto antes
        try:
            os.remove(tmp_path)
            tmp_path = None
        except OSError:
            pass

        salt = secrets.token_bytes(16)
        derived = hashlib_mod.pbkdf2_hmac("sha256", passphrase.encode(), salt, 600000)[:32]
        fernet_key = b64_mod.urlsafe_b64encode(derived)
        ct = Fernet(fernet_key).encrypt(db_bytes)
        encrypted_blob = salt + ct

        # Escreve o blob criptografado em outro tempfile só para servir
        enc_tmp = tempfile.NamedTemporaryFile(suffix=".db.enc", delete=False)
        enc_tmp.write(encrypted_blob)
        enc_tmp.close()
        enc_path = enc_tmp.name

        @after_this_request
        def _cleanup_encrypted(response):
            try:
                os.remove(enc_path)
            except OSError:
                pass
            return response

        try:
            log_admin_action("emergency_backup_download", details=f"size_plain={len(db_bytes)}B size_enc={len(encrypted_blob)}B encrypted=1")
        except Exception:
            pass
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(
            enc_path,
            as_attachment=True,
            download_name=f"atendeia-backup-{timestamp}.db.enc",
            mimetype="application/octet-stream",
        )
    except Exception as e:
        err_id = secrets.token_hex(6)
        safe_log(f"[EMERGENCY_BACKUP] Erro id={err_id}: {e}", level="ERROR")
        try:
            log_admin_action("emergency_backup_error", details=f"err_id={err_id}")
        except Exception:
            pass
        return f"Falha ao gerar backup. Consulte os logs (id={err_id}).", 500
    finally:
        # Garante remoção do .db em claro mesmo em erro
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


@app.route("/admin/recrypt-status", methods=["GET","POST"])
@admin_required
def admin_recrypt_status():
    """Inspecionar estado da migração SECRET_KEY → DATA_ENCRYPTION_KEY e rodar manualmente."""
    msg = ""
    msg_class = "alert-success"
    result = None

    if request.method == "POST":
        action = request.form.get("action", "")
        if action == "run_recrypt":
            result = migrate_recrypt_to_new_data_key()
            if result and result.get("skipped"):
                msg = f"⏭️ Migração pulada: {result.get('reason', '')}"
                msg_class = "alert-info"
            elif result and "error" in result:
                msg = f"❌ Erro: {result.get('error')}"
                msg_class = "alert-error"
            elif result:
                tot_migrated = result.get("users_migrated", 0) + result.get("settings_migrated", 0)
                tot_failed = result.get("users_failed", 0) + result.get("settings_failed", 0)
                if tot_migrated:
                    msg = f"✅ Migrados: {result.get('users_migrated',0)} campos users + {result.get('settings_migrated',0)} settings"
                elif tot_failed:
                    msg = f"⚠️ {tot_failed} valor(es) falharam na migração — verifique logs"
                    msg_class = "alert-error"
                else:
                    msg = "✅ Tudo já está na chave nova, nada a migrar"
                    msg_class = "alert-info"

    data_key_set = bool(os.getenv("DATA_ENCRYPTION_KEY", "").strip())
    secret_key_set = bool(os.getenv("SECRET_KEY", "").strip())
    keys_equal = (os.getenv("DATA_ENCRYPTION_KEY", "").strip() ==
                  os.getenv("SECRET_KEY", "").strip()) if data_key_set and secret_key_set else False
    current_source = getattr(_get_fernet, "_source", "?")
    _get_fernet()  # garante populado
    current_source = getattr(_get_fernet, "_source", "?")

    alert = f'<div class="alert {msg_class}">{msg}</div>' if msg else ""

    # Detalhes do resultado
    result_html = ""
    if result and not result.get("skipped"):
        result_html = f"""
        <div class="card" style="margin-top:20px;padding:20px">
            <h3 style="font-size:16px;margin-bottom:12px">Resultado da última execução</h3>
            <table style="width:100%;font-size:14px;line-height:1.8">
                <tr><td>Campos users migrados</td><td style="text-align:right">{result.get('users_migrated', 0)}</td></tr>
                <tr><td>Campos users já na chave nova</td><td style="text-align:right">{result.get('users_already_new', 0)}</td></tr>
                <tr><td>Campos users com falha</td><td style="text-align:right;color:var(--red)">{result.get('users_failed', 0)}</td></tr>
                <tr><td>Settings migrados</td><td style="text-align:right">{result.get('settings_migrated', 0)}</td></tr>
                <tr><td>Settings já na chave nova</td><td style="text-align:right">{result.get('settings_already_new', 0)}</td></tr>
                <tr><td>Settings com falha</td><td style="text-align:right;color:var(--red)">{result.get('settings_failed', 0)}</td></tr>
            </table>
            {f'<p style="margin-top:14px;color:var(--red);font-size:13px">Falhas: {esc(", ".join(result.get("failures", [])[:10]))}</p>' if result.get("failures") else ""}
        </div>
        """

    explain = """
    <div class="card" style="padding:20px;margin-bottom:20px;background:rgba(99,102,241,0.05)">
        <h3 style="font-size:16px;margin-bottom:10px">💡 Como funciona</h3>
        <p style="font-size:13px;line-height:1.7;color:var(--text2)">
            A migração só faz alguma coisa se você tiver setado <code>DATA_ENCRYPTION_KEY</code>
            no Railway com valor diferente da <code>SECRET_KEY</code>. Ela passa em todos os
            tokens criptografados (em <code>users</code> e <code>system_settings</code>) e:
        </p>
        <ol style="margin-top:8px;color:var(--text2);font-size:13px;line-height:1.7">
            <li>Tenta decifrar com a chave nova — se já funciona, pula.</li>
            <li>Senão, tenta decifrar com a chave antiga (SECRET_KEY) — se funciona, recriptografa com a nova e salva.</li>
            <li>Se falhar com as duas, NÃO sobrescreve (dado pode estar corrompido) — só conta como falha e loga.</li>
        </ol>
        <p style="font-size:13px;line-height:1.7;color:var(--text2);margin-top:8px">
            🛡️ <strong>Idempotente:</strong> pode rodar várias vezes sem causar dano. Rodada automaticamente no startup.
        </p>
    </div>
    """

    content = f"""<div class="container">
        <div class="page-header"><h1>Recriptografia 🔐</h1><p>Migração SECRET_KEY → DATA_ENCRYPTION_KEY</p></div>
        {alert}
        {explain}

        <div class="card" style="padding:20px;margin-bottom:20px">
            <h3 style="font-size:16px;margin-bottom:12px">Estado atual das chaves</h3>
            <table style="width:100%;font-size:14px;line-height:1.9">
                <tr><td>SECRET_KEY definida no Railway</td><td style="text-align:right">{"✅ Sim" if secret_key_set else "❌ Não"}</td></tr>
                <tr><td>DATA_ENCRYPTION_KEY definida</td><td style="text-align:right">{"✅ Sim" if data_key_set else "❌ Não (usando SECRET_KEY)"}</td></tr>
                <tr><td>Chaves idênticas</td><td style="text-align:right">{"⚠️ Sim — sem benefício de separar" if keys_equal else "✅ Não (corretamente separadas)" if data_key_set else "N/A"}</td></tr>
                <tr><td>Chave em uso agora</td><td style="text-align:right"><strong>{esc(current_source)}</strong></td></tr>
            </table>
        </div>

        <div class="card" style="padding:20px">
            <h3 style="font-size:16px;margin-bottom:12px">Executar migração manualmente</h3>
            <p style="font-size:13px;color:var(--text2);margin-bottom:14px">
                A migração também roda automaticamente no startup. Use isto para verificar
                ou para forçar uma nova varredura.
            </p>
            <form method="POST">{csrf_field()}
                <input type="hidden" name="action" value="run_recrypt">
                <button type="submit" class="btn btn-primary">🔐 Rodar migração agora</button>
            </form>
        </div>

        {result_html}
    </div>"""
    return admin_html("Recriptografia", content)


@app.route("/admin/backups", methods=["GET","POST"])
@admin_required
def admin_backups():
    """Painel de backups: listar, executar agora, ver detalhes de integridade."""
    msg = ""
    msg_class = "alert-success"

    # Ações
    if request.method == "POST":
        action = request.form.get("action", "")
        if action == "run_backup":
            result = perform_database_backup()
            if result and result.get("ok"):
                msg = f"✅ Backup criado: {result['size_mb']}MB, criptografado: {'sim' if result['encrypted'] else 'não'}, removidos antigos: {result['removed_old']}"
                log_admin_action("backup_manual", details=f"size={result['size_mb']}MB")
            else:
                msg = "❌ Falha ao executar backup. Veja os logs do Railway."
                msg_class = "alert-error"
        elif action == "run_retention":
            result = perform_retention_cleanup()
            if result and result.get("ok"):
                msg = f"✅ Retenção: {result['messages_deleted']} msgs, {result['conversations_deleted']} convs, {result['audit_log_deleted']} audit, {result['verification_codes_deleted']} códigos expirados"
                log_admin_action("retention_manual")
            else:
                msg = "❌ Falha ao executar retenção. Veja os logs do Railway."
                msg_class = "alert-error"

    # Listar backups existentes
    backups_info = []
    try:
        if os.path.isdir(BACKUP_DIR):
            entries = sorted(os.listdir(BACKUP_DIR), reverse=True)
            for entry in entries:
                if entry.endswith(".meta.json"):
                    continue
                full = os.path.join(BACKUP_DIR, entry)
                if not os.path.isfile(full):
                    continue
                size_bytes = os.path.getsize(full)
                mtime = datetime.fromtimestamp(os.path.getmtime(full))
                meta = {}
                meta_path = full + ".meta.json"
                if os.path.isfile(meta_path):
                    try:
                        with open(meta_path) as f:
                            meta = json.load(f)
                    except Exception:
                        meta = {}
                backups_info.append({
                    "name": entry,
                    "size_mb": round(size_bytes / (1024 * 1024), 2),
                    "created_at": mtime.strftime("%Y-%m-%d %H:%M:%S"),
                    "encrypted": meta.get("encrypted", entry.endswith(".enc")),
                    "sha256": (meta.get("sha256_unencrypted", "") or "")[:16] + "..." if meta.get("sha256_unencrypted") else "—",
                })
    except Exception as e:
        msg = f"Erro listando backups: {e}"
        msg_class = "alert-error"

    backup_rows = ""
    if backups_info:
        for b in backups_info:
            enc_badge = '<span class="badge badge-green">🔒 Criptografado</span>' if b["encrypted"] else '<span class="badge badge-orange">⚠️ Não criptografado</span>'
            backup_rows += f"""<tr>
                <td style="font-family:monospace;font-size:12px">{esc(b['name'])}</td>
                <td>{b['size_mb']} MB</td>
                <td>{b['created_at']}</td>
                <td>{enc_badge}</td>
                <td style="font-family:monospace;font-size:11px;color:var(--text3)">{esc(b['sha256'])}</td>
            </tr>"""
    else:
        backup_rows = '<tr><td colspan="5" style="text-align:center;color:var(--text2);padding:30px">Nenhum backup ainda. Execute o primeiro abaixo.</td></tr>'

    alert = f'<div class="alert {msg_class}">{msg}</div>' if msg else ""

    # Detecta dinamicamente qual chave está sendo usada para criptografia
    _get_fernet()  # garante que _source foi populado
    crypto_key_source = getattr(_get_fernet, "_source", "desconhecida")

    # Info sobre próxima execução
    next_hour = int(os.getenv("MAINTENANCE_HOUR_UTC", "6"))
    now_utc = datetime.utcnow()
    next_run = now_utc.replace(hour=next_hour, minute=0, second=0, microsecond=0)
    if next_run <= now_utc:
        next_run = next_run + timedelta(days=1)
    next_run_brt = next_run - timedelta(hours=3)

    retention_table = """
    <table style="width:100%;font-size:14px;line-height:2">
        <tr><td>Plano <strong>Starter</strong></td><td style="text-align:right">90 dias</td></tr>
        <tr><td>Plano <strong>Profissional</strong></td><td style="text-align:right">180 dias</td></tr>
        <tr><td>Plano <strong>Business</strong></td><td style="text-align:right">365 dias</td></tr>
        <tr><td>Plano <strong>Agência</strong></td><td style="text-align:right">730 dias</td></tr>
    </table>"""

    content = f"""<div class="container">
        <div class="page-header"><h1>Backups e Manutenção 💾</h1><p>Backup diário automatizado + retenção de dados LGPD</p></div>
        {alert}

        <div class="grid-2" style="margin-bottom:20px">
            <div class="card">
                <div class="card-header"><span class="card-title">⏰ Próxima execução automática</span></div>
                <div style="padding:20px">
                    <p style="font-size:24px;font-weight:700;margin-bottom:4px">{next_run_brt.strftime('%H:%M')} BRT</p>
                    <p style="color:var(--text2);font-size:13px">{next_run.strftime('%Y-%m-%d %H:%M')} UTC</p>
                    <p style="color:var(--text2);font-size:13px;margin-top:8px">Backup + retenção rodam juntos diariamente.</p>
                </div>
            </div>
            <div class="card">
                <div class="card-header"><span class="card-title">🗑️ Retenção por plano</span></div>
                <div style="padding:20px">{retention_table}</div>
            </div>
        </div>

        <div class="card" style="margin-bottom:20px">
            <div class="card-header">
                <span class="card-title">🎯 Executar manualmente</span>
            </div>
            <div style="padding:20px;display:flex;gap:12px;flex-wrap:wrap">
                <form method="POST" style="margin:0">{csrf_field()}
                    <input type="hidden" name="action" value="run_backup">
                    <button type="submit" class="btn btn-primary">📦 Rodar backup agora</button>
                </form>
                <form method="POST" style="margin:0"
                      onsubmit="return confirm('Vai APAGAR mensagens antigas conforme a retenção do plano de cada usuário. Continuar?')">{csrf_field()}
                    <input type="hidden" name="action" value="run_retention">
                    <button type="submit" class="btn btn-secondary">🧹 Rodar limpeza de retenção agora</button>
                </form>
            </div>
        </div>

        <div class="card">
            <div class="card-header">
                <span class="card-title">Backups existentes ({len(backups_info)})</span>
                <span style="font-size:13px;color:var(--text2)">Rotação: últimos {BACKUP_RETENTION_DAYS} dias</span>
            </div>
            <div class="table-wrap">
                <table>
                    <thead><tr>
                        <th>Arquivo</th>
                        <th>Tamanho</th>
                        <th>Criado em</th>
                        <th>Segurança</th>
                        <th>SHA-256</th>
                    </tr></thead>
                    <tbody>{backup_rows}</tbody>
                </table>
            </div>
            <div style="padding:14px 20px;color:var(--text2);font-size:12px;border-top:1px solid var(--border)">
                📁 Pasta no servidor: <code style="background:var(--bg2);padding:2px 6px;border-radius:4px">{esc(BACKUP_DIR)}</code><br>
                🔐 Criptografia: Fernet (AES-128 + HMAC-SHA256). Chave em uso: <strong>{esc(crypto_key_source)}</strong>.<br>
                ✅ Integridade: cada backup tem SHA-256 calculado antes da criptografia.
            </div>
        </div>
    </div>"""
    return admin_html("Backups", content)


# ─── ADMIN: CONFIGURAÇÕES DE API ───────────────────────────────
@app.route("/admin/api-settings", methods=["GET", "POST"])
@admin_required
def admin_api_settings():
    msg = ""
    if request.method == "POST":
        keys = ["ANTHROPIC_API_KEY", "GROQ_API_KEY", "OPENAI_API_KEY", "MERCADOPAGO_ACCESS_TOKEN", "MP_WEBHOOK_SECRET", "WHATSAPP_VERIFY_TOKEN", "WHATSAPP_APP_SECRET"]
        for key in keys:
            value = request.form.get(key, "").strip()
            if value:
                set_setting(key, value)
        base_url = request.form.get("BASE_URL", "").strip()
        if base_url:
            set_setting("BASE_URL", base_url)
        smtp_email = request.form.get("SMTP_EMAIL", "").strip()
        smtp_password = request.form.get("SMTP_PASSWORD", "").strip()
        smtp_host = request.form.get("SMTP_HOST", "").strip()
        smtp_port = request.form.get("SMTP_PORT", "").strip()
        if smtp_email: set_setting("SMTP_EMAIL", smtp_email)
        if smtp_password: set_setting("SMTP_PASSWORD", smtp_password)
        if smtp_host: set_setting("SMTP_HOST", smtp_host)
        if smtp_port: set_setting("SMTP_PORT", smtp_port)
        ai_engine_val = request.form.get("AI_ENGINE", "").strip()
        if ai_engine_val: set_setting("AI_ENGINE", ai_engine_val)
        resend_key = request.form.get("RESEND_API_KEY", "").strip()
        resend_from = request.form.get("RESEND_FROM_EMAIL", "").strip()
        if resend_key: set_setting("RESEND_API_KEY", resend_key)
        if resend_from: set_setting("RESEND_FROM_EMAIL", resend_from)
        msg = '<div class="alert alert-success">Configurações de API salvas!</div>'

    anthropic_key = get_setting("ANTHROPIC_API_KEY")
    groq_key = get_setting("GROQ_API_KEY")
    openai_key = get_setting("OPENAI_API_KEY")
    mp_token = get_setting("MERCADOPAGO_ACCESS_TOKEN")
    mp_webhook_secret = get_setting("MP_WEBHOOK_SECRET", "")
    wa_verify = get_setting("WHATSAPP_VERIFY_TOKEN", "meu_token_verificacao")
    wa_app_secret = get_setting("WHATSAPP_APP_SECRET", "")
    base_url = get_setting("BASE_URL", "http://localhost:8080")
    smtp_email = get_setting("SMTP_EMAIL", "contato@atendente.online")
    smtp_password = get_setting("SMTP_PASSWORD")
    smtp_host = get_setting("SMTP_HOST", "smtp.hostinger.com")
    smtp_port = get_setting("SMTP_PORT", "465")
    ai_engine = get_setting("AI_ENGINE", "claude")
    resend_key = get_setting("RESEND_API_KEY")
    resend_from = get_setting("RESEND_FROM_EMAIL", "atendente.online <onboarding@resend.dev>")

    def mask(key):
        if not key: return ""
        return key[:8] + "..." + key[-4:] if len(key) > 16 else key[:4] + "..."

    content = f"""<div class="container">
        <div class="page-header fade-in"><h1>Configurações de API 🔑</h1><p>Configure todas as chaves de API do sistema</p></div>
        {msg}
        <form method="POST">{csrf_field()}
        <div class="grid-2">
            <div class="card fade-in fade-in-1">
                <div class="card-header"><span class="card-title">IA e Transcrição</span></div>
                <div class="form-group">
                    <label class="form-label">Motor de IA</label>
                    <select name="AI_ENGINE" class="form-input" style="background:#2a2a3a;border:2px solid var(--accent)">
                        <option value="claude" {'selected' if ai_engine == 'claude' else ''}>Claude (Anthropic) — Recomendado</option>
                        <option value="openai" {'selected' if ai_engine == 'openai' else ''}>ChatGPT (OpenAI)</option>
                    </select>
                    <small style="color:var(--text3)">Motor atual: <strong style="color:var(--accent2)">{'Claude' if ai_engine == 'claude' else 'ChatGPT'}</strong></small>
                </div>
                <div class="form-group">
                    <label class="form-label">Anthropic API Key (Claude)</label>
                    <input type="text" name="ANTHROPIC_API_KEY" class="form-input" placeholder="sk-ant-..." value="" autocomplete="off"
                        style="background:#2a2a3a;border:2px solid {'var(--green)' if anthropic_key else 'var(--red)'}">
                    <small style="color:var(--text3)">{'✅ Configurada: ' + mask(anthropic_key) if anthropic_key else '❌ Não configurada'}</small>
                </div>
                <div class="form-group">
                    <label class="form-label">Groq API Key (Transcrição de áudio)</label>
                    <input type="text" name="GROQ_API_KEY" class="form-input" placeholder="gsk_..." value="" autocomplete="off"
                        style="background:#2a2a3a;border:2px solid {'var(--green)' if groq_key else 'var(--red)'}">
                    <small style="color:var(--text3)">{'✅ Configurada: ' + mask(groq_key) if groq_key else '❌ Não configurada — áudios não serão transcritos'}</small>
                </div>
                <div class="form-group">
                    <label class="form-label">OpenAI API Key (ChatGPT + fallback áudio)</label>
                    <input type="text" name="OPENAI_API_KEY" class="form-input" placeholder="sk-..." value="" autocomplete="off"
                        style="background:#2a2a3a;border:2px solid {'var(--green)' if openai_key else 'var(--orange)'}">
                    <small style="color:var(--text3)">{'✅ Configurada: ' + mask(openai_key) if openai_key else '⬜ Configure para usar ChatGPT como motor de IA'}</small>
                </div>
            </div>
            <div class="card fade-in fade-in-2">
                <div class="card-header"><span class="card-title">Pagamentos e WhatsApp</span></div>
                <div class="form-group">
                    <label class="form-label">Mercado Pago Access Token</label>
                    <input type="text" name="MERCADOPAGO_ACCESS_TOKEN" class="form-input" placeholder="APP_USR-..." value="" autocomplete="off"
                        style="background:#2a2a3a;border:2px solid {'var(--green)' if mp_token and mp_token != 'TEST-xxxx' else 'var(--orange)'}">
                    <small style="color:var(--text3)">{'✅ Configurado: ' + mask(mp_token) if mp_token and mp_token != 'TEST-xxxx' else '⚠️ Não configurado — checkout simulado'}</small>
                </div>
                <div class="form-group">
                    <label class="form-label">🔐 MP Webhook Secret (assinatura)</label>
                    <input type="password" name="MP_WEBHOOK_SECRET" class="form-input" placeholder="{'••••••••' if mp_webhook_secret else 'Chave secreta para validar webhooks'}" value="" autocomplete="off"
                        style="background:#2a2a3a;border:2px solid {'var(--green)' if mp_webhook_secret else 'var(--orange)'}">
                    <small style="color:var(--text3)">{'✅ Configurado — webhooks de pagamento serão validados' if mp_webhook_secret else '⚠️ Configure em MP > Webhooks > Secret signature (previne fraude de pagamento)'}</small>
                </div>
                <div class="form-group">
                    <label class="form-label">WhatsApp Verify Token</label>
                    <input type="text" name="WHATSAPP_VERIFY_TOKEN" class="form-input" value="{esc(wa_verify)}" autocomplete="off"
                        style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.08)">
                    <small style="color:var(--text3)">Token usado na verificação do webhook do Meta</small>
                </div>
                <div class="form-group">
                    <label class="form-label">WhatsApp App Secret (segurança)</label>
                    <input type="password" name="WHATSAPP_APP_SECRET" class="form-input" placeholder="{'••••••••' if wa_app_secret else 'Meta App Secret'}" autocomplete="off"
                        style="background:#2a2a3a;border:2px solid {'var(--green)' if wa_app_secret else 'var(--orange)'}">
                    <small style="color:var(--text3)">{'✅ Configurado — assinaturas do webhook serão validadas' if wa_app_secret else '⚠️ Configure para validar assinatura do Meta (previne webhooks falsos)'}</small>
                </div>
                <div class="form-group">
                    <label class="form-label">URL Base do Sistema</label>
                    <input type="text" name="BASE_URL" class="form-input" value="{esc(base_url)}" autocomplete="off"
                        style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.08)">
                    <small style="color:var(--text3)">URL pública (ex: https://seudominio.com)</small>
                </div>
            </div>
        </div>
        <button type="submit" class="btn btn-primary btn-lg">Salvar todas as configurações</button>
        </form>

        <div class="card fade-in fade-in-3" style="margin-top:32px">
            <div class="card-header"><span class="card-title">Email (Resend) — Verificação de conta</span></div>
            <form method="POST">{csrf_field()}
            <div class="grid-2">
                <div class="form-group">
                    <label class="form-label">Resend API Key</label>
                    <input type="text" name="RESEND_API_KEY" class="form-input" placeholder="re_..." value="" autocomplete="off"
                        style="background:#2a2a3a;border:2px solid {'var(--green)' if resend_key else 'var(--red)'}">
                    <small style="color:var(--text3)">{'✅ Configurada: ' + mask(resend_key) if resend_key else '❌ Não configurada — emails não serão enviados'}</small>
                </div>
                <div class="form-group">
                    <label class="form-label">Email remetente</label>
                    <input type="text" name="RESEND_FROM_EMAIL" class="form-input" value="{resend_from}" autocomplete="off"
                        style="background:#2a2a3a;border:1px solid rgba(255,255,255,0.08)">
                    <small style="color:var(--text3)">Use "onboarding@resend.dev" (grátis) ou configure domínio no Resend</small>
                </div>
            </div>
            <button type="submit" class="btn btn-primary">Salvar configurações de email</button>
            </form>
        </div>

        <div class="card fade-in fade-in-4" style="margin-top:32px">
            <div class="card-header"><span class="card-title">Onde conseguir as chaves</span></div>
            <div style="color:var(--text2);font-size:14px;line-height:2">
                <p><strong style="color:var(--text)">Anthropic (Claude):</strong> <a href="https://console.anthropic.com/settings/keys" target="_blank">console.anthropic.com/settings/keys</a></p>
                <p><strong style="color:var(--text)">Groq (Áudio):</strong> <a href="https://console.groq.com/keys" target="_blank">console.groq.com/keys</a> — praticamente grátis!</p>
                <p><strong style="color:var(--text)">OpenAI (Fallback):</strong> <a href="https://platform.openai.com/api-keys" target="_blank">platform.openai.com/api-keys</a></p>
                <p><strong style="color:var(--text)">Mercado Pago:</strong> <a href="https://www.mercadopago.com.br/developers/panel/app" target="_blank">mercadopago.com.br/developers/panel/app</a></p>
            </div>
        </div>
    </div>"""
    return admin_html("Configurações de API", content)


# ─── ADMIN: EXPORTAR DADOS ────────────────────────────────────
@app.route("/admin/export/<string:data_type>")
@admin_required
def admin_export(data_type):
    db = get_db()
    if data_type == "users":
        rows = db.execute("SELECT id,name,email,company,phone,plan,plan_status,msgs_used,msgs_limit,created_at,last_login FROM users").fetchall()
        csv = "id,nome,email,empresa,telefone,plano,status,msgs_usadas,msgs_limite,cadastro,ultimo_login\n"
        csv += "\n".join(",".join(csv_safe(r[k]) for k in r.keys()) for r in rows)
    elif data_type == "payments":
        rows = db.execute("SELECT p.id,u.name,u.email,p.amount,p.status,p.plan,p.mp_payment_id,p.created_at FROM payments p JOIN users u ON p.user_id=u.id").fetchall()
        csv = "id,cliente,email,valor,status,plano,mp_id,data\n"
        csv += "\n".join(",".join(csv_safe(r[k]) for k in r.keys()) for r in rows)
    else:
        return "Tipo inválido", 400

    output = io.BytesIO(("\ufeff" + csv).encode("utf-8"))
    output.seek(0)
    return send_file(output, mimetype="text/csv", as_attachment=True, download_name=f"atendeia_{data_type}_{datetime.now().strftime('%Y%m%d')}.csv")


def migrate_recrypt_to_new_data_key():
    """⚙️ MIGRAÇÃO CRÍTICA: recriptografa dados existentes quando DATA_ENCRYPTION_KEY
    é ativada pela primeira vez (chave nova ≠ chave antiga SECRET_KEY).

    Estratégia segura:
    1. Constrói Fernet_OLD (chave SECRET_KEY) e Fernet_NEW (chave DATA_ENCRYPTION_KEY)
    2. Se as chaves derivadas forem idênticas (caso em que DATA_ENCRYPTION_KEY == SECRET_KEY),
       não há nada a fazer — retorna.
    3. Para cada valor com prefixo `fer:v1:`:
       a. Tenta decifrar com chave NOVA. Se OK, já está migrado, pula.
       b. Tenta decifrar com chave ANTIGA. Se OK, recriptografa com NOVA e salva.
       c. Se falhar com as DUAS, NÃO sobrescreve (dado pode estar corrompido) — só conta como falha.

    Idempotente: rodar múltiplas vezes é seguro. Só re-processa o que ainda está com chave antiga.

    Cobre: users (USER_ENCRYPTED_FIELDS) e system_settings (SENSITIVE_SETTINGS)."""

    data_key = os.getenv("DATA_ENCRYPTION_KEY", "").strip()
    if not data_key:
        # DATA_ENCRYPTION_KEY não setada → nada a migrar (sistema usa SECRET_KEY como antes)
        return {"skipped": True, "reason": "DATA_ENCRYPTION_KEY não definida"}

    secret_key = os.getenv("SECRET_KEY", "").strip()
    if not secret_key:
        # Sem SECRET_KEY antiga não dá pra decifrar nada existente. Mas se o sistema sempre
        # funcionou só com DATA_ENCRYPTION_KEY (instalação nova), não há legado pra migrar.
        return {"skipped": True, "reason": "SECRET_KEY não definida — nada a migrar"}

    if data_key == secret_key:
        return {"skipped": True, "reason": "DATA_ENCRYPTION_KEY == SECRET_KEY, sem migração necessária"}

    fernet_old = _build_fernet_from_key(secret_key)
    fernet_new = _build_fernet_from_key(data_key)

    if fernet_old is None or fernet_new is None:
        safe_log("[RECRYPT] Não foi possível construir Fernet antigo/novo — abortando", level="ERROR")
        return {"skipped": True, "reason": "Falha ao construir Fernet"}

    # Helper interno: tenta decifrar valor com fernet específico
    def _try_decrypt(fernet, encrypted_str):
        if not isinstance(encrypted_str, str) or not encrypted_str.startswith("fer:v1:"):
            return None
        try:
            return fernet.decrypt(encrypted_str[7:].encode("ascii")).decode("utf-8")
        except Exception:
            return None

    def _re_encrypt(plaintext):
        """Re-criptografa com a chave NOVA explicitamente (não usa _encrypt_value pra não
        depender do cache de _get_fernet)."""
        try:
            token = fernet_new.encrypt(plaintext.encode("utf-8"))
            return f"fer:v1:{token.decode('ascii')}"
        except Exception:
            return None

    results = {
        "users_migrated": 0,
        "users_already_new": 0,
        "users_failed": 0,
        "settings_migrated": 0,
        "settings_already_new": 0,
        "settings_failed": 0,
        "failures": [],
    }

    try:
        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row

        # ─── (1) USERS: campos criptografados ───
        try:
            cols = list(USER_ENCRYPTED_FIELDS)
            rows = db_conn.execute(f"SELECT id, {', '.join(cols)} FROM users").fetchall()
            for row in rows:
                uid = row["id"]
                updates = {}
                for col in cols:
                    val = row[col]
                    if not val or not isinstance(val, str) or not val.startswith("fer:v1:"):
                        continue
                    # Tenta chave nova primeiro
                    plain_new = _try_decrypt(fernet_new, val)
                    if plain_new is not None:
                        results["users_already_new"] += 1
                        continue
                    # Tenta chave antiga
                    plain_old = _try_decrypt(fernet_old, val)
                    if plain_old is None:
                        results["users_failed"] += 1
                        results["failures"].append(f"users.id={uid} col={col}")
                        safe_log(f"[RECRYPT] Falha em users.id={uid} col={col} (corrompido?)", level="ERROR")
                        continue
                    new_token = _re_encrypt(plain_old)
                    if new_token:
                        updates[col] = new_token
                if updates:
                    set_clause = ", ".join(f"{c}=?" for c in updates.keys())
                    params = list(updates.values()) + [uid]
                    db_conn.execute(f"UPDATE users SET {set_clause} WHERE id=?", params)
                    results["users_migrated"] += len(updates)
        except Exception as e:
            safe_log(f"[RECRYPT] Erro processando users: {e}", level="ERROR")

        # ─── (2) SYSTEM_SETTINGS: chaves sensíveis ───
        try:
            sensitive_keys = list(SENSITIVE_SETTINGS) if 'SENSITIVE_SETTINGS' in globals() else []
            for key in sensitive_keys:
                r = db_conn.execute("SELECT value FROM system_settings WHERE key=?", (key,)).fetchone()
                if not r or not r["value"]:
                    continue
                val = r["value"]
                if not isinstance(val, str) or not val.startswith("fer:v1:"):
                    # Texto puro ou XOR legado: deixa pra migrate_encrypt_existing_secrets() cuidar
                    continue
                plain_new = _try_decrypt(fernet_new, val)
                if plain_new is not None:
                    results["settings_already_new"] += 1
                    continue
                plain_old = _try_decrypt(fernet_old, val)
                if plain_old is None:
                    results["settings_failed"] += 1
                    results["failures"].append(f"system_settings.key={key}")
                    safe_log(f"[RECRYPT] Falha em system_settings.{key}", level="ERROR")
                    continue
                new_token = _re_encrypt(plain_old)
                if new_token:
                    db_conn.execute("UPDATE system_settings SET value=? WHERE key=?", (new_token, key))
                    results["settings_migrated"] += 1
        except Exception as e:
            safe_log(f"[RECRYPT] Erro processando system_settings: {e}", level="ERROR")

        db_conn.commit()
        db_conn.close()

        total_migrated = results["users_migrated"] + results["settings_migrated"]
        total_failed = results["users_failed"] + results["settings_failed"]

        if total_migrated:
            safe_log(f"[RECRYPT] ✅ Migração concluída: {results['users_migrated']} campo(s) users, "
                     f"{results['settings_migrated']} setting(s). Já-na-nova: "
                     f"{results['users_already_new']}+{results['settings_already_new']}.")
        elif total_failed:
            safe_log(f"[RECRYPT] ⚠️ {total_failed} valor(es) não puderam ser migrados. Verifique logs.", level="ERROR")
        else:
            safe_log(f"[RECRYPT] Nada a migrar. Tudo já está na chave nova.")

        return results
    except Exception as e:
        safe_log(f"[RECRYPT] Erro fatal na migração: {e}", level="ERROR")
        return {"error": str(e)}


def migrate_encrypt_existing_secrets():
    """Migra segredos: texto puro → Fernet, formato antigo XOR → Fernet"""
    fernet = _get_fernet()
    if not fernet:
        is_dev = os.getenv("FLASK_ENV", "").lower() == "development"
        if not is_dev:
            safe_log("[MIGRATION] ❌ Criptografia indisponível — migração abortada", level="ERROR")
            return
        safe_log("[MIGRATION] ⚠️ DEV: criptografia indisponível, migração pulada", level="WARN")
        return
    try:
        db_conn = sqlite3.connect(DATABASE)
        db_conn.row_factory = sqlite3.Row
        migrated = 0
        for key in SENSITIVE_SETTINGS:
            row = db_conn.execute("SELECT value FROM system_settings WHERE key=?", (key,)).fetchone()
            if not row or not row["value"]:
                continue
            current = row["value"]
            # Já está em formato Fernet novo, pula
            if current.startswith("fer:v1:"):
                continue
            # Está no formato antigo XOR, descriptografa e re-criptografa com Fernet
            if current.startswith("enc:v1:"):
                plaintext = _decrypt_legacy(current)
                if plaintext:
                    encrypted = _encrypt_value(plaintext)
                    if encrypted.startswith("fer:v1:"):
                        db_conn.execute("UPDATE system_settings SET value=? WHERE key=?", (encrypted, key))
                        safe_log(f"[MIGRATION] Re-criptografado (XOR → Fernet): {key}")
                        migrated += 1
                continue
            # Texto puro → criptografa com Fernet
            encrypted = _encrypt_value(current)
            if encrypted.startswith("fer:v1:"):
                db_conn.execute("UPDATE system_settings SET value=? WHERE key=?", (encrypted, key))
                safe_log(f"[MIGRATION] Criptografado (plaintext → Fernet): {key}")
                migrated += 1
        db_conn.commit()
        db_conn.close()
        if migrated:
            safe_log(f"[MIGRATION] ✅ {migrated} segredo(s) migrado(s) para Fernet")
    except Exception as e:
        safe_log(f"[MIGRATION] Erro: {e}", level="ERROR")


def check_production_requirements():
    """Verifica requisitos obrigatórios de produção. Aborta se algo crítico faltar.

    Secrets de webhook (MP/WhatsApp) também são exigidos por padrão, com escape hatches:
    - SKIP_MP_REQUIREMENT=1 → pula MP_WEBHOOK_SECRET (use se não usa MP)
    - SKIP_WA_REQUIREMENT=1 → pula WHATSAPP_APP_SECRET (use se não usa WhatsApp)
    """
    is_dev = os.getenv("FLASK_ENV", "").lower() == "development"
    errors = []

    # 1. cryptography deve estar instalada em produção
    try:
        from cryptography.fernet import Fernet
    except ImportError:
        errors.append("cryptography>=42.0 não instalada — segredos não podem ser criptografados")

    # 2. SECRET_KEY deve ser fixo (não aleatório) em produção
    if not os.getenv("SECRET_KEY"):
        errors.append("SECRET_KEY não configurado — sessões e segredos criptografados ficam inválidos após restart")

    # 3. ADMIN_PASSWORD deve estar configurado
    if not os.getenv("ADMIN_PASSWORD"):
        errors.append("ADMIN_PASSWORD não configurado — painel admin inacessível")

    # 4. requests é obrigatório para webhooks, mídia, IA e pagamentos
    try:
        import requests  # noqa: F401
    except ImportError:
        errors.append("requests não instalado — integrações externas não funcionam")

    # 5. Secrets de webhook — fail-fast em produção em vez de só rejeitar no request.
    # Aceita tanto env var quanto settings do banco (NOTA: settings só consulta após init_db,
    # então em produção o RECOMENDADO é configurar via env var; aqui só verificamos env).
    skip_mp = os.getenv("SKIP_MP_REQUIREMENT", "").strip() == "1"
    if not skip_mp and not os.getenv("MP_WEBHOOK_SECRET"):
        errors.append("MP_WEBHOOK_SECRET não configurado — webhooks Mercado Pago serão rejeitados (defina SKIP_MP_REQUIREMENT=1 se não usa MP)")

    skip_wa = os.getenv("SKIP_WA_REQUIREMENT", "").strip() == "1"
    if not skip_wa and not os.getenv("WHATSAPP_APP_SECRET"):
        errors.append("WHATSAPP_APP_SECRET não configurado — webhooks WhatsApp serão rejeitados (defina SKIP_WA_REQUIREMENT=1 se não usa WhatsApp)")

    # 6. DATA_ENCRYPTION_KEY separada — recomendação forte para SaaS multi-cliente.
    # Sem ela, o sistema cai em SECRET_KEY (que também assina cookies de sessão); se a
    # SECRET_KEY vazar, vaza simultaneamente sessões E dados criptografados em repouso.
    # Não aborta para não quebrar deploys existentes — emite warning evidente.
    # Para tornar obrigatório, defina REQUIRE_SEPARATE_DATA_KEY=1.
    require_data_key = os.getenv("REQUIRE_SEPARATE_DATA_KEY", "").strip() == "1"
    if not os.getenv("DATA_ENCRYPTION_KEY"):
        msg = ("DATA_ENCRYPTION_KEY não configurada — fallback para SECRET_KEY. "
               "Para SaaS multi-cliente, recomenda-se chave separada (gere com: "
               "python3 -c 'import secrets; print(secrets.token_hex(32))'). "
               "Defina REQUIRE_SEPARATE_DATA_KEY=1 para tornar obrigatório.")
        if require_data_key:
            errors.append(msg)
        elif not is_dev:
            safe_log(f"[BOOT] ⚠️ {msg}", level="WARN")

    if errors and not is_dev:
        print("="*70)
        print("  🚨 ERRO CRÍTICO: requisitos de produção não atendidos")
        print("="*70)
        for e in errors:
            print(f"  ❌ {e}")
        print("="*70)
        print("  Para resolver:")
        print("  1. Adicione 'cryptography>=42.0' em requirements.txt")
        print("  2. No Railway → Variables, defina:")
        print("     SECRET_KEY    (gere com: python3 -c 'import secrets; print(secrets.token_hex(32))')")
        print("     ADMIN_PASSWORD (sua senha de admin)")
        print()
        print("  Para rodar em modo desenvolvimento, defina FLASK_ENV=development")
        print("="*70)
        raise SystemExit(1)
    elif errors and is_dev:
        print("[DEV] Avisos de produção (ignorados em FLASK_ENV=development):")
        for e in errors:
            print(f"  ⚠️ {e}")


# ═══════════════════════════════════════════════════════════════
#  INIT & RUN
# ═══════════════════════════════════════════════════════════════

# Verifica requisitos ANTES de iniciar o banco
check_production_requirements()

# Inicializa o banco sempre (necessário para gunicorn no Railway)
init_db()
# IMPORTANTE: rodar PRIMEIRO a recriptografia (caso DATA_ENCRYPTION_KEY tenha sido ativada
# enquanto havia dados antigos criptografados com SECRET_KEY).
# Esta chamada é idempotente e segura — não faz nada se DATA_ENCRYPTION_KEY não estiver setada
# ou se já estiver tudo migrado.
try:
    migrate_recrypt_to_new_data_key()
except Exception as e:
    safe_log(f"[RECRYPT] Migração falhou: {e}", level="ERROR")
migrate_encrypt_existing_secrets()
migrate_encrypt_user_tokens()

# Guard multi-worker: em produção (Railway com gunicorn) cada worker importa o módulo
# e dispararia o scheduler — rodando backup/retenção N vezes. Em DEV (FLASK_ENV=development)
# ou single-worker, sempre roda. Em produção, exige SCHEDULER_LEADER=1 num único worker.
# Defina SCHEDULER_LEADER=1 apenas em UM worker (ex.: usando --preload + condição no gunicorn
# config, ou rodando um worker dedicado com essa env var).
def _is_scheduler_leader():
    if os.getenv("FLASK_ENV", "").strip() == "development":
        return True
    if os.getenv("SCHEDULER_LEADER", "").strip() == "1":
        return True
    # Heurística: se WEB_CONCURRENCY/GUNICORN_WORKERS não definidos ou == 1, é single-worker
    workers = os.getenv("WEB_CONCURRENCY", os.getenv("GUNICORN_WORKERS", "1")).strip()
    return workers == "1"


if _is_scheduler_leader():
    # Inicia scheduler de posts da agência (em background)
    try:
        start_social_scheduler()
    except Exception as e:
        safe_log(f"[SCHEDULER] Não foi possível iniciar: {e}", level="ERROR")

    # LGPD: Inicia scheduler diário de backup + retenção (em background)
    try:
        start_daily_maintenance_scheduler()
        safe_log("[MAINTENANCE] Scheduler diário ativo (backup + retenção LGPD)")
    except Exception as e:
        safe_log(f"[MAINTENANCE] Não foi possível iniciar: {e}", level="ERROR")
else:
    safe_log("[SCHEDULER] Worker não-líder — schedulers não iniciados (defina SCHEDULER_LEADER=1 em um worker)")


# ─── FAVICON (logo na aba do navegador) ───────────────────
@app.route("/favicon.ico")
def favicon():
    """Serve o logo como favicon para aparecer na aba do navegador"""
    import base64 as b64
    from flask import Response
    try:
        png_bytes = b64.b64decode(LOGO_NAV_B64)
        return Response(png_bytes, mimetype="image/png", headers={
            "Cache-Control": "public, max-age=86400"
        })
    except Exception as e:
        safe_log(f"[FAVICON] Erro: {e}", level="ERROR")
        return Response(status=404)


@app.route("/apple-touch-icon.png")
@app.route("/apple-touch-icon-precomposed.png")
def apple_touch_icon():
    """Ícone para iOS quando salvar como atalho na tela inicial"""
    import base64 as b64
    from flask import Response
    try:
        png_bytes = b64.b64decode(LOGO_NAV_B64)
        return Response(png_bytes, mimetype="image/png", headers={
            "Cache-Control": "public, max-age=86400"
        })
    except Exception as e:
        safe_log(f"[APPLE ICON] Erro: {e}", level="ERROR")
        return Response(status=404)


@app.route("/icon-192.png")
def icon_192():
    """Ícone PWA"""
    import base64 as b64
    from flask import Response
    try:
        png_bytes = b64.b64decode(LOGO_NAV_B64)
        return Response(png_bytes, mimetype="image/png", headers={
            "Cache-Control": "public, max-age=86400"
        })
    except:
        return Response(status=404)


if __name__ == "__main__":
    port = int(os.getenv("PORT", 8080))
    print("\n" + "="*60)
    print("  ⚡ atendente.online — Sistema rodando!")
    print(f"  📍 Painel cliente: http://localhost:{port}")
    print(f"  🔐 Painel admin:   http://localhost:{port}/admin/login")
    print(f"  👤 Admin email:    {ADMIN_EMAIL}")
    print(f"  🔑 Admin senha:    {'✅ Configurada' if ADMIN_PASSWORD else '❌ NÃO CONFIGURADA — defina ADMIN_PASSWORD'}")
    print("  ─────────────────────────────────────")
    print(f"  🤖 Claude (IA):    {'✅ Configurada' if get_setting('ANTHROPIC_API_KEY') else '❌ Configure no admin → APIs'}")
    print(f"  🎤 Groq (Áudio):   {'✅ Configurada' if get_setting('GROQ_API_KEY') else '❌ Configure no admin → APIs'}")
    print(f"  🎤 OpenAI:         {'✅ Configurada' if get_setting('OPENAI_API_KEY') else '⬜ Opcional'}")
    print("="*60 + "\n")
    app.run(debug=os.getenv("FLASK_ENV")=="development", host="0.0.0.0", port=port)
